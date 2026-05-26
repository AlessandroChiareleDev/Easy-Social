import openpyxl
import os

file_path = r"C:\Users\xandao\Downloads\RELATORIO_SOLUCOES_AGOSTO_2025_ERROS_S1210 resposta erro 456.xlsx"
sheet_name = "1210"

if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    exit(1)

wb = openpyxl.load_workbook(file_path, data_only=True)
if sheet_name not in wb.sheetnames:
    print(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
    exit(1)

ws = wb[sheet_name]

# Dimensions
print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")

# Header
header = [cell.value for cell in ws[1]]
print(f"Header: {header}")

# Identify columns of interest
keywords = ["recibo", "id", "cpf", "status", "erro", "per", "data", "dt"]
cols_of_interest = []
cpf_col_idx = -1

for idx, col_name in enumerate(header):
    if col_name:
        col_name_lower = str(col_name).lower()
        if any(kw in col_name_lower for kw in keywords):
            cols_of_interest.append((idx, col_name))
        if "cpf" in col_name_lower:
            cpf_col_idx = idx

print(f"Columns of interest: {[c[1] for c in cols_of_interest]}")

# Targeted CPFs
target_cpfs = ["00391802984", "00001931946", "391802984", "1931946"]

def format_cpf(val):
    if val is None: return ""
    return str(val).strip().zfill(11)

for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
    row_cpf = format_cpf(row[cpf_col_idx]) if cpf_col_idx != -1 else ""
    
    match = False
    if row_cpf in target_cpfs:
        match = True
    else:
        # Fallback check in all columns if cpf_col_idx failed or to be sure
        for val in row:
            if val and any(t in str(val) for t in target_cpfs):
                match = True
                break
    
    if match:
        print(f"\n--- Row {row_idx} ---")
        for idx, val in enumerate(row):
            if val is not None and str(val).strip() != "":
                col_name = header[idx] if idx < len(header) else f"Col_{idx+1}"
                print(f"{col_name}: {val}")
