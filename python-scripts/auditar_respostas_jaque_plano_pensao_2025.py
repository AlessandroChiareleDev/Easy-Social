from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
ORIG_DIR = ROOT / "relatorio_ana" / "JAQUE_PLANO_PENSAO_2025_FEV_OUT"
RESP_DIR = ROOT / "relatorio_ana" / "RESPOSTAS_JAQUE_PLANO_PENSAO_2025_FEV_OUT"
ERRORS_CSV = ROOT / "relatorio_ana" / "ERROS_S1210_SOLUCOES_GERAL" / "erros_classificados.csv"
OUT_DIR = ROOT / "relatorio_ana" / "AUDITORIA_RESPOSTAS_JAQUE_PLANO_PENSAO_2025"

MONTHS = {
    "FEVEREIRO": "2025-02",
    "MARCO": "2025-03",
    "MARCO": "2025-03",
    "ABRIL": "2025-04",
    "MAIO": "2025-05",
    "JUNHO": "2025-06",
    "JULHO": "2025-07",
    "AGOSTO": "2025-08",
    "SETEMBRO": "2025-09",
    "OUTUBRO": "2025-10",
}

PLAN_CATEGORY = "plano de saude coletivo ausente"
PENSION_CATEGORY = "pensao alimenticia - beneficiarios ausentes"


def norm_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text)


def only_digits(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\D", "", str(value))


def normalize_cpf(value: Any) -> str:
    digits = only_digits(value)
    if not digits:
        return ""
    if len(digits) > 11:
        digits = digits[-11:]
    return digits.zfill(11)


def parse_number(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value))
        except InvalidOperation:
            return None
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("R$", "").replace(" ", "")
    text = re.sub(r"[^0-9,.-]", "", text)
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    elif text.count(".") > 1:
        parts = text.split(".")
        text = "".join(parts[:-1]) + "." + parts[-1]
    try:
        return Decimal(text)
    except InvalidOperation:
        return None


def month_from_name(path: Path) -> str | None:
    name = norm_text(path.stem).upper()
    for token, per_apur in MONTHS.items():
        if token in name:
            return per_apur
    return None


def category_to_type(category: str) -> str | None:
    normalized = norm_text(category)
    if PLAN_CATEGORY in normalized:
        return "PLANO"
    if PENSION_CATEGORY in normalized:
        return "PENSAO"
    return None


def load_target_universe() -> dict[tuple[str, str, str], dict[str, Any]]:
    targets: dict[tuple[str, str, str], dict[str, Any]] = {}
    with ERRORS_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            tipo = category_to_type(row.get("categoria", ""))
            if not tipo:
                continue
            per_apur = str(row.get("per_apur") or "").strip()
            cpf = normalize_cpf(row.get("cpf"))
            if not per_apur or not cpf:
                continue
            targets[(per_apur, cpf, tipo)] = {**row, "cpf": cpf, "tipo": tipo}
    return targets


def find_header(ws) -> tuple[int, dict[str, int], list[str]] | None:
    for row_index in range(1, min(ws.max_row, 30) + 1):
        values = [ws.cell(row_index, col).value for col in range(1, ws.max_column + 1)]
        headers = [str(value).strip() if value is not None else "" for value in values]
        normalized = [norm_text(header) for header in headers]
        has_cpf = any(header in {"cpf", "cpf normalizado"} for header in normalized)
        has_known = any(
            ("operadora" in header or "ans" in header or "beneficiario" in header or "valor deduzido" in header)
            for header in normalized
        )
        if has_cpf and has_known:
            header_map: dict[str, int] = {}
            for index, header in enumerate(normalized):
                if header and header not in header_map:
                    header_map[header] = index
            return row_index, header_map, headers
    return None


def get_by_contains(row: list[Any], header_map: dict[str, int], *needles: str) -> Any:
    for header, index in header_map.items():
        if all(needle in header for needle in needles):
            return row[index] if index < len(row) else None
    return None


def sheet_type(sheet_name: str, headers: list[str]) -> str | None:
    combined = norm_text(sheet_name + " " + " ".join(headers))
    if "beneficiario" in combined or "pensao" in combined or "valor deduzido" in combined:
        return "PENSAO"
    if "plano" in combined or "operadora" in combined or "registro ans" in combined:
        return "PLANO"
    return None


def row_cpf(row: list[Any], header_map: dict[str, int]) -> str:
    for header in ("cpf normalizado", "cpf"):
        if header in header_map:
            value = row[header_map[header]] if header_map[header] < len(row) else None
            cpf = normalize_cpf(value)
            if cpf:
                return cpf
    return ""


def validate_plan(row: list[Any], header_map: dict[str, int]) -> tuple[bool, str, dict[str, Any]]:
    cnpj = get_by_contains(row, header_map, "cnpj", "operadora")
    ans = get_by_contains(row, header_map, "registro", "ans")
    valor = get_by_contains(row, header_map, "valor", "titular")
    parsed = parse_number(valor)
    data = {"cnpj_operadora": cnpj, "registro_ans": ans, "valor_titular": str(parsed) if parsed is not None else valor}
    missing = []
    if not only_digits(cnpj):
        missing.append("cnpj_operadora")
    if not str(ans or "").strip():
        missing.append("registro_ans")
    if parsed is None:
        missing.append("valor_titular")
    return not missing, ";".join(missing), data


def validate_pension(row: list[Any], header_map: dict[str, int]) -> tuple[bool, str, dict[str, Any]]:
    beneficiaries: list[dict[str, Any]] = []
    for number in range(1, 9):
        cpf_benef = get_by_contains(row, header_map, "cpf beneficiario", str(number))
        tipo = get_by_contains(row, header_map, "tipo rendimento", str(number))
        percentual = get_by_contains(row, header_map, "percentual", str(number))
        valor = get_by_contains(row, header_map, "valor deduzido", str(number))
        cpf_norm = normalize_cpf(cpf_benef)
        parsed_value = parse_number(valor)
        if cpf_norm or str(tipo or "").strip() or parsed_value is not None:
            beneficiaries.append(
                {
                    "cpf_beneficiario": cpf_norm,
                    "tipo_rendimento": str(tipo or "").strip(),
                    "percentual": str(percentual or "").strip(),
                    "valor_deduzido": str(parsed_value) if parsed_value is not None else valor,
                    "valido": bool(cpf_norm and str(tipo or "").strip() and parsed_value is not None),
                }
            )
    valid = [item for item in beneficiaries if item["valido"]]
    if valid:
        return True, "", {"beneficiarios": valid}
    return False, "beneficiario_tipo_ou_valor", {"beneficiarios": beneficiaries}


def scan_workbooks(folder: Path) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    entries: list[dict[str, Any]] = []
    sheet_summary: list[dict[str, Any]] = []
    for path in sorted(folder.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        per_apur = month_from_name(path)
        workbook = load_workbook(path, data_only=True, read_only=True)
        try:
            for ws in workbook.worksheets:
                header = find_header(ws)
                if not header:
                    sheet_summary.append(
                        {"file": path.name, "per_apur": per_apur, "sheet": ws.title, "tipo": None, "linhas": 0, "validas": 0, "obs": "sem_header"}
                    )
                    continue
                header_row, header_map, headers = header
                tipo = sheet_type(ws.title, headers)
                total_rows = 0
                valid_rows = 0
                for values in ws.iter_rows(min_row=header_row + 1, values_only=True):
                    row = list(values)
                    cpf = row_cpf(row, header_map)
                    if not cpf:
                        continue
                    total_rows += 1
                    if tipo == "PLANO":
                        valid, reason, data = validate_plan(row, header_map)
                    elif tipo == "PENSAO":
                        valid, reason, data = validate_pension(row, header_map)
                    else:
                        valid, reason, data = False, "tipo_desconhecido", {}
                    if valid:
                        valid_rows += 1
                    entries.append(
                        {
                            "file": path.name,
                            "per_apur": per_apur,
                            "sheet": ws.title,
                            "tipo": tipo,
                            "cpf": cpf,
                            "valid": valid,
                            "invalid_reason": reason,
                            "prefer_score": (2 if any(token in norm_text(ws.title) for token in ("ok", "resposta", "preenchido")) else 0) + (1 if valid else 0),
                            "data": data,
                        }
                    )
                sheet_summary.append(
                    {
                        "file": path.name,
                        "per_apur": per_apur,
                        "sheet": ws.title,
                        "tipo": tipo,
                        "headers": headers,
                        "linhas": total_rows,
                        "validas": valid_rows,
                    }
                )
        finally:
            workbook.close()
    return entries, sheet_summary


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    targets = load_target_universe()
    response_entries, response_sheets = scan_workbooks(RESP_DIR)
    original_entries, original_sheets = scan_workbooks(ORIG_DIR)

    response_by_key: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    all_response_by_key: dict[tuple[str, str, str], list[dict[str, Any]]] = defaultdict(list)
    for entry in response_entries:
        if not entry.get("per_apur") or not entry.get("tipo"):
            continue
        key = (entry["per_apur"], entry["cpf"], entry["tipo"])
        all_response_by_key[key].append(entry)
        if entry["valid"]:
            response_by_key[key].append(entry)

    chosen_rows: list[dict[str, Any]] = []
    missing_rows: list[dict[str, Any]] = []
    coverage_rows: list[dict[str, Any]] = []
    for key, target in sorted(targets.items()):
        candidates = sorted(response_by_key.get(key, []), key=lambda item: item["prefer_score"], reverse=True)
        any_rows = all_response_by_key.get(key, [])
        covered = bool(candidates)
        chosen = candidates[0] if candidates else None
        coverage_rows.append(
            {
                "per_apur": key[0],
                "cpf": key[1],
                "tipo": key[2],
                "coberto": covered,
                "linhas_resposta": len(any_rows),
                "linhas_validas": len(candidates),
                "arquivo_escolhido": chosen.get("file") if chosen else "",
                "aba_escolhida": chosen.get("sheet") if chosen else "",
            }
        )
        if chosen:
            chosen_rows.append(
                {
                    "per_apur": key[0],
                    "cpf": key[1],
                    "tipo": key[2],
                    "file": chosen["file"],
                    "sheet": chosen["sheet"],
                    "data_json": json.dumps(chosen["data"], ensure_ascii=False, default=str),
                }
            )
        else:
            reasons = Counter(entry.get("invalid_reason") or "sem_linha" for entry in any_rows)
            missing_rows.append(
                {
                    "per_apur": key[0],
                    "cpf": key[1],
                    "tipo": key[2],
                    "linhas_resposta": len(any_rows),
                    "motivos_invalidos": "; ".join(f"{reason}:{count}" for reason, count in reasons.items()),
                    "erro_mensagem": target.get("erro_mensagem", ""),
                }
            )

    target_keys = set(targets)
    extras = [
        {"per_apur": key[0], "cpf": key[1], "tipo": key[2], "validas": len(entries), "arquivos": "; ".join(sorted({item["file"] for item in entries}))}
        for key, entries in sorted(response_by_key.items())
        if key not in target_keys
    ]
    duplicates = [
        {"per_apur": key[0], "cpf": key[1], "tipo": key[2], "validas": len(entries), "arquivos_abas": "; ".join(sorted({f"{item['file']}::{item['sheet']}" for item in entries}))}
        for key, entries in sorted(response_by_key.items())
        if key in target_keys and len(entries) > 1
    ]

    original_keys = {(entry["per_apur"], entry["cpf"], entry["tipo"]) for entry in original_entries if entry.get("per_apur") and entry.get("tipo") and entry.get("cpf")}
    response_any_keys = set(all_response_by_key)
    original_missing_any = sorted(original_keys - response_any_keys)
    original_extra_any = sorted(response_any_keys - original_keys)

    by_month_type: dict[tuple[str, str], dict[str, int]] = defaultdict(lambda: {"total": 0, "coberto": 0, "faltante": 0})
    for row in coverage_rows:
        bucket = by_month_type[(row["per_apur"], row["tipo"])]
        bucket["total"] += 1
        if row["coberto"]:
            bucket["coberto"] += 1
        else:
            bucket["faltante"] += 1
    coverage_summary = [
        {
            "per_apur": per_apur,
            "tipo": tipo,
            "total": values["total"],
            "coberto": values["coberto"],
            "faltante": values["faltante"],
            "pct": round(values["coberto"] / values["total"] * 100, 2) if values["total"] else 0,
        }
        for (per_apur, tipo), values in sorted(by_month_type.items())
    ]

    invalid_reasons = Counter(entry.get("invalid_reason") for entry in response_entries if entry.get("tipo") and not entry.get("valid"))
    total_targets = len(targets)
    total_covered = sum(1 for row in coverage_rows if row["coberto"])
    summary = {
        "total_alvo_atual_plano_pensao": total_targets,
        "total_coberto_valido": total_covered,
        "total_faltante_valido": total_targets - total_covered,
        "cobertura_pct": round(total_covered / total_targets * 100, 2) if total_targets else 0,
        "resolve_100_porcento_alvo_atual": total_covered == total_targets,
        "total_linhas_resposta": len(response_entries),
        "total_linhas_resposta_validas": sum(1 for item in response_entries if item.get("valid")),
        "extras_validos_fora_alvo_atual": len(extras),
        "duplicados_validos_no_alvo_atual": len(duplicates),
        "invalid_reasons": dict(invalid_reasons),
        "originais_total_chaves": len(original_keys),
        "originais_sem_linha_na_resposta": len(original_missing_any),
        "resposta_com_linha_fora_originais": len(original_extra_any),
        "arquivos_resposta": sorted(path.name for path in RESP_DIR.glob("*.xlsx") if not path.name.startswith("~$")),
        "arquivos_originais": sorted(path.name for path in ORIG_DIR.glob("*.xlsx") if not path.name.startswith("~$")),
        "cobertura_por_mes_tipo": coverage_summary,
    }

    (OUT_DIR / "auditoria_final_resumo.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    (OUT_DIR / "estrutura_workbooks_final.json").write_text(
        json.dumps({"originais": original_sheets, "respostas": response_sheets}, ensure_ascii=False, indent=2, default=str),
        encoding="utf-8",
    )
    write_csv(OUT_DIR / "cobertura_final_por_mes_tipo.csv", coverage_summary, ["per_apur", "tipo", "total", "coberto", "faltante", "pct"])
    write_csv(OUT_DIR / "cobertura_final_por_cpf.csv", coverage_rows, ["per_apur", "cpf", "tipo", "coberto", "linhas_resposta", "linhas_validas", "arquivo_escolhido", "aba_escolhida"])
    write_csv(OUT_DIR / "faltantes_final.csv", missing_rows, ["per_apur", "cpf", "tipo", "linhas_resposta", "motivos_invalidos", "erro_mensagem"])
    write_csv(OUT_DIR / "respostas_validas_final.csv", chosen_rows, ["per_apur", "cpf", "tipo", "file", "sheet", "data_json"])
    write_csv(OUT_DIR / "extras_validos_final.csv", extras, ["per_apur", "cpf", "tipo", "validas", "arquivos"])
    write_csv(OUT_DIR / "duplicados_validos_final.csv", duplicates, ["per_apur", "cpf", "tipo", "validas", "arquivos_abas"])
    write_csv(
        OUT_DIR / "originais_sem_resposta_linha_final.csv",
        [{"per_apur": key[0], "cpf": key[1], "tipo": key[2]} for key in original_missing_any],
        ["per_apur", "cpf", "tipo"],
    )

    lines = [
        "# Auditoria final das respostas da Jaque - Plano e Pensao",
        "",
        "## Resultado principal",
        "",
        f"Universo atual de plano + pensao: **{total_targets}** CPF-mes-tipo.",
        f"Cobertura valida encontrada nas respostas: **{total_covered}** de **{total_targets}** (**{summary['cobertura_pct']}%**).",
        f"Resolve 100% do alvo atual? **{'SIM' if summary['resolve_100_porcento_alvo_atual'] else 'NAO'}**.",
        "",
        "## Cobertura por mes e tipo",
        "",
        "| Mes | Tipo | Total | Coberto | Faltante | % |",
        "|---|---|---:|---:|---:|---:|",
    ]
    for row in coverage_summary:
        lines.append(f"| {row['per_apur']} | {row['tipo']} | {row['total']} | {row['coberto']} | {row['faltante']} | {row['pct']}% |")
    lines.extend(
        [
            "",
            "## Observacoes de estrutura",
            "",
            f"Arquivos originais: {len(summary['arquivos_originais'])}. Arquivos de resposta: {len(summary['arquivos_resposta'])}.",
            f"Chaves dos originais sem nenhuma linha correspondente na resposta: {summary['originais_sem_linha_na_resposta']}.",
            f"Linhas validas duplicadas dentro do alvo atual: {summary['duplicados_validos_no_alvo_atual']}.",
            f"Linhas validas fora do alvo atual: {summary['extras_validos_fora_alvo_atual']}.",
            "",
            "## Arquivos gerados",
            "",
            "- auditoria_final_resumo.json",
            "- cobertura_final_por_mes_tipo.csv",
            "- cobertura_final_por_cpf.csv",
            "- faltantes_final.csv",
            "- respostas_validas_final.csv",
            "- extras_validos_final.csv",
            "- duplicados_validos_final.csv",
            "- estrutura_workbooks_final.json",
        ]
    )
    (OUT_DIR / "RELATORIO_AUDITORIA_FINAL_RESPOSTAS_JAQUE.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())