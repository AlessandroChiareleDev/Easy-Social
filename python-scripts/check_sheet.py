import openpyxl

wb = openpyxl.load_workbook(
    r'..\backend\uploads\1774459449290-Relat__rio_DIRF_2025.xlsx',
    read_only=True,
    data_only=True
)
print('Sheets:', wb.sheetnames)

ws = wb['Base Ficha Financeira 2025']
print(f'Min row: {ws.min_row}, Max row: {ws.max_row}')
print(f'Min col: {ws.min_column}, Max col: {ws.max_column}')

rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
print('First 3 rows:')
for r in rows:
    print(r)

wb.close()
