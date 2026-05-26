from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

import psycopg2.extras
from lxml import etree
from openpyxl import load_workbook


ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
BACKEND_V2 = Path(r"C:\Users\xandao\Documents\GitHub\Easy-eSocial-v2\backend")
if str(ROOT / "python-scripts") not in sys.path:
    sys.path.insert(0, str(ROOT / "python-scripts"))
if str(BACKEND_V2) not in sys.path:
    sys.path.insert(0, str(BACKEND_V2))

import enviar_correcao_agosto_jaque as envio_base  # noqa: E402
import reenviar_agosto_sem_mudanca_solucoes as reenvio_base  # noqa: E402
from app import db, esocial_client, tenant  # noqa: E402
from app.xml_diff import eventos_iguais  # noqa: E402
from app.xml_s1210 import _gerar_id  # noqa: E402
from app.xml_signer import S1010XMLSigner  # noqa: E402


EMPRESA_ID = 2
PER_APUR = "2025-08"
CONFIRM_TOKEN = "CORRIGIR_RECIBOS_FALTANTES_18"
OUT_DIR = ROOT / "relatorio_ana" / "REENVIO_AGOSTO_XML_IDENTICO"
PLANILHA = OUT_DIR / "RECIBOS_FALTANTES_18_SOLUCOES_AGOSTO_2025.xlsx"
XML_DIR = OUT_DIR / "xml_unsigned_recibos_faltantes_18"
MANIFEST = OUT_DIR / "manifest_corrigir_recibos_faltantes_18.json"


def normalize_cpf(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits.zfill(11) if digits else ""


def clean_receipt(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace(" ", "")


def receipt_like(value: str) -> bool:
    return bool(re.fullmatch(r"1\.1\.\d{10,}", value or ""))


def load_candidates() -> list[dict[str, Any]]:
    if not PLANILHA.exists():
        raise RuntimeError(f"planilha nao encontrada: {PLANILHA}")
    workbook = load_workbook(PLANILHA, data_only=True, read_only=True)
    sheet = workbook["RECIBOS_18"] if "RECIBOS_18" in workbook.sheetnames else workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise RuntimeError("planilha vazia")
    headers = {str(value).strip(): index for index, value in enumerate(rows[0]) if value is not None}
    required = ["CPF", "Recibo correto", "Recibo testado", "Item erro"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise RuntimeError(f"colunas ausentes na planilha: {missing}")

    candidates: list[dict[str, Any]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        cpf = normalize_cpf(row[headers["CPF"]] if headers["CPF"] < len(row) else None)
        if not cpf:
            continue
        correct_receipt = clean_receipt(row[headers["Recibo correto"]] if headers["Recibo correto"] < len(row) else None)
        tested_receipt = clean_receipt(row[headers["Recibo testado"]] if headers["Recibo testado"] < len(row) else None)
        error_item = row[headers["Item erro"]] if headers["Item erro"] < len(row) else None
        if not receipt_like(correct_receipt):
            raise RuntimeError(f"linha {row_number} CPF {cpf}: recibo correto ausente ou invalido: {correct_receipt!r}")
        if tested_receipt and correct_receipt == tested_receipt:
            raise RuntimeError(f"linha {row_number} CPF {cpf}: recibo correto igual ao recibo que ja falhou")
        try:
            error_item_id = int(str(error_item).strip()) if error_item is not None and str(error_item).strip() else None
        except ValueError as exc:
            raise RuntimeError(f"linha {row_number} CPF {cpf}: Item erro invalido: {error_item!r}") from exc
        candidates.append(
            {
                "cpf": cpf,
                "chosen_receipt": correct_receipt,
                "receipt_used_wrong": tested_receipt,
                "error_item_id": error_item_id,
                "sheet_row": row_number,
            }
        )
    workbook.close()

    if len(candidates) != 18:
        raise RuntimeError(f"esperado 18 CPFs preenchidos; encontrado {len(candidates)}")
    seen: set[str] = set()
    duplicates: list[str] = []
    for candidate in candidates:
        if candidate["cpf"] in seen:
            duplicates.append(candidate["cpf"])
        seen.add(candidate["cpf"])
    if duplicates:
        raise RuntimeError(f"CPFs duplicados na planilha: {duplicates}")
    return sorted(candidates, key=lambda item: item["cpf"])


def current_error_rows(conn, cpfs: list[str]) -> dict[str, dict[str, Any]]:
    internal_empresa_id = tenant.internal_empresa_id(EMPRESA_ID)
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        cursor.execute(
            """
            SELECT DISTINCT ON (it.cpf)
                   it.id AS item_id,
                   it.timeline_envio_id AS envio_id,
                   it.cpf,
                   it.status,
                   it.erro_codigo,
                   it.erro_mensagem,
                   it.nr_recibo_anterior,
                   it.nr_recibo_novo,
                   it.versao_anterior_id,
                   it.xml_enviado_oid,
                   it.criado_em
              FROM timeline_envio_item it
              JOIN timeline_envio te ON te.id = it.timeline_envio_id
              JOIN timeline_mes tm ON tm.id = te.timeline_mes_id
             WHERE tm.empresa_id = %s
               AND tm.per_apur = %s
               AND it.tipo_evento = 'S-1210'
               AND it.cpf = ANY(%s)
             ORDER BY it.cpf, it.criado_em DESC NULLS LAST, it.id DESC
            """,
            (internal_empresa_id, PER_APUR, cpfs),
        )
        rows = {row["cpf"]: dict(row) for row in cursor.fetchall()}
    missing = [cpf for cpf in cpfs if cpf not in rows]
    if missing:
        raise RuntimeError(f"CPFs sem linha atual: {missing}")
    return rows


def assert_current_receipt_error(row: dict[str, Any], candidate: dict[str, Any]) -> None:
    if candidate.get("error_item_id") and int(row["item_id"]) != int(candidate["error_item_id"]):
        raise RuntimeError(
            f"CPF {candidate['cpf']} mudou de item atual: {row['item_id']} != {candidate['error_item_id']}"
        )
    if row.get("status") != "erro_esocial" or str(row.get("erro_codigo")) != "401":
        raise RuntimeError(f"CPF {candidate['cpf']} nao esta no erro 401 atual: {row}")
    message = str(row.get("erro_mensagem") or "")
    if "459" not in message:
        raise RuntimeError(f"CPF {candidate['cpf']} erro atual nao e 459: {message}")
    if candidate.get("receipt_used_wrong") and str(row.get("nr_recibo_anterior") or "") != candidate["receipt_used_wrong"]:
        raise RuntimeError(
            f"CPF {candidate['cpf']} recibo errado mudou: "
            f"{row.get('nr_recibo_anterior')} != {candidate['receipt_used_wrong']}"
        )
    if not row.get("versao_anterior_id"):
        raise RuntimeError(f"CPF {candidate['cpf']} sem versao_anterior_id")
    if not row.get("xml_enviado_oid"):
        raise RuntimeError(f"CPF {candidate['cpf']} sem xml_enviado_oid no erro atual")


def prepare_xml(conn, row: dict[str, Any], candidate: dict[str, Any]) -> tuple[bytes, dict[str, Any]]:
    xml_error = reenvio_base.read_xml_lo(conn, int(row["xml_enviado_oid"]))
    inner = reenvio_base.inner_s1210(xml_error)
    event_node = inner.xpath('./*[local-name()="evtPgtos"]')[0]
    old_id = event_node.get("Id") or ""
    cpf = reenvio_base.only_digits(inner.xpath('string(//*[local-name()="ideBenef"]/*[local-name()="cpfBenef"])'))
    if cpf != candidate["cpf"]:
        raise RuntimeError(f"XML do erro e de CPF {cpf}, esperado {candidate['cpf']}")

    ide_evento = event_node.xpath('./*[local-name()="ideEvento"]')[0]
    old_receipt = ide_evento.xpath('string(./*[local-name()="nrRecibo"])').strip()
    expected_old = str(row.get("nr_recibo_anterior") or "")
    if expected_old and old_receipt != expected_old:
        raise RuntimeError(f"CPF {cpf} XML nao contem recibo atual esperado: {old_receipt} != {expected_old}")

    tp_insc_text = inner.xpath('string(//*[local-name()="ideEmpregador"]/*[local-name()="tpInsc"])').strip()
    nr_insc = inner.xpath('string(//*[local-name()="ideEmpregador"]/*[local-name()="nrInsc"])').strip()
    event_node.set("Id", _gerar_id(int(tp_insc_text), nr_insc))
    reenvio_base.set_child(ide_evento, "indRetif", "2")
    reenvio_base.set_child(ide_evento, "nrRecibo", candidate["chosen_receipt"], after_tag="indRetif")

    if inner.xpath('string(//*[local-name()="ideEvento"]/*[local-name()="tpAmb"])').strip() != "1":
        raise RuntimeError(f"CPF {cpf} nao esta em tpAmb producao")
    if inner.xpath('string(//*[local-name()="ideEvento"]/*[local-name()="perApur"])').strip() != PER_APUR:
        raise RuntimeError(f"CPF {cpf} perApur divergente")

    xml_new = etree.tostring(inner, xml_declaration=True, encoding="UTF-8", pretty_print=False)
    if not eventos_iguais(xml_error, xml_new):
        raise RuntimeError(f"CPF {cpf} corpo canonico mudou")
    return xml_new, {
        "cpf": cpf,
        "source_error_item_id": row["item_id"],
        "source_error_envio_id": row["envio_id"],
        "source_xml_oid": row["xml_enviado_oid"],
        "old_id": old_id,
        "new_id": event_node.get("Id"),
        "wrong_receipt": old_receipt,
        "chosen_receipt": candidate["chosen_receipt"],
        "sheet_row": candidate["sheet_row"],
    }


def generate_manifest() -> dict[str, Any]:
    XML_DIR.mkdir(parents=True, exist_ok=True)
    candidates = load_candidates()
    conn = db.connect(empresa_id=EMPRESA_ID)
    generated: list[dict[str, Any]] = []
    try:
        rows = current_error_rows(conn, [item["cpf"] for item in candidates])
        seen_ids: set[str] = set()
        for candidate in candidates:
            row = rows[candidate["cpf"]]
            assert_current_receipt_error(row, candidate)
            xml_new, meta = prepare_xml(conn, row, candidate)
            if meta["new_id"] in seen_ids:
                raise RuntimeError(f"Id duplicado: {meta['new_id']}")
            seen_ids.add(str(meta["new_id"]))
            xml_path = XML_DIR / f"S1210_{PER_APUR}_{candidate['cpf']}_fix459_unsigned.xml"
            xml_path.write_bytes(xml_new)
            generated.append(
                {
                    "cpf": candidate["cpf"],
                    "xml": str(xml_path),
                    "evento_id": row["versao_anterior_id"],
                    "nr_recibo": candidate["chosen_receipt"],
                    "id_evento": meta["new_id"],
                    "validation": meta,
                }
            )
    finally:
        conn.close()

    manifest = {
        "empresa_id": EMPRESA_ID,
        "per_apur": PER_APUR,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "total": len(generated),
        "regra": "corrigir os 18 recibos faltantes preenchidos na planilha; alterar somente Id/nrRecibo/assinatura",
        "planilha": str(PLANILHA),
        "targets": generated,
    }
    MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return manifest


def sign_targets(targets: list[dict[str, Any]], senha: str) -> list[dict[str, Any]]:
    pfx_data = envio_base.DEFAULT_CERT.read_bytes()
    signed: list[dict[str, Any]] = []
    for target in targets:
        xml_assinado = S1010XMLSigner.assinar(Path(target["xml"]).read_bytes(), pfx_data, senha)
        signed_id = esocial_client._extrair_id(xml_assinado)
        if signed_id != target["id_evento"]:
            raise RuntimeError(f"Id assinado divergente para CPF {target['cpf']}: {signed_id}")
        signed.append({**target, "xml_assinado": xml_assinado, "id_evento_assinado": signed_id})
    return signed


def create_timeline(conn, total: int) -> tuple[int, int]:
    internal_empresa_id = tenant.internal_empresa_id(EMPRESA_ID)
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
        cursor.execute(
            "SELECT id FROM timeline_mes WHERE empresa_id=%s AND per_apur=%s",
            (internal_empresa_id, PER_APUR),
        )
        month = cursor.fetchone()
        if not month:
            raise RuntimeError(f"timeline_mes nao existe para empresa={EMPRESA_ID} per_apur={PER_APUR}")
        month_id = int(month["id"])
        cursor.execute(
            "SELECT COALESCE(MAX(sequencia), 0)+1 AS prox FROM timeline_envio WHERE timeline_mes_id=%s",
            (month_id,),
        )
        sequence = int(cursor.fetchone()["prox"])
        cursor.execute(
            """
            INSERT INTO timeline_envio
              (timeline_mes_id, sequencia, tipo, status,
               iniciado_em, total_tentados, total_sucesso, total_erro, resumo)
            VALUES
              (%s, %s, 'envio_massa', 'em_andamento', now(), %s, 0, 0, %s)
            RETURNING id
            """,
            (
                month_id,
                sequence,
                total,
                psycopg2.extras.Json(
                    {
                        "rotulo": "corrigir_recibos_faltantes_18_solucoes",
                        "empresa_id_externo": EMPRESA_ID,
                        "per_apur": PER_APUR,
                        "ambiente": "producao",
                        "origem": str(MANIFEST),
                        "planilha": str(PLANILHA),
                    }
                ),
            ),
        )
        envio_id = int(cursor.fetchone()["id"])
    conn.commit()
    return envio_id, month_id


def execute() -> dict[str, Any]:
    manifest = generate_manifest()
    targets = manifest["targets"]
    if len(targets) != 18:
        raise RuntimeError(f"esperado enviar 18 correcoes; encontrados {len(targets)}")
    senha = reenvio_base.read_password()
    signed = sign_targets(targets, senha)
    conn_db = db.connect(empresa_id=EMPRESA_ID)
    conn_w = db.connect(empresa_id=EMPRESA_ID)
    try:
        envio_base._verificar_estado_atual(conn_db, signed)
        envio_id, month_id = create_timeline(conn_db, len(signed))
        print(f"=> correcao recibos faltantes 18: envio_id={envio_id} timeline_mes={month_id} targets={len(signed)}")
        item_ids = envio_base._criar_items(conn_db, envio_id, signed)
        envio_base._persistir_xmls_assinados(conn_db, conn_w, signed, item_ids)
        print(f"=> XMLs assinados gravados e vinculados: {len(item_ids)}")

        sucesso_total = 0
        erro_total = 0
        protocolos: list[str] = []
        histograma: dict[str, int] = {}
        for index in range(0, len(signed), envio_base.CFG_LOTE_MAX):
            batch = signed[index:index + envio_base.CFG_LOTE_MAX]
            print(f"\n>> lote {index // envio_base.CFG_LOTE_MAX + 1} ({len(batch)} eventos)")
            result = envio_base._processar_lote(
                batch,
                item_ids,
                cert_path=envio_base.DEFAULT_CERT,
                senha=senha,
                cnpj="09445502000109",
                conn_db=conn_db,
                conn_w=conn_w,
            )
            sucesso_total += int(result["sucesso"])
            erro_total += int(result["erro"])
            if result.get("protocolo"):
                protocolos.append(str(result["protocolo"]))
            for code, count in dict(result.get("histograma") or {}).items():
                histograma[str(code)] = histograma.get(str(code), 0) + int(count)

        envio_base._atualizar_envio(
            conn_db,
            envio_id,
            status="concluido",
            sucesso=sucesso_total,
            erro=erro_total,
            resumo_extra={
                "rotulo_final": "corrigir_recibos_faltantes_18_solucoes",
                "protocolos": protocolos,
                "histograma_erros": histograma,
                "manifest": str(MANIFEST),
                "planilha": str(PLANILHA),
            },
        )
        print("\n=== RESUMO CORRECAO RECIBOS FALTANTES 18 ===")
        print(f"envio_id  : {envio_id}")
        print(f"protocolos: {protocolos}")
        print(f"sucesso   : {sucesso_total}")
        print(f"erro      : {erro_total}")
        print(f"histograma: {histograma}")
        return {
            "envio_id": envio_id,
            "sucesso": sucesso_total,
            "erro": erro_total,
            "protocolos": protocolos,
            "histograma": histograma,
            "manifest": str(MANIFEST),
        }
    finally:
        conn_db.close()
        conn_w.close()


def dry_run() -> dict[str, Any]:
    manifest = generate_manifest()
    return {
        "ok": True,
        "dry_run": True,
        "manifest": str(MANIFEST),
        "total": manifest["total"],
        "primeiros_cpfs": [target["cpf"] for target in manifest["targets"][:10]],
        "ultimos_cpfs": [target["cpf"] for target in manifest["targets"][-10:]],
    }


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--confirmar", default="")
    args = parser.parse_args()
    if not args.execute:
        print(json.dumps(dry_run(), ensure_ascii=False, indent=2, default=str))
        return 0
    if args.confirmar != CONFIRM_TOKEN:
        raise SystemExit(f"Para executar, use --confirmar {CONFIRM_TOKEN}")
    print(json.dumps(execute(), ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())