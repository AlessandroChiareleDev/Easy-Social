"""Import + Envio Lote 2 Julho/2025 APPA.

Etapas:
  1. Parse XLSX (Lote para Envio + Assistência Médica)
  2. Insere s1210_xlsx, s1210_cpf_scope, s1210_operadoras
  3. Envia via Repo API (lotes de 50)

Uso:
  python envio_lote2_julho.py --dry-run   # apenas analisa, sem gravar
  python envio_lote2_julho.py             # import + envio real
"""
from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
import time
from datetime import datetime, timezone
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

import openpyxl
import psycopg2
import psycopg2.extras
import requests

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)
from db_config import DB_CONFIG

# ── Configuração ──────────────────────────────────────────────────────────
XLSX_PATH = Path(r"C:\Users\xandao\Downloads\07_Julho_lote 002_APPA.xlsx")
PER_APUR  = "2025-07"
LOTE_NUM  = 2
EMPRESA_ID = 1
ABA_GERAL  = "Lote para Envio"
ABA_OPER   = "Assistência Médica"

API_URL    = "http://localhost:8000/api/s1210-repo/enviar-lote-cpfs"
BATCH_SIZE = 50
TIMEOUT    = 600
OUTDIR     = Path(ROOT) / "saida_lote2_julho"
OUTDIR.mkdir(exist_ok=True)

# ── Helpers ───────────────────────────────────────────────────────────────
def clean_cpf(v) -> str | None:
    if v is None: return None
    s = re.sub(r"\D", "", str(v))
    return s if len(s) == 11 else None

# ── 1. Parse XLSX ─────────────────────────────────────────────────────────
def parse_xlsx():
    print(f"Lendo XLSX: {XLSX_PATH.name}")
    wb = openpyxl.load_workbook(str(XLSX_PATH), read_only=True, data_only=True)

    # ── Aba geral: lista de CPFs do lote ──
    ws = wb[ABA_GERAL]
    cpfs_scope: list[dict] = []  # {cpf, nome, matricula, row_number, raw_row}
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i == 1: continue  # header
        cpf = clean_cpf(row[8])  # col I
        if not cpf: continue
        cpfs_scope.append({
            "cpf": cpf,
            "nome": None,
            "matricula": str(row[6]) if row[6] else None,  # CodigoFuncionario
            "row_number": i,
            "raw_row": json.dumps(list(row), default=str),
        })

    print(f"  [{ABA_GERAL}] {len(cpfs_scope)} CPFs")

    # ── Aba operadoras ──
    ws2 = wb[ABA_OPER]
    oper_rows: list[dict] = []
    for i, row in enumerate(ws2.iter_rows(values_only=True), 1):
        if i == 1: continue
        cpf = clean_cpf(row[0])
        if not cpf: continue
        cnpj = str(row[17] or "").strip()
        ans  = str(row[18] or "").strip()
        vlr  = row[21]
        cod  = row[13]
        nome_ev = str(row[14] or "")
        # Ignora informativos
        if cnpj.lower() in ("informativo", "", "none", "nan"): continue
        if ans.lower() in ("informativo", "", "none", "nan"): continue
        cnpj_d = re.sub(r"\D", "", cnpj)
        ans_d  = re.sub(r"\D", "", ans)
        if not cnpj_d or not ans_d: continue
        try:
            centavos = int(float(vlr or 0))
        except Exception:
            continue
        if centavos <= 0: continue
        oper_rows.append({
            "cpf": cpf,
            "rubrica_origem": str(cod or ""),
            "cnpj_operadora": cnpj_d,
            "reg_ans": ans_d,
            "nome_operadora": nome_ev,
            "valor": centavos,
            "raw_row": json.dumps(list(row), default=str),
        })

    print(f"  [{ABA_OPER}] {len(oper_rows)} entradas de operadoras válidas")
    return cpfs_scope, oper_rows

# ── 2. Import DB ──────────────────────────────────────────────────────────
def importar_db(cpfs_scope: list[dict], oper_rows: list[dict], dry_run: bool) -> int:
    """Insere xlsx, scope e operadoras. Retorna xlsx_id."""
    stat = XLSX_PATH.stat()
    sha256 = hashlib.sha256(XLSX_PATH.read_bytes()).hexdigest()

    with psycopg2.connect(**DB_CONFIG) as conn:
        cur = conn.cursor()

        # Verifica se já existe
        cur.execute("SELECT id FROM s1210_xlsx WHERE sha256=%s", (sha256,))
        row = cur.fetchone()
        if row:
            xlsx_id = row[0]
            print(f"  s1210_xlsx já existe (id={xlsx_id}) — pulando insert")
        else:
            if not dry_run:
                totais = {str(LOTE_NUM) + "_LOTE": len(cpfs_scope)}
                cur.execute(
                    """INSERT INTO s1210_xlsx
                         (empresa_id, per_apur, nome_arquivo, tamanho_bytes, sha256,
                          storage_path, aba_geral, aba_operadoras, uploaded_at,
                          parse_ok, totais_json)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s) RETURNING id""",
                    (EMPRESA_ID, PER_APUR, XLSX_PATH.name, stat.st_size, sha256,
                     f"local/{PER_APUR}/{sha256[:12]}.xlsx",
                     ABA_GERAL, ABA_OPER,
                     datetime.now(timezone.utc),
                     True,
                     json.dumps(totais)),
                )
                xlsx_id = cur.fetchone()[0]
                print(f"  s1210_xlsx inserido (id={xlsx_id})")
            else:
                xlsx_id = 0
                print(f"  [DRY] s1210_xlsx seria inserido")

        # scope
        if not dry_run:
            cur.execute(
                "DELETE FROM s1210_cpf_scope WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s",
                (EMPRESA_ID, PER_APUR, LOTE_NUM)
            )
            print(f"  s1210_cpf_scope delete existentes: {cur.rowcount}")
            psycopg2.extras.execute_batch(
                cur,
                """INSERT INTO s1210_cpf_scope
                     (xlsx_id, empresa_id, per_apur, lote_num, cpf, nome, matricula, row_number, raw_row)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                [(xlsx_id, EMPRESA_ID, PER_APUR, LOTE_NUM,
                  r["cpf"], r["nome"], r["matricula"], r["row_number"], r["raw_row"])
                 for r in cpfs_scope],
                page_size=500,
            )
            print(f"  s1210_cpf_scope inseridos: {len(cpfs_scope)}")
        else:
            print(f"  [DRY] s1210_cpf_scope: {len(cpfs_scope)} CPFs")

        # operadoras
        if not dry_run:
            cur.execute(
                "DELETE FROM s1210_operadoras WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s",
                (EMPRESA_ID, PER_APUR, LOTE_NUM)
            )
            print(f"  s1210_operadoras delete existentes: {cur.rowcount}")
            psycopg2.extras.execute_batch(
                cur,
                """INSERT INTO s1210_operadoras
                     (xlsx_id, empresa_id, per_apur, lote_num, cpf, rubrica_origem,
                      cnpj_operadora, reg_ans, nome_operadora, valor, raw_row)
                   VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                [(xlsx_id, EMPRESA_ID, PER_APUR, LOTE_NUM,
                  r["cpf"], r["rubrica_origem"], r["cnpj_operadora"], r["reg_ans"],
                  r["nome_operadora"], r["valor"], r["raw_row"])
                 for r in oper_rows],
                page_size=500,
            )
            print(f"  s1210_operadoras inseridos: {len(oper_rows)}")
        else:
            print(f"  [DRY] s1210_operadoras: {len(oper_rows)} linhas")

        if not dry_run:
            conn.commit()

    return xlsx_id

# ── 3. Carregar dados do DB para envio ────────────────────────────────────
def carregar_planos() -> dict[str, list[dict]]:
    sql = """
        SELECT cpf, cnpj_operadora, MAX(reg_ans) AS reg_ans, SUM(valor)::BIGINT AS centavos
          FROM s1210_operadoras
         WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
           AND cnpj_operadora IS NOT NULL
         GROUP BY cpf, cnpj_operadora
         ORDER BY cpf, cnpj_operadora
    """
    out: dict[str, list[dict]] = {}
    with psycopg2.connect(**DB_CONFIG) as conn, conn.cursor() as cur:
        cur.execute(sql, (EMPRESA_ID, PER_APUR, LOTE_NUM))
        for cpf, cnpj, ans, centavos in cur.fetchall():
            cents = int(centavos or 0)
            if cents <= 0: continue
            out.setdefault(cpf, []).append({
                "cnpjOper": cnpj,
                "regANS": ans or "",
                "vlrSaudeTit": f"{cents / 100:.2f}",
            })
    return out

def carregar_cpfs_pendentes() -> list[str]:
    with psycopg2.connect(**DB_CONFIG) as conn, conn.cursor() as cur:
        cur.execute(
            """
            WITH lv AS (
              SELECT DISTINCT ON (cpf) cpf, status
                FROM s1210_cpf_envios
               WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
               ORDER BY cpf, enviado_em DESC NULLS LAST, id DESC
            )
            SELECT s.cpf
              FROM s1210_cpf_scope s
              LEFT JOIN lv ON lv.cpf=s.cpf
             WHERE s.empresa_id=%s AND s.per_apur=%s AND s.lote_num=%s
               AND (lv.status IS NULL OR lv.status NOT IN ('ok', 'na'))
             ORDER BY s.cpf
            """,
            (EMPRESA_ID, PER_APUR, LOTE_NUM, EMPRESA_ID, PER_APUR, LOTE_NUM),
        )
        return [row[0] for row in cur.fetchall()]

# ── 4. Envio ──────────────────────────────────────────────────────────────
def enviar_bloco(cpfs: list[str], planos: dict) -> dict:
    ps_slice = {cpf: planos[cpf] for cpf in cpfs if cpf in planos}
    payload: dict = {
        "per_apur": PER_APUR,
        "lote_num": LOTE_NUM,
        "cpfs": cpfs,
        "confirmar_producao": True,
    }
    if ps_slice:
        payload["plan_saude_por_cpf"] = ps_slice
    t0 = time.time()
    r = requests.post(API_URL, json=payload, timeout=TIMEOUT)
    r.raise_for_status()
    body = r.json()
    body["_client_elapsed_s"] = round(time.time() - t0, 1)
    return body

# ── Main ──────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--batch", type=int, default=BATCH_SIZE)
    ap.add_argument("--max", type=int, default=None)
    ap.add_argument("--skip-import", action="store_true", help="Pula etapa de import DB")
    ap.add_argument("--continuar-com-erros", action="store_true", help="Não para em bloco 100% erro")
    ap.add_argument("--workers", type=int, default=1, help="Workers paralelos (default 1)")
    args = ap.parse_args()

    print(f"=== Envio Julho 2025 L2 — {XLSX_PATH.name} ===\n")

    # ── Parse ──
    cpfs_scope, oper_rows = parse_xlsx()
    cpfs_set = {r["cpf"] for r in cpfs_scope}
    planos_xlsx = {}
    for r in oper_rows:
        planos_xlsx.setdefault(r["cpf"], [])
    sem_plano_xlsx = [c for c in cpfs_set if c not in planos_xlsx]
    print(f"\nResumo XLSX:")
    print(f"  CPFs no lote:         {len(cpfs_scope)}")
    print(f"  Com plano real:       {len(planos_xlsx)}")
    print(f"  Sem plano (informativo): {len(sem_plano_xlsx)}")

    if args.dry_run:
        print("\n[DRY-RUN] Nenhuma gravação realizada.")
        importar_db(cpfs_scope, oper_rows, dry_run=True)
        return

    # ── Import DB ──
    if not args.skip_import:
        print("\n--- Importando para DB ---")
        importar_db(cpfs_scope, oper_rows, dry_run=False)
    else:
        print("\n--- Skip import (--skip-import) ---")

    # ── Envio ──
    planos_db = carregar_planos()
    cpfs = carregar_cpfs_pendentes()
    if args.max:
        cpfs = cpfs[:args.max]

    print(f"\n--- Envio ---")
    print(f"  Pendentes no DB:  {len(cpfs)}")
    print(f"  Com plano no DB:  {len(planos_db)}")
    sem_plano_envio = [c for c in cpfs if c not in planos_db]
    print(f"  Sem plano entre pendentes: {len(sem_plano_envio)}")
    if sem_plano_envio:
        print(f"  (serão enviados sem planSaude — rubricas informativas)")

    batches = [cpfs[i:i + args.batch] for i in range(0, len(cpfs), args.batch)]
    total_ok = total_err = 0
    t_global = time.time()

    if args.workers > 1:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading
        lock = threading.Lock()

        def _job(idx_bloco):
            i, bloco = idx_bloco
            try:
                body = enviar_bloco(bloco, planos_db)
            except Exception as e:
                return (i, None, f"REDE {type(e).__name__}: {e}")
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            (OUTDIR / f"resp_bloco{i:03d}_{ts}.json").write_text(
                json.dumps(body, ensure_ascii=False, indent=2), encoding="utf-8")
            ok_n  = sum(1 for r in body.get("resultados", []) if r.get("sucesso"))
            err_n = sum(1 for r in body.get("resultados", []) if not r.get("sucesso"))
            erros = [(r.get("cpf"), r.get("etapa"), str(r.get("erro",""))[:80])
                     for r in body.get("resultados", []) if not r.get("sucesso")]
            return (i, (ok_n, err_n, erros, body.get("_client_elapsed_s", 0)), None)

        with ThreadPoolExecutor(max_workers=args.workers) as ex:
            futs = {ex.submit(_job, (i, b)): i for i, b in enumerate(batches, 1)}
            for fut in as_completed(futs):
                i, res, err_msg = fut.result()
                if err_msg:
                    print(f"[bloco {i}] FALHA REDE: {err_msg}")
                    continue
                ok_n, err_n, erros, dt_b = res
                with lock:
                    total_ok += ok_n; total_err += err_n
                    feitos = total_ok + total_err
                    dt_total = time.time() - t_global
                    taxa = feitos / dt_total * 60 if dt_total else 0
                    eta = (len(cpfs) - feitos) / max(taxa, 0.01)
                    print(f"[bloco {i:>3}/{len(batches)}] ok={ok_n} err={err_n} tempo={dt_b}s | "
                          f"acum ok={total_ok} err={total_err} | "
                          f"{taxa:.0f} CPF/min ETA={eta:.1f}min")
                    for cpf, etapa, msg in erros[:2]:
                        print(f"   ERR {cpf} {etapa}: {msg}")
    else:
        for idx, bloco in enumerate(batches, 1):
            print(f"\n=== bloco {idx}/{len(batches)} — {len(bloco)} CPFs ===")
            try:
                body = enviar_bloco(bloco, planos_db)
            except Exception as e:
                print(f"  ERRO HTTP: {e}")
                break

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            out_path = OUTDIR / f"resp_bloco{idx:03d}_{ts}.json"
            out_path.write_text(json.dumps(body, ensure_ascii=False, indent=2), encoding="utf-8")

            ok  = sum(1 for r in body.get("resultados", []) if r.get("sucesso"))
            err = sum(1 for r in body.get("resultados", []) if not r.get("sucesso"))
            total_ok += ok
            total_err += err
            print(f"  ok={ok} err={err} tempo={body.get('_client_elapsed_s')}s | acum ok={total_ok} err={total_err}")

            if err and not ok and not args.continuar_com_erros:
                print("  [AVISO] bloco 100% erro — parando (use --continuar-com-erros para ignorar)")
                break

    dt_total = time.time() - t_global
    print(f"\n=== FIM === ok={total_ok} err={total_err} tempo={dt_total:.0f}s")

if __name__ == "__main__":
    main()
