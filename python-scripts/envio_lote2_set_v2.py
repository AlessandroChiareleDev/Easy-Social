"""
Reenvio Lote 2 Setembro/2025 APPA — usando XLSX da Ana com plan_saude correto.

CORRECAO do envio anterior (envio_lote2_setembro_complemento.py):
  - O envio anterior deixou backend montar plan_saude da tabela s1210_operadoras
    (suja: rubrica 774 sindical entrava como plano coletivo).
  - eSocial rejeitou 54/57 com codigo 8 ('Plano de saude coletivo deve ser preenchido').
  - Como rejeicao na fase de processamento NAO altera recibo, o recibo do XLSX
    de recibos AINDA e vigente. Da pra reenviar.

Fonte definitiva agora:
  - Recibos: BuscarRecibo_Set2025_L2_COMPLEMENTO_20260506_1628_retorno.xlsx (col 'CPF (limpo)' + 'RECIBO ATIVO eSocial')
  - planSaude: Correção Lote 2 _erros  certo.xlsx (filtra CNPJ 14d + ANS numerico)

Filtro do XLSX da Ana:
  - col 14: 'CNPJ OPERADORA ' precisa ter 14 digitos
  - col 15: 'ANS ' precisa ser numerico
  - col 16: 'NATUREZA RUBRICA' deve ser 9219 (plano coletivo empresarial) -- garantia extra
  Linhas com cnpj='-' / 'Informativo' (rubricas 774 e 9281) sao ignoradas.

Uso:
  python envio_lote2_set_v2.py --dry-run-1   # mostra payload do 1o CPF
  python envio_lote2_set_v2.py --max 1       # envia 1 CPF teste
  python envio_lote2_set_v2.py               # tudo
"""
from __future__ import annotations
import argparse, json, os, sys, time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

import requests
ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

PER_APUR = "2025-09"
LOTE_NUM = 2
EMPRESA_ID = 1
API_URL = "http://localhost:8000/api/s1210-repo/enviar-lote-cpfs"
BATCH_SIZE = 50
TIMEOUT = 600

XLSX_RECIBOS = Path.home()/"Downloads"/"BuscarRecibo_Set2025_L2_COMPLEMENTO_20260506_1628_retorno.xlsx"
ABA_RECIBOS = "Set2025 L2 complemento"
COL_CPF_R = 2
COL_REC_R = 8

XLSX_ANA = Path.home()/"Downloads"/"Correção Lote 2 _erros  certo.xlsx"
ABA_ANA = "Planilha1"
# colunas (0-based) confirmadas via inspecao:
COL_CPF_A = 8        # '874.987.157-91'
COL_COD_EVT = 10     # 516, 9281, 631, 774, 775
COL_CNPJ = 14        # CNPJ OPERADORA (com pontuacao ou '-')
COL_ANS = 15         # numero ANS ou '-' ou 'Informativo'
COL_NAT = 16         # NATUREZA RUBRICA (9219 valido, 9299 sindical)
COL_VLR = 18         # ValorEvento (em centavos, int)

OUTDIR = os.path.join(ROOT, "saida_lote2_set_v2")
os.makedirs(OUTDIR, exist_ok=True)


def _digits(s) -> str:
    return "".join(c for c in str(s or "") if c.isdigit())


def carregar_recibos() -> dict[str, str]:
    from openpyxl import load_workbook
    wb = load_workbook(XLSX_RECIBOS, read_only=True, data_only=True)
    ws = wb[ABA_RECIBOS]
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        cpf = _digits(r[COL_CPF_R]).zfill(11)
        rec = str(r[COL_REC_R] or "").strip()
        if len(cpf) == 11 and rec.startswith("1."):
            out[cpf] = rec
    return out


def carregar_plan_saude_ana() -> tuple[dict[str, list[dict]], dict[str, list[tuple]]]:
    """Retorna (plan_saude_por_cpf, descartados_por_cpf)."""
    from openpyxl import load_workbook
    wb = load_workbook(XLSX_ANA, read_only=True, data_only=True)
    ws = wb[ABA_ANA]
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    agg: dict[str, dict[str, dict]] = defaultdict(lambda: defaultdict(lambda: {"reg_ans": "", "vlr_cent": 0, "rubricas": []}))
    descartados: dict[str, list[tuple]] = defaultdict(list)
    for r in rows:
        cpf = _digits(r[COL_CPF_A]).zfill(11)
        if len(cpf) != 11:
            continue
        cnpj = _digits(r[COL_CNPJ])
        ans = _digits(r[COL_ANS])
        nat = r[COL_NAT]
        try:
            vlr = int(r[COL_VLR] or 0)
        except (TypeError, ValueError):
            vlr = 0
        cod = r[COL_COD_EVT]
        if len(cnpj) != 14 or not ans:
            descartados[cpf].append((cod, str(r[COL_CNPJ]), str(r[COL_ANS]), nat, vlr))
            continue
        if str(nat) != "9219":
            descartados[cpf].append((cod, cnpj, ans, nat, vlr))
            continue
        agg[cpf][cnpj]["reg_ans"] = ans
        agg[cpf][cnpj]["vlr_cent"] += vlr
        agg[cpf][cnpj]["rubricas"].append(cod)

    out: dict[str, list[dict]] = {}
    for cpf, cnpjs in agg.items():
        items = []
        for cnpj, d in cnpjs.items():
            items.append({
                "cnpjOper": cnpj,
                "regANS": d["reg_ans"],
                "vlrSaudeTit": round(d["vlr_cent"]/100, 2),
            })
        out[cpf] = items
    return out, descartados


def enviar_bloco(cpfs, recibos, plan_saude):
    payload = {
        "per_apur": PER_APUR,
        "lote_num": LOTE_NUM,
        "cpfs": cpfs,
        "confirmar_producao": True,
        "recibo_override_por_cpf": {c: recibos[c] for c in cpfs if c in recibos},
        "plan_saude_por_cpf":     {c: plan_saude[c] for c in cpfs if c in plan_saude},
    }
    t0 = time.time()
    r = requests.post(API_URL, json=payload, timeout=TIMEOUT)
    r.raise_for_status()
    body = r.json()
    body["_client_elapsed_s"] = round(time.time()-t0, 1)
    return body


def resumir(body):
    ok = err = 0
    erros = []
    for det in body.get("resultados", []):
        if det.get("sucesso"): ok += 1
        else:
            err += 1
            ocorr = det.get("ocorrencias") or []
            ocstr = "; ".join(f'{o.get("codigo")}:{(o.get("descricao") or "")[:80]}' for o in ocorr)
            erros.append((det.get("cpf",""), str(det.get("codigo_resposta","")), ocstr))
    return ok, err, erros


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--max", type=int, default=None)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--dry-run-1", action="store_true", help="Mostra payload do 1o CPF")
    ap.add_argument("--cpf", default=None, help="Forca 1 CPF especifico")
    ap.add_argument("--batch", type=int, default=BATCH_SIZE)
    args = ap.parse_args()

    recibos = carregar_recibos()
    plan_saude, descartados = carregar_plan_saude_ana()
    cpfs = sorted(set(recibos.keys()))
    print(f"[recibos]    {len(recibos)} CPFs")
    print(f"[plan_saude] {len(plan_saude)} CPFs com itens validos no XLSX da Ana")

    sem_ps = [c for c in cpfs if c not in plan_saude]
    if sem_ps:
        print(f"[ATENCAO] {len(sem_ps)} CPFs com recibo MAS sem plan_saude (cnpj/ans validos): {sem_ps[:8]}{'...' if len(sem_ps)>8 else ''}")

    if args.cpf:
        cpfs = [args.cpf]
    elif args.max:
        cpfs = cpfs[:args.max]

    if args.dry_run_1:
        c = cpfs[0]
        print(f"\n=== DRY-RUN payload do CPF {c} ===")
        payload = {
            "per_apur": PER_APUR, "lote_num": LOTE_NUM, "cpfs": [c],
            "confirmar_producao": True,
            "recibo_override_por_cpf": {c: recibos[c]},
            "plan_saude_por_cpf": {c: plan_saude.get(c, [])},
        }
        print(json.dumps(payload, indent=2, ensure_ascii=False))
        if c in descartados:
            print(f"\n[descartados {c}]")
            for d in descartados[c][:5]:
                print(" ", d)
        return

    if args.dry_run:
        print(f"[dry-run] {len(cpfs)} CPFs, batches: {[len(cpfs[i:i+args.batch]) for i in range(0,len(cpfs),args.batch)]}")
        for c in cpfs[:3]:
            print(f"  {c}: rec={recibos.get(c)} plan={plan_saude.get(c)}")
        return

    total_ok = total_err = 0
    batches = [cpfs[i:i+args.batch] for i in range(0,len(cpfs),args.batch)]
    t0 = time.time()
    for i, bloco in enumerate(batches, 1):
        print(f"\n=== Bloco {i}/{len(batches)} ({len(bloco)}) ===")
        try:
            body = enviar_bloco(bloco, recibos, plan_saude)
        except Exception as e:
            print(f"FALHA HTTP: {e}"); break
        ok, err, erros = resumir(body)
        total_ok += ok; total_err += err
        print(f"  ok={ok} err={err} elapsed={body.get('_client_elapsed_s')}s")
        for cpf, cod, oc in erros[:5]:
            print(f"    {cpf} cod={cod} {oc}")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(OUTDIR, f"resp_b{i:02d}_{ts}.json")
        with open(path,"w",encoding="utf-8") as f:
            json.dump(body, f, ensure_ascii=False, indent=2)
        print(f"  saved: {path}")
    print(f"\n=== TOTAL ok={total_ok} err={total_err} em {round(time.time()-t0,1)}s ===")


if __name__ == "__main__":
    main()
