"""
Reenvio Lote 2 Setembro/2025 APPA — v3 — para os 4 CPFs que ficaram pendentes.

Esses 4 CPFs não vieram no XLSX BuscarRecibo_Set2025_L2_COMPLEMENTO. Recibos foram
extraídos diretamente do ZIP master (xmls do e social mes a mes / 09-set2025.zip)
parseando os S-1210 originais — pegando o evento mais recente (dhProcessamento mais
novo) de cada CPF.

CPFs alvo:
  77542975749 → recibo 1.1.0000000035299426157  (ideDmDev=01513073)
  79308899715 → recibo 1.1.0000000035299426775  (ideDmDev=01513065+01513100)
  80209734787 → recibo 1.1.0000000035299426989  (ideDmDev=01513073)
  84177926787 → recibo 1.1.0000000035299427217  (ideDmDev=01513065+01513100)

planSaude do XLSX da Ana 'Assistencia Medica Lote 2092025 (1).xlsx' aba
'Assistencia Médica - Lote 2', filtrando CNPJ 14d + ANS numerico + natureza 9219.

Uso:
  python envio_lote2_set_v3.py --dry-run-1   # mostra payload do 1o
  python envio_lote2_set_v3.py --max 1       # envia 1
  python envio_lote2_set_v3.py               # envia os 4
"""
from __future__ import annotations
import argparse, json, os, sys, time
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

import requests
ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

PER_APUR = "2025-09"
LOTE_NUM = 2
EMPRESA_ID = 1
API_URL = "http://localhost:8000/api/s1210-repo/enviar-lote-cpfs"
TIMEOUT = 600

# Recibos ATIVOS — XLSX 'Erro Final Lote 2 - 092025.xlsx' enviado pela Ana 06/05 21:57
# (recibos do ZIP master e do banco eram obsoletos — esses 4 foram retificados externamente)
RECIBOS_ATIVOS = {
    "77542975749": "1.1.0000000040407380475",
    "79308899715": "1.1.0000000040407389717",
    "80209734787": "1.1.0000000040407391313",
    "84177926787": "1.1.0000000040409573514",
}

XLSX_ANA = Path.home()/"Downloads"/"Assistencia Medica Lote 2092025 (1).xlsx"
ABA_ANA = "Assistencia Médica - Lote 2"
COL_CPF = 8
COL_COD = 10
COL_CNPJ = 14
COL_ANS = 15
COL_NAT = 16
COL_VLR = 18

OUTDIR = os.path.join(ROOT, "saida_lote2_set_v3")
os.makedirs(OUTDIR, exist_ok=True)


def _digits(s) -> str:
    return "".join(c for c in str(s or "") if c.isdigit())


def carregar_plan_saude() -> dict[str, list[dict]]:
    from openpyxl import load_workbook
    wb = load_workbook(XLSX_ANA, read_only=True, data_only=True)
    ws = wb[ABA_ANA]
    agg = defaultdict(lambda: defaultdict(lambda: {"reg_ans":"", "vlr_cent":0}))
    for r in ws.iter_rows(min_row=2, values_only=True):
        cpf = _digits(r[COL_CPF]).zfill(11)
        if cpf not in RECIBOS_ATIVOS: continue
        cnpj = _digits(r[COL_CNPJ])
        ans  = _digits(r[COL_ANS])
        nat  = r[COL_NAT]
        try: vlr = int(r[COL_VLR] or 0)
        except (TypeError, ValueError): vlr = 0
        if len(cnpj)!=14 or not ans or str(nat)!="9219":
            continue
        agg[cpf][cnpj]["reg_ans"] = ans
        agg[cpf][cnpj]["vlr_cent"] += vlr
    out = {}
    for cpf, cnpjs in agg.items():
        out[cpf] = [
            {"cnpjOper": cnpj, "regANS": d["reg_ans"], "vlrSaudeTit": round(d["vlr_cent"]/100, 2)}
            for cnpj, d in cnpjs.items()
        ]
    return out


def enviar(cpfs, recibos, plan_saude):
    payload = {
        "per_apur": PER_APUR,
        "lote_num": LOTE_NUM,
        "cpfs": cpfs,
        "confirmar_producao": True,
        "recibo_override_por_cpf": {c: recibos[c] for c in cpfs if c in recibos},
        "plan_saude_por_cpf":     {c: plan_saude[c] for c in cpfs if c in plan_saude},
    }
    t0 = time.time()
    r = requests.post(API_URL, json=payload, timeout=TIMEOUT)
    r.raise_for_status()
    body = r.json()
    body["_client_elapsed_s"] = round(time.time()-t0, 1)
    return body


def resumir(body):
    ok = err = 0
    erros = []
    for det in body.get("resultados", []):
        if det.get("sucesso"): ok += 1
        else:
            err += 1
            ocorr = det.get("ocorrencias") or []
            ocstr = "; ".join(f'{o.get("codigo")}:{(o.get("descricao") or "")[:120]}' for o in ocorr)
            erros.append((det.get("cpf",""), str(det.get("codigo_resposta","")), ocstr))
    return ok, err, erros


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--max", type=int, default=None)
    ap.add_argument("--dry-run-1", action="store_true")
    ap.add_argument("--cpf", default=None)
    args = ap.parse_args()

    plan_saude = carregar_plan_saude()
    cpfs = sorted(RECIBOS_ATIVOS.keys())
    print(f"[recibos]    {len(RECIBOS_ATIVOS)} CPFs (extraidos do zip set2025)")
    print(f"[plan_saude] {len(plan_saude)} CPFs com itens validos")
    falt = [c for c in cpfs if c not in plan_saude]
    if falt:
        print(f"[ERRO] CPFs sem plan_saude: {falt}"); return

    if args.cpf:
        cpfs = [args.cpf]
    elif args.max:
        cpfs = cpfs[:args.max]

    if args.dry_run_1:
        c = cpfs[0]
        payload = {"per_apur":PER_APUR,"lote_num":LOTE_NUM,"cpfs":[c],
                   "confirmar_producao":True,
                   "recibo_override_por_cpf":{c:RECIBOS_ATIVOS[c]},
                   "plan_saude_por_cpf":{c:plan_saude[c]}}
        print(json.dumps(payload, indent=2, ensure_ascii=False)); return

    print(f"\n=== Enviando {len(cpfs)} CPFs ===")
    try:
        body = enviar(cpfs, RECIBOS_ATIVOS, plan_saude)
    except Exception as e:
        print(f"FALHA HTTP: {e}"); return
    ok, err, erros = resumir(body)
    print(f"  ok={ok} err={err} elapsed={body.get('_client_elapsed_s')}s")
    for c, cod, oc in erros:
        print(f"    {c} cod={cod} {oc}")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out = os.path.join(OUTDIR, f"resp_{ts}.json")
    with open(out, "w", encoding="utf-8") as f:
        json.dump(body, f, indent=2, ensure_ascii=False)
    print(f"\nResposta salva: {out}")


if __name__ == "__main__":
    main()
