"""
Envio Lote 2 Setembro/2025 APPA — COMPLEMENTO (57 CPFs com recibos vigentes).

Adaptado de envio_lote3_agosto.py para:
  - per_apur=2025-09, lote_num=2, empresa_id=1
  - Restringe aos 57 CPFs vindos do XLSX da Ana
    "BuscarRecibo_Set2025_L2_COMPLEMENTO_20260506_1628_retorno.xlsx"
  - recibo_override_por_cpf vem da coluna "RECIBO ATIVO eSocial (preencher)" (col 8)
  - workers=1, batch=50, sequencial (são só 57 CPFs -> 2 POSTs)

Uso:
  python envio_lote2_setembro_complemento.py --dry-run        # mostra o que faria
  python envio_lote2_setembro_complemento.py --max 1          # 1 CPF teste
  python envio_lote2_setembro_complemento.py                  # tudo (50 + 7)
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

import requests

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

PER_APUR = "2025-09"
LOTE_NUM = 2
EMPRESA_ID = 1
API_URL = "http://localhost:8000/api/s1210-repo/enviar-lote-cpfs"
BATCH_SIZE = 50
TIMEOUT = 600

XLSX_RECIBOS = Path.home() / "Downloads" / "BuscarRecibo_Set2025_L2_COMPLEMENTO_20260506_1628_retorno.xlsx"
ABA = "Set2025 L2 complemento"
COL_CPF = 2          # 'CPF (limpo)' (11 digitos)
COL_RECIBO = 8       # 'RECIBO ATIVO eSocial (preencher)'

OUTDIR = os.path.join(ROOT, "saida_lote2_set_complemento")
os.makedirs(OUTDIR, exist_ok=True)


def carregar_recibos_xlsx() -> dict[str, str]:
    from openpyxl import load_workbook
    wb = load_workbook(XLSX_RECIBOS, read_only=True, data_only=True)
    if ABA not in wb.sheetnames:
        raise SystemExit(f"Aba '{ABA}' nao encontrada. Abas: {wb.sheetnames}")
    ws = wb[ABA]
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf_raw = row[COL_CPF]
        rec_raw = row[COL_RECIBO]
        if not cpf_raw or not rec_raw:
            continue
        cpf = "".join(ch for ch in str(cpf_raw) if ch.isdigit()).zfill(11)
        rec = str(rec_raw).strip()
        if len(cpf) == 11 and rec.startswith("1."):
            out[cpf] = rec
    return out


def enviar_bloco(cpfs: list[str], recibos_override: dict[str, str]) -> dict:
    payload = {
        "per_apur": PER_APUR,
        "lote_num": LOTE_NUM,
        "cpfs": cpfs,
        "confirmar_producao": True,
        "recibo_override_por_cpf": {c: recibos_override[c] for c in cpfs if c in recibos_override},
    }
    t0 = time.time()
    r = requests.post(API_URL, json=payload, timeout=TIMEOUT)
    r.raise_for_status()
    dt = time.time() - t0
    body = r.json()
    body["_client_elapsed_s"] = round(dt, 1)
    return body


def resumir(body: dict) -> tuple[int, int, list[tuple[str, str, str]]]:
    ok = err = 0
    erros: list[tuple[str, str, str]] = []
    for det in body.get("resultados", []):
        cpf = det.get("cpf", "")
        if det.get("sucesso"):
            ok += 1
        else:
            err += 1
            erros.append((cpf, str(det.get("codigo_resposta", "")),
                          str(det.get("descricao_resposta") or det.get("erro") or "")[:160]))
    return ok, err, erros


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--max", type=int, default=None)
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--batch", type=int, default=BATCH_SIZE)
    args = ap.parse_args()

    recibos = carregar_recibos_xlsx()
    cpfs = sorted(recibos.keys())
    print(f"[xlsx] {len(cpfs)} CPFs com recibo override carregados de {XLSX_RECIBOS.name}")
    if not cpfs:
        print("nada a enviar"); return

    if args.max:
        cpfs = cpfs[: args.max]
        print(f"[--max] reduzido a {len(cpfs)} CPFs")

    if args.dry_run:
        print(f"\n[dry-run] per_apur={PER_APUR} lote={LOTE_NUM} empresa={EMPRESA_ID}")
        print(f"[dry-run] batches: {[len(cpfs[i:i+args.batch]) for i in range(0, len(cpfs), args.batch)]}")
        print(f"[dry-run] amostra (5 primeiros):")
        for c in cpfs[:5]:
            print(f"   {c} -> {recibos[c]}")
        return

    total_ok = total_err = 0
    batches = [cpfs[i:i + args.batch] for i in range(0, len(cpfs), args.batch)]
    t_global = time.time()
    for i, bloco in enumerate(batches, 1):
        print(f"\n=== Bloco {i}/{len(batches)} ({len(bloco)} CPFs) ===")
        try:
            body = enviar_bloco(bloco, recibos)
        except Exception as e:
            print(f"  FALHA HTTP: {e}")
            break
        ok, err, erros = resumir(body)
        total_ok += ok; total_err += err
        print(f"  ok={ok} err={err} elapsed={body.get('_client_elapsed_s')}s")
        if erros:
            print("  primeiros erros:")
            for cpf, cod, desc in erros[:5]:
                print(f"    {cpf} cod={cod} {desc}")
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(OUTDIR, f"resp_b{i:02d}_{ts}.json")
        with open(path, "w", encoding="utf-8") as f:
            json.dump(body, f, ensure_ascii=False, indent=2)
        print(f"  resposta salva: {path}")

    dt = round(time.time() - t_global, 1)
    print(f"\n=== TOTAL ok={total_ok} err={total_err} em {dt}s ({len(batches)} blocos) ===")


if __name__ == "__main__":
    main()
