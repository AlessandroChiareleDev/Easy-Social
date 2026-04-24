"""
Envio Lote 3 Maio/2025 APPA (empresa_id=1) — receita PC1-15-v2.

Fluxo:
1. Le CPFs de s1210_cpf_scope (populado previamente via upload XLSX).
2. Filtra CPFs que ainda NAO tem envio OK (via DISTINCT ON ult status).
3. Fatia em blocos de BATCH_SIZE (50) e faz POST sequencial no endpoint
   /api/s1210-repo/enviar-lote-cpfs, 1 POST por bloco, sem threads.
4. Salva JSON de resposta de cada bloco em saida_lote3_maio/resp_<ts>.json.
5. Parada automatica se acontecer:
   - timeout de rede (sem retry)
   - erro HTTP != 2xx
   - codigo_resposta != '201' e != erro de negocio conhecido (459/861)

Uso:
  # Dry run — so mostra CPFs que seriam enviados, nao envia
  python envio_lote3_maio.py --dry-run

  # Rodar so 1 CPF (teste inicial)
  python envio_lote3_maio.py --max 1

  # Rodar primeiros 10
  python envio_lote3_maio.py --max 10

  # Rodar tudo (batches de 50)
  python envio_lote3_maio.py

  # Forcar override de recibo a partir de XLSX da Ana (quando houver)
  python envio_lote3_maio.py --recibos-xlsx "C:\\...\\Lote3_Erros_maio.xlsx" --aba Mai_2025 --col-recibo 2

Ref: Mensagem-PC1-15-v2.md
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from datetime import datetime

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

import requests

ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

import psycopg2
import psycopg2.extras
from db_config import DB_CONFIG

PER_APUR = "2025-08"
LOTE_NUM = 3
EMPRESA_ID = 1
API_URL = "http://localhost:8000/api/s1210-repo/enviar-lote-cpfs"
BATCH_SIZE = 50
TIMEOUT = 600
OUTDIR = os.path.join(ROOT, "saida_lote3_agosto")
os.makedirs(OUTDIR, exist_ok=True)


def carregar_cpfs_pendentes() -> list[str]:
    """CPFs no scope que ainda NAO tem envio com status='ok'."""
    with psycopg2.connect(**DB_CONFIG) as c:
        with c.cursor() as cur:
            cur.execute(
                """
                WITH lv AS (
                  SELECT DISTINCT ON (cpf) cpf, status
                  FROM s1210_cpf_envios
                  WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
                  ORDER BY cpf, enviado_em DESC NULLS LAST
                )
                SELECT s.cpf
                FROM s1210_cpf_scope s
                LEFT JOIN lv ON lv.cpf = s.cpf
                WHERE s.empresa_id=%s AND s.per_apur=%s AND s.lote_num=%s
                  AND (lv.status IS NULL OR lv.status <> 'ok')
                ORDER BY s.cpf
                """,
                (EMPRESA_ID, PER_APUR, LOTE_NUM, EMPRESA_ID, PER_APUR, LOTE_NUM),
            )
            return [r[0] for r in cur.fetchall()]


def carregar_recibos_override_db() -> dict[str, str]:
    """Para cada CPF, pega o nr_recibo_novo do ULTIMO envio OK no mesmo per_apur.
    Isto cobre o caso 'Lote 1 ja retificou esse CPF em Maio -> recibo do ZIP invalido'.
    """
    sql = """
        SELECT DISTINCT ON (cpf) cpf, nr_recibo_novo
          FROM s1210_cpf_envios
         WHERE empresa_id=%s AND per_apur=%s AND status='ok'
           AND nr_recibo_novo IS NOT NULL
         ORDER BY cpf, enviado_em DESC
    """
    with psycopg2.connect(**DB_CONFIG) as c, c.cursor() as cur:
        cur.execute(sql, (EMPRESA_ID, PER_APUR))
        return {cpf: rec for cpf, rec in cur.fetchall()}


def carregar_recibos_override(xlsx: str, aba: str, col_cpf: int, col_recibo: int) -> dict[str, str]:
    from openpyxl import load_workbook
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb[aba]
    out: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cpf_raw = row[col_cpf]
        rec_raw = row[col_recibo]
        if cpf_raw is None or rec_raw is None:
            continue
        cpf = "".join(ch for ch in str(cpf_raw) if ch.isdigit()).zfill(11)
        rec = str(rec_raw).strip()
        if len(cpf) == 11 and rec.startswith("1."):
            out[cpf] = rec
    return out


def carregar_plan_saude_xlsx(xlsx: str, aba: str = "Assistencia Medica") -> dict[str, list[dict]]:
    """Le aba 'Assistencia Medica' do XLSX Ana (corrigida 24/04/2026 c/ cabecalhos).
    Colunas:
      [8]  CNPJ operadora (ou '-')
      [9]  Registro ANS   (ou '-')
      [11] CodigoEvento
      [17] ValorEvento (centavos)
      [29] CPF
    Regra: linhas com CNPJ 14d + ANS numerico -> agrega em plan_saude.
    Linhas com '-' nao entram (verba transitoria/informativa 9911/9299).
    Multiplas rubricas mesmo CNPJ -> SOMA valor no mesmo detPlanSaude.
    """
    from openpyxl import load_workbook
    from collections import defaultdict

    # aba normalizada — pegar 'Assistencia/Assitencia Medica' tolerante a typo/acento
    import unicodedata
    def _norm(s: str) -> str:
        return "".join(c for c in unicodedata.normalize("NFKD", s.lower()) if not unicodedata.combining(c))
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    alvo = None
    for sn in wb.sheetnames:
        n = _norm(sn)
        if "med" in n and ("assist" in n or "assit" in n):
            alvo = sn; break
    if not alvo:
        raise ValueError(f"Aba Assistencia Medica nao encontrada em {wb.sheetnames}")
    ws = wb[alvo]

    def _d(s): return "".join(c for c in str(s or "") if c.isdigit())

    agg: dict[str, dict[str, dict]] = defaultdict(lambda: defaultdict(lambda: {"reg_ans": "", "vlr_cent": 0}))
    for r in ws.iter_rows(min_row=2, values_only=True):
        cpf = _d(r[6]).zfill(11)
        if len(cpf) != 11:
            continue
        cnpj = _d(r[11])
        ans = _d(r[12])
        valor = r[15] or 0
        if len(cnpj) != 14 or not ans or ans == "0":
            continue
        agg[cpf][cnpj]["reg_ans"] = ans
        try:
            agg[cpf][cnpj]["vlr_cent"] += int(valor)
        except (TypeError, ValueError):
            pass

    out: dict[str, list[dict]] = {}
    for cpf, cnpjs in agg.items():
        entries = []
        for cnpj, d in cnpjs.items():
            entries.append({
                "cnpjOper": cnpj,
                "regANS": d["reg_ans"],
                "vlrSaudeTit": round(d["vlr_cent"] / 100, 2),
            })
        out[cpf] = entries
    return out


def enviar_bloco(cpfs: list[str], recibos_override: dict[str, str], plan_saude_override: dict[str, list[dict]] | None = None) -> dict:
    payload = {
        "per_apur": PER_APUR,
        "lote_num": LOTE_NUM,
        "cpfs": cpfs,
        "confirmar_producao": True,
    }
    override_slice = {c: recibos_override[c] for c in cpfs if c in recibos_override}
    if override_slice:
        payload["recibo_override_por_cpf"] = override_slice
    if plan_saude_override:
        ps_slice = {c: plan_saude_override[c] for c in cpfs if c in plan_saude_override}
        if ps_slice:
            payload["plan_saude_por_cpf"] = ps_slice

    t0 = time.time()
    r = requests.post(API_URL, json=payload, timeout=TIMEOUT)
    r.raise_for_status()
    dt = time.time() - t0
    body = r.json()
    body["_client_elapsed_s"] = round(dt, 1)
    return body


def resumir(body: dict) -> tuple[int, int, list[tuple[str, str, str]]]:
    ok = err = 0
    erros: list[tuple[str, str, str]] = []
    for det in body.get("resultados", []):
        cpf = det.get("cpf", "")
        if det.get("sucesso"):
            ok += 1
        else:
            err += 1
            erros.append((cpf, str(det.get("codigo_resposta", "")), str(det.get("descricao_resposta") or det.get("erro") or "")[:120]))
    return ok, err, erros


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--max", type=int, default=None, help="Limite de CPFs (debug)")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--batch", type=int, default=BATCH_SIZE)
    ap.add_argument("--recibos-xlsx", default=None)
    ap.add_argument("--aba", default=None)
    ap.add_argument("--col-cpf", type=int, default=0)
    ap.add_argument("--col-recibo", type=int, default=1)
    ap.add_argument("--plan-saude-xlsx", default=None, help="XLSX Ana (aba Assistencia Medica) c/ CNPJ+ANS")
    ap.add_argument("--workers", type=int, default=1, help="Numero de workers paralelos (default 1=sequencial)")
    args = ap.parse_args()

    cpfs = carregar_cpfs_pendentes()
    print(f"[scope pendente] {len(cpfs)} CPFs em {PER_APUR} lote={LOTE_NUM}")
    if not cpfs:
        print("nada a enviar"); return
    if args.max:
        cpfs = cpfs[: args.max]
        print(f"[--max] reduzido a {len(cpfs)} CPFs")

    recibos_override: dict[str, str] = {}
    # 1) DB: pega recibo_novo do ultimo envio OK por CPF (cobre Lote 1 Maio já retificado)
    recibos_override.update(carregar_recibos_override_db())
    print(f"[override DB] {len(recibos_override)} recibos vindos de s1210_cpf_envios")

    # 2) XLSX opcional: sobrescreve overrides do DB quando informado
    if args.recibos_xlsx:
        if not args.aba:
            print("ERRO: --aba obrigatorio quando usa --recibos-xlsx"); return
        xlsx_over = carregar_recibos_override(args.recibos_xlsx, args.aba, args.col_cpf, args.col_recibo)
        print(f"[override XLSX] {len(xlsx_over)} recibos carregados de {args.recibos_xlsx} (substitui DB)")
        recibos_override.update(xlsx_over)

    # 3) plan_saude_por_cpf do XLSX Ana (aba Assistencia Medica com CNPJ+ANS)
    plan_saude_override: dict[str, list[dict]] = {}
    if args.plan_saude_xlsx:
        plan_saude_override = carregar_plan_saude_xlsx(args.plan_saude_xlsx)
        print(f"[plan_saude] {len(plan_saude_override)} CPFs com CNPJ+ANS carregados de {args.plan_saude_xlsx}")

    if args.dry_run:
        print("[dry-run] primeiros 5 CPFs:", cpfs[:5])
        if recibos_override:
            sample = {c: recibos_override[c] for c in cpfs[:5] if c in recibos_override}
            print("[dry-run] overrides amostra:", sample)
        return

    total_ok = total_err = 0
    batches = [cpfs[i : i + args.batch] for i in range(0, len(cpfs), args.batch)]
    t_global = time.time()

    if args.workers > 1:
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading
        lock = threading.Lock()
        progresso = {"feitos": 0, "ok": 0, "err": 0, "ultimo_marco": 0, "t_marco": t_global}

        def _job(idx_bloco):
            i, bloco = idx_bloco
            try:
                body = enviar_bloco(bloco, recibos_override, plan_saude_override)
            except requests.exceptions.RequestException as e:
                return (i, None, f"REDE {type(e).__name__}: {e}")
            ok, err, erros = resumir(body)
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            with open(os.path.join(OUTDIR, f"resp_bloco{i:03d}_{ts}.json"), "w", encoding="utf-8") as f:
                json.dump(body, f, ensure_ascii=False, indent=2)
            return (i, (ok, err, erros, body.get("_client_elapsed_s", 0)), None)

        with ThreadPoolExecutor(max_workers=args.workers) as ex:
            futs = {ex.submit(_job, (i, b)): i for i, b in enumerate(batches, 1)}
            for fut in as_completed(futs):
                i, res, err_msg = fut.result()
                if err_msg:
                    print(f"[bloco {i}] FALHA REDE: {err_msg}")
                    continue
                ok, err, erros, dt_b = res
                with lock:
                    total_ok += ok; total_err += err
                    progresso["feitos"] += ok + err
                    progresso["ok"] += ok; progresso["err"] += err
                    feitos = progresso["feitos"]
                    print(f"[bloco {i:>3}] ok={ok} err={err} tempo={dt_b}s | acumulado ok={total_ok} err={total_err}")
                    for cpf, cod, desc in erros[:3]:
                        print(f"   ERR {cpf} cod={cod} desc={desc[:80]}")
                    # marco a cada 500
                    if feitos // 500 > progresso["ultimo_marco"] // 500:
                        progresso["ultimo_marco"] = feitos
                        agora = time.time()
                        dt_seg = agora - progresso["t_marco"]
                        cpfs_no_marco = feitos - (progresso["ultimo_marco"] - 500 if progresso["ultimo_marco"] >= 500 else 0)
                        # taxa global
                        dt_total = agora - t_global
                        taxa_global = feitos / dt_total * 60 if dt_total else 0
                        taxa_erro = total_err / max(feitos, 1) * 100
                        eta_min = (len(cpfs) - feitos) / max(taxa_global, 0.01)
                        print(f"\n>>> MARCO {feitos}/{len(cpfs)} | "
                              f"ok={total_ok} err={total_err} "
                              f"taxa_erro={taxa_erro:.2f}% | "
                              f"velocidade={taxa_global:.1f} CPF/min | "
                              f"ETA={eta_min:.1f} min\n")
                        progresso["t_marco"] = agora
    else:
        for idx, bloco in enumerate(batches, 1):
            print(f"\n=== bloco {idx}/{len(batches)} — {len(bloco)} CPFs ===")
            try:
                body = enviar_bloco(bloco, recibos_override, plan_saude_override)
            except requests.exceptions.RequestException as e:
                print(f"FALHA DE REDE no bloco {idx}: {type(e).__name__}: {e}")
                print("PARANDO. Rode de novo depois de resolver.")
                return
            ok, err, erros = resumir(body)
            total_ok += ok; total_err += err
            dt = body.get("_client_elapsed_s", "?")
            print(f"bloco {idx}: ok={ok} err={err} tempo={dt}s")
            for cpf, cod, desc in erros[:10]:
                print(f"  ERR {cpf} cod={cod} desc={desc}")
            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            with open(os.path.join(OUTDIR, f"resp_bloco{idx:03d}_{ts}.json"), "w", encoding="utf-8") as f:
                json.dump(body, f, ensure_ascii=False, indent=2)

    t_dt = time.time() - t_global
    print(f"\n========= TOTAL =========")
    print(f"CPFs enviados: {total_ok + total_err}")
    print(f"  ok:  {total_ok}")
    print(f"  err: {total_err}")
    taxa_erro_final = total_err / max(total_ok+total_err, 1) * 100
    print(f"  taxa_erro: {taxa_erro_final:.2f}%")
    print(f"tempo total: {t_dt:.1f}s ({(total_ok+total_err)/(t_dt/60):.1f} CPFs/min)")


if __name__ == "__main__":
    main()
