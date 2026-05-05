"""
S-1210 Missão APPA — rotas FastAPI

Fonte de escopo: 3 XLSX da Ana em C:\\Users\\xandao\\Downloads
Fonte de dados:  3 ZIPs do eSocial em C:\\Users\\xandao\\Downloads

Endpoints:
- GET  /api/esocial/s1210-missao/fontes              → valida existência dos 6 arquivos
- POST /api/esocial/s1210-missao/carregar            → parseia XLSX+ZIP, retorna resumo 3×4
- GET  /api/esocial/s1210-missao/compartimento/{mes}/{lote} → lista CPFs de um compartimento
- POST /api/esocial/s1210-missao/testar-um-cpf       → retifica 1 CPF em PRODUÇÃO e retorna resultado

⚠️ Regras inegociáveis:
- NUNCA usar explorador_eventos como escopo
- NUNCA enviar sem OK explícito do usuário (esse endpoint exige body.confirmar=True)
- NUNCA mexer em S-1200
"""
from __future__ import annotations

import os
import re
import time
import json
import logging
import zipfile
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, HTTPException
from pydantic import BaseModel
from openpyxl import load_workbook

router = APIRouter(prefix="/api/esocial/s1210-missao", tags=["s1210-missao"])
log = logging.getLogger("s1210-missao")

# ── Constantes da missão ────────────────────────────────────────────
DOWNLOADS = Path(os.environ.get("USERPROFILE", r"C:\Users\xandao")) / "Downloads"

FONTES = {
    "2025-02": {
        "xlsx": DOWNLOADS / "02. Fevereiro_2025_APPA certa.xlsx",
        "zip":  DOWNLOADS / "xmls do e social mes a mes" / "02-fev2025.zip",
        "aba_geral": "Geral Para Envio_Lotes",
        "aba_oper":  "Operadoras_012025",
        "total_lote1": 9472,
    },
    "2025-03": {
        "xlsx": DOWNLOADS / "03. Marco_2025_APPA.xlsx",
        "zip":  DOWNLOADS / "xmls do e social mes a mes" / "03-marc2025.zip",
        "aba_geral": "Geral Para envio de Lotes",
        "aba_oper":  "Operadora 022025",
        "total_lote1": 8165,
    },
    "2025-04": {
        "xlsx": DOWNLOADS / "04. Abril_2025_APPA.xlsx",
        "zip":  DOWNLOADS / "xmls do e social mes a mes" / "04-abril2025.zip",
        "aba_geral": "Geral Envio para Lotes",
        "aba_oper":  "Operadoras 032025",
        "total_lote1": 7142,
    },
    "2025-05": {
        "xlsx": DOWNLOADS / "1º Lote Maio Junho e Julho 2025 com cpf.xlsx",
        "zip":  DOWNLOADS / "xmls do e social mes a mes" / "05-maio.zip",
        "aba_geral": "052025",
        "aba_oper":  "",
        "total_lote1": 8724,
    },
    "2025-06": {
        "xlsx": DOWNLOADS / "07 Julho_lote 003_APPA.xlsx",
        "zip":  DOWNLOADS / "xmls do e social mes a mes" / "06-Jun2025.zip",
        "aba_geral": "Lote para Envio",
        "aba_oper":  "",
        "total_lote1": 0,
    },
    "2025-07": {
        "xlsx": DOWNLOADS / "08 Agosto_lote 003_APPA.xlsx",
        "zip":  DOWNLOADS / "xmls do e social mes a mes" / "07- Jul2025.zip",
        "aba_geral": "Lote para Envio",
        "aba_oper":  "",
        "total_lote1": 0,
    },
    "2025-08": {
        "xlsx": DOWNLOADS / "1º Lote Agosto Setembro e Outubro 2025.xlsx",
        "zip":  DOWNLOADS / "xmls do e social mes a mes" / "08- ago2025.zip",
        "aba_geral": "AGOSTO ",  # aba tem espaco no fim no XLSX original
        "aba_oper":  "",
        "total_lote1": 5095,
    },
    "2025-09": {
        "xlsx": DOWNLOADS / "1º Lote Agosto Setembro e Outubro 2025.xlsx",
        "zip":  DOWNLOADS / "xmls do e social mes a mes" / "09-set2025.zip",
        "aba_geral": "SETEMBRO",
        "aba_oper":  "",
        "total_lote1": 5847,
    },
    "2025-10": {
        "xlsx": DOWNLOADS / "1º Lote Agosto Setembro e Outubro 2025.xlsx",
        "zip":  DOWNLOADS / "xmls do e social mes a mes" / "10-out2025.zip",
        "aba_geral": "OUTUBRO",
        "aba_oper":  "",
        "total_lote1": 5395,
    },
    "2025-11": {
        "xlsx": DOWNLOADS / "1º Lote novembro 2025.xlsx",
        "zip":  DOWNLOADS / "xmls do e social mes a mes" / "11-nov2025.zip",
        "aba_geral": "Planilha1",
        "aba_oper":  "",
        "total_lote1": 4873,
        "col_lote": 7,
        "col_cpf": 8,
    },
    "2025-12": {
        "xlsx": DOWNLOADS / "1º Lote Dezembro 2025.xlsx",
        "zip":  DOWNLOADS / "xmls do e social mes a mes" / "12-dez2025.zip",
        "aba_geral": "Planilha1",
        "aba_oper":  "",
        "total_lote1": 5083,
        "col_lote": 0,   # col A = '1º Lote'
        "col_cpf": 8,    # col I = CPF
    },
}

# Cache em memória (evita reparsear XLSX/ZIP a cada request)
# Chave: (mes, lote) → lista de CPFs
_CACHE_XLSX: dict[str, dict[str, list[str]]] = {}
# Chave: (mes, cpf) → dict com nr_recibo + dados do S-1210 original
_CACHE_RECIBOS: dict[str, dict[str, dict]] = {}


# ── Helpers — XLSX ──────────────────────────────────────────────────

def _norm_lote(v) -> str:
    """Normaliza '1º Lote' / '1º' / '1' → '1_LOTE'. Retorna '' se não for 1-4."""
    s = str(v or "").strip()
    for ch in s:
        if ch.isdigit() and ch in "1234":
            return f"{ch}_LOTE"
    return ""


def _parse_xlsx_escopo(mes: str) -> dict[str, list[str]]:
    """
    Parseia XLSX da Ana, aba 'Geral…', e retorna dict lote→lista_cpfs.
    Valida total do Lote 1 contra FONTES[mes]['total_lote1'].
    """
    fonte = FONTES[mes]
    path = fonte["xlsx"]
    if not path.exists():
        raise FileNotFoundError(f"XLSX não encontrado: {path}")

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        aba = fonte["aba_geral"]
        if aba not in wb.sheetnames:
            raise ValueError(
                f"Aba '{aba}' não encontrada em {path.name}. "
                f"Abas existentes: {wb.sheetnames}"
            )
        ws = wb[aba]
        lotes: dict[str, list[str]] = {"1_LOTE": [], "2_LOTE": [], "3_LOTE": [], "4_LOTE": []}
        vistos: set[str] = set()  # CPF deduplica por arquivo todo

        # Permite override de colunas em FONTES (ex: novembro tem 1 col extra)
        col_lote = fonte.get("col_lote", 6)
        col_cpf = fonte.get("col_cpf", 7)
        min_cols = max(col_lote, col_cpf) + 1

        it = ws.iter_rows(values_only=True)
        next(it, None)  # pula header

        for row in it:
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            if len(row) < min_cols:
                continue
            lote_key = _norm_lote(row[col_lote])
            if not lote_key:
                continue
            cpf_raw = str(row[col_cpf] or "")
            cpf = "".join(ch for ch in cpf_raw if ch.isdigit())
            if len(cpf) == 0:
                continue
            cpf = cpf.zfill(11)
            if len(cpf) != 11:
                continue
            chave = f"{lote_key}:{cpf}"
            if chave in vistos:
                continue
            vistos.add(chave)
            lotes[lote_key].append(cpf)
    finally:
        wb.close()

    # Validação (soft): avisa no log mas não aborta se divergir em ±1
    total_esperado = fonte["total_lote1"]
    total_real = len(lotes["1_LOTE"])
    if total_real != total_esperado:
        log.warning(
            f"[{mes}] Lote 1 parser={total_real} esperado={total_esperado} "
            f"(diff={total_real - total_esperado})"
        )

    return lotes


# ── Helpers — ZIP (busca S-1210 e extrai dados) ─────────────────────

_RE_CPF       = re.compile(r"<cpfBenef>(\d+)</cpfBenef>")
_RE_PER_APUR  = re.compile(r"<perApur>(\d{4}-\d{2})</perApur>")
_RE_IND_RETIF = re.compile(r"<indRetif>(\d)</indRetif>")
_RE_NR_RECIBO_RETORNO = re.compile(
    r"</evento>(.*?)</retornoProcessamentoDownload>", re.DOTALL
)
_RE_NR_RECIBO_TAG = re.compile(r"<nrRecibo[^>]*>([^<]+)</nrRecibo>")
_RE_DH_PROC = re.compile(r"<dhProcessamento[^>]*>([^<]+)</dhProcessamento>")
_RE_INFO_PGTO = re.compile(r"<infoPgto>(.*?)</infoPgto>", re.DOTALL)
_RE_PGTO_CAMPO = {
    "dtPgto":   re.compile(r"<dtPgto[^>]*>([^<]+)</dtPgto>"),
    "tpPgto":   re.compile(r"<tpPgto[^>]*>([^<]+)</tpPgto>"),
    "perRef":   re.compile(r"<perRef[^>]*>([^<]+)</perRef>"),
    "ideDmDev": re.compile(r"<ideDmDev[^>]*>([^<]+)</ideDmDev>"),
    "vrLiq":    re.compile(r"<vrLiq[^>]*>([^<]+)</vrLiq>"),
}
_RE_INFO_IR_CR = re.compile(r"<infoIRCR>(.*?)</infoIRCR>", re.DOTALL)
_RE_IR_TP_CR = re.compile(r"<tpCR[^>]*>([^<]+)</tpCR>")
_RE_IR_VR_CR = re.compile(r"<vrCR[^>]*>([^<]+)</vrCR>")
# Blocos aninhados dentro de infoIRCR — precisam ser replicados na retificação
# senão o eSocial rejeita com "Grupo 'xxx' deve ser preenchido" (cod 8).
_RE_DED_DEPEN = re.compile(r"<dedDepen>(.*?)</dedDepen>", re.DOTALL)
_RE_PEN_ALIM  = re.compile(r"<penAlim>(.*?)</penAlim>",  re.DOTALL)
_RE_DEP_TPR   = re.compile(r"<tpRend[^>]*>([^<]+)</tpRend>")
_RE_DEP_CPF   = re.compile(r"<cpfDep[^>]*>([^<]+)</cpfDep>")
_RE_DEP_VLR   = re.compile(r"<vlrDedDep[^>]*>([^<]+)</vlrDedDep>")
_RE_PEN_VLR   = re.compile(r"<vlrDedPenAlim[^>]*>([^<]+)</vlrDedPenAlim>")


def _extract_s1210_from_xml(xml_str: str) -> Optional[dict]:
    """
    Extrai de um XML S-1210 (de retornoProcessamentoDownload) os campos necessários
    para montar uma retificação. Retorna None se não for S-1210 válido com recibo.
    """
    if "evtPgtos" not in xml_str:
        return None

    cpf_m = _RE_CPF.search(xml_str)
    per_m = _RE_PER_APUR.search(xml_str)
    if not cpf_m or not per_m:
        return None

    # nrRecibo do PRÓPRIO evento fica dentro de <retornoProcessamentoDownload>
    # (após o </evento>). Se não tiver, o evento não foi aceito.
    retorno_m = _RE_NR_RECIBO_RETORNO.search(xml_str)
    if not retorno_m:
        return None
    retorno_section = retorno_m.group(1)
    rec_m = _RE_NR_RECIBO_TAG.search(retorno_section)
    if not rec_m:
        return None
    nr_recibo = rec_m.group(1).strip()

    # dhProcessamento (para escolher o mais recente)
    dh_m = _RE_DH_PROC.search(retorno_section) or _RE_DH_PROC.search(xml_str)
    dh_proc = dh_m.group(1).strip() if dh_m else ""

    # indRetif do evento
    ind_m = _RE_IND_RETIF.search(xml_str)
    ind_retif = ind_m.group(1) if ind_m else "1"

    # infoPgto[]
    pgtos = []
    for pgto_block in _RE_INFO_PGTO.findall(xml_str):
        d = {}
        for campo, regex in _RE_PGTO_CAMPO.items():
            m = regex.search(pgto_block)
            if m:
                d[campo] = m.group(1).strip()
        if d.get("dtPgto") and d.get("tpPgto") and d.get("ideDmDev") and d.get("vrLiq"):
            pgtos.append(d)

    # infoIRCR (pode ter múltiplos; replicamos todos os blocos aninhados)
    info_ir = []
    for ircr_block in _RE_INFO_IR_CR.findall(xml_str):
        tp_m = _RE_IR_TP_CR.search(ircr_block)
        vr_m = _RE_IR_VR_CR.search(ircr_block)
        if tp_m:
            d = {"tpCR": tp_m.group(1).strip()}
            if vr_m:
                d["vrCR"] = vr_m.group(1).strip()
            # dedDepen[] — dedução de dependente para IR
            deps = []
            for dep_block in _RE_DED_DEPEN.findall(ircr_block):
                tpr = _RE_DEP_TPR.search(dep_block)
                cpfd = _RE_DEP_CPF.search(dep_block)
                vlr = _RE_DEP_VLR.search(dep_block)
                if tpr and cpfd and vlr:
                    deps.append({
                        "tpRend": tpr.group(1).strip(),
                        "cpfDep": cpfd.group(1).strip(),
                        "vlrDedDep": vlr.group(1).strip(),
                    })
            if deps:
                d["dedDepen"] = deps
            # penAlim[] — pensão alimentícia (um bloco por beneficiário × tpRend)
            pens = []
            for pen_block in _RE_PEN_ALIM.findall(ircr_block):
                tpr = _RE_DEP_TPR.search(pen_block)
                cpfd = _RE_DEP_CPF.search(pen_block)
                vlr = _RE_PEN_VLR.search(pen_block)
                if tpr and cpfd and vlr:
                    pens.append({
                        "tpRend": tpr.group(1).strip(),
                        "cpfDep": cpfd.group(1).strip(),
                        "vlrDedPenAlim": vlr.group(1).strip(),
                    })
            if pens:
                d["penAlim"] = pens
            info_ir.append(d)

    return {
        "cpf": cpf_m.group(1).zfill(11),
        "per_apur": per_m.group(1),
        "ind_retif": ind_retif,
        "nr_recibo": nr_recibo,
        "dh_proc": dh_proc,
        "info_pgtos": pgtos,
        "info_ir_cr": info_ir,
    }


def _indexar_zip(mes: str) -> dict[str, dict]:
    """
    Percorre o ZIP do mês e indexa S-1210 por CPF, escolhendo o recibo
    com dhProcessamento mais recente.
    Retorna dict: cpf → {nr_recibo, per_apur, info_pgtos, info_ir_cr, dh_proc, ...}
    """
    fonte = FONTES[mes]
    path = fonte["zip"]
    if not path.exists():
        raise FileNotFoundError(f"ZIP não encontrado: {path}")

    idx: dict[str, dict] = {}
    with zipfile.ZipFile(str(path), "r") as zf:
        names = [n for n in zf.namelist() if "S-1210" in n and n.endswith(".xml")]
        log.info(f"[{mes}] ZIP tem {len(names)} S-1210 XMLs")
        for i, name in enumerate(names):
            if i % 5000 == 0 and i > 0:
                log.info(f"[{mes}] indexando {i}/{len(names)}...")
            try:
                raw = zf.read(name).decode("utf-8", errors="replace")
            except Exception as e:
                log.warning(f"[{mes}] falha lendo {name}: {e}")
                continue
            info = _extract_s1210_from_xml(raw)
            if not info:
                continue
            cpf = info["cpf"]
            atual = idx.get(cpf)
            # Escolhe o mais recente por dh_proc
            if (atual is None) or (info["dh_proc"] > atual["dh_proc"]):
                idx[cpf] = info
    log.info(f"[{mes}] indexou {len(idx)} CPFs com S-1210 + nrRecibo")
    return idx


# ── Endpoints ───────────────────────────────────────────────────────

@router.get("/fontes")
def get_fontes():
    """Valida existência dos 6 arquivos fonte."""
    out = []
    for mes, fonte in FONTES.items():
        out.append({
            "mes": mes,
            "xlsx_nome": fonte["xlsx"].name,
            "xlsx_ok": fonte["xlsx"].exists(),
            "xlsx_mb": round(fonte["xlsx"].stat().st_size / 1024 / 1024, 1) if fonte["xlsx"].exists() else 0,
            "zip_nome": fonte["zip"].name,
            "zip_ok": fonte["zip"].exists(),
            "zip_mb": round(fonte["zip"].stat().st_size / 1024 / 1024, 1) if fonte["zip"].exists() else 0,
            "total_lote1_esperado": fonte["total_lote1"],
        })
    return {"downloads": str(DOWNLOADS), "fontes": out}


@router.post("/carregar")
def carregar(indexar_zips: bool = False, forcar: bool = False):
    """
    Parseia os 3 XLSX e retorna totais por mes × lote.
    - Usa cache em memória: 2ª chamada é instantânea (a menos que forcar=True).
    - Se indexar_zips=True, também processa os 3 ZIPs (demora ~10-30 min total).
    """
    resumo = []
    for mes in FONTES.keys():
        try:
            if forcar or mes not in _CACHE_XLSX:
                lotes = _parse_xlsx_escopo(mes)
                _CACHE_XLSX[mes] = lotes
            else:
                lotes = _CACHE_XLSX[mes]
        except Exception as e:
            log.exception(f"Erro parseando XLSX {mes}")
            return {"erro": f"XLSX {mes}: {e}"}

        mes_resumo = {"mes": mes, "lotes": {}}
        # Conta ja_feito/erro por lote juntando pipeline_runs(per_apur) x pipeline_cpf_results
        contagens = _contar_processados(mes, lotes)
        for lote_key, cpfs in lotes.items():
            feito = contagens[lote_key]["ok"]
            err = contagens[lote_key]["erro"]
            mes_resumo["lotes"][lote_key] = {
                "total": len(cpfs),
                "pendente": max(0, len(cpfs) - feito - err),
                "ja_feito": feito,
                "erro": err,
            }
        resumo.append(mes_resumo)

    # Opcional: indexar ZIPs e cruzar com escopo pra marcar ja_feito
    if indexar_zips:
        for mes in FONTES.keys():
            try:
                idx = _indexar_zip(mes)
                _CACHE_RECIBOS[mes] = idx
            except Exception as e:
                log.exception(f"Erro indexando ZIP {mes}")
                return {"erro": f"ZIP {mes}: {e}"}
            # cruzamento: quantos CPFs do escopo têm recibo no ZIP
            lotes = _CACHE_XLSX[mes]
            for item in resumo:
                if item["mes"] != mes:
                    continue
                for lote_key, cpfs in lotes.items():
                    com_recibo = sum(1 for c in cpfs if c in idx)
                    item["lotes"][lote_key]["com_recibo_no_zip"] = com_recibo

    return {"resumo": resumo, "indexado_zip": indexar_zips}


@router.get("/compartimento/{mes}/{lote}")
def compartimento(mes: str, lote: str, limit: int = 50, offset: int = 0):
    """Lista CPFs de um compartimento (mes × lote)."""
    if mes not in FONTES:
        raise HTTPException(400, f"mes inválido: {mes}")
    lote_key = _norm_lote(lote) or lote
    if lote_key not in ("1_LOTE", "2_LOTE", "3_LOTE", "4_LOTE"):
        raise HTTPException(400, f"lote inválido: {lote}")

    if mes not in _CACHE_XLSX:
        _CACHE_XLSX[mes] = _parse_xlsx_escopo(mes)

    cpfs = _CACHE_XLSX[mes][lote_key]
    total = len(cpfs)
    idx = _CACHE_RECIBOS.get(mes, {})

    items = []
    for cpf in cpfs[offset:offset + limit]:
        rec_info = idx.get(cpf)
        items.append({
            "cpf": cpf,
            "nr_recibo_original": rec_info["nr_recibo"] if rec_info else None,
            "tem_s1210_no_zip": rec_info is not None,
        })
    return {"mes": mes, "lote": lote_key, "total": total, "offset": offset, "items": items}


# ── Envio em produção ───────────────────────────────────────────────

class TestarUmCpfReq(BaseModel):
    mes: str                    # "2025-02" | "2025-03" | "2025-04"
    lote: str = "1"             # "1" | "2" | "3" | "4"
    indice: int = 0             # qual CPF da lista (0 = primeiro)
    confirmar_producao: bool    # deve ser True explicitamente
    nr_recibo_override: Optional[str] = None  # sobrescreve o recibo extraído do ZIP (ex: recibo ativo achado no banco)


@router.post("/testar-um-cpf")
def testar_um_cpf(req: TestarUmCpfReq):
    """
    Pega 1 CPF do compartimento (mes × lote), busca S-1210 original no ZIP,
    monta retif sem plano de saúde (lote 1) e envia em PRODUÇÃO.

    ⚠️ Só executa se req.confirmar_producao=True.
    """
    resultado = _testar_um_cpf_impl(req)
    # persiste o resultado pra aparecer no painel (best-effort)
    lote_key = _norm_lote(req.lote) or req.lote
    lote_num = int(lote_key[0]) if lote_key and lote_key[0].isdigit() else 0
    # anexa snapshots se tiver s1210 no resultado (só o sucesso tem isso — mas tudo bem)
    if resultado.get("cpf") and not resultado.get("pulado"):
        try:
            _persistir_cpf_result(req.mes, lote_num, resultado)
        except Exception as e:
            log.warning(f"persist testar-um-cpf falhou: {e}")
    return resultado


def _testar_um_cpf_impl(req: TestarUmCpfReq):
    if not req.confirmar_producao:
        raise HTTPException(400, "confirmar_producao=true é obrigatório para envio em produção")
    if req.mes not in FONTES:
        raise HTTPException(400, f"mes inválido: {req.mes}")

    lote_key = _norm_lote(req.lote) or req.lote
    if lote_key not in ("1_LOTE", "2_LOTE", "3_LOTE", "4_LOTE"):
        raise HTTPException(400, f"lote inválido: {req.lote}")
    lote_num = int(lote_key[0]) if lote_key and lote_key[0].isdigit() else 0

    # 1) Escopo — garantir XLSX parseada
    if req.mes not in _CACHE_XLSX:
        try:
            _CACHE_XLSX[req.mes] = _parse_xlsx_escopo(req.mes)
        except Exception as e:
            raise HTTPException(500, f"Erro parseando XLSX: {e}")

    cpfs = _CACHE_XLSX[req.mes][lote_key]
    if req.indice >= len(cpfs):
        raise HTTPException(400, f"indice {req.indice} fora do range (lote tem {len(cpfs)} CPFs)")
    cpf_alvo = cpfs[req.indice]
    log.info(f"🎯 CPF alvo: {cpf_alvo} ({req.mes} lote {lote_key} indice {req.indice})")

    # 2) Buscar S-1210 original no ZIP (SÓ desse CPF — rápido)
    s1210_original = _buscar_s1210_unico(req.mes, cpf_alvo)
    if not s1210_original:
        return {
            "sucesso": False,
            "etapa": "buscar_recibo",
            "cpf": cpf_alvo,
            "mes": req.mes,
            "lote": lote_key,
            "erro": f"Nenhum S-1210 com nrRecibo encontrado no ZIP para CPF {cpf_alvo}",
        }

    recibo_usado = s1210_original['nr_recibo']
    recibo_zip = s1210_original['nr_recibo']
    if req.nr_recibo_override:
        recibo_usado = req.nr_recibo_override.strip()
        log.info(f"🔄 Override de recibo: ZIP={recibo_zip} → usando={recibo_usado}")
    log.info(f"📎 Recibo a enviar: {recibo_usado} "
             f"(perApur={s1210_original['per_apur']}, pgtos={len(s1210_original['info_pgtos'])})")

    # 3) Carregar certificado A1
    try:
        cnpj, pfx_data, senha = _load_cert_ativo()
    except Exception as e:
        raise HTTPException(500, f"Erro carregando certificado ativo: {e}")

    # 4) Montar XML retif (LOTE 1 = SEM planSaude)
    from esocial.xml_s1210 import S1210XMLGenerator
    from esocial.xml_signer import S1010XMLSigner as XMLSigner
    from esocial.soap_builder import SOAPEnvelopeBuilder
    from esocial.esocial_client import ESocialClient

    empregador = {"tpInsc": 1, "nrInsc": cnpj}

    # info_ir_complem: usa o primeiro tpCR do original (ou None se não houver)
    info_ir_complem = None
    if s1210_original["info_ir_cr"]:
        info_ir_complem = {"infoIRCR": s1210_original["info_ir_cr"]}

    # Lote 3 SEMPRE é retificação (indRetif=2)
    # Lote 1/2/4: usa do original ou força 2 se erro de duplicidade
    ind_retif_usar = "2"  # Lote 3 MISSAO sempre retif
    nr_recibo_para_retif = recibo_usado

    # Montar plan_saude apenas para Lote 2/3 (com plano de saúde)
    plan_saude_usar = None
    if lote_key in ("2_LOTE", "3_LOTE"):
        # Busca plan_saude agregado por CNPJ na tabela s1210_operadoras
        try:
            plan_saude_usar = _buscar_plan_saude_por_cpf(cpf_alvo, s1210_original['per_apur'], lote_num)
        except Exception as e:
            log.warning(f"Erro buscando plan_saude para {cpf_alvo}: {e}")
            plan_saude_usar = None

    try:
        xml_bytes = S1210XMLGenerator.gerar(
            empregador=empregador,
            beneficiario={"cpfBenef": cpf_alvo},
            info_pgtos=s1210_original["info_pgtos"],
            per_apur=s1210_original["per_apur"],
            ind_retif=ind_retif_usar,
            nr_recibo=nr_recibo_para_retif,
            info_ir_complem=info_ir_complem,
            plan_saude=plan_saude_usar,
            tp_amb="1",       # PRODUÇÃO
        )
    except Exception as e:
        return {
            "sucesso": False,
            "etapa": "gerar_xml",
            "cpf": cpf_alvo,
            "erro": str(e),
            "nr_recibo_original": s1210_original["nr_recibo"],
        }

    # 5) Assinar + Envelopar + Enviar PROD + Pollar
    try:
        xml_assinado = XMLSigner.assinar(xml_bytes, pfx_data, senha)
    except Exception as e:
        return {
            "sucesso": False, "etapa": "assinar_xml", "cpf": cpf_alvo,
            "erro": str(e), "nr_recibo_original": s1210_original["nr_recibo"],
        }

    try:
        soap = SOAPEnvelopeBuilder.montar_envio(
            [xml_assinado], empregador, empregador.copy(), grupo="3"  # periódicos
        )
        url_envio = SOAPEnvelopeBuilder.url_envio(producao=True)
        resultado = ESocialClient.enviar_lote(soap, pfx_data, senha, url=url_envio)
    except Exception as e:
        return {
            "sucesso": False, "etapa": "enviar_soap", "cpf": cpf_alvo,
            "erro": str(e), "nr_recibo_original": s1210_original["nr_recibo"],
        }

    if not resultado.get("sucesso"):
        return {
            "sucesso": False,
            "etapa": "envio_rejeitado",
            "cpf": cpf_alvo,
            "mes": req.mes,
            "lote": lote_key,
            "codigo_resposta_envio": resultado.get("codigo_resposta"),
            "descricao_envio": resultado.get("descricao"),
            "erro": resultado.get("erro") or resultado.get("descricao"),
            "nr_recibo_original": s1210_original["nr_recibo"],
        }

    protocolo = resultado.get("protocolo")
    log.info(f"✅ Aceito p/ processamento. Protocolo: {protocolo}")

    # 6) Pollar consulta
    try:
        url_consulta = SOAPEnvelopeBuilder.url_consulta(producao=True)
        for attempt in range(15):
            time.sleep(5)
            consulta = ESocialClient.consultar_lote(
                protocolo, pfx_data, senha, url=url_consulta
            )
            if consulta.get("eventos"):
                evt = consulta["eventos"][0]
                nr_recibo_novo = evt.get("nr_recibo")
                codigo = evt.get("codigo_resposta", "?")
                descricao = evt.get("descricao", "")
                ocorr = evt.get("ocorrencias", []) or []

                if nr_recibo_novo:
                    return {
                        "sucesso": True,
                        "etapa": "processado",
                        "cpf": cpf_alvo,
                        "mes": req.mes,
                        "lote": lote_key,
                        "protocolo": protocolo,
                        "nr_recibo_zip": recibo_zip,
                        "nr_recibo_usado": recibo_usado,
                        "nr_recibo_original": recibo_usado,
                        "nr_recibo_novo": nr_recibo_novo,
                        "codigo_resposta": codigo,
                        "descricao": descricao,
                        "ocorrencias": ocorr,
                    }
                # processado mas sem recibo = rejeitado
                return {
                    "sucesso": False,
                    "etapa": "processamento_rejeitado",
                    "cpf": cpf_alvo,
                    "mes": req.mes,
                    "lote": lote_key,
                    "protocolo": protocolo,
                    "nr_recibo_zip": recibo_zip,
                    "nr_recibo_usado": recibo_usado,
                    "nr_recibo_original": recibo_usado,
                    "codigo_resposta": codigo,
                    "descricao": descricao,
                    "ocorrencias": ocorr,
                    "erro": f"Código {codigo}: {descricao}",
                }
            elif consulta.get("codigo_resposta") == "101":
                log.info(f"⏳ Ainda processando ({attempt+1}/15)...")
                continue
    except Exception as e:
        return {
            "sucesso": False,
            "etapa": "consulta",
            "cpf": cpf_alvo,
            "protocolo": protocolo,
            "nr_recibo_original": s1210_original["nr_recibo"],
            "erro": str(e),
        }

    return {
        "sucesso": False,
        "etapa": "timeout",
        "cpf": cpf_alvo,
        "protocolo": protocolo,
        "nr_recibo_original": s1210_original["nr_recibo"],
        "erro": "Timeout após 15 tentativas de consulta",
    }


# ── Helpers internos de envio ───────────────────────────────────────

def _buscar_s1210_unico(mes: str, cpf: str) -> Optional[dict]:
    """
    Percorre o ZIP do mês procurando S-1210 de UM CPF específico.
    Retorna o mais recente (dhProcessamento) ou None.
    Otimizado: para ao encontrar todos os matches do CPF (no filename tem o nrInsc mas não CPF).
    """
    # Se já indexado em cache, usa
    if mes in _CACHE_RECIBOS:
        return _CACHE_RECIBOS[mes].get(cpf)

    fonte = FONTES[mes]
    path = fonte["zip"]
    if not path.exists():
        raise FileNotFoundError(f"ZIP não encontrado: {path}")

    best: Optional[dict] = None
    with zipfile.ZipFile(str(path), "r") as zf:
        names = [n for n in zf.namelist() if "S-1210" in n and n.endswith(".xml")]
        log.info(f"[{mes}] buscando CPF {cpf} em {len(names)} S-1210 XMLs...")
        for name in names:
            try:
                raw = zf.read(name).decode("utf-8", errors="replace")
            except Exception:
                continue
            # Filtro rápido: CPF precisa aparecer no XML
            if cpf not in raw:
                continue
            info = _extract_s1210_from_xml(raw)
            if not info or info["cpf"] != cpf:
                continue
            if (best is None) or (info["dh_proc"] > best["dh_proc"]):
                best = info
    return best


def _load_cert_ativo():
    """Carrega certificado A1 ativo do DB LOCAL. Retorna (cnpj, pfx_bytes, senha_plain)."""
    import sys as _sys
    import os as _os
    # Garantir que db_config seja importável
    base = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
    if base not in _sys.path:
        _sys.path.insert(0, base)
    from db_config import LOCAL_DB_CONFIG
    import psycopg2

    from esocial.certificate_manager import CertificateManager

    conn = psycopg2.connect(**LOCAL_DB_CONFIG)
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT cnpj, arquivo_path, senha_encrypted "
                "FROM certificados_a1 WHERE ativo = TRUE LIMIT 1"
            )
            row = cur.fetchone()
    finally:
        conn.close()
    if not row:
        raise RuntimeError("Nenhum certificado A1 ativo no DB LOCAL (tabela certificados_a1)")
    cnpj, cert_path, senha_enc = row
    senha = CertificateManager.decrypt_password(senha_enc)
    with open(cert_path, "rb") as f:
        pfx_data = f.read()
    return cnpj, pfx_data, senha


def _buscar_plan_saude_por_cpf(cpf: str, per_apur: str, lote_num: int) -> Optional[list[dict]]:
    """
    Busca plan_saude agregado por CNPJ para um CPF em um lote.
    Retorna lista de dicts: [{cnpjOper, regANS, vlrSaudeTit}, ...]
    ou None se não houver dados.
    
    Usada para Lote 2/3 que têm plano de saúde com operadora.
    """
    import sys as _sys
    import os as _os
    base = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
    if base not in _sys.path:
        _sys.path.insert(0, base)
    from db_config import DB_CONFIG
    import psycopg2
    import psycopg2.extras

    conn = psycopg2.connect(**DB_CONFIG)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute("""
                SELECT cnpj_operadora,
                       MAX(reg_ans) AS reg_ans,
                       SUM(valor)::BIGINT AS soma_centavos
                  FROM s1210_operadoras
                 WHERE per_apur   = %s
                   AND cpf        = %s
                   AND lote_num   = %s
              GROUP BY cnpj_operadora
              ORDER BY cnpj_operadora
            """, (per_apur, cpf, lote_num))
            rows = cur.fetchall()
    finally:
        conn.close()

    if not rows:
        return None

    result = []
    for r in rows:
        centavos = int(r["soma_centavos"] or 0)
        if centavos <= 0:
            continue
        vlr = f"{centavos / 100:.2f}"
        result.append({
            "cnpjOper": r["cnpj_operadora"],
            "regANS": r["reg_ans"] or "",
            "vlrSaudeTit": vlr,
        })
    
    return result if result else None


# ── Helpers de persistência na pipeline ─────────────────────────────

def _db_connect():
    """Abre conexão com Supabase (DB_CONFIG)."""
    import sys as _sys
    import os as _os
    base = _os.path.dirname(_os.path.dirname(_os.path.abspath(__file__)))
    if base not in _sys.path:
        _sys.path.insert(0, base)
    from db_config import DB_CONFIG
    import psycopg2
    return psycopg2.connect(**DB_CONFIG)


def _ensure_run_id(per_apur: str, lote_num: int) -> int:
    """
    Garante um pipeline_runs para esta missão e retorna o id (INT).
    Usa o MESMO run por (per_apur, lote_num) — não cria um novo a cada clique,
    assim o painel agrega todos os envios da missão.
    """
    conn = _db_connect()
    try:
        with conn.cursor() as cur:
            # Procura run existente DESTA missão (após cutoff)
            cur.execute(
                """
                SELECT id FROM pipeline_runs
                 WHERE per_apur = %s AND lote_atual = %s
                   AND started_at >= %s::timestamptz
                 ORDER BY started_at DESC LIMIT 1
                """,
                (per_apur, lote_num, MISSAO_CUTOFF_UTC),
            )
            row = cur.fetchone()
            if row:
                return row[0]
            # Cria novo
            cur.execute(
                """
                INSERT INTO pipeline_runs
                  (per_apur, status, total_cpfs, cpfs_ok, cpfs_erro, cpfs_ignorados,
                   s1298_done, s1299_done, lote_atual, total_lotes, started_at)
                VALUES (%s, 'rodando', 0, 0, 0, 0, false, false, %s, 4, NOW())
                RETURNING id
                """,
                (per_apur, lote_num),
            )
            new_id = cur.fetchone()[0]
        conn.commit()
        return new_id
    finally:
        conn.close()


def _persistir_cpf_result(per_apur: str, lote_num: int, resultado: dict) -> None:
    """Persiste 1 CPF em pipeline_cpf_results (best-effort, não aborta o fluxo)."""
    try:
        run_id = _ensure_run_id(per_apur, lote_num)
    except Exception as e:
        log.warning(f"⚠️  falha criando/achando pipeline_runs: {e}")
        return

    cpf = resultado.get("cpf") or "?"
    status = "ok" if resultado.get("sucesso") else "erro"
    nr_orig = resultado.get("nr_recibo_usado") or resultado.get("nr_recibo_original")
    nr_novo = resultado.get("nr_recibo_novo")
    erro = resultado.get("erro") or resultado.get("descricao")
    pagamentos = resultado.get("pagamentos_snapshot")
    info_ir = resultado.get("info_ir_snapshot")

    try:
        conn = _db_connect()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO pipeline_cpf_results
                      (run_id, cpf, status, nr_recibo_original, nr_recibo_novo,
                       pagamentos, info_ir_cr, erro_descricao, lote_num, processed_at)
                    VALUES (%s, %s, %s, %s, %s, %s::jsonb, %s::jsonb, %s, %s, NOW())
                    """,
                    (
                        run_id,
                        cpf,
                        status,
                        nr_orig,
                        nr_novo,
                        json.dumps(pagamentos) if pagamentos else None,
                        json.dumps(info_ir) if info_ir else None,
                        erro,
                        lote_num,
                    ),
                )
                # atualiza contadores no pipeline_runs
                if status == "ok":
                    cur.execute(
                        "UPDATE pipeline_runs SET cpfs_ok = COALESCE(cpfs_ok,0)+1 WHERE id=%s",
                        (run_id,),
                    )
                else:
                    cur.execute(
                        "UPDATE pipeline_runs SET cpfs_erro = COALESCE(cpfs_erro,0)+1 WHERE id=%s",
                        (run_id,),
                    )
                # Espelha em s1210_cpf_envios — fonte da v_s1210_contadores
                # (view lê "último envio por CPF", então o status aqui alimenta
                # o painel "Por Lote" / "Compartimento" do repositório).
                try:
                    cur.execute(
                        """
                        INSERT INTO s1210_cpf_envios
                          (empresa_id, per_apur, cpf, lote_num, status,
                           nr_recibo_usado, nr_recibo_novo, protocolo,
                           codigo_resposta, descricao_resposta, erro_descricao,
                           pagamentos, info_ir, duracao_ms)
                        VALUES (1, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s,
                                %s::jsonb, %s::jsonb, %s)
                        """,
                        (
                            per_apur, cpf, lote_num, status,
                            nr_orig, nr_novo,
                            resultado.get("protocolo"),
                            resultado.get("codigo_resposta"),
                            resultado.get("descricao"),
                            erro,
                            json.dumps(pagamentos) if pagamentos else None,
                            json.dumps(info_ir) if info_ir else None,
                            resultado.get("t_ms"),
                        ),
                    )
                except Exception as e:
                    log.warning(f"⚠️  falha insert s1210_cpf_envios: {e}")
            conn.commit()
        finally:
            conn.close()
    except Exception as e:
        log.warning(f"⚠️  falha persist pipeline_cpf_results: {e}")


MISSAO_CUTOFF_UTC = "2026-04-21 10:00:00+00"
# ⚠️ Runs antes desse cutoff são HISTÓRICO (GPT anterior, escopo errado vindo de
# explorador_eventos, bug [106] mascarado). Não apagar — só não mostrar na tela
# da missão nova (call 20/04/2026).


def _contar_processados(mes: str, lotes_cpfs: dict[str, list[str]]) -> dict[str, dict]:
    out = {lk: {"ok": 0, "erro": 0} for lk in lotes_cpfs.keys()}
    try:
        conn = _db_connect()
    except Exception as e:
        log.warning(f"⚠️  _contar_processados DB erro: {e}")
        return out
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT DISTINCT ON (r.cpf) r.cpf, r.status
                  FROM pipeline_cpf_results r
                  JOIN pipeline_runs pr ON pr.id = r.run_id
                 WHERE pr.per_apur = %s
                   AND r.status IN ('ok','erro')
                   AND r.processed_at >= %s::timestamptz
                 ORDER BY r.cpf, r.processed_at DESC
                """,
                (mes, MISSAO_CUTOFF_UTC),
            )
            done: dict[str, str] = {c: s for c, s in cur.fetchall()}
    finally:
        conn.close()

    for lote_key, cpfs in lotes_cpfs.items():
        for cpf in cpfs:
            st = done.get(cpf)
            if st == "ok":
                out[lote_key]["ok"] += 1
            elif st == "erro":
                out[lote_key]["erro"] += 1
    return out




class BatchStartReq(BaseModel):
    mes: str
    lote: str = "1"
    offset: int = 0
    limit: Optional[int] = None
    confirmar_producao: bool = False


@router.post("/batch/start")
def batch_start(req: BatchStartReq):
    """Inicia envio em LOTE em PRODUÇÃO."""
    if not req.confirmar_producao:
        raise HTTPException(400, "confirmar_producao=true é obrigatório")
    if req.mes not in FONTES:
        raise HTTPException(400, f"mes inválido: {req.mes}")
    lote_key = _norm_lote(req.lote) or req.lote
    if lote_key not in ("1_LOTE", "2_LOTE", "3_LOTE", "4_LOTE"):
        raise HTTPException(400, f"lote inválido: {req.lote}")

    if req.mes not in _CACHE_XLSX:
        try:
            _CACHE_XLSX[req.mes] = _parse_xlsx_escopo(req.mes)
        except Exception as e:
            raise HTTPException(500, f"Erro parseando XLSX: {e}")

    cpfs = _CACHE_XLSX[req.mes][lote_key]
    if req.offset >= len(cpfs):
        raise HTTPException(400, f"offset {req.offset} fora do range ({len(cpfs)} CPFs)")

    from esocial import s1210_batch
    try:
        state = s1210_batch.start_batch(req.mes, lote_key, cpfs, req.offset, req.limit)
    except RuntimeError as e:
        raise HTTPException(409, str(e))
    return state


@router.post("/batch/pause")
def batch_pause():
    from esocial import s1210_batch
    try:
        return s1210_batch.pause_batch()
    except RuntimeError as e:
        raise HTTPException(409, str(e))


@router.post("/batch/resume")
def batch_resume():
    from esocial import s1210_batch
    try:
        return s1210_batch.resume_batch()
    except RuntimeError as e:
        raise HTTPException(409, str(e))


@router.post("/batch/stop")
def batch_stop():
    from esocial import s1210_batch
    try:
        return s1210_batch.stop_batch()
    except RuntimeError as e:
        raise HTTPException(409, str(e))


@router.get("/batch/status")
def batch_status(since_seq: int = 0, log_limit: int = 200):
    from esocial import s1210_batch
    return s1210_batch.get_status(since_seq=since_seq, log_limit=log_limit)

