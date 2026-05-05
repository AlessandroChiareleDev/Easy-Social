"""
RepositÃ³rio S-1210 â€” rotas FastAPI para a tela /repositorio-s1210

Diferente de s1210_missao_routes.py (que lÃª XLSX direto do Downloads
em cada request), este mÃ³dulo persiste:

  â€¢ XLSX oficiais da Ana em disco (pasta de storage local por ora; o
    campo s1210_xlsx.storage_path guarda o caminho e facilita migrar
    para bucket Supabase Storage sem mudar API).
  â€¢ Escopo CPF Ã— Lote Ã— MÃªs em s1210_cpf_scope.
  â€¢ Contadores consumidos pela tela vÃªm da view v_s1210_contadores.

Endpoints desta fatia mÃ­nima:
  GET  /api/s1210-repo/overview                           â†’ 3 meses Ã— 4 lotes (contadores)
  GET  /api/s1210-repo/xlsx                               â†’ lista XLSX ingeridas
  POST /api/s1210-repo/xlsx/ingest                        â†’ upload + parse + popula scope
  GET  /api/s1210-repo/por-lote/{lote_num}/{per_apur}     â†’ lista CPFs (stub p/ nÃ­vel 2)

NÃ£o contÃ©m:
  â€¢ Envio ao eSocial (fica numa prÃ³xima fatia)
  â€¢ Parser de operadoras (aba Operadoras â€” prÃ³xima fatia)
  â€¢ Vertente B Mensal (prÃ³xima fatia)
"""
from __future__ import annotations

import hashlib
import json
import logging
import os
from io import BytesIO
from pathlib import Path
from typing import Optional

import psycopg2
import psycopg2.extras
from fastapi import APIRouter, File, Form, HTTPException, UploadFile
from fastapi.responses import Response
from openpyxl import load_workbook
from pydantic import BaseModel

from db_config import DB_CONFIG
from esocial.tenant import connect_for_empresa, DEFAULT_EMPRESA_ID
# Reusa o parser jÃ¡ testado da missÃ£o legada
from esocial.s1210_missao_routes import _parse_xlsx_escopo, FONTES

log = logging.getLogger("s1210-repo")
router = APIRouter(prefix="/api/s1210-repo", tags=["s1210-repo"])

# Empresa padrÃ£o (APPA) enquanto nÃ£o implementamos multi-tenant na tela nova.

# Storage local (mesmo contrato que bucket Supabase â€” path relativo).
# Quando migrarmos para Supabase Storage, sÃ³ troca a implementaÃ§Ã£o de
# _save_to_storage / _read_from_storage; a API pÃºblica nÃ£o muda.
STORAGE_ROOT = Path(__file__).resolve().parent.parent.parent / "backend" / "uploads" / "s1210-xlsx"
STORAGE_ROOT.mkdir(parents=True, exist_ok=True)


def _db(empresa_id: int | None = None):
    """Conexao roteada por empresa_id (default = APPA/Supabase)."""
    return connect_for_empresa(empresa_id if empresa_id is not None else DEFAULT_EMPRESA_ID)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# Worker paralelo (ProcessPool) â€” gera + assina UM S-1210
# Declarado no topo do mÃ³dulo para ser picklable em Windows.
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _worker_build_sign(args: tuple) -> dict:
    """
    Executa em subprocesso. Recebe dados jÃ¡ resolvidos do main
    (sem precisar do ZIP nem do banco) e devolve xml_assinado + evt_id.

    args = (
        seq: int,
        cpf: str,
        empregador: dict,
        per_apur_zip: str,
        info_pgtos: list,
        info_ir_cr: list | None,
        ind_retif: str,
        nr_recibo_usado: str,
        pfx_data: bytes,
        senha: str,
        tp_amb: str,
        plan_saude: dict | None,
    )
    """
    (seq, cpf, empregador, per_apur_zip, info_pgtos, info_ir_cr,
     ind_retif, nr_recibo_usado, pfx_data, senha, tp_amb, plan_saude) = args
    try:
        from esocial.xml_s1210 import S1210XMLGenerator
        from esocial.xml_signer import S1010XMLSigner as XMLSigner
        import re as _re

        info_ir_complem = None
        if info_ir_cr:
            info_ir_complem = {"infoIRCR": info_ir_cr}

        xml_bytes = S1210XMLGenerator.gerar(
            empregador=empregador,
            beneficiario={"cpfBenef": cpf},
            info_pgtos=info_pgtos,
            per_apur=per_apur_zip,
            ind_retif=ind_retif,
            nr_recibo=nr_recibo_usado,
            info_ir_complem=info_ir_complem,
            plan_saude=plan_saude,
            seq=seq,
            tp_amb=tp_amb,
        )
        xml_assinado = XMLSigner.assinar(xml_bytes, pfx_data, senha)
        xml_str = xml_assinado.decode("utf-8") if isinstance(xml_assinado, bytes) else xml_assinado
        m = _re.search(r'Id="([^"]+)"', xml_str)
        if not m:
            return {"cpf": cpf, "ok": False, "etapa": "extrair_id",
                    "erro": "NÃ£o foi possÃ­vel extrair Id do XML assinado"}
        return {"cpf": cpf, "ok": True,
                "xml_assinado": xml_assinado, "evt_id": m.group(1),
                "nr_recibo_usado": nr_recibo_usado}
    except Exception as e:
        # Identifica etapa pelo tipo de falha, sem perder contexto
        etapa = "assinar_xml"
        try:
            # se chegou no signer, tem xml_bytes local; senÃ£o foi no gerar
            if "xml_bytes" not in locals():
                etapa = "gerar_xml"
        except Exception:
            pass
        return {"cpf": cpf, "ok": False, "etapa": etapa,
                "erro": str(e), "nr_recibo_usado": nr_recibo_usado}


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# Storage (hoje: disco local; amanhÃ£: bucket Supabase â€” mesma API)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
def _save_to_storage(empresa_id: int, per_apur: str, sha256: str, data: bytes) -> str:
    """Grava bytes e devolve storage_path relativo."""
    rel = f"{empresa_id}/{per_apur}/{sha256}.xlsx"
    abs_path = STORAGE_ROOT / rel
    abs_path.parent.mkdir(parents=True, exist_ok=True)
    if not abs_path.exists():
        abs_path.write_bytes(data)
    return rel


def _storage_abs(storage_path: str) -> Path:
    return STORAGE_ROOT / storage_path


def _ensure_lote1_codfunc_table(conn) -> None:
    with conn.cursor() as cur:
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS s1210_lote1_codfunc_scope (
                id BIGSERIAL PRIMARY KEY,
                empresa_id INT NOT NULL,
                per_apur VARCHAR(7) NOT NULL,
                codigo_empresa VARCHAR(32),
                codigo_lote VARCHAR(64),
                codigo_filial VARCHAR(64),
                codigo_funcionario VARCHAR(64) NOT NULL,
                cpf VARCHAR(11),
                concatenar VARCHAR(128),
                lote_label VARCHAR(64),
                source_filename VARCHAR(255),
                source_sha256 VARCHAR(64),
                created_at TIMESTAMPTZ DEFAULT NOW()
            )
            """
        )
        cur.execute(
            """
            ALTER TABLE s1210_lote1_codfunc_scope
            ADD COLUMN IF NOT EXISTS cpf VARCHAR(11)
            """
        )
        cur.execute(
            """
            CREATE INDEX IF NOT EXISTS idx_s1210_lote1_codfunc_scope_per
            ON s1210_lote1_codfunc_scope (empresa_id, per_apur)
            """
        )


def _parse_lote1_codfunc_xlsx(data: bytes) -> dict[str, list[dict]]:
    """
    Parseia XLSX no layout novo (abas MMYYYY com CodigoFuncionario).
    Retorna dict per_apur -> rows deduplicadas por (codigo_empresa|codigo_funcionario).
    Se existir coluna CPF, preserva o valor normalizado com apenas digitos.
    """
    wb = load_workbook(filename=BytesIO(data), read_only=True, data_only=True)
    try:
        out: dict[str, list[dict]] = {}
        for ws in wb.worksheets:
            title = (ws.title or "").strip()
            if len(title) != 6 or not title.isdigit():
                continue
            mm = int(title[:2])
            yyyy = int(title[2:])
            if yyyy < 2000 or yyyy > 2100 or mm < 1 or mm > 12:
                continue
            per_apur = f"{yyyy:04d}-{mm:02d}"

            it = ws.iter_rows(values_only=True)
            next(it, None)

            seen: set[str] = set()
            rows: list[dict] = []
            for row in it:
                if not row:
                    continue
                vals = ["" if c is None else str(c).strip() for c in row[:8]]
                if all(v == "" for v in vals):
                    continue

                codigo_empresa = vals[0]
                codigo_lote = vals[2]
                codigo_filial = vals[3]
                codigo_func = vals[4]
                cpf = "".join(ch for ch in vals[7] if ch.isdigit()) if len(vals) > 7 else ""
                concatenar = vals[5]
                lote_label = vals[6]

                if not codigo_func:
                    continue
                key = f"{codigo_empresa}|{codigo_func}"
                if key in seen:
                    continue
                seen.add(key)

                rows.append(
                    {
                        "codigo_empresa": codigo_empresa,
                        "codigo_lote": codigo_lote,
                        "codigo_filial": codigo_filial,
                        "codigo_funcionario": codigo_func,
                        "cpf": cpf,
                        "concatenar": concatenar,
                        "lote_label": lote_label,
                    }
                )

            if rows:
                out[per_apur] = rows
        return out
    finally:
        wb.close()


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# S-5002 â€” Ã­ndice em memÃ³ria (CPF â†’ lista de dicts) por per_apur.
# O eSocial devolve o IR consolidado neste evento. Usamos para
# enriquecer /detalhe-cpf quando o S-1210 vem sem <vrCR>.
# ConstruÃ§Ã£o preguiÃ§osa: 1Âª chamada escaneia o ZIP inteiro (~30s p/ 32k
# XMLs) e cacheia para as chamadas seguintes.
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
import re as _re_s5002
import zipfile as _zipfile_s5002
import threading as _threading_s5002

_S5002_INDEX: dict[str, dict[str, list[dict]]] = {}   # per_apur â†’ cpf â†’ [records]
_S5002_LOCK = _threading_s5002.Lock()


def _parse_s5002_xml(raw: str) -> Optional[dict]:
    """Extrai campos relevantes de um XML S-5002. Retorna None se nÃ£o bater."""
    m_cpf = _re_s5002.search(r"<cpfBenef>([^<]+)</cpfBenef>", raw)
    m_per = _re_s5002.search(r"<perApur>([^<]+)</perApur>", raw)
    m_id = _re_s5002.search(r'Id="([^"]+)"', raw)
    if not (m_cpf and m_per and m_id):
        return None
    # consolidApurMen: pega o bloco â€” os campos todos ali juntos
    rec = {
        "cpf": m_cpf.group(1),
        "per_apur": m_per.group(1),
        "id": m_id.group(1),
        "nr_recibo": "1.1." + m_id.group(1)[-19:],
        "CRMen": None,
        "vlrRendTrib": None,
        "vlrPrevOficial": None,
        "vlrCRMen": None,   # <-- IRRF retido (valor do imposto)
        "infoIR": [],       # lista {tpInfoIR, valor}
        "vazio": False,
    }
    mc = _re_s5002.search(
        r"<consolidApurMen>.*?<CRMen>([^<]+)</CRMen>"
        r".*?<vlrRendTrib>([^<]+)</vlrRendTrib>"
        r".*?<vlrPrevOficial>([^<]+)</vlrPrevOficial>"
        r".*?<vlrCRMen>([^<]+)</vlrCRMen>",
        raw, _re_s5002.DOTALL,
    )
    if mc:
        rec["CRMen"] = mc.group(1)
        rec["vlrRendTrib"] = mc.group(2)
        rec["vlrPrevOficial"] = mc.group(3)
        rec["vlrCRMen"] = mc.group(4)
    else:
        rec["vazio"] = True
    for tp, v in _re_s5002.findall(
        r"<infoIR><tpInfoIR>([^<]+)</tpInfoIR><valor>([^<]+)</valor></infoIR>",
        raw,
    ):
        rec["infoIR"].append({"tpInfoIR": tp, "valor": v})
    return rec


def _build_s5002_index(per_apur: str) -> dict[str, list[dict]]:
    """LÃª o ZIP do perÃ­odo e indexa S-5002 por CPF. Chamada com lock."""
    fonte = FONTES.get(per_apur)
    if not fonte:
        return {}
    zpath = fonte["zip"]
    if not zpath.exists():
        log.warning(f"S-5002 index: ZIP ausente {zpath}")
        return {}
    idx: dict[str, list[dict]] = {}
    try:
        with _zipfile_s5002.ZipFile(str(zpath), "r") as zf:
            nomes = [n for n in zf.namelist() if "S-5002" in n and n.endswith(".xml")]
            log.info(f"S-5002 index {per_apur}: escaneando {len(nomes)} XMLs...")
            for n in nomes:
                try:
                    raw = zf.read(n).decode("utf-8", errors="replace")
                except Exception:
                    continue
                rec = _parse_s5002_xml(raw)
                if not rec:
                    continue
                if rec["per_apur"] != per_apur:
                    continue
                idx.setdefault(rec["cpf"], []).append(rec)
        log.info(f"S-5002 index {per_apur}: {len(idx)} CPFs indexados")
    except Exception as e:
        log.error(f"S-5002 index {per_apur}: falha {e}")
    return idx


def _buscar_s5002_por_cpf(per_apur: str, cpf: str) -> list[dict]:
    """Retorna lista de S-5002 do CPF no perÃ­odo (ordenada por nr_recibo desc)."""
    global _S5002_INDEX
    if per_apur not in _S5002_INDEX:
        with _S5002_LOCK:
            if per_apur not in _S5002_INDEX:
                _S5002_INDEX[per_apur] = _build_s5002_index(per_apur)
    regs = _S5002_INDEX.get(per_apur, {}).get(cpf, [])
    # ordena por nr_recibo (string) desc â€” recibos maiores = mais novos
    return sorted(regs, key=lambda r: r.get("nr_recibo", ""), reverse=True)


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# GET /overview  â€”  3 meses Ã— 4 lotes com contadores
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
class OverviewLote(BaseModel):
    per_apur: str
    lote_num: int
    total: int
    ok: int
    erro: int
    enviando: int
    pendente: int
    na: int = 0
    tem_xlsx: bool


class OverviewResponse(BaseModel):
    empresa_id: int
    meses: list[str]  # ordem canÃ´nica dos 3 meses da missÃ£o
    por_lote: dict[int, list[OverviewLote]]  # lote_num â†’ 3 itens (um por mÃªs)


class OverviewAnualCelula(BaseModel):
    per_apur: str
    lote_num: int
    total: int
    ok: int
    erro: int
    enviando: int
    pendente: int
    na: int = 0
    tem_xlsx: bool
    estado: str


class OverviewAnualMes(BaseModel):
    per_apur: str
    lotes: list[OverviewAnualCelula]
    fechado: bool = False


class OverviewAnualResponse(BaseModel):
    empresa_id: int
    ano: int
    meses: list[OverviewAnualMes]


class IngestLote1CodfuncResponse(BaseModel):
    empresa_id: int
    source_filename: str
    source_sha256: str
    meses: dict[str, int]


@router.get("/overview", response_model=OverviewResponse)
def overview(empresa_id: int = DEFAULT_EMPRESA_ID):
    """
    Retorna grid 4 lotes Ã— 3 meses. Lotes sem XLSX ingerida aparecem
    com total=0 e tem_xlsx=false â€” isso sinaliza Ã  tela que o card deve
    mostrar â€œingerir XLSXâ€.
    """
    meses = ["2025-02", "2025-03", "2025-04"]
    conn = _db(empresa_id)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Quais (per_apur) tÃªm XLSX ingerida
            cur.execute(
                """
                SELECT DISTINCT per_apur FROM s1210_xlsx
                 WHERE empresa_id = %s AND parse_ok = TRUE
                """,
                (empresa_id,),
            )
            meses_com_xlsx = {r["per_apur"] for r in cur.fetchall()}

            # Contadores da view
            cur.execute(
                """
                SELECT per_apur, lote_num, total, ok, erro, enviando, na, pendente
                  FROM v_s1210_contadores
                 WHERE empresa_id = %s
                """,
                (empresa_id,),
            )
            contadores = {(r["per_apur"], r["lote_num"]): r for r in cur.fetchall()}
    finally:
        conn.close()

    por_lote: dict[int, list[OverviewLote]] = {1: [], 2: [], 3: [], 4: []}
    for lote in (1, 2, 3, 4):
        for mes in meses:
            c = contadores.get((mes, lote))
            por_lote[lote].append(
                OverviewLote(
                    per_apur=mes,
                    lote_num=lote,
                    total=c["total"] if c else 0,
                    ok=c["ok"] if c else 0,
                    erro=c["erro"] if c else 0,
                    enviando=c["enviando"] if c else 0,
                    pendente=c["pendente"] if c else 0,
                    na=c["na"] if c else 0,
                    tem_xlsx=mes in meses_com_xlsx,
                )
            )
    return OverviewResponse(empresa_id=empresa_id, meses=meses, por_lote=por_lote)


@router.get("/anual/overview", response_model=OverviewAnualResponse)
def overview_anual(
    ano: int = 2025,
    empresa_id: int = DEFAULT_EMPRESA_ID,
):
    """
    Retorna visÃ£o anual S-1210 para 11 meses (fev..dez) Ã— 4 lotes.

    Regras:
      - meses sem qualquer dado retornam estado=sem_dados;
      - meses/lotes com dados atuais (contadores/histÃ³rico) retornam valores reais;
      - estado operacional Ã© derivado dos contadores agregados.
    """
    if ano < 2000 or ano > 2100:
        raise HTTPException(400, f"ano invÃ¡lido: {ano}")

    meses = [f"{ano}-{m:02d}" for m in range(1, 13)]
    conn = _db(empresa_id)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT DISTINCT per_apur
                  FROM s1210_xlsx
                 WHERE empresa_id = %s
                   AND parse_ok = TRUE
                   AND per_apur LIKE %s
                """,
                (empresa_id, f"{ano}-%"),
            )
            meses_com_xlsx = {r["per_apur"] for r in cur.fetchall()}

            cur.execute(
                """
                SELECT per_apur, lote_num, total, ok, erro, enviando, na, pendente
                  FROM v_s1210_contadores
                 WHERE empresa_id = %s
                   AND per_apur LIKE %s
                """,
                (empresa_id, f"{ano}-%"),
            )
            contadores = {(r["per_apur"], r["lote_num"]): r for r in cur.fetchall()}

            _ensure_lote1_codfunc_table(conn)
            cur.execute(
                """
                SELECT per_apur,
                       COUNT(DISTINCT codigo_funcionario) AS total_codfunc,
                       COUNT(DISTINCT cpf) FILTER (WHERE cpf IS NOT NULL AND cpf <> '') AS total_cpf
                  FROM s1210_lote1_codfunc_scope
                 WHERE empresa_id = %s
                   AND per_apur LIKE %s
                 GROUP BY per_apur
                """,
                (empresa_id, f"{ano}-%"),
            )
            total_codfunc_l1 = {
                r["per_apur"]: {
                    "codfunc": int(r["total_codfunc"]),
                    "cpf": int(r["total_cpf"]),
                }
                for r in cur.fetchall()
            }
            meses_com_codfunc = set(total_codfunc_l1.keys())

            # Meses fechados (S-1299 confirmado)
            meses_fechados: set[str] = set()
            try:
                cur.execute(
                    """
                    SELECT per_apur FROM s1299_fechamento_status
                     WHERE empresa_id = %s AND fechado = TRUE
                    """,
                    (empresa_id,),
                )
                meses_fechados = {r["per_apur"] for r in cur.fetchall()}
            except Exception:
                meses_fechados = set()
    finally:
        conn.close()

    out_meses: list[OverviewAnualMes] = []
    # APPA (empresa_id=1) usa o padrao historico de 4 lotes fixos.
    # Demais empresas (Solucoes em diante) usam lotes dinamicos:
    # mostra apenas os lotes que existem em scope; se nao houver nenhum,
    # mostra 1 lote vazio (padrao novo do sistema).
    APPA_ID = 1
    for mes in meses:
        if empresa_id == APPA_ID:
            lotes_a_mostrar = [1, 2, 3, 4]
        else:
            existentes = sorted({
                int(lote)
                for (per, lote) in contadores.keys()
                if per == mes
            })
            if not existentes:
                lotes_a_mostrar = [1]
            else:
                lotes_a_mostrar = existentes
        lotes: list[OverviewAnualCelula] = []
        for lote in lotes_a_mostrar:
            c = contadores.get((mes, lote))
            total = int(c["total"]) if c else 0
            ok = int(c["ok"]) if c else 0
            erro = int(c["erro"]) if c else 0
            enviando = int(c["enviando"]) if c else 0
            pendente = int(c["pendente"]) if c else 0
            na = int(c["na"]) if c else 0
            tem_xlsx = (mes in meses_com_xlsx) or (mes in meses_com_codfunc)

            # Fallback anual: meses novos com Lote 1 em layout por cÃ³digo de funcionÃ¡rio
            # ainda sem mapeamento para CPF.
            if lote == 1 and total == 0 and ok == 0 and erro == 0 and enviando == 0 and pendente == 0 and na == 0:
                fallback_l1 = total_codfunc_l1.get(mes, {"codfunc": 0, "cpf": 0})
                cod_total = fallback_l1.get("codfunc", 0)
                cpf_total = fallback_l1.get("cpf", 0)
                if cod_total > 0:
                    total = cpf_total or cod_total
                    pendente = total
                    estado = "pronto_para_processar" if cpf_total > 0 else "aguardando_mapeamento"
                else:
                    estado = "sem_dados"
            elif total == 0 and ok == 0 and erro == 0 and enviando == 0 and pendente == 0 and na == 0:
                estado = "sem_dados"
            elif enviando > 0:
                estado = "processando"
            elif pendente > 0:
                estado = "pronto_para_processar"
            elif erro > 0:
                estado = "concluido_com_erros"
            else:
                estado = "concluido"

            lotes.append(
                OverviewAnualCelula(
                    per_apur=mes,
                    lote_num=lote,
                    total=total,
                    ok=ok,
                    erro=erro,
                    enviando=enviando,
                    pendente=pendente,
                    na=na,
                    tem_xlsx=tem_xlsx,
                    estado=estado,
                )
            )
        out_meses.append(OverviewAnualMes(per_apur=mes, lotes=lotes, fechado=(mes in meses_fechados)))

    return OverviewAnualResponse(empresa_id=empresa_id, ano=ano, meses=out_meses)


@router.post("/anual/ingest-lote1-codfunc", response_model=IngestLote1CodfuncResponse)
async def ingest_anual_lote1_codfunc(
    empresa_id: int = Form(DEFAULT_EMPRESA_ID),
    file: UploadFile = File(...),
):
    """
    Ingesta layout novo (1Âº lote por cÃ³digo de funcionÃ¡rio) em staging.
    NÃ£o tenta converter para CPF automaticamente.
    """
    data = await file.read()
    if not data:
        raise HTTPException(400, "arquivo vazio")

    sha = hashlib.sha256(data).hexdigest()
    parsed = _parse_lote1_codfunc_xlsx(data)
    if not parsed:
        raise HTTPException(400, "Nenhuma aba MMYYYY vÃ¡lida encontrada no arquivo")

    conn = _db()
    try:
        _ensure_lote1_codfunc_table(conn)
        with conn.cursor() as cur:
            meses = sorted(parsed.keys())
            cur.execute(
                """
                DELETE FROM s1210_lote1_codfunc_scope
                 WHERE empresa_id = %s
                   AND per_apur = ANY(%s)
                """,
                (empresa_id, meses),
            )

            total_por_mes: dict[str, int] = {}
            for per_apur, rows in parsed.items():
                values = [
                    (
                        empresa_id,
                        per_apur,
                        r.get("codigo_empresa"),
                        r.get("codigo_lote"),
                        r.get("codigo_filial"),
                        r.get("codigo_funcionario"),
                        r.get("cpf"),
                        r.get("concatenar"),
                        r.get("lote_label"),
                        file.filename or "arquivo.xlsx",
                        sha,
                    )
                    for r in rows
                ]
                psycopg2.extras.execute_values(
                    cur,
                    """
                    INSERT INTO s1210_lote1_codfunc_scope
                    (empresa_id, per_apur, codigo_empresa, codigo_lote, codigo_filial,
                     codigo_funcionario, cpf, concatenar, lote_label, source_filename, source_sha256)
                    VALUES %s
                    """,
                    values,
                )
                total_por_mes[per_apur] = len(rows)

        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    return IngestLote1CodfuncResponse(
        empresa_id=empresa_id,
        source_filename=file.filename or "arquivo.xlsx",
        source_sha256=sha,
        meses=total_por_mes,
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# GET /xlsx â€” lista XLSX ingeridas por empresa
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@router.get("/xlsx")
def listar_xlsx(empresa_id: int = DEFAULT_EMPRESA_ID):
    conn = _db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT id, per_apur, nome_arquivo, tamanho_bytes, sha256,
                       aba_geral, aba_operadoras, uploaded_at, parse_ok,
                       parse_erro, totais_json
                  FROM s1210_xlsx
                 WHERE empresa_id = %s
                 ORDER BY per_apur, uploaded_at DESC
                """,
                (empresa_id,),
            )
            rows = cur.fetchall()
    finally:
        conn.close()
    return {"xlsx": rows}


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# POST /xlsx/ingest  â€” upload + parse + popular scope
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
class IngestResponse(BaseModel):
    xlsx_id: int
    per_apur: str
    sha256: str
    totais: dict[str, int]
    storage_path: str
    ja_existia: bool


@router.post("/xlsx/ingest", response_model=IngestResponse)
async def ingest_xlsx(
    per_apur: str = Form(..., description="'2025-02' | '2025-03' | '2025-04'"),
    aba_geral: str = Form(...),
    aba_operadoras: Optional[str] = Form(None),
    empresa_id: int = Form(DEFAULT_EMPRESA_ID),
    file: UploadFile = File(...),
):
    """
    Recebe 1 XLSX, calcula hash, grava em storage local, parseia a aba
    Geral e popula s1210_cpf_scope. Idempotente: se o mesmo hash jÃ¡
    foi ingerido para (empresa,mÃªs), retorna o registro existente.
    """
    if per_apur not in ("2025-02", "2025-03", "2025-04"):
        raise HTTPException(400, f"per_apur invÃ¡lido: {per_apur}")

    data = await file.read()
    if not data:
        raise HTTPException(400, "arquivo vazio")
    sha = hashlib.sha256(data).hexdigest()
    storage_path = _save_to_storage(empresa_id, per_apur, sha, data)

    conn = _db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # IdempotÃªncia
            cur.execute(
                """
                SELECT id, per_apur, sha256, storage_path, parse_ok, totais_json
                  FROM s1210_xlsx
                 WHERE empresa_id = %s AND per_apur = %s AND sha256 = %s
                """,
                (empresa_id, per_apur, sha),
            )
            existing = cur.fetchone()
            if existing and existing["parse_ok"]:
                return IngestResponse(
                    xlsx_id=existing["id"],
                    per_apur=existing["per_apur"],
                    sha256=existing["sha256"],
                    totais=existing["totais_json"] or {},
                    storage_path=existing["storage_path"],
                    ja_existia=True,
                )

            # Insere ou reusa registro pendente
            if existing:
                xlsx_id = existing["id"]
            else:
                cur.execute(
                    """
                    INSERT INTO s1210_xlsx
                      (empresa_id, per_apur, nome_arquivo, tamanho_bytes,
                       sha256, storage_path, aba_geral, aba_operadoras)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
                    RETURNING id
                    """,
                    (
                        empresa_id, per_apur, file.filename, len(data),
                        sha, storage_path, aba_geral, aba_operadoras,
                    ),
                )
                xlsx_id = cur.fetchone()["id"]
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(500, f"erro gravando s1210_xlsx: {e}")

    # Parseia o XLSX (reusa FONTES p/ saber aba, mas sobrescreve path
    # temporariamente â€” o parser atual lÃª do disco Downloads). Mais
    # simples: parseamos direto aqui com openpyxl.
    try:
        totais = _parse_and_populate_scope(
            conn=conn,
            xlsx_id=xlsx_id,
            empresa_id=empresa_id,
            per_apur=per_apur,
            xlsx_abs_path=_storage_abs(storage_path),
            aba_geral=aba_geral,
        )
    except Exception as e:
        log.exception("parse falhou")
        with conn.cursor() as cur:
            cur.execute(
                "UPDATE s1210_xlsx SET parse_ok=FALSE, parse_erro=%s WHERE id=%s",
                (str(e)[:500], xlsx_id),
            )
        conn.commit()
        conn.close()
        raise HTTPException(500, f"parse falhou: {e}")

    with conn.cursor() as cur:
        cur.execute(
            "UPDATE s1210_xlsx SET parse_ok=TRUE, totais_json=%s WHERE id=%s",
            (json.dumps(totais), xlsx_id),
        )
    conn.commit()
    conn.close()

    return IngestResponse(
        xlsx_id=xlsx_id,
        per_apur=per_apur,
        sha256=sha,
        totais=totais,
        storage_path=storage_path,
        ja_existia=False,
    )


def _parse_and_populate_scope(
    conn,
    xlsx_id: int,
    empresa_id: int,
    per_apur: str,
    xlsx_abs_path: Path,
    aba_geral: str,
) -> dict[str, int]:
    """Parseia a aba Geral do XLSX e popula s1210_cpf_scope (idempotente)."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_abs_path, read_only=True, data_only=True)
    try:
        if aba_geral not in wb.sheetnames:
            raise ValueError(
                f"aba '{aba_geral}' nÃ£o encontrada. abas={wb.sheetnames}"
            )
        ws = wb[aba_geral]
        rows_para_inserir: list[tuple] = []
        vistos: set[str] = set()

        def _norm_lote(v) -> Optional[int]:
            s = str(v or "").strip()
            for ch in s:
                if ch.isdigit() and ch in "1234":
                    return int(ch)
            return None

        it = ws.iter_rows(values_only=True)
        next(it, None)  # header
        for idx, row in enumerate(it, start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            if len(row) < 8:
                continue
            lote = _norm_lote(row[6])
            if lote is None:
                continue
            cpf_raw = str(row[7] or "")
            cpf = "".join(ch for ch in cpf_raw if ch.isdigit())
            if not cpf:
                continue
            cpf = cpf.zfill(11)
            if len(cpf) != 11:
                continue
            if cpf in vistos:
                continue
            vistos.add(cpf)
            # Nome/matrÃ­cula: tentar colunas frequentes; nÃ£o obrigatÃ³rio
            nome = None
            matricula = None
            raw = {f"c{i}": (str(v) if v is not None else None) for i, v in enumerate(row)}
            rows_para_inserir.append((
                xlsx_id, empresa_id, per_apur, cpf, nome, matricula,
                lote, idx, json.dumps(raw),
            ))
    finally:
        wb.close()

    totais = {"1_LOTE": 0, "2_LOTE": 0, "3_LOTE": 0, "4_LOTE": 0}
    with conn.cursor() as cur:
        # Limpa scope anterior dessa (empresa, per_apur) â€” a XLSX mais
        # recente Ã© a verdade.
        cur.execute(
            "DELETE FROM s1210_cpf_scope WHERE empresa_id=%s AND per_apur=%s",
            (empresa_id, per_apur),
        )
        if rows_para_inserir:
            psycopg2.extras.execute_values(
                cur,
                """
                INSERT INTO s1210_cpf_scope
                  (xlsx_id, empresa_id, per_apur, cpf, nome, matricula,
                   lote_num, row_number, raw_row)
                VALUES %s
                """,
                rows_para_inserir,
                template="(%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb)",
            )
        for r in rows_para_inserir:
            totais[f"{r[6]}_LOTE"] = totais.get(f"{r[6]}_LOTE", 0) + 1
    return totais


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# GET /por-lote/{lote}/{per_apur}  â€” lista CPFs de um compartimento
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@router.get("/por-lote/{lote_num}/{per_apur}")
def listar_cpfs_compartimento(
    lote_num: int,
    per_apur: str,
    empresa_id: int = DEFAULT_EMPRESA_ID,
    limit: int = 200,
    offset: int = 0,
    status: Optional[str] = None,
    q: Optional[str] = None,
):
    if lote_num not in (1, 2, 3, 4):
        raise HTTPException(400, "lote_num invÃ¡lido")
    if status not in (None, "todos", "ok", "erro", "pendente", "enviando", "na"):
        raise HTTPException(400, "status invÃ¡lido")
    # Normaliza busca: dÃ­gitos puros viram padrÃ£o exato em cpf/matricula;
    # qualquer outra coisa cai em ILIKE no nome/matrÃ­cula/cpf.
    q_clean = (q or "").strip()
    q_has = bool(q_clean)
    q_digits = "".join(ch for ch in q_clean if ch.isdigit())
    conn = _db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Fallback temporario: Lote 1 pode existir apenas na tabela de staging
            # (codigo_funcionario) enquanto o mapeamento CPF nao foi entregue.
            if lote_num == 1:
                cur.execute(
                    """
                    SELECT COUNT(*) AS c
                      FROM s1210_cpf_scope
                     WHERE empresa_id=%s AND per_apur=%s AND lote_num=1
                    """,
                    (empresa_id, per_apur),
                )
                tem_scope_cpf = (cur.fetchone() or {}).get("c", 0) > 0

                if not tem_scope_cpf:
                    status_norm = status or "todos"
                    if status_norm in ("ok", "erro", "enviando"):
                        return {
                            "total": 0,
                            "totais": {"total": 0, "ok": 0, "erro": 0, "enviando": 0, "pendente": 0},
                            "cpfs": [],
                            "offset": offset,
                            "limit": limit,
                            "status_filtro": status_norm,
                            "q": q_clean or None,
                        }

                    busca_where_l1 = ""
                    busca_params_l1: list = []
                    if q_has:
                        like_q = f"%{q_clean}%"
                        like_digits = f"%{q_digits}%" if q_digits else None
                        if like_digits:
                            busca_where_l1 = (
                                " AND (codigo_funcionario ILIKE %s "
                                "   OR COALESCE(cpf,'') ILIKE %s "
                                "   OR COALESCE(concatenar,'') ILIKE %s "
                                "   OR COALESCE(codigo_empresa,'') ILIKE %s "
                                "   OR COALESCE(codigo_filial,'') ILIKE %s) "
                            )
                            busca_params_l1 = [like_digits, like_digits, like_q, like_q, like_q]
                        else:
                            busca_where_l1 = (
                                " AND (COALESCE(concatenar,'') ILIKE %s "
                                "   OR COALESCE(codigo_empresa,'') ILIKE %s "
                                "   OR COALESCE(codigo_filial,'') ILIKE %s "
                                "   OR COALESCE(cpf,'') ILIKE %s "
                                "   OR codigo_funcionario ILIKE %s) "
                            )
                            busca_params_l1 = [like_q, like_q, like_q, like_q, like_q]

                    sql_l1 = f"""
                        WITH base AS (
                            SELECT DISTINCT ON (codigo_funcionario)
                                   codigo_funcionario,
                                   cpf,
                                   codigo_empresa,
                                   codigo_filial,
                                   concatenar,
                                   lote_label,
                                   cpf
                              FROM s1210_lote1_codfunc_scope
                             WHERE empresa_id=%s AND per_apur=%s
                             {busca_where_l1}
                             ORDER BY codigo_funcionario
                        )
                        SELECT
                            COALESCE(NULLIF(b.cpf, ''), ''::text) AS cpf,
                            CONCAT('Funcionario ', b.codigo_funcionario)::text AS nome,
                            b.codigo_funcionario::text AS matricula,
                            1::int AS lote_num,
                            'pendente'::text AS status,
                            NULL::text AS nr_recibo_novo,
                            NULL::text AS nr_recibo_usado,
                            NULL::text AS erro_descricao,
                            NULL::timestamptz AS enviado_em,
                            NULL::int AS codigo_resposta,
                            NULL::text AS descricao_resposta,
                            NULL::text AS nr_recibo_zip,
                               CASE WHEN COALESCE(NULLIF(b.cpf, ''), '') = ''
                                   THEN b.codigo_funcionario::text
                                   ELSE NULL::text END AS identificador,
                               CASE WHEN COALESCE(NULLIF(b.cpf, ''), '') = ''
                                   THEN 'codfunc_scope'::text
                                   ELSE 'cpf_codfunc_scope'::text END AS origem
                        FROM base b
                        ORDER BY b.codigo_funcionario
                        LIMIT %s OFFSET %s
                    """

                    params_l1 = [empresa_id, per_apur, *busca_params_l1, limit, offset]
                    cur.execute(sql_l1, params_l1)
                    rows_l1 = cur.fetchall()

                    count_l1 = f"""
                        WITH base AS (
                            SELECT DISTINCT ON (codigo_funcionario)
                                   codigo_funcionario
                              FROM s1210_lote1_codfunc_scope
                             WHERE empresa_id=%s AND per_apur=%s
                             {busca_where_l1}
                             ORDER BY codigo_funcionario
                        )
                        SELECT COUNT(*) AS c FROM base
                    """
                    cur.execute(count_l1, [empresa_id, per_apur, *busca_params_l1])
                    total_l1 = (cur.fetchone() or {}).get("c", 0)

                    return {
                        "total": total_l1,
                        "totais": {
                            "total": total_l1,
                            "ok": 0,
                            "erro": 0,
                            "enviando": 0,
                            "pendente": total_l1,
                        },
                        "cpfs": rows_l1,
                        "offset": offset,
                        "limit": limit,
                        "status_filtro": status_norm,
                        "q": q_clean or None,
                    }

            # Totais do compartimento inteiro (via view jÃ¡ existente)
            cur.execute(
                """SELECT total, ok, erro, enviando, na, pendente
                     FROM v_s1210_contadores
                    WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s""",
                (empresa_id, per_apur, lote_num),
            )
            tot_row = cur.fetchone() or {
                "total": 0, "ok": 0, "erro": 0, "enviando": 0, "na": 0, "pendente": 0,
            }

            # Filtro opcional por status (server-side para nÃ£o depender da pÃ¡gina)
            status_where = ""
            if status and status != "todos":
                if status == "pendente":
                    status_where = " AND COALESCE(u.status,'pendente') IN ('pendente') "
                else:
                    status_where = f" AND COALESCE(u.status,'pendente') = '{status}' "

            # Filtro de busca livre (cpf/nome/matrÃ­cula)
            busca_where = ""
            busca_params: list = []
            if q_has:
                if q_digits and len(q_digits) >= 3:
                    # dÃ­gitos: tenta cpf/matrÃ­cula por LIKE
                    busca_where = (
                        " AND (s.cpf ILIKE %s OR s.matricula ILIKE %s "
                        "      OR COALESCE(s.nome,'') ILIKE %s) "
                    )
                    like = f"%{q_digits}%"
                    busca_params = [like, like, f"%{q_clean}%"]
                else:
                    busca_where = (
                        " AND (COALESCE(s.nome,'') ILIKE %s "
                        "      OR s.matricula ILIKE %s) "
                    )
                    like = f"%{q_clean}%"
                    busca_params = [like, like]

            sql = f"""
                WITH ult AS (
                    SELECT DISTINCT ON (cpf) cpf, status, nr_recibo_novo,
                           nr_recibo_usado, erro_descricao, enviado_em,
                           codigo_resposta, descricao_resposta
                      FROM s1210_cpf_envios
                     WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
                     ORDER BY cpf, enviado_em DESC
                )
                SELECT s.cpf, s.nome, s.matricula, s.lote_num,
                       COALESCE(u.status,'pendente') AS status,
                       u.nr_recibo_novo, u.nr_recibo_usado,
                       u.erro_descricao, u.enviado_em,
                       u.codigo_resposta, u.descricao_resposta,
                      r.nr_recibo_zip,
                      COALESCE(NULLIF(s.cpf, ''), NULLIF(s.matricula, ''), 'sem-id') AS identificador,
                      'cpf_scope'::text AS origem
                  FROM s1210_cpf_scope s
                  LEFT JOIN ult u ON u.cpf = s.cpf
                  LEFT JOIN s1210_cpf_recibo r
                         ON r.empresa_id = s.empresa_id
                        AND r.per_apur   = s.per_apur
                        AND r.cpf        = s.cpf
                 WHERE s.empresa_id=%s AND s.per_apur=%s AND s.lote_num=%s
                 {status_where}
                 {busca_where}
                 ORDER BY
                    CASE COALESCE(u.status,'pendente')
                        WHEN 'ok'       THEN 1
                        WHEN 'erro'     THEN 2
                        WHEN 'enviando' THEN 3
                        WHEN 'na'       THEN 5
                        ELSE 4
                    END,
                    u.enviado_em ASC NULLS LAST,
                    s.cpf
                 LIMIT %s OFFSET %s
                """
            params = [
                empresa_id, per_apur, lote_num,
                empresa_id, per_apur, lote_num,
                *busca_params,
                limit, offset,
            ]
            cur.execute(sql, params)
            rows = cur.fetchall()

            # Total considerando filtros (para paginacao).
            # BUG #6 fix: antes, quando so havia filtro de status (sem q), o total vinha da
            # view v_s1210_contadores, que pode divergir do CTE `ult` usado na listagem
            # (ex.: CPFs removidos de cpf_envios). Agora sempre conta sobre a mesma CTE.
            if q_has or (status and status != "todos"):
                count_sql = f"""
                    WITH ult AS (
                        SELECT DISTINCT ON (cpf) cpf, status
                          FROM s1210_cpf_envios
                         WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
                         ORDER BY cpf, enviado_em DESC
                    )
                    SELECT COUNT(*) AS c
                      FROM s1210_cpf_scope s
                      LEFT JOIN ult u ON u.cpf = s.cpf
                     WHERE s.empresa_id=%s AND s.per_apur=%s AND s.lote_num=%s
                     {status_where}
                     {busca_where}
                """
                cur.execute(count_sql, [
                    empresa_id, per_apur, lote_num,
                    empresa_id, per_apur, lote_num,
                    *busca_params,
                ])
                total_filtrado = cur.fetchone()["c"]
            else:
                total_filtrado = tot_row["total"]
    finally:
        conn.close()
    return {
        "total": total_filtrado,
        "totais": tot_row,
        "cpfs": rows,
        "offset": offset,
        "limit": limit,
        "status_filtro": status or "todos",
        "q": q_clean or None,
    }


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# GET /codigos-agregados/{lote}/{per_apur}
# Retorna a contagem de cÃ³digos (cdResposta/codOcorrencia) considerando
# o compartimento INTEIRO, nÃ£o apenas a pÃ¡gina carregada no frontend.
# Usado pelo funil "Filtrar por cÃ³digo" da tela de compartimento.
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@router.get("/codigos-agregados/{lote_num}/{per_apur}")
def codigos_agregados_compartimento(
    lote_num: int,
    per_apur: str,
    empresa_id: int = DEFAULT_EMPRESA_ID,
):
    import re, json as _json
    if lote_num not in (1, 2, 3, 4):
        raise HTTPException(400, "lote_num invÃ¡lido")
    conn = _db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                WITH ult AS (
                    SELECT DISTINCT ON (cpf) cpf, status, codigo_resposta,
                           descricao_resposta, erro_descricao
                      FROM s1210_cpf_envios
                     WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
                     ORDER BY cpf, enviado_em DESC NULLS LAST
                )
                SELECT COALESCE(u.status,'pendente') AS status,
                       u.codigo_resposta, u.descricao_resposta, u.erro_descricao
                  FROM s1210_cpf_scope s
                  LEFT JOIN ult u ON u.cpf = s.cpf
                 WHERE s.empresa_id=%s AND s.per_apur=%s AND s.lote_num=%s
                """,
                (empresa_id, per_apur, lote_num,
                 empresa_id, per_apur, lote_num),
            )
            rows = cur.fetchall()
    finally:
        conn.close()

    # Agrega no formato que o frontend consome:
    # chave = "<cdResposta>/<codOcorrencia>" (igual a codigoChave() do .vue)
    # BUG #3 fix: CPF com varias ocorrencias (ex.: [861, 726]) vira 1 bucket por ocorrencia.
    # Antes so contava a primeira, escondendo codigos 726, 1043 etc. do funil.
    _re_oc = re.compile(r'ocorrencias=(\[.*\])')
    def _parse_ocorrencias(txt):
        """Retorna lista de (codigo, descricao) de TODAS as ocorrencias."""
        if not txt:
            return []
        m = _re_oc.search(txt)
        if not m:
            return []
        try:
            arr = _json.loads(m.group(1))
            if isinstance(arr, list):
                out = []
                for item in arr:
                    if not isinstance(item, dict):
                        continue
                    cod = str(item.get('codigo') or '')
                    desc = str(item.get('descricao') or '')
                    if cod:
                        out.append((cod, desc))
                return out
        except Exception:
            return []
        return []

    buckets = {}
    total_cpfs = 0
    for r in rows:
        status = r['status']
        if status == 'pendente' or status == 'enviando':
            continue
        total_cpfs += 1
        if status == 'ok':
            chave = f"{r.get('codigo_resposta') or '201'}/"
            desc_default = r.get('descricao_resposta') or 'Sucesso'
            tipo = 'ok'
            b = buckets.get(chave)
            if b:
                b['qtd'] += 1
            else:
                buckets[chave] = {
                    'chave': chave,
                    'descricao': (desc_default or '')[:120],
                    'qtd': 1,
                    'tipo': tipo,
                }
        elif status == 'erro':
            cd_resp = r.get('codigo_resposta') or 'erro'
            ocs = _parse_ocorrencias(r.get('erro_descricao'))
            if not ocs:
                # sem ocorrencias parseadas — mantem chave sem sufixo
                chave = f"{cd_resp}/"
                desc_default = r.get('descricao_resposta') or r.get('erro_descricao') or '—'
                b = buckets.get(chave)
                if b:
                    b['qtd'] += 1
                else:
                    buckets[chave] = {
                        'chave': chave,
                        'descricao': (desc_default or '')[:120],
                        'qtd': 1,
                        'tipo': 'err',
                    }
            else:
                # 1 linha do CPF entra em 1 bucket por ocorrencia
                for cod, desc in ocs:
                    chave = f"{cd_resp}/{cod}"
                    b = buckets.get(chave)
                    if b:
                        b['qtd'] += 1
                    else:
                        buckets[chave] = {
                            'chave': chave,
                            'descricao': (desc or r.get('descricao_resposta') or '')[:120],
                            'qtd': 1,
                            'tipo': 'err',
                        }

    lista = sorted(buckets.values(), key=lambda x: -x['qtd'])
    # total = numero de CPFs (nao soma de buckets, pois CPF com N ocorrencias conta N vezes em buckets)
    return {'codigos': lista, 'total': total_cpfs}


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# POST /enviar-cpf  â€” envia UM CPF em PRODUÃ‡ÃƒO e persiste resultado
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
class EnviarCpfReq(BaseModel):
    cpf: str
    per_apur: str
    lote_num: int
    confirmar_producao: bool = False


@router.post("/enviar-cpf")
def enviar_cpf(req: EnviarCpfReq):
    """
    Envia UM CPF S-1210 em produÃ§Ã£o e grava resultado em s1210_cpf_envios.

    âš ï¸ Exige confirmar_producao=True. NÃ£o depende de XLSX carregada em
    memÃ³ria: resolve o Ã­ndice do CPF no prÃ³prio s1210_cpf_scope para
    reusar o pipeline jÃ¡ testado (_testar_um_cpf_impl).
    """
    if not req.confirmar_producao:
        raise HTTPException(400, "confirmar_producao=true Ã© obrigatÃ³rio para envio em produÃ§Ã£o")
    if req.lote_num not in (1, 2, 3, 4):
        raise HTTPException(400, "lote_num invÃ¡lido")

    from esocial.s1210_missao_routes import (
        _testar_um_cpf_impl, TestarUmCpfReq, _parse_xlsx_escopo, _CACHE_XLSX,
        _buscar_s1210_unico,
    )
    from esocial.s1210_batch import _buscar_recibo_ativo

    # Garante XLSX parseada (usa a legada, que lÃª do Downloads).
    if req.per_apur not in _CACHE_XLSX:
        _CACHE_XLSX[req.per_apur] = _parse_xlsx_escopo(req.per_apur)

    lote_key = f"{req.lote_num}_LOTE"
    cpfs_lote = _CACHE_XLSX[req.per_apur].get(lote_key, [])
    try:
        indice = cpfs_lote.index(req.cpf)
    except ValueError:
        raise HTTPException(
            404,
            f"CPF {req.cpf} nÃ£o estÃ¡ no lote {req.lote_num} de {req.per_apur} (XLSX da Ana)",
        )

    # Chain walk â€” busca recibo ATIVO em pipeline_cpf_results antes de enviar.
    # Sem isso, o recibo do ZIP pode estar obsoleto (ocorrÃªncia 459 do eSocial).
    recibo_override = None
    try:
        s1210_zip = _buscar_s1210_unico(req.per_apur, req.cpf)
        if s1210_zip:
            recibo_ativo, fonte, n_cand = _buscar_recibo_ativo(req.cpf, s1210_zip)
            log.info(
                f"ðŸ”— chain walk {req.cpf}: fonte={fonte} candidatos={n_cand} "
                f"zip={s1210_zip['nr_recibo']} ativo={recibo_ativo}"
            )
            if fonte == "cadeia":
                recibo_override = recibo_ativo
    except Exception as e:
        log.warning(f"chain walk falhou para {req.cpf}: {e}")

    log.info(f"â–¶ï¸  enviar-cpf: {req.cpf} {req.per_apur} lote={req.lote_num} indice={indice} override={recibo_override}")

    # Marca como 'enviando' enquanto roda.
    conn = _db()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                INSERT INTO s1210_cpf_envios
                    (empresa_id, per_apur, lote_num, cpf, status, descricao_resposta, enviado_em)
                VALUES (%s, %s, %s, %s, 'enviando', 'envio iniciado', NOW())
                """,
                (DEFAULT_EMPRESA_ID, req.per_apur, req.lote_num, req.cpf),
            )
        conn.commit()
    finally:
        conn.close()

    # Delega para o pipeline existente.
    sub_req = TestarUmCpfReq(
        mes=req.per_apur,
        lote=str(req.lote_num),
        indice=indice,
        confirmar_producao=True,
        nr_recibo_override=recibo_override,
    )
    try:
        resultado = _testar_um_cpf_impl(sub_req)
    except HTTPException:
        raise
    except Exception as e:
        log.exception("enviar-cpf falhou")
        resultado = {
            "sucesso": False,
            "etapa": "excecao",
            "cpf": req.cpf,
            "erro": str(e),
        }

    # Persiste resultado final em s1210_cpf_envios.
    status = "ok" if resultado.get("sucesso") else "erro"
    descricao = resultado.get("descricao") or ""
    erro_desc = resultado.get("erro") or ""
    codigo = str(resultado.get("codigo_resposta") or "")
    nr_recibo_novo = resultado.get("nr_recibo_novo")
    nr_recibo_usado = resultado.get("nr_recibo_usado") or resultado.get("nr_recibo_original")
    etapa = resultado.get("etapa") or ""
    ocorrencias = resultado.get("ocorrencias")
    protocolo = resultado.get("protocolo")

    # CompÃµe erro final com etapa + ocorrÃªncias, se houver.
    if status == "erro":
        partes = [p for p in [etapa, erro_desc, descricao] if p]
        if ocorrencias:
            partes.append("ocorrencias=" + json.dumps(ocorrencias, ensure_ascii=False))
        erro_final = " | ".join(partes) or "erro desconhecido"
    else:
        erro_final = None

    conn = _db()
    try:
        with conn.cursor() as cur:
            # Tenta ATUALIZAR o marcador 'enviando' mais recente deste CPF
            # (evita acumular 2 linhas "enviando" + "ok" na timeline).
            cur.execute(
                """
                UPDATE s1210_cpf_envios
                   SET status=%s,
                       codigo_resposta=%s,
                       descricao_resposta=%s,
                       nr_recibo_novo=%s,
                       nr_recibo_usado=%s,
                       protocolo=%s,
                       erro_descricao=%s,
                       enviado_em=NOW()
                 WHERE id = (
                     SELECT id FROM s1210_cpf_envios
                      WHERE empresa_id=%s AND per_apur=%s
                        AND lote_num=%s  AND cpf=%s
                        AND status='enviando'
                      ORDER BY enviado_em DESC
                      LIMIT 1
                 )
                """,
                (
                    status, codigo, descricao,
                    nr_recibo_novo, nr_recibo_usado, protocolo,
                    erro_final,
                    DEFAULT_EMPRESA_ID, req.per_apur, req.lote_num, req.cpf,
                ),
            )
            if cur.rowcount == 0:
                # Fallback: nÃ£o havia marcador 'enviando' â€” insere uma linha nova.
                cur.execute(
                    """
                    INSERT INTO s1210_cpf_envios
                        (empresa_id, per_apur, lote_num, cpf,
                         status, codigo_resposta, descricao_resposta,
                         nr_recibo_novo, nr_recibo_usado, protocolo,
                         erro_descricao, enviado_em)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                    """,
                    (
                        DEFAULT_EMPRESA_ID, req.per_apur, req.lote_num, req.cpf,
                        status, codigo, descricao,
                        nr_recibo_novo, nr_recibo_usado, protocolo,
                        erro_final,
                    ),
                )
        conn.commit()
    finally:
        conn.close()

    return resultado


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# POST /enviar-lote-cpfs  â€” envia atÃ© 50 CPFs num ÃšNICO lote eSocial
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# MotivaÃ§Ã£o: enviar CPF-a-CPF com CONCURRENCY=3 no frontend gerava:
#   â€¢ cdResposta 401 + ocorrÃªncia 1089 ("enviado ao mesmo tempo em
#     mais de um lote") â€” simultaneidade do prÃ³prio bot.
#   â€¢ Overhead de ~5â€“15s por CPF (envio + polling).
# Enviando vÃ¡rios CPFs num Ãºnico <envioLoteEventos> (eSocial permite
# atÃ© 50) resolvemos os dois: 1 polling sÃ³ + sem simultaneidade.
#
# Regras especiais aplicadas aqui (vs. /enviar-cpf):
#   â€¢ OcorrÃªncia 543 ("jÃ¡ existe na base") â†’ marcado como status='ok'
#     com descricao "idempotente â€” jÃ¡ existia no AN".
#   â€¢ OcorrÃªncia 1089 ("enviado ao mesmo tempo") â†’ marcado como status
#     'erro_retry' para o player reintentar mais tarde. (Ainda Ã© erro
#     enquanto nÃ£o for reenviado com sucesso.)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
class EnviarLoteCpfsReq(BaseModel):
    per_apur: str
    lote_num: int
    cpfs: list[str]
    confirmar_producao: bool = False
    plan_saude_override: Optional[dict] = None  # {cnpjOper, regANS, vlrSaudeTit} aplicado a todos
    plan_saude_por_cpf: Optional[dict[str, list[dict]]] = None  # {cpf: [{cnpjOper, regANS, vlrSaudeTit}, ...]} â€” usado no Lote 2
    recibo_override_por_cpf: Optional[dict[str, str]] = None  # {cpf: nr_recibo} â€” forÃ§a recibo, bypass chain walk
    forcar_inclusao: bool = False  # se True envia ind_retif=1 sem nrRecibo (caso evento foi excluido externamente)


_MAX_CPFS_POR_LOTE = 50


@router.post("/enviar-lote-cpfs")
def enviar_lote_cpfs(req: EnviarLoteCpfsReq):
    """
    Envia atÃ© 50 CPFs em UM Ãºnico lote eSocial (envioLoteEventos).

    Retorna lista de resultados por CPF com status final aplicado:
      - ok (inclui ok_idempotente por ocorrÃªncia 543)
      - erro (inclui erro_retry por ocorrÃªncia 1089 â€” player reintenta)
    """
    import time as _time

    if not req.confirmar_producao:
        raise HTTPException(400, "confirmar_producao=true Ã© obrigatÃ³rio")
    if req.lote_num not in (1, 2, 3, 4):
        raise HTTPException(400, "lote_num invÃ¡lido")
    cpfs_norm: list[str] = []
    for c in req.cpfs or []:
        cc = (c or "").strip().replace(".", "").replace("-", "")
        if len(cc) == 11 and cc.isdigit() and cc not in cpfs_norm:
            cpfs_norm.append(cc)
    if not cpfs_norm:
        raise HTTPException(400, "lista de cpfs vazia ou invÃ¡lida")
    if len(cpfs_norm) > _MAX_CPFS_POR_LOTE:
        raise HTTPException(400, f"mÃ¡ximo {_MAX_CPFS_POR_LOTE} CPFs por lote")

    t0 = _time.time()

    from esocial.s1210_missao_routes import (
        _parse_xlsx_escopo, _CACHE_XLSX, _buscar_s1210_unico, _load_cert_ativo,
        _CACHE_RECIBOS, _indexar_zip,
    )
    from esocial.s1210_batch import _buscar_recibo_ativo
    from esocial.xml_s1210 import S1210XMLGenerator
    from esocial.xml_signer import S1010XMLSigner as XMLSigner
    from esocial.soap_builder import SOAPEnvelopeBuilder
    from esocial.esocial_client import ESocialClient

    # Garante que o ZIP do mÃªs esteja indexado em memÃ³ria (cpf â†’ s1210).
    # Primeira chamada paga o custo (~1-3 min); seguintes ficam instantÃ¢neas.
    if req.per_apur not in _CACHE_RECIBOS:
        t_idx = _time.time()
        try:
            _CACHE_RECIBOS[req.per_apur] = _indexar_zip(req.per_apur)
            print(f"[s1210-batch] indexou ZIP {req.per_apur}: "
                  f"{len(_CACHE_RECIBOS[req.per_apur])} CPFs Â· {_time.time()-t_idx:.1f}s",
                  flush=True)
        except Exception as e:
            raise HTTPException(500, f"Falha indexando ZIP {req.per_apur}: {e}")


    # Resultado por CPF â€” comeÃ§a todo None, preenchido nas etapas.
    resultados: dict[str, dict] = {cpf: {"cpf": cpf} for cpf in cpfs_norm}

    # Marca tudo como 'enviando' no banco logo de cara (BULK INSERT â€” O3).
    import psycopg2.extras as _pgx
    conn = _db()
    try:
        with conn.cursor() as cur:
            _pgx.execute_values(
                cur,
                """
                INSERT INTO s1210_cpf_envios
                    (empresa_id, per_apur, lote_num, cpf, status, descricao_resposta, enviado_em)
                VALUES %s
                """,
                [
                    (DEFAULT_EMPRESA_ID, req.per_apur, req.lote_num, cpf,
                     "enviando", "envio em lote iniciado")
                    for cpf in cpfs_norm
                ],
                template="(%s, %s, %s, %s, %s, %s, NOW())",
                page_size=200,
            )
        conn.commit()
    finally:
        conn.close()

    # Garante XLSX parseada
    if req.per_apur not in _CACHE_XLSX:
        _CACHE_XLSX[req.per_apur] = _parse_xlsx_escopo(req.per_apur)
    cpfs_lote_xlsx = set(_CACHE_XLSX[req.per_apur].get(f"{req.lote_num}_LOTE", []))

    # Fallback: aceita CPFs vindos de s1210_cpf_scope (caminho oficial p/ Lote 2/3/4
    # cujo XLSX nao esta em FONTES). Une os dois conjuntos.
    try:
        conn_scope = _db()
        try:
            with conn_scope.cursor() as cur_sc:
                cur_sc.execute(
                    "SELECT cpf FROM s1210_cpf_scope "
                    "WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s",
                    (DEFAULT_EMPRESA_ID, req.per_apur, req.lote_num),
                )
                cpfs_lote_xlsx |= {r[0] for r in cur_sc.fetchall()}
        finally:
            conn_scope.close()
    except Exception as _e:
        log.warning(f"[scope-fallback] falha consultando s1210_cpf_scope: {_e}")

    # Carrega cert uma vez
    try:
        cnpj, pfx_data, senha = _load_cert_ativo()
    except Exception as e:
        # falha total â€” marca todos como erro
        for cpf in cpfs_norm:
            resultados[cpf].update({
                "sucesso": False,
                "etapa": "cert_load",
                "erro": f"Erro carregando certificado: {e}",
            })
        _persistir_resultados_batch(req.per_apur, req.lote_num, resultados)
        raise HTTPException(500, f"Erro carregando certificado: {e}")

    empregador = {"tpInsc": 1, "nrInsc": cnpj}

    # â”€â”€ Fase 1a: prÃ©-processa dados do ZIP por CPF (paralelo I/O-bound) â”€â”€
    #           monta lista de tarefas pro pool de build+sign.
    t_fase1a = _time.time()

    def _preparar_um(tupla):
        seq, cpf = tupla
        if cpf not in cpfs_lote_xlsx:
            return {"cpf": cpf, "ok": False, "etapa": "validar_scope",
                    "erro": f"CPF {cpf} nÃ£o estÃ¡ no lote {req.lote_num} de {req.per_apur}"}
        s1210_zip = _buscar_s1210_unico(req.per_apur, cpf)
        if not s1210_zip:
            return {"cpf": cpf, "ok": False, "etapa": "buscar_recibo",
                    "erro": f"Nenhum S-1210 com nrRecibo encontrado no ZIP para CPF {cpf}"}
        recibo_usado = s1210_zip["nr_recibo"]
        # Override manual tem prioridade absoluta (bypassa chain walk)
        override = (req.recibo_override_por_cpf or {}).get(cpf)
        if override:
            recibo_usado = override
        else:
            try:
                recibo_ativo, fonte, _ = _buscar_recibo_ativo(cpf, s1210_zip)
                if fonte == "cadeia" and recibo_ativo:
                    recibo_usado = recibo_ativo
            except Exception as e:
                log.warning(f"chain walk falhou para {cpf}: {e}")
        return {
            "cpf": cpf, "ok": True, "seq": seq,
            "recibo_usado": recibo_usado,
            "per_apur_zip": s1210_zip["per_apur"],
            "info_pgtos": s1210_zip["info_pgtos"],
            "info_ir_cr": s1210_zip["info_ir_cr"],
        }

    from concurrent.futures import ThreadPoolExecutor
    tarefas: list[tuple] = []  # args pro worker de build+sign
    entradas = list(enumerate(cpfs_norm, start=1))  # (seq, cpf)
    # Chain walk abre 1 conn no Supabase pooler por CPF; limitamos a 5 threads
    # pra nÃ£o estourar o limite de conexÃµes concorrentes (pooler transaction mode).
    n_prep = min(5, len(entradas)) or 1
    with ThreadPoolExecutor(max_workers=n_prep) as pool:
        for prep in pool.map(_preparar_um, entradas):
            cpf = prep["cpf"]
            if not prep.get("ok"):
                resultados[cpf].update({
                    "sucesso": False,
                    "etapa": prep.get("etapa") or "prep",
                    "erro": prep.get("erro") or "falha no prep",
                })
                continue
            resultados[cpf]["nr_recibo_usado"] = prep["recibo_usado"]
            # Lote 2/3: plan_saude por CPF (lista de entradas, 1 por CNPJ operadora)
            # Lote 1: plan_saude = None  |  Lote 4: plan_saude_override (se informado)
            plan_saude_cpf: list[dict] | dict | None = None
            if req.plan_saude_por_cpf and cpf in req.plan_saude_por_cpf:
                plan_saude_cpf = req.plan_saude_por_cpf[cpf]
            elif req.plan_saude_override:
                plan_saude_cpf = req.plan_saude_override
            _ind_retif = "1" if req.forcar_inclusao else "2"
            _nr_rec = None if req.forcar_inclusao else prep["recibo_usado"]
            tarefas.append((
                prep["seq"], cpf, empregador,
                prep["per_apur_zip"],
                prep["info_pgtos"],
                prep["info_ir_cr"],
                _ind_retif,                 # ind_retif
                _nr_rec,
                pfx_data, senha,
                "1",                        # tp_amb
                plan_saude_cpf,             # plan_saude (por CPF ou override global)
            ))
    print(f"[s1210-batch] Fase 1a (prep tarefas paralelo): {_time.time()-t_fase1a:.1f}s Â· tarefas={len(tarefas)} Â· workers={n_prep}", flush=True)

    # â”€â”€ Fase 1b: build + sign em paralelo (ProcessPool) â€” O1 â”€â”€
    xmls_por_cpf: dict[str, bytes] = {}
    id_por_cpf: dict[str, str] = {}

    if tarefas:
        t_fase1b = _time.time()
        from concurrent.futures import ThreadPoolExecutor
        # Threads: sign (OpenSSL) e etree serializaÃ§Ã£o liberam GIL â†’ paralelismo real.
        # Evita overhead de spawn do ProcessPool no Windows (~seg por worker).
        n_workers = min(16, len(tarefas))
        with ThreadPoolExecutor(max_workers=n_workers) as pool:
            for res in pool.map(_worker_build_sign, tarefas, chunksize=1):
                cpf = res["cpf"]
                if res.get("ok"):
                    xmls_por_cpf[cpf] = res["xml_assinado"]
                    id_por_cpf[cpf] = res["evt_id"]
                    resultados[cpf]["evt_id"] = res["evt_id"]
                    resultados[cpf]["nr_recibo_usado"] = res.get("nr_recibo_usado")
                else:
                    resultados[cpf].update({
                        "sucesso": False,
                        "etapa": res.get("etapa") or "gerar_xml",
                        "erro": res.get("erro") or "falha no worker",
                    })
                    if res.get("nr_recibo_usado"):
                        resultados[cpf]["nr_recibo_usado"] = res["nr_recibo_usado"]
        dur_1b = _time.time() - t_fase1b
        print(f"[s1210-batch] Fase 1b (build+sign paralelo): {dur_1b:.1f}s Â· workers={n_workers} Â· tarefas={len(tarefas)}", flush=True)


    if not xmls_por_cpf:
        # Nenhum XML vÃ¡lido â€” sÃ³ persistir os erros e retornar.
        _persistir_resultados_batch(req.per_apur, req.lote_num, resultados)
        return {
            "resumo": {"ok": 0, "erro": len(cpfs_norm), "ok_idempotente": 0, "erro_retry": 0},
            "resultados": list(resultados.values()),
            "duracao_ms": int((_time.time() - t0) * 1000),
        }

    log.info(f"ðŸ“¤ Enviando lote com {len(xmls_por_cpf)} XMLs ao eSocial (de {len(cpfs_norm)} CPFs pedidos)")

    # â”€â”€ Fase 2: montar SOAP Ãºnico e enviar â”€â”€
    t_fase2 = _time.time()
    try:
        soap = SOAPEnvelopeBuilder.montar_envio(
            list(xmls_por_cpf.values()), empregador, empregador.copy(), grupo="3"
        )
        url_envio = SOAPEnvelopeBuilder.url_envio(producao=True)
        envio = ESocialClient.enviar_lote(soap, pfx_data, senha, url=url_envio)
        print(f"[s1210-batch] Fase 2 (SOAP envio): {_time.time()-t_fase2:.1f}s", flush=True)
    except Exception as e:
        for cpf in xmls_por_cpf:
            resultados[cpf].update({
                "sucesso": False,
                "etapa": "enviar_soap",
                "erro": str(e),
            })
        _persistir_resultados_batch(req.per_apur, req.lote_num, resultados)
        raise HTTPException(502, f"Erro enviando lote SOAP: {e}")

    if not envio.get("sucesso"):
        # Lote REJEITADO como um todo (ex.: XML mal-formado no esquema do lote).
        # Todos os CPFs do lote caem como erro.
        erro_lote = envio.get("erro") or envio.get("descricao") or "lote rejeitado"
        for cpf in xmls_por_cpf:
            resultados[cpf].update({
                "sucesso": False,
                "etapa": "envio_rejeitado",
                "codigo_resposta": envio.get("codigo_resposta"),
                "descricao": envio.get("descricao"),
                "erro": erro_lote,
            })
        _persistir_resultados_batch(req.per_apur, req.lote_num, resultados)
        return {
            "resumo": {
                "ok": 0,
                "erro": len(cpfs_norm),
                "ok_idempotente": 0,
                "erro_retry": 0,
            },
            "resultados": list(resultados.values()),
            "duracao_ms": int((_time.time() - t0) * 1000),
        }

    protocolo = envio.get("protocolo")
    log.info(f"âœ… Lote aceito. Protocolo: {protocolo}. Pollando consultaâ€¦")

    # ── Fase 3: pollar consulta (adaptativo — robusto v2) ──
    # ANTES: parava no 1º 'eventos' não-vazio e marcava resto como timeout.
    # AGORA: acumula e SO sai quando todos os CPFs do batch retornaram
    # (ou esgota ~5min). Evita marcar como timeout um evento que o eSocial
    # processou com sucesso — causa raiz do código 459 em retries.
    eventos_retorno: list[dict] = []
    eventos_acumulados: dict[str, dict] = {}
    t_poll = _time.time()
    total_cpfs = len(xmls_por_cpf)
    try:
        url_consulta = SOAPEnvelopeBuilder.url_consulta(producao=True)
        import time as __t
        # Esperas: 1s→5s adaptativo, depois 5s fixo. Total máx ≈ 5min.
        waits = [1.0, 1.5, 2.0, 3.0, 4.0] + [5.0] * 55
        for attempt, wait in enumerate(waits, start=1):
            __t.sleep(wait)
            cons = ESocialClient.consultar_lote(protocolo, pfx_data, senha, url=url_consulta)
            if cons.get("eventos"):
                for evt in cons["eventos"]:
                    eid = evt.get("id")
                    if eid:
                        eventos_acumulados[eid] = evt
                cobertura = len(eventos_acumulados)
                if cobertura >= total_cpfs:
                    print(f"[s1210-batch] Fase 3 (poll) COMPLETO {cobertura}/{total_cpfs} em {_time.time()-t_poll:.1f}s · attempt {attempt}", flush=True)
                    break
                else:
                    log.info(f"⏳ poll parcial {cobertura}/{total_cpfs} (attempt {attempt}/{len(waits)})")
                    continue
            elif cons.get("codigo_resposta") == "101":
                log.info(f"⏳ Lote ainda processando ({attempt}/{len(waits)})…")
                continue
            else:
                log.warning(f"consulta sem eventos e cd={cons.get('codigo_resposta')}")
                if eventos_acumulados:
                    continue
                break
        eventos_retorno = list(eventos_acumulados.values())
        if len(eventos_retorno) < total_cpfs:
            log.warning(f"poll exauriu: {len(eventos_retorno)}/{total_cpfs} CPFs em {_time.time()-t_poll:.1f}s")
    except Exception as e:
        for cpf in xmls_por_cpf:
            resultados[cpf].update({
                "sucesso": False,
                "etapa": "consulta",
                "protocolo": protocolo,
                "erro": str(e),
            })
        _persistir_resultados_batch(req.per_apur, req.lote_num, resultados)
        raise HTTPException(502, f"Erro consultando lote: {e}")

    # â”€â”€ Fase 4: mapear eventos de volta para CPFs â”€â”€
    id_to_cpf = {v: k for k, v in id_por_cpf.items()}
    cpfs_sem_retorno = set(xmls_por_cpf.keys())

    for evt in eventos_retorno:
        evt_id = evt.get("id")
        cpf = id_to_cpf.get(evt_id)
        if not cpf:
            log.warning(f"Evento retornado com Id={evt_id} nÃ£o mapeia a nenhum CPF do batch")
            continue
        cpfs_sem_retorno.discard(cpf)

        codigo = evt.get("codigo_resposta")
        descricao = evt.get("descricao") or ""
        nr_recibo_novo = evt.get("nr_recibo")
        ocorr = evt.get("ocorrencias") or []

        # Detecta ocorrÃªncia 543 (idempotente) e 1089 (simultaneidade)
        oc_codes = {str(o.get("codigo")) for o in ocorr if o.get("codigo")}
        tem_543 = "543" in oc_codes
        tem_1089 = "1089" in oc_codes

        if nr_recibo_novo:
            resultados[cpf].update({
                "sucesso": True,
                "etapa": "processado",
                "protocolo": protocolo,
                "codigo_resposta": codigo,
                "descricao": descricao,
                "nr_recibo_novo": nr_recibo_novo,
                "ocorrencias": ocorr,
            })
        elif tem_543:
            # Tratamento especial: 543 = jÃ¡ existe â†’ sucesso idempotente
            resultados[cpf].update({
                "sucesso": True,
                "etapa": "idempotente",
                "protocolo": protocolo,
                "codigo_resposta": codigo,
                "descricao": "idempotente â€” jÃ¡ existia no AN (ocorrÃªncia 543)",
                "ocorrencias": ocorr,
                "idempotente": True,
            })
        elif tem_1089:
            # Tratamento especial: 1089 = simultaneidade â†’ marcar como erro_retry
            resultados[cpf].update({
                "sucesso": False,
                "etapa": "processamento_rejeitado",
                "protocolo": protocolo,
                "codigo_resposta": codigo,
                "descricao": descricao,
                "ocorrencias": ocorr,
                "retry": True,  # sinaliza pro player
                "erro": f"CÃ³digo {codigo}: {descricao} (1089 â€” agendar retry)",
            })
        else:
            resultados[cpf].update({
                "sucesso": False,
                "etapa": "processamento_rejeitado",
                "protocolo": protocolo,
                "codigo_resposta": codigo,
                "descricao": descricao,
                "ocorrencias": ocorr,
                "erro": f"CÃ³digo {codigo}: {descricao}",
            })

    # CPFs que enviamos mas nÃ£o voltaram no retorno (timeout parcial)
    for cpf in cpfs_sem_retorno:
        resultados[cpf].update({
            "sucesso": False,
            "etapa": "timeout",
            "protocolo": protocolo,
            "erro": "Timeout â€” evento nÃ£o retornou na consulta",
            "retry": True,
        })

    _persistir_resultados_batch(req.per_apur, req.lote_num, resultados)

    # â”€â”€ Resumo â”€â”€
    ok = sum(1 for r in resultados.values() if r.get("sucesso"))
    idempotente = sum(1 for r in resultados.values() if r.get("idempotente"))
    erro_retry = sum(1 for r in resultados.values() if r.get("retry"))
    erro = sum(1 for r in resultados.values() if not r.get("sucesso"))

    return {
        "protocolo": protocolo,
        "resumo": {
            "ok": ok,
            "ok_idempotente": idempotente,
            "erro": erro,
            "erro_retry": erro_retry,
            "enviados": len(xmls_por_cpf),
            "total": len(cpfs_norm),
        },
        "resultados": list(resultados.values()),
        "duracao_ms": int((_time.time() - t0) * 1000),
    }


def _persistir_resultados_batch(per_apur: str, lote_num: int, resultados: dict[str, dict]):
    """
    Atualiza s1210_cpf_envios com o resultado final de cada CPF do batch.
    O4: 1 bulk UPDATE via VALUES() + fallback bulk INSERT para quem nÃ£o
    tinha linha 'enviando' (edge case raro).
    """
    import psycopg2.extras as _pgx

    # Monta linhas (cpf, status, codigo, descricao, nr_novo, nr_usado, protocolo, erro)
    rows = []
    for cpf, r in resultados.items():
        status = "ok" if r.get("sucesso") else "erro"
        codigo = str(r.get("codigo_resposta") or "") or None
        descricao = r.get("descricao") or ""
        erro_desc = r.get("erro") or ""
        etapa = r.get("etapa") or ""
        nr_recibo_novo = r.get("nr_recibo_novo")
        nr_recibo_usado = r.get("nr_recibo_usado")
        protocolo = r.get("protocolo")
        ocorrencias = r.get("ocorrencias")

        if status == "erro":
            partes = [p for p in [etapa, erro_desc, descricao] if p]
            if ocorrencias:
                partes.append("ocorrencias=" + json.dumps(ocorrencias, ensure_ascii=False))
            erro_final = " | ".join(partes) or "erro desconhecido"
        else:
            erro_final = None

        rows.append((cpf, status, codigo, descricao,
                     nr_recibo_novo, nr_recibo_usado, protocolo, erro_final))

    if not rows:
        return

    # Injeta empresa_id/per_apur/lote_num em cada linha para simplificar o SQL.
    rows_full = [
        (DEFAULT_EMPRESA_ID, per_apur, lote_num, *r) for r in rows
    ]

    conn = _db()
    try:
        with conn.cursor() as cur:
            sql = """
                WITH v(empresa_id, per_apur, lote_num, cpf, status_novo, codigo,
                       descricao, nr_novo, nr_usado, protocolo, erro) AS (
                    VALUES %s
                ),
                alvo AS (
                    SELECT DISTINCT ON (e.cpf) e.id, e.cpf
                      FROM s1210_cpf_envios e
                      JOIN v
                        ON v.cpf        = e.cpf
                       AND v.empresa_id = e.empresa_id
                       AND v.per_apur   = e.per_apur
                       AND v.lote_num   = e.lote_num
                     WHERE e.status = 'enviando'
                     ORDER BY e.cpf, e.enviado_em DESC NULLS LAST
                )
                UPDATE s1210_cpf_envios u
                   SET status             = v.status_novo,
                       codigo_resposta    = v.codigo,
                       descricao_resposta = v.descricao,
                       nr_recibo_novo     = v.nr_novo,
                       nr_recibo_usado    = v.nr_usado,
                       protocolo          = v.protocolo,
                       erro_descricao     = v.erro,
                       enviado_em         = NOW()
                  FROM v
                  JOIN alvo ON alvo.cpf = v.cpf
                 WHERE u.id = alvo.id
                RETURNING u.cpf
            """
            _pgx.execute_values(
                cur, sql, rows_full,
                template="(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                page_size=200,
                fetch=True,
            )
            atualizados = {row[0] for row in cur.fetchall()}

            # Fallback: quem nÃ£o tinha linha 'enviando' (raro) entra por INSERT.
            faltantes = [r for r in rows if r[0] not in atualizados]
            if faltantes:
                _pgx.execute_values(
                    cur,
                    """
                    INSERT INTO s1210_cpf_envios
                        (empresa_id, per_apur, lote_num, cpf,
                         status, codigo_resposta, descricao_resposta,
                         nr_recibo_novo, nr_recibo_usado, protocolo,
                         erro_descricao, enviado_em)
                    VALUES %s
                    """,
                    [
                        (DEFAULT_EMPRESA_ID, per_apur, lote_num, cpf,
                         status, codigo, descricao,
                         nr_novo, nr_usado, protocolo, erro)
                        for (cpf, status, codigo, descricao,
                             nr_novo, nr_usado, protocolo, erro) in faltantes
                    ],
                    template="(%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())",
                    page_size=200,
                )
        conn.commit()
    finally:
        conn.close()


def _persistir_resultados_batch_empresa(per_apur: str, lote_num: int,
                                        resultados: dict[str, dict], empresa_id: int):
    """Wrapper com empresa_id explÃ­cito (nÃ£o usado pela rota atual)."""
    _persistir_resultados_batch(per_apur, lote_num, resultados)


# MantÃ©m os argumentos do bulk UPDATE completos (empresa_id, per_apur, lote_num)
def _persistir_resultados_batch_v2(per_apur: str, lote_num: int,
                                    resultados: dict[str, dict]):
    return _persistir_resultados_batch(per_apur, lote_num, resultados)





# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# GET /detalhe-cpf/{lote_num}/{per_apur}/{cpf}
#   Junta tudo que podemos mostrar sobre UM CPF:
#     â€¢ Dados do scope (nome, matrÃ­cula)
#     â€¢ S-1210 do ZIP (pagamentos, IR, recibo original)
#     â€¢ Chain walk (recibo ativo apÃ³s retificaÃ§Ãµes passadas)
#     â€¢ HistÃ³rico de envios em s1210_cpf_envios
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
_TP_PGTO_LABELS = {
    "1": "Pagamento de remuneraÃ§Ã£o â€” folha mensal",
    "2": "RemuneraÃ§Ã£o de perÃ­odo anterior",
    "3": "Pagamento de fÃ©rias",
    "4": "Pagamento de 13Âº salÃ¡rio",
    "5": "Pagamento de benefÃ­cio previdenciÃ¡rio",
    "6": "Pagamento de ajuste do 13Âº salÃ¡rio",
    "7": "Pagamento de rescisÃ£o",
    "8": "Pagamento de PLR",
    "9": "Pagamento de RRA",
}
_TP_CR_LABELS = {
    "0561": "IRRF â€” Trabalho assalariado",
    "0588": "IRRF â€” 13Âº salÃ¡rio",
    "3533": "IRRF â€” Rendimentos de aposentadoria e pensÃµes",
    "5204": "IRRF â€” RescisÃ£o de contrato de trabalho",
    "5936": "IRRF â€” Plano de previdÃªncia complementar",
}
# Labels do <tpInfoIR> dentro do S-5002 (Anexo I tabela 2 do leiaute)
_TP_INFO_IR_LABELS = {
    "11": "Rendimento tributÃ¡vel",
    "12": "PrevidÃªncia oficial",
    "13": "Dependentes",
    "14": "PensÃ£o alimentÃ­cia",
    "31": "IRRF retido",
    "43": "13Âº salÃ¡rio tributÃ¡vel",
    "7900": "DeduÃ§Ãµes (prev/dep/pensÃ£o)",
    "9901": "Outros rendimentos",
    "9903": "Ajuda de custo",
}


@router.get("/detalhe-cpf/{lote_num}/{per_apur}/{cpf}")
def detalhe_cpf(
    lote_num: int,
    per_apur: str,
    cpf: str,
    empresa_id: int = DEFAULT_EMPRESA_ID,
):
    """Retorna pacote completo de detalhes de um CPF (pagamentos, IR, recibos, histÃ³rico)."""
    if lote_num not in (1, 2, 3, 4):
        raise HTTPException(400, "lote_num invÃ¡lido")
    cpf = (cpf or "").strip().replace(".", "").replace("-", "")
    if len(cpf) != 11 or not cpf.isdigit():
        raise HTTPException(400, "CPF invÃ¡lido")

    from esocial.s1210_missao_routes import _buscar_s1210_unico
    from esocial.s1210_batch import _buscar_recibo_ativo

    # 1) Scope (nome, matrÃ­cula, lote) + histÃ³rico de envios
    scope_row = None
    envios: list[dict] = []
    conn = _db()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """SELECT cpf, nome, matricula, lote_num, per_apur
                     FROM s1210_cpf_scope
                    WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s AND cpf=%s""",
                (empresa_id, per_apur, lote_num, cpf),
            )
            scope_row = cur.fetchone()
            cur.execute(
                """SELECT status, codigo_resposta, descricao_resposta,
                          nr_recibo_usado, nr_recibo_novo, protocolo,
                          erro_descricao, enviado_em
                     FROM s1210_cpf_envios
                    WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s AND cpf=%s
                    ORDER BY enviado_em DESC""",
                (empresa_id, per_apur, lote_num, cpf),
            )
            envios = list(cur.fetchall() or [])
    finally:
        conn.close()

    if not scope_row:
        raise HTTPException(404, f"CPF {cpf} nÃ£o estÃ¡ no lote {lote_num} de {per_apur}")

    # 2) S-1210 do ZIP â€” pagamentos, IR, recibo original
    zip_data = None
    zip_erro = None
    try:
        zip_data = _buscar_s1210_unico(per_apur, cpf)
    except Exception as e:
        zip_erro = str(e)
        log.warning(f"detalhe-cpf: falha ao ler ZIP para {cpf}: {e}")

    # 3) Chain walk â€” recibo ativo
    recibo_ativo = None
    recibo_fonte = None
    cadeia_n = 0
    if zip_data:
        try:
            recibo_ativo, recibo_fonte, cadeia_n = _buscar_recibo_ativo(cpf, zip_data)
        except Exception as e:
            log.warning(f"detalhe-cpf: chain walk falhou para {cpf}: {e}")

    # 4) Enriquece pagamentos com label + formato bonito
    pagamentos = []
    total_liquido = 0.0
    if zip_data and zip_data.get("info_pgtos"):
        for p in zip_data["info_pgtos"]:
            tp = str(p.get("tpPgto") or "")
            try:
                vr = float(str(p.get("vrLiq", "0")).replace(",", "."))
                total_liquido += vr
            except (ValueError, TypeError):
                vr = None
            pagamentos.append({
                "dt_pgto": p.get("dtPgto"),
                "tp_pgto": tp,
                "tp_pgto_label": _TP_PGTO_LABELS.get(tp, f"Tipo {tp}"),
                "per_ref": p.get("perRef"),
                "ide_dm_dev": p.get("ideDmDev"),
                "vr_liq": vr,
                "vr_liq_raw": p.get("vrLiq"),
            })

    # 5) IR complementar com label
    ir_entries = []
    if zip_data and zip_data.get("info_ir_cr"):
        for ir in zip_data["info_ir_cr"]:
            tp = str(ir.get("tpCR") or "")
            try:
                vr = float(str(ir.get("vrCR", "0")).replace(",", "."))
            except (ValueError, TypeError):
                vr = None
            ir_entries.append({
                "tp_cr": tp,
                "tp_cr_label": _TP_CR_LABELS.get(tp, f"CÃ³digo {tp}"),
                "vr_cr": vr,
                "vr_cr_raw": ir.get("vrCR"),
            })

    # 5b) S-5002 (consolidado de IRRF gerado pela RFB)
    #     Quando o S-1210 vem sem <vrCR>, o IR real estÃ¡ aqui.
    s5002_list: list[dict] = []
    s5002_ativo: Optional[dict] = None
    ir_retido_s5002: Optional[float] = None
    try:
        s5002_recs = _buscar_s5002_por_cpf(per_apur, cpf)
        for r in s5002_recs:
            def _f(v):
                if v is None:
                    return None
                try:
                    return float(str(v).replace(",", "."))
                except (ValueError, TypeError):
                    return None
            s5002_list.append({
                "nr_recibo": r.get("nr_recibo"),
                "id": r.get("id"),
                "vazio": r.get("vazio", False),
                "cr_men": r.get("CRMen"),
                "vlr_rend_trib": _f(r.get("vlrRendTrib")),
                "vlr_prev_oficial": _f(r.get("vlrPrevOficial")),
                "vlr_ir_retido": _f(r.get("vlrCRMen")),   # IRRF retido
                "info_ir": [
                    {
                        "tp_info_ir": it.get("tpInfoIR"),
                        "tp_info_ir_label": _TP_INFO_IR_LABELS.get(
                            str(it.get("tpInfoIR")),
                            f"tpInfoIR {it.get('tpInfoIR')}",
                        ),
                        "valor": _f(it.get("valor")),
                    }
                    for it in (r.get("infoIR") or [])
                ],
            })
        # Ativo: o S-5002 preferencialmente vinculado ao nr_recibo_ativo
        # (o eSocial regenera S-5002 a cada retif). Se nÃ£o bater, usa o mais novo nÃ£o-vazio.
        ativo_nr = recibo_ativo or (zip_data.get("nr_recibo") if zip_data else None)
        for s in s5002_list:
            if ativo_nr and s["nr_recibo"] == ativo_nr:
                s5002_ativo = s
                break
        if not s5002_ativo:
            for s in s5002_list:
                if not s["vazio"]:
                    s5002_ativo = s
                    break
        if s5002_ativo:
            ir_retido_s5002 = s5002_ativo.get("vlr_ir_retido")
    except Exception as e:
        log.warning(f"detalhe-cpf: S-5002 falhou para {cpf}/{per_apur}: {e}")

    # 5c) IR efetivo (o que a UI deve exibir como "IR retido")
    #     Prioridade: vrCR do S-1210 (se presente e > 0) â†’ vlrCRMen do S-5002.
    ir_efetivo_valor: Optional[float] = None
    ir_efetivo_fonte: Optional[str] = None
    for e in ir_entries:
        if e.get("vr_cr") is not None and e["vr_cr"] > 0:
            ir_efetivo_valor = e["vr_cr"]
            ir_efetivo_fonte = "S-1210"
            break
    if ir_efetivo_valor is None and ir_retido_s5002 is not None:
        ir_efetivo_valor = ir_retido_s5002
        ir_efetivo_fonte = "S-5002"
    if ir_efetivo_valor is None and ir_entries:
        # fallback: primeiro vrCR mesmo que 0
        ir_efetivo_valor = ir_entries[0].get("vr_cr")
        ir_efetivo_fonte = "S-1210"

    # 6) Ãšltimo envio (resumo)
    ultimo = envios[0] if envios else None

    return {
        "cpf": cpf,
        "nome": scope_row["nome"],
        "matricula": scope_row["matricula"],
        "lote_num": lote_num,
        "per_apur": per_apur,
        # â€” XML/ZIP â€”
        "zip_encontrado": bool(zip_data),
        "zip_erro": zip_erro,
        "ind_retif_original": zip_data.get("ind_retif") if zip_data else None,
        "dh_processamento": zip_data.get("dh_proc") if zip_data else None,
        "nr_recibo_zip": zip_data.get("nr_recibo") if zip_data else None,
        # â€” Chain walk â€”
        "nr_recibo_ativo": recibo_ativo,
        "recibo_fonte": recibo_fonte,       # 'zip' ou 'cadeia'
        "cadeia_candidatos": cadeia_n,
        # â€” ConteÃºdo do evento â€”
        "pagamentos": pagamentos,
        "total_vr_liq": round(total_liquido, 2) if pagamentos else None,
        "info_ir": ir_entries,
        # â€” IR do S-5002 (consolidado pela RFB) â€”
        "s5002_list": s5002_list,
        "s5002_ativo": s5002_ativo,
        "ir_efetivo_valor": ir_efetivo_valor,
        "ir_efetivo_fonte": ir_efetivo_fonte,
        # â€” Status atual + histÃ³rico â€”
        "status_atual": (ultimo["status"] if ultimo else "pendente"),
        "ultimo_envio": ultimo,
        "historico_envios": envios,
        "qtd_envios": len(envios),
        # â€” Metadados do evento (fixos) â€”
        "empregador_cnpj_raiz": "05969071",   # APPA
        "tp_amb": "1",                         # produÃ§Ã£o
        "proc_emi": "1",
        "ver_proc": "EasySocial_1.0",
    }


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# GET /xml-cpf/{lote_num}/{per_apur}/{cpf}
# Retorna o XML S-1210 original desse CPF (do ZIP do eSocial) como
# download (application/xml). Ãštil para inspeÃ§Ã£o manual.
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@router.get("/xml-cpf/{lote_num}/{per_apur}/{cpf}")
def baixar_xml_cpf(lote_num: int, per_apur: str, cpf: str):
    import zipfile
    from esocial.s1210_missao_routes import FONTES, _extract_s1210_from_xml

    if lote_num not in (1, 2, 3, 4):
        raise HTTPException(400, "lote_num invÃ¡lido")
    cpf = (cpf or "").strip().replace(".", "").replace("-", "")
    if len(cpf) != 11 or not cpf.isdigit():
        raise HTTPException(400, "CPF invÃ¡lido")
    if per_apur not in FONTES:
        raise HTTPException(400, f"per_apur {per_apur} sem ZIP configurado")

    zpath = FONTES[per_apur]["zip"]
    if not zpath.exists():
        raise HTTPException(404, f"ZIP nÃ£o encontrado em {zpath}")

    melhor_xml: Optional[str] = None
    melhor_dh: str = ""
    with zipfile.ZipFile(str(zpath), "r") as zf:
        for name in zf.namelist():
            if "S-1210" not in name or not name.endswith(".xml"):
                continue
            try:
                raw = zf.read(name).decode("utf-8", errors="replace")
            except Exception:
                continue
            if cpf not in raw:
                continue
            info = _extract_s1210_from_xml(raw)
            if not info or info.get("cpf") != cpf:
                continue
            dh = str(info.get("dh_proc") or "")
            if melhor_xml is None or dh > melhor_dh:
                melhor_xml = raw
                melhor_dh = dh

    if not melhor_xml:
        raise HTTPException(404, f"Nenhum S-1210 encontrado para CPF {cpf} em {per_apur}")

    fname = f"S-1210_{cpf}_{per_apur}.xml"
    return Response(
        content=melhor_xml.encode("utf-8"),
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
# DELETE /xlsx/{id}  â€” remove XLSX + scope (uso administrativo)
# â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
@router.delete("/xlsx/{xlsx_id}")
def remover_xlsx(xlsx_id: int, empresa_id: int = DEFAULT_EMPRESA_ID):
    conn = _db()
    try:
        with conn.cursor() as cur:
            cur.execute(
                "DELETE FROM s1210_xlsx WHERE id=%s AND empresa_id=%s RETURNING per_apur, storage_path",
                (xlsx_id, empresa_id),
            )
            row = cur.fetchone()
        conn.commit()
    finally:
        conn.close()
    if not row:
        raise HTTPException(404, "xlsx nÃ£o encontrada")
    return {"removido": xlsx_id, "per_apur": row[0], "storage_path": row[1]}

# ════════════════════════════════════════════════════════════════════
# Endpoints "Não Aplica" (N/A) — marca CPFs que não devem ser enviados
# ao eSocial (demitidos, sem base IRRF, folha paga em outra
# competência, etc.). Insere registro em s1210_cpf_envios com
# status='na' e motivo em erro_descricao.
# ════════════════════════════════════════════════════════════════════
class MarcarNAItem(BaseModel):
    cpf: str
    motivo: str | None = None


class MarcarNARequest(BaseModel):
    per_apur: str
    lote_num: int
    itens: list[MarcarNAItem]


class MarcarNAResponse(BaseModel):
    per_apur: str
    lote_num: int
    marcados: int
    ignorados_fora_escopo: int


@router.post("/marcar-na", response_model=MarcarNAResponse)
def marcar_na(req: MarcarNARequest, empresa_id: int = DEFAULT_EMPRESA_ID):
    """
    Marca uma lista de CPFs com status='na' (Não Aplica). Apenas CPFs
    presentes no escopo (s1210_cpf_scope) do compartimento são marcados.
    """
    if req.lote_num not in (1, 2, 3, 4):
        raise HTTPException(400, "lote_num inválido")
    if not req.itens:
        return MarcarNAResponse(per_apur=req.per_apur, lote_num=req.lote_num, marcados=0, ignorados_fora_escopo=0)

    cpfs_norm = []
    motivos = {}
    for it in req.itens:
        c = "".join(ch for ch in (it.cpf or "") if ch.isdigit()).zfill(11)
        if len(c) == 11:
            cpfs_norm.append(c)
            motivos[c] = (it.motivo or "").strip() or "marcado como N/A"

    conn = _db()
    marcados = 0
    fora = 0
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT cpf FROM s1210_cpf_scope
                    WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
                      AND cpf = ANY(%s)""",
                (empresa_id, req.per_apur, req.lote_num, cpfs_norm),
            )
            no_escopo = {r[0] for r in cur.fetchall()}
            fora = len(set(cpfs_norm) - no_escopo)
            for c in no_escopo:
                cur.execute(
                    """INSERT INTO s1210_cpf_envios
                        (empresa_id, per_apur, cpf, lote_num, status,
                         erro_descricao, enviado_em)
                       VALUES (%s, %s, %s, %s, 'na', %s, NOW())""",
                    (empresa_id, req.per_apur, c, req.lote_num, motivos.get(c)),
                )
                marcados += 1
        conn.commit()
    finally:
        conn.close()
    return MarcarNAResponse(
        per_apur=req.per_apur, lote_num=req.lote_num,
        marcados=marcados, ignorados_fora_escopo=fora,
    )


class DesmarcarNARequest(BaseModel):
    per_apur: str
    lote_num: int
    cpfs: list[str]


@router.post("/desmarcar-na")
def desmarcar_na(req: DesmarcarNARequest, empresa_id: int = DEFAULT_EMPRESA_ID):
    """
    Remove a marcação 'na' (apaga apenas as linhas de envio com status='na'
    para os CPFs informados — preserva histórico de outras tentativas).
    """
    cpfs_norm = ["".join(ch for ch in c if ch.isdigit()).zfill(11) for c in req.cpfs]
    cpfs_norm = [c for c in cpfs_norm if len(c) == 11]
    if not cpfs_norm:
        return {"removidos": 0}
    conn = _db()
    try:
        with conn.cursor() as cur:
            cur.execute(
                """DELETE FROM s1210_cpf_envios
                    WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
                      AND status='na' AND cpf = ANY(%s)""",
                (empresa_id, req.per_apur, req.lote_num, cpfs_norm),
            )
            n = cur.rowcount
        conn.commit()
    finally:
        conn.close()
    return {"removidos": n}


# ════════════════════════════════════════════════════════════════════
# GET /cpfs-do-mes  → lista CPFs de um mês/lote com status atual
# ════════════════════════════════════════════════════════════════════
@router.get("/cpfs-do-mes")
def cpfs_do_mes(
    per_apur: str,
    empresa_id: int = DEFAULT_EMPRESA_ID,
    lote_num: Optional[int] = None,
):
    """Retorna CPFs do escopo do mês com status atual (último envio) e flag tem_xml.

    tem_xml = True quando raw_row contém 'xml_path' (caso da empresa Soluções,
    onde indexamos os XMLs de retorno do explorador de eventos).
    """
    conn = _db(empresa_id)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            params: list = [empresa_id, per_apur]
            where_lote = ""
            if lote_num is not None:
                where_lote = " AND s.lote_num = %s"
                params.append(lote_num)
            cur.execute(
                f"""
                WITH ult AS (
                    SELECT DISTINCT ON (empresa_id, per_apur, lote_num, cpf)
                           empresa_id, per_apur, lote_num, cpf,
                           status, nr_recibo_usado, nr_recibo_novo,
                           descricao_resposta, erro_descricao, enviado_em
                      FROM s1210_cpf_envios
                     WHERE empresa_id = %s AND per_apur = %s
                     ORDER BY empresa_id, per_apur, lote_num, cpf,
                              enviado_em DESC NULLS LAST, id DESC
                )
                SELECT s.cpf, s.nome, s.matricula, s.lote_num, s.row_number,
                       (s.raw_row ? 'xml_path') AS tem_xml,
                       (s.raw_row ->> 'nr_recibo') AS nr_recibo_xml,
                       u.status,
                       u.nr_recibo_usado,
                       u.nr_recibo_novo,
                       u.descricao_resposta,
                       u.erro_descricao,
                       u.enviado_em
                  FROM s1210_cpf_scope s
                  LEFT JOIN ult u
                    ON u.empresa_id = s.empresa_id
                   AND u.per_apur = s.per_apur
                   AND u.lote_num = s.lote_num
                   AND u.cpf = s.cpf
                 WHERE s.empresa_id = %s AND s.per_apur = %s{where_lote}
                 ORDER BY s.lote_num, s.row_number NULLS LAST, s.cpf
                """,
                (empresa_id, per_apur, *params),
            )
            rows = cur.fetchall()
    finally:
        conn.close()
    # Normaliza: cpf vem como char(11), pode ter padding
    out = []
    for r in rows:
        d = dict(r)
        d["cpf"] = (d.get("cpf") or "").strip()
        # status efetivo: 'pendente' quando não há envio
        if not d.get("status"):
            d["status"] = "pendente"
        out.append(d)
    return {"empresa_id": empresa_id, "per_apur": per_apur, "total": len(out), "cpfs": out}


# ════════════════════════════════════════════════════════════════════
# GET /xml-cpf  → baixa o XML de retorno (S-1210) de um CPF específico
# ════════════════════════════════════════════════════════════════════
@router.get("/xml-cpf")
def baixar_xml_cpf(
    per_apur: str,
    cpf: str,
    empresa_id: int = DEFAULT_EMPRESA_ID,
):
    """Retorna o XML armazenado em raw_row.xml_path (Soluções)."""
    cpf_norm = "".join(ch for ch in cpf if ch.isdigit()).zfill(11)
    if len(cpf_norm) != 11:
        raise HTTPException(status_code=400, detail="cpf inválido")

    conn = _db(empresa_id)
    try:
        with conn.cursor() as cur:
            cur.execute(
                """
                SELECT raw_row ->> 'xml_path'
                  FROM s1210_cpf_scope
                 WHERE empresa_id=%s AND per_apur=%s AND cpf=%s
                 LIMIT 1
                """,
                (empresa_id, per_apur, cpf_norm),
            )
            row = cur.fetchone()
    finally:
        conn.close()

    if not row or not row[0]:
        raise HTTPException(status_code=404, detail="XML não disponível para este CPF")

    xml_path = Path(row[0])
    if not xml_path.is_file():
        raise HTTPException(status_code=404, detail=f"arquivo não encontrado: {xml_path}")

    try:
        data = xml_path.read_bytes()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"erro lendo XML: {e}")

    fname = f"{cpf_norm}_{per_apur}_S-1210.xml"
    return Response(
        content=data,
        media_type="application/xml",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ════════════════════════════════════════════════════════════════════
# POST /anual/dividir-lote → divide CPFs de um lote em 2 lotes iguais
# ════════════════════════════════════════════════════════════════════
class DividirLoteRequest(BaseModel):
    per_apur: str
    lote_origem: int = 1
    lote_destino: int = 2


@router.post("/anual/dividir-lote")
def dividir_lote(req: DividirLoteRequest, empresa_id: int = DEFAULT_EMPRESA_ID):
    """Move metade dos CPFs do lote_origem para lote_destino.

    Bloqueia APPA (empresa_id=1) — fluxo legado usa 4 lotes fixos.
    """
    if empresa_id == 1:
        raise HTTPException(400, "Operação não disponível para APPA (lotes fixos)")
    if req.lote_origem == req.lote_destino:
        raise HTTPException(400, "lote_origem e lote_destino devem ser diferentes")
    if not (1 <= req.lote_destino <= 4):
        raise HTTPException(400, "lote_destino fora do intervalo 1..4")

    conn = _db(empresa_id)
    try:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT cpf FROM s1210_cpf_scope WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s ORDER BY row_number NULLS LAST, cpf",
                (empresa_id, req.per_apur, req.lote_origem),
            )
            cpfs = [r[0] for r in cur.fetchall()]
            if not cpfs:
                raise HTTPException(404, f"lote {req.lote_origem} vazio em {req.per_apur}")

            # destino nao pode ja existir
            cur.execute(
                "SELECT COUNT(*) FROM s1210_cpf_scope WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s",
                (empresa_id, req.per_apur, req.lote_destino),
            )
            if cur.fetchone()[0] > 0:
                raise HTTPException(409, f"lote_destino {req.lote_destino} já existe em {req.per_apur}")

            metade = len(cpfs) // 2
            mover = cpfs[metade:]  # segunda metade vai pro novo lote

            cur.execute(
                """UPDATE s1210_cpf_scope
                      SET lote_num = %s
                    WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
                      AND cpf = ANY(%s)""",
                (req.lote_destino, empresa_id, req.per_apur, req.lote_origem, mover),
            )
            movidos_scope = cur.rowcount

            # tambem move historico de envios pra manter consistencia da view
            cur.execute(
                """UPDATE s1210_cpf_envios
                      SET lote_num = %s
                    WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
                      AND cpf = ANY(%s)""",
                (req.lote_destino, empresa_id, req.per_apur, req.lote_origem, mover),
            )
            movidos_envios = cur.rowcount
        conn.commit()
    finally:
        conn.close()

    return {
        "ok": True,
        "per_apur": req.per_apur,
        "lote_origem": req.lote_origem,
        "lote_destino": req.lote_destino,
        "total_origem_antes": len(cpfs),
        "movidos_scope": movidos_scope,
        "movidos_envios": movidos_envios,
        "restantes_origem": len(cpfs) - movidos_scope,
    }


# ════════════════════════════════════════════════════════════════════
# POST /anual/unir-lotes → desfaz divisão (move tudo do lote N pro 1)
# ════════════════════════════════════════════════════════════════════
class UnirLotesRequest(BaseModel):
    per_apur: str
    lote_destino: int = 1
    lote_origem: int = 2


@router.post("/anual/unir-lotes")
def unir_lotes(req: UnirLotesRequest, empresa_id: int = DEFAULT_EMPRESA_ID):
    """Move todos os CPFs do lote_origem para o lote_destino (default: 2 → 1)."""
    if empresa_id == 1:
        raise HTTPException(400, "Operação não disponível para APPA (lotes fixos)")
    if req.lote_origem == req.lote_destino:
        raise HTTPException(400, "lote_origem e lote_destino devem ser diferentes")

    conn = _db(empresa_id)
    try:
        with conn.cursor() as cur:
            # checa colisao de cpfs entre os dois lotes (UNIQUE empresa_id, per_apur, cpf)
            cur.execute(
                """SELECT COUNT(*) FROM (
                       SELECT cpf FROM s1210_cpf_scope
                        WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
                       INTERSECT
                       SELECT cpf FROM s1210_cpf_scope
                        WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
                   ) x""",
                (empresa_id, req.per_apur, req.lote_origem,
                 empresa_id, req.per_apur, req.lote_destino),
            )
            colisoes = cur.fetchone()[0]
            if colisoes > 0:
                raise HTTPException(
                    409,
                    f"{colisoes} CPFs estão duplicados entre lote {req.lote_origem} e {req.lote_destino}",
                )

            cur.execute(
                """UPDATE s1210_cpf_scope SET lote_num=%s
                    WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s""",
                (req.lote_destino, empresa_id, req.per_apur, req.lote_origem),
            )
            movidos_scope = cur.rowcount
            cur.execute(
                """UPDATE s1210_cpf_envios SET lote_num=%s
                    WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s""",
                (req.lote_destino, empresa_id, req.per_apur, req.lote_origem),
            )
            movidos_envios = cur.rowcount
        conn.commit()
    finally:
        conn.close()

    return {
        "ok": True,
        "per_apur": req.per_apur,
        "movidos_scope": movidos_scope,
        "movidos_envios": movidos_envios,
    }