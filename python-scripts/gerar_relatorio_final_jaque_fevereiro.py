from __future__ import annotations

import csv
import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
BACKEND_V2 = Path(r"C:\Users\xandao\Documents\GitHub\Easy-eSocial-v2\backend")
if str(BACKEND_V2) not in sys.path:
    sys.path.insert(0, str(BACKEND_V2))

from app import db, tenant  # noqa: E402


EMPRESA_ID = 2
PER_APUR = "2025-02"
PER_LABEL = "Fevereiro/2025"
OUT_DIR = ROOT / "relatorio_ana" / "relatório final Jaque"
OUT_XLSX = OUT_DIR / "2025-02_fevereiro_relatorio_final_jaque.xlsx"
SAVED_XLSX = OUT_XLSX
MANIFEST = ROOT / "relatorio_ana" / "CORRECAO_JAQUE_PLANO_PENSAO_2025" / PER_APUR / "manifest_correcao_jaque_2025-02.json"
PENDENCIAS_JAQUE = ROOT / "relatorio_ana" / "CORRECAO_JAQUE_PLANO_PENSAO_2025" / PER_APUR / "pendencias_finais_fevereiro.csv"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SUBHEADER_FILL = PatternFill("solid", fgColor="D9EAF7")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
NOTE_BORDER = Border(
    left=Side(style="thin", color="D6B656"),
    right=Side(style="thin", color="D6B656"),
    top=Side(style="thin", color="D6B656"),
    bottom=Side(style="thin", color="D6B656"),
)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def cpf_mask(cpf: str | None) -> str:
    digits = re.sub(r"\D", "", str(cpf or ""))
    if len(digits) != 11:
        return str(cpf or "")
    return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"


def compact_message(message: str | None, limit: int = 240) -> str:
    text = re.sub(r"\s+", " ", str(message or "")).strip()
    return text if len(text) <= limit else text[: limit - 1].rstrip() + "…"


def extract_dependente(message: str | None) -> str:
    match = re.search(r"CPF do dependente\s+(\d{11})\s+inválido", str(message or ""), re.IGNORECASE)
    return match.group(1) if match else ""


def load_manifest_scope() -> tuple[set[str], dict[str, dict[str, str]]]:
    data = json.loads(MANIFEST.read_text(encoding="utf-8"))
    scope: set[str] = set()
    missing: dict[str, dict[str, str]] = {}
    for item in data.get("targets") or []:
        cpf = str(item.get("cpf") or "")
        if cpf:
            scope.add(cpf)
    for item in data.get("missing_valid_responses") or []:
        cpf = str(item.get("cpf") or "")
        if cpf:
            scope.add(cpf)
            missing[cpf] = {key: str(value or "") for key, value in item.items()}
    if PENDENCIAS_JAQUE.exists():
        with PENDENCIAS_JAQUE.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                cpf = str(row.get("cpf") or "")
                if cpf:
                    scope.add(cpf)
    return scope, missing


def latest_s1210_errors() -> list[dict[str, Any]]:
    internal_empresa_id = tenant.internal_empresa_id(EMPRESA_ID)
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(
                """
                WITH latest AS (
                    SELECT DISTINCT ON (it.cpf)
                           it.cpf, it.status, it.erro_codigo, it.erro_mensagem,
                           it.nr_recibo_anterior, it.nr_recibo_novo,
                           it.id AS item_id, te.id AS envio_id, it.criado_em
                      FROM timeline_envio_item it
                      JOIN timeline_envio te ON te.id = it.timeline_envio_id
                      JOIN timeline_mes tm ON tm.id = te.timeline_mes_id
                     WHERE tm.empresa_id = %s
                       AND tm.per_apur = %s
                       AND it.tipo_evento = 'S-1210'
                     ORDER BY it.cpf, it.criado_em DESC NULLS LAST, it.id DESC
                )
                SELECT * FROM latest
                 WHERE status = 'erro_esocial'
                 ORDER BY cpf
                """,
                (internal_empresa_id, PER_APUR),
            )
            return [dict(row) for row in cursor.fetchall()]
    finally:
        conn.close()


def classify(message: str | None) -> str:
    text = str(message or "")
    if "Plano de saúde coletivo" in text:
        return "Plano de saúde sem dados"
    if "pensão alimentícia" in text:
        return "Pensão alimentícia exigida"
    if "1861" in text or "CPF do dependente" in text:
        return "CPF dependente inválido"
    if "459" in text or "157" in text or "recibo de entrega" in text or "mesmo tipo" in text or "mesmo objeto" in text:
        return "Recibo anterior inválido"
    return "Outros"


def action_for(category: str) -> tuple[str, str]:
    if category == "Plano de saúde sem dados":
        return "Jaque", "Enviar CNPJ da operadora, registro ANS e valor titular válidos."
    if category == "Pensão alimentícia exigida":
        return "Jaque/Alex", "Confirmar se há pensão real; se houver, informar beneficiário, tipo e valor maior que zero."
    if category == "CPF dependente inválido":
        return "Jaque/DP", "Confirmar CPF do dependente e dados para cadastro/declaração no S-1210."
    if category == "Recibo anterior inválido":
        return "Alex", "Informar o recibo ativo correto do S-1210, mesmo CPF, mesmo período e mesmo objeto."
    return "A definir", "Analisar mensagem do eSocial."


def build_rows(errors: list[dict[str, Any]], jaque_scope: set[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for item in errors:
        cpf = str(item.get("cpf") or "")
        category = classify(item.get("erro_mensagem"))
        owner, action = action_for(category)
        scope = "Pacote Jaque" if cpf in jaque_scope else "Fora do pacote Jaque"
        rows.append(
            {
                "cpf": cpf,
                "cpf_formatado": cpf_mask(cpf),
                "escopo": scope,
                "tipo_erro": category,
                "codigo": str(item.get("erro_codigo") or ""),
                "cpf_dependente": extract_dependente(item.get("erro_mensagem")),
                "cpf_dependente_formatado": cpf_mask(extract_dependente(item.get("erro_mensagem"))),
                "responsavel": owner,
                "acao": action,
                "envio_id": item.get("envio_id"),
                "item_id": item.get("item_id"),
                "mensagem": compact_message(item.get("erro_mensagem"), 320),
            }
        )
    return rows


def add_note(ws, title: str, body: str, start_col: int = 8, start_row: int = 1, width_cols: int = 5, height_rows: int = 11) -> None:
    end_col = start_col + width_cols - 1
    end_row = start_row + height_rows - 1
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    title_cell = ws.cell(start_row, start_col, title)
    title_cell.fill = HEADER_FILL
    title_cell.font = HEADER_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=end_row, end_column=end_col)
    body_cell = ws.cell(start_row + 1, start_col, body)
    body_cell.fill = NOTE_FILL
    body_cell.border = NOTE_BORDER
    body_cell.alignment = Alignment(wrap_text=True, vertical="top")
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).border = NOTE_BORDER
    for col in range(start_col, end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    for row in range(start_row + 1, end_row + 1):
        ws.row_dimensions[row].height = 28


def setup_sheet(ws, title: str | None = None) -> None:
    ws.sheet_view.showGridLines = False
    if title:
        ws["A1"] = title
        ws["A1"].font = Font(size=15, bold=True, color="1F4E78")


def write_table(ws, start_row: int, headers: list[str], rows: list[list[Any]], table_name: str) -> int:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(start_row, col_idx, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    for row_idx, row_values in enumerate(rows, start=start_row + 1):
        for col_idx, value in enumerate(row_values, start=1):
            cell = ws.cell(row_idx, col_idx, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFD")
    end_row = start_row + max(len(rows), 1)
    end_col = len(headers)
    if rows:
        ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
        table = Table(displayName=table_name, ref=ref)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        table.tableStyleInfo = style
        ws.add_table(table)
    ws.freeze_panes = ws.cell(start_row + 1, 1)
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(end_col)}{end_row}"
    return end_row


def autosize(ws, widths: dict[int, int] | None = None) -> None:
    widths = widths or {}
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        if col_idx in widths:
            ws.column_dimensions[letter].width = widths[col_idx]
            continue
        max_len = 10
        for cell in ws[letter]:
            value = str(cell.value or "")
            max_len = max(max_len, min(len(value) + 2, 42))
        ws.column_dimensions[letter].width = max_len


def write_summary(wb: Workbook, rows: list[dict[str, Any]]) -> None:
    ws = wb.active
    ws.title = "Resumo"
    setup_sheet(ws, f"Relatório final Jaque - {PER_LABEL}")
    counts_scope = Counter(row["escopo"] for row in rows)
    counts_type = Counter(row["tipo_erro"] for row in rows)
    summary_rows = [
        ["Total de erros no Front V2", len(rows)],
        ["Dentro do pacote Jaque", counts_scope.get("Pacote Jaque", 0)],
        ["Fora do pacote Jaque", counts_scope.get("Fora do pacote Jaque", 0)],
        ["Plano de saúde", counts_type.get("Plano de saúde sem dados", 0)],
        ["Pensão alimentícia", counts_type.get("Pensão alimentícia exigida", 0)],
        ["CPF dependente inválido", counts_type.get("CPF dependente inválido", 0)],
        ["Recibo anterior inválido", counts_type.get("Recibo anterior inválido", 0)],
    ]
    write_table(ws, 3, ["Indicador", "Quantidade"], summary_rows, "ResumoFevereiro")
    add_note(
        ws,
        "Como ler este mês",
        "O Front V2 mostra 10 erros porque ele considera o último status de todos os S-1210 de fevereiro. "
        "Dentro do pacote tratado com a última resposta da Jaque ficaram 5 pendências. "
        "As outras 5 são pendências operacionais antigas ou fora da última tabela enviada. "
        "As abas seguintes separam os erros por causa para facilitar a correção.",
        start_col=5,
        start_row=3,
        width_cols=6,
        height_rows=10,
    )
    ws["A13"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A13"].font = Font(italic=True, color="666666")
    autosize(ws, {1: 34, 2: 16})


def table_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [
            row["cpf_formatado"],
            row["escopo"],
            row["tipo_erro"],
            row["cpf_dependente_formatado"] if row["cpf_dependente"] else "",
            row["responsavel"],
            row["acao"],
            row["mensagem"],
        ]
        for row in rows
    ]


def write_error_sheet(wb: Workbook, name: str, title: str, rows: list[dict[str, Any]], note: str, table_name: str) -> None:
    ws = wb.create_sheet(name)
    setup_sheet(ws, title)
    headers = ["CPF", "Escopo", "Tipo", "CPF dependente", "Responsável", "Próxima ação", "Mensagem curta"]
    write_table(ws, 3, headers, table_rows(rows), table_name)
    add_note(ws, "Por que não passou", note, start_col=9, start_row=3, width_cols=5, height_rows=12)
    autosize(ws, {1: 17, 2: 22, 3: 25, 4: 18, 5: 16, 6: 38, 7: 52})


def build_workbook(rows: list[dict[str, Any]]) -> None:
    global SAVED_XLSX
    wb = Workbook()
    write_summary(wb, rows)

    jaque_rows = [row for row in rows if row["escopo"] == "Pacote Jaque"]
    fora_rows = [row for row in rows if row["escopo"] != "Pacote Jaque"]
    write_error_sheet(
        wb,
        "Front V2 - 10",
        "Todos os erros atuais do Front V2 - Fevereiro/2025",
        rows,
        "Esta aba é a fotografia do Front V2: último status pendente por CPF. "
        "Ela junta pendências da última tabela da Jaque com pendências antigas de recibo/dependente.",
        "FrontV2Dez",
    )
    write_error_sheet(
        wb,
        "Pacote Jaque - 5",
        "Pendências do pacote Jaque - Fevereiro/2025",
        jaque_rows,
        "Estes 5 CPFs são os que sobraram após aplicar as respostas válidas da Jaque. "
        "Plano: faltou dado obrigatório. Pensão: a resposta veio zerada ou incompatível com a regra do eSocial.",
        "PacoteJaqueCinco",
    )
    write_error_sheet(
        wb,
        "Fora pacote - 5",
        "Pendências fora do pacote Jaque - Fevereiro/2025",
        fora_rows,
        "Estes 5 aparecem no Front V2, mas não vieram como correção válida na última tabela da Jaque. "
        "Dois dependem de recibo ativo correto; três dependem de CPF/dados de dependente.",
        "ForaPacoteCinco",
    )

    notes = {
        "Plano de saúde sem dados": "A planilha não trouxe CNPJ da operadora, registro ANS e/ou valor titular válidos. Sem estes campos, o S-1210 não monta o grupo de plano de saúde exigido pelo eSocial.",
        "Pensão alimentícia exigida": "A tentativa com valor zero foi rejeitada pelo schema. A tentativa sem o grupo também foi rejeitada pelo eSocial, que continua exigindo beneficiário de pensão para estes CPFs.",
        "CPF dependente inválido": "O CPF do dependente indicado na mensagem não está aceito pelo eSocial para o trabalhador: precisa estar no RET ou ser declarado corretamente no próprio S-1210.",
        "Recibo anterior inválido": "A retificação precisa apontar para o recibo ativo do mesmo evento S-1210, mesmo CPF e mesmo período. Se o recibo informado for de outro tipo/objeto, o eSocial retorna erro 157; se estiver excluído/retificado, retorna erro 459.",
    }
    sheet_names = {
        "Plano de saúde sem dados": "Plano",
        "Pensão alimentícia exigida": "Pensao",
        "CPF dependente inválido": "Dependente",
        "Recibo anterior inválido": "Recibo",
    }
    table_names = {
        "Plano de saúde sem dados": "ErrosPlano",
        "Pensão alimentícia exigida": "ErrosPensao",
        "CPF dependente inválido": "ErrosDependente",
        "Recibo anterior inválido": "ErrosRecibo",
    }
    for category in sheet_names:
        category_rows = [row for row in rows if row["tipo_erro"] == category]
        write_error_sheet(wb, sheet_names[category], category, category_rows, notes[category], table_names[category])

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(OUT_XLSX)
        SAVED_XLSX = OUT_XLSX
    except PermissionError:
        fallback = OUT_DIR / f"{OUT_XLSX.stem}_atualizado_{datetime.now().strftime('%Y%m%d_%H%M%S')}{OUT_XLSX.suffix}"
        wb.save(fallback)
        SAVED_XLSX = fallback


def main() -> int:
    jaque_scope, _missing = load_manifest_scope()
    errors = latest_s1210_errors()
    rows = build_rows(errors, jaque_scope)
    build_workbook(rows)
    print(json.dumps({"ok": True, "per_apur": PER_APUR, "arquivo": str(SAVED_XLSX), "total_erros": len(rows), "por_tipo": dict(Counter(row["tipo_erro"] for row in rows)), "por_escopo": dict(Counter(row["escopo"] for row in rows))}, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())