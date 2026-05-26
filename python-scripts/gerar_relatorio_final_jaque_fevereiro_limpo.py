from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import psycopg2.extras
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
BACKEND_V2 = Path(r"C:\Users\xandao\Documents\GitHub\Easy-eSocial-v2\backend")
if str(BACKEND_V2) not in sys.path:
    sys.path.insert(0, str(BACKEND_V2))

from app import db, tenant  # noqa: E402


EMPRESA_ID = 2
PER_APUR = "2025-02"
OUT_DIR = ROOT / "relatorio_ana" / "relatório final Jaque"
OUT_XLSX = OUT_DIR / "2025-02_relatorio_final_jaque.xlsx"
FEB_RESPONSE_XLSX = ROOT / "relatorio_ana" / "RESPOSTAS_JAQUE_PLANO_PENSAO_2025_FEV_OUT" / "SOLUCOES_FEVEREIRO_2025_JAQUE_PLANO_PENSAO_PREENCHER.xlsx"

PLAN_CPFS = ["30729903877", "33163519814", "42852580802"]
PENSION_CPFS = ["28710124829", "71985883104"]
DEPENDENT_CPFS = ["10423668455", "38434368862", "81529368553"]
RECEIPT_CPFS = ["36832724810", "93564139249"]
ALL_CPFS = PLAN_CPFS + PENSION_CPFS + DEPENDENT_CPFS + RECEIPT_CPFS

RECEIPTS_TESTED = {
    "36832724810": "1.1.0000000031405347921",
    "93564139249": "1.1.0000000031331770805",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SOFT_FILL = PatternFill("solid", fgColor="EAF3F8")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
THIN = Side(style="thin", color="D9E2F3")
NOTE_SIDE = Side(style="thin", color="D6B656")
TABLE_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NOTE_BORDER = Border(left=NOTE_SIDE, right=NOTE_SIDE, top=NOTE_SIDE, bottom=NOTE_SIDE)


def cpf_digits(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits.zfill(11) if digits else ""


def cpf_mask(cpf: str) -> str:
    digits = cpf_digits(cpf)
    if len(digits) != 11:
        return str(cpf or "")
    return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def short_error(message: str, limit: int = 230) -> str:
    text = clean_text(message)
    return text if len(text) <= limit else text[: limit - 1].rstrip() + "…"


def extract_dependents(message: str) -> list[str]:
    return re.findall(r"CPF do dependente\s+(\d{11})\s+inválido", message or "", flags=re.IGNORECASE)


def latest_errors() -> dict[str, dict[str, Any]]:
    internal_empresa_id = tenant.internal_empresa_id(EMPRESA_ID)
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(
                """
                WITH latest AS (
                    SELECT DISTINCT ON (it.cpf)
                           it.cpf, it.status, it.erro_codigo, it.erro_mensagem,
                           it.nr_recibo_anterior, it.nr_recibo_novo,
                           it.id AS item_id, te.id AS envio_id, it.criado_em
                      FROM timeline_envio_item it
                      JOIN timeline_envio te ON te.id = it.timeline_envio_id
                      JOIN timeline_mes tm ON tm.id = te.timeline_mes_id
                     WHERE tm.empresa_id = %s
                       AND tm.per_apur = %s
                       AND it.tipo_evento = 'S-1210'
                       AND it.cpf = ANY(%s)
                     ORDER BY it.cpf, it.criado_em DESC NULLS LAST, it.id DESC
                )
                SELECT * FROM latest ORDER BY cpf
                """,
                (internal_empresa_id, PER_APUR, ALL_CPFS),
            )
            return {str(row["cpf"]): dict(row) for row in cursor.fetchall()}
    finally:
        conn.close()


def response_rows() -> tuple[dict[str, dict[str, Any]], dict[str, dict[str, Any]]]:
    wb = load_workbook(FEB_RESPONSE_XLSX, data_only=True)
    plan: dict[str, dict[str, Any]] = {}
    pension: dict[str, dict[str, Any]] = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header_idx = None
        for idx, row in enumerate(rows[:20]):
            if any("CPF" in clean_text(value).upper() for value in row):
                header_idx = idx
                break
        if header_idx is None:
            continue
        headers = [clean_text(value) or f"col_{idx + 1}" for idx, value in enumerate(rows[header_idx])]
        for row in rows[header_idx + 1:]:
            data = {headers[idx]: row[idx] if idx < len(row) else None for idx in range(len(headers))}
            cpfs_in_row = {cpf_digits(value) for key, value in data.items() if "CPF" in key.upper()}
            for cpf in PLAN_CPFS:
                if cpf in cpfs_in_row:
                    plan[cpf] = data
            for cpf in PENSION_CPFS:
                if cpf in cpfs_in_row:
                    pension[cpf] = data
    return plan, pension


def value(data: dict[str, Any], key: str) -> Any:
    return data.get(key)


def add_title(ws, title: str, subtitle: str | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(size=15, bold=True, color="1F4E78")
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, italic=True, color="666666")


def write_table(ws, start_row: int, headers: list[str], rows: list[list[Any]], name: str) -> int:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = TABLE_BORDER
    for row_idx, row_values in enumerate(rows, start_row + 1):
        for col, cell_value in enumerate(row_values, 1):
            cell = ws.cell(row_idx, col, cell_value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = TABLE_BORDER
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFD")
    end_row = start_row + max(len(rows), 1)
    if rows:
        table = Table(displayName=name, ref=f"A{start_row}:{get_column_letter(len(headers))}{end_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    ws.freeze_panes = ws.cell(start_row + 1, 1)
    return end_row


def add_note(ws, text: str, start_col: int, start_row: int = 3, cols: int = 5, rows: int = 8) -> None:
    end_col = start_col + cols - 1
    end_row = start_row + rows - 1
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    title = ws.cell(start_row, start_col, "Por que não passou")
    title.fill = HEADER_FILL
    title.font = HEADER_FONT
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=end_row, end_column=end_col)
    body = ws.cell(start_row + 1, start_col, text)
    body.fill = NOTE_FILL
    body.border = NOTE_BORDER
    body.alignment = Alignment(wrap_text=True, vertical="top")
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).border = NOTE_BORDER
    for col in range(start_col, end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 17
    for row in range(start_row + 1, end_row + 1):
        ws.row_dimensions[row].height = 26


def autosize(ws, widths: dict[int, int]) -> None:
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 18)


def build_summary(ws, counts: Counter[str]) -> None:
    add_title(ws, "Relatório final Jaque - Fevereiro/2025", "Arquivo limpo, somente pendências finais por tipo de correção.")
    ws.merge_cells("A4:F8")
    msg = ws["A4"]
    msg.value = (
        "Fevereiro ficou com 10 pendências atuais no S-1210. "
        "Para a devolução da Jaque, os 5 pontos principais são 3 planos de saúde sem dados válidos "
        "e 2 pensões com informação zerada/incompatível. As outras 5 pendências estão separadas porque dependem "
        "de CPF de dependente ou recibo ativo correto. As abas abaixo já estão no formato de cobrança da informação."
    )
    msg.fill = SOFT_FILL
    msg.alignment = Alignment(wrap_text=True, vertical="top")
    msg.border = TABLE_BORDER
    for row in range(4, 9):
        ws.row_dimensions[row].height = 26
        for col in range(1, 7):
            ws.cell(row, col).border = TABLE_BORDER
            ws.cell(row, col).fill = SOFT_FILL
    summary = [
        ["Plano de saúde", counts["plano"], "Jaque precisa reenviar CNPJ operadora, Registro ANS e valor titular."],
        ["Pensão alimentícia", counts["pensao"], "Jaque precisa confirmar beneficiário e valor real; zero não passa."],
        ["CPF dependente inválido", counts["dependente_trabalhador"], "Precisa validar/corrigir os CPFs de dependente informados pelo eSocial."],
        ["Recibo", counts["recibo"], "Os recibos testados retornaram 157; precisa recibo do mesmo objeto S-1210."],
    ]
    write_table(ws, 11, ["Tipo", "Qtd. trabalhadores", "Próxima ação"], summary, "FevereiroResumo")
    ws["A18"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A18"].font = Font(italic=True, color="666666")
    autosize(ws, {1: 26, 2: 16, 3: 70})


def build_plan(ws, plan_data: dict[str, dict[str, Any]], latest: dict[str, dict[str, Any]]) -> None:
    add_title(ws, "Plano de saúde - Fevereiro/2025")
    headers = ["CPF", "CPF Normalizado", "Nome Trabalhador", "CNPJ Operadora", "Registro ANS", "Valor Titular Descontado em Folha", "O que falta corrigir"]
    rows = []
    for cpf in PLAN_CPFS:
        data = plan_data.get(cpf, {})
        rows.append([
            value(data, "CPF") or cpf_mask(cpf),
            cpf,
            clean_text(value(data, "Nome Trabalhador")),
            value(data, "CNPJ Operadora"),
            value(data, "Registro ANS"),
            value(data, "Valor Titular Descontado em Folha"),
            "CNPJ precisa ser numérico válido; Registro ANS e valor titular precisam estar preenchidos.",
        ])
    write_table(ws, 3, headers, rows, "PlanoSaude")
    add_note(ws, "O eSocial recusou porque o grupo de plano de saúde coletivo não pôde ser montado. Na resposta recebida, CNPJ veio como texto 'ok' e Registro ANS/valor titular ficaram vazios.", 9)
    autosize(ws, {1: 18, 2: 17, 3: 34, 4: 20, 5: 18, 6: 24, 7: 46})


def build_pension(ws, pension_data: dict[str, dict[str, Any]], latest: dict[str, dict[str, Any]]) -> None:
    add_title(ws, "Pensão alimentícia - Fevereiro/2025")
    headers = ["CPF", "CPF Normalizado", "Nome Trabalhador", "CPF Beneficiário 1", "Tipo Rendimento 1", "Percentual 1", "Valor Deduzido 1", "O que falta corrigir"]
    rows = []
    for cpf in PENSION_CPFS:
        data = pension_data.get(cpf, {})
        rows.append([
            value(data, "CPF") or cpf_mask(cpf),
            cpf,
            clean_text(value(data, "Nome Trabalhador")),
            value(data, "CPF Beneficiario 1"),
            value(data, "Tipo Rendimento 1"),
            value(data, "Percentual 1"),
            value(data, "Valor Deduzido 1"),
            "Valor deduzido não pode ser zero. Confirmar se há pensão e informar valor correto maior que zero.",
        ])
    write_table(ws, 3, headers, rows, "PensaoAlimenticia")
    add_note(ws, "A resposta trouxe beneficiário, mas valor/percentual zero. O XML com zero não passa no schema, e o eSocial também recusou a retirada do grupo de pensão para estes CPFs.", 10)
    autosize(ws, {1: 18, 2: 17, 3: 34, 4: 20, 5: 18, 6: 14, 7: 16, 8: 48})


def build_dependent(ws, latest: dict[str, dict[str, Any]]) -> None:
    add_title(ws, "CPF dependente inválido - Fevereiro/2025")
    headers = ["CPF Trabalhador", "Nome Trabalhador", "CPF Dependente", "Erro eSocial", "Informação necessária"]
    rows = []
    for cpf in DEPENDENT_CPFS:
        item = latest.get(cpf, {})
        dependents = extract_dependents(str(item.get("erro_mensagem") or "")) or [""]
        for dep in dependents:
            rows.append([
                cpf_mask(cpf),
                "",
                cpf_mask(dep) if dep else "",
                short_error(str(item.get("erro_mensagem") or ""), 180),
                "Confirmar CPF correto do dependente e dados para cadastro/declaração no S-1210.",
            ])
    write_table(ws, 3, headers, rows, "DependenteInvalido")
    add_note(ws, "O eSocial não aceitou o CPF do dependente. O CPF precisa existir no RET do eSocial ou ser informado corretamente no próprio evento, conforme a regra do S-1210.", 7)
    autosize(ws, {1: 18, 2: 30, 3: 18, 4: 54, 5: 42})


def build_receipt(ws, latest: dict[str, dict[str, Any]]) -> None:
    add_title(ws, "Recibo - Fevereiro/2025")
    headers = ["CPF", "Recibo testado", "Resultado", "O que precisa agora", "Erro eSocial"]
    rows = []
    for cpf in RECEIPT_CPFS:
        item = latest.get(cpf, {})
        rows.append([
            cpf_mask(cpf),
            RECEIPTS_TESTED.get(cpf, item.get("nr_recibo_anterior") or ""),
            "Não resolveu: retornou 401/157",
            "Enviar recibo ativo do mesmo S-1210, mesmo CPF, mesmo período e mesmo objeto/demonstrativo.",
            short_error(str(item.get("erro_mensagem") or ""), 220),
        ])
    write_table(ws, 3, headers, rows, "Recibo")
    add_note(ws, "Os dois recibos informados foram testados. O erro mudou para 157, indicando que o recibo não corresponde ao mesmo objeto do S-1210 que está sendo retificado.", 7)
    autosize(ws, {1: 18, 2: 28, 3: 28, 4: 52, 5: 56})


def build_workbook() -> dict[str, Any]:
    plan_data, pension_data = response_rows()
    latest = latest_errors()
    counts = Counter({
        "plano": len(PLAN_CPFS),
        "pensao": len(PENSION_CPFS),
        "dependente_trabalhador": len(DEPENDENT_CPFS),
        "recibo": len(RECEIPT_CPFS),
    })
    wb = Workbook()
    wb.remove(wb.active)
    build_summary(wb.create_sheet("Fevereiro"), counts)
    build_plan(wb.create_sheet("Plano de saude"), plan_data, latest)
    build_pension(wb.create_sheet("Pensao alimenticia"), pension_data, latest)
    build_dependent(wb.create_sheet("Dependente invalido"), latest)
    build_receipt(wb.create_sheet("Recibo"), latest)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_XLSX)
    return {
        "arquivo": str(OUT_XLSX),
        "abas": wb.sheetnames,
        "contagens": dict(counts),
        "plano_linhas": len(plan_data),
        "pensao_linhas": len(pension_data),
    }


def main() -> int:
    print(json.dumps(build_workbook(), ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())