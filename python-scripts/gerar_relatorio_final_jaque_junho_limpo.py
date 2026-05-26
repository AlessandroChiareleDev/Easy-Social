from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import psycopg2.extras
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
BACKEND_V2 = Path(r"C:\Users\xandao\Documents\GitHub\Easy-eSocial-v2\backend")
if str(BACKEND_V2) not in sys.path:
    sys.path.insert(0, str(BACKEND_V2))

from app import db, tenant  # noqa: E402


EMPRESA_ID = 2
PER_APUR = "2025-06"
OUT_DIR = ROOT / "relatorio_ana" / "relatório final Jaque"
OUT_XLSX = OUT_DIR / "2025-06_relatorio_final_jaque.xlsx"
RESPONSE_DIR = ROOT / "relatorio_ana" / "RESPOSTAS_JAQUE_PLANO_PENSAO_2025_FEV_OUT"
RESPONSE_GLOB = "SOLUCOES_JUNHO_2025_JAQUE_PLANO_PENSAO_PREENCHER*.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SOFT_FILL = PatternFill("solid", fgColor="EAF3F8")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9E2F3")
NOTE_SIDE = Side(style="thin", color="D6B656")
TABLE_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NOTE_BORDER = Border(left=NOTE_SIDE, right=NOTE_SIDE, top=NOTE_SIDE, bottom=NOTE_SIDE)


def cpf_digits(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits.zfill(11)[-11:] if digits else ""


def cpf_mask(value: Any) -> str:
    digits = cpf_digits(value)
    if len(digits) != 11:
        return str(value or "")
    return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def short_error(message: str, limit: int = 230) -> str:
    text = clean_text(message)
    return text if len(text) <= limit else text[: limit - 1].rstrip() + "..."


def response_workbook_path() -> Path:
    files = sorted(RESPONSE_DIR.glob(RESPONSE_GLOB))
    if not files:
        raise FileNotFoundError(f"Planilha de resposta de junho nao encontrada: {RESPONSE_GLOB}")
    return files[-1]


def latest_errors() -> list[dict[str, Any]]:
    internal_empresa_id = tenant.internal_empresa_id(EMPRESA_ID)
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(
                """
                WITH latest AS (
                    SELECT DISTINCT ON (it.cpf)
                           it.cpf, it.status, it.erro_codigo, it.erro_mensagem,
                           it.nr_recibo_anterior, it.nr_recibo_novo,
                           it.id AS item_id, te.id AS envio_id, it.criado_em
                      FROM timeline_envio_item it
                      JOIN timeline_envio te ON te.id = it.timeline_envio_id
                      JOIN timeline_mes tm ON tm.id = te.timeline_mes_id
                     WHERE tm.empresa_id = %s
                       AND tm.per_apur = %s
                       AND it.tipo_evento = 'S-1210'
                     ORDER BY it.cpf, it.criado_em DESC NULLS LAST, it.id DESC
                )
                SELECT * FROM latest
                 WHERE status <> 'sucesso'
                   AND COALESCE(erro_codigo, '') <> '202'
                 ORDER BY cpf
                """,
                (internal_empresa_id, PER_APUR),
            )
            return [dict(row) for row in cursor.fetchall()]
    finally:
        conn.close()


def is_plan_error(message: str) -> bool:
    lowered = clean_text(message).lower()
    return "plano de saúde coletivo" in lowered or "plano de saude coletivo" in lowered


def response_rows(cpfs: set[str]) -> dict[str, dict[str, Any]]:
    workbook_path = response_workbook_path()
    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        output: dict[str, dict[str, Any]] = {}
        for ws in workbook.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header_index = None
            for index, row in enumerate(rows[:20]):
                if any("CPF" in clean_text(value).upper() for value in row):
                    header_index = index
                    break
            if header_index is None:
                continue
            headers = [clean_text(value) or f"col_{index + 1}" for index, value in enumerate(rows[header_index])]
            for row in rows[header_index + 1 :]:
                data = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
                cpfs_in_row = {cpf_digits(value) for key, value in data.items() if "CPF" in key.upper()}
                for cpf in cpfs_in_row & cpfs:
                    output.setdefault(cpf, data)
        return output
    finally:
        workbook.close()


def add_title(ws, title: str, subtitle: str | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(size=15, bold=True, color="1F4E78")
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, italic=True, color="666666")


def write_table(ws, start_row: int, headers: list[str], rows: list[list[Any]], name: str) -> int:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = TABLE_BORDER
    for row_index, row_values in enumerate(rows, start_row + 1):
        for col, cell_value in enumerate(row_values, 1):
            cell = ws.cell(row_index, col, cell_value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = TABLE_BORDER
            if row_index % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFD")
    end_row = start_row + max(len(rows), 1)
    if rows:
        table = Table(displayName=name, ref=f"A{start_row}:{get_column_letter(len(headers))}{end_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    ws.freeze_panes = ws.cell(start_row + 1, 1)
    return end_row


def add_note(ws, text: str, start_col: int, start_row: int = 3, cols: int = 5, rows: int = 8) -> None:
    end_col = start_col + cols - 1
    end_row = start_row + rows - 1
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    title = ws.cell(start_row, start_col, "Por que não passou")
    title.fill = HEADER_FILL
    title.font = HEADER_FONT
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=end_row, end_column=end_col)
    body = ws.cell(start_row + 1, start_col, text)
    body.fill = NOTE_FILL
    body.border = NOTE_BORDER
    body.alignment = Alignment(wrap_text=True, vertical="top")
    for row_index in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row_index, col).border = NOTE_BORDER
    for col in range(start_col, end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 17
    for row_index in range(start_row + 1, end_row + 1):
        ws.row_dimensions[row_index].height = 26


def autosize(ws, widths: dict[int, int]) -> None:
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 18)


def build_summary(ws, counts: Counter[str]) -> None:
    add_title(ws, "Relatório final Jaque - Junho/2025", "Pendências finais após as correções possíveis do mês.")
    ws.merge_cells("A4:F8")
    msg = ws["A4"]
    msg.value = (
        "Junho ficou com 5 pendências atuais no S-1210. Foram resolvidas as respostas válidas de plano/pensão "
        "e os dependentes 1861 corrigíveis com infoDep. Restaram somente planos de saúde sem dados válidos "
        "na resposta da Jaque: a planilha traz 'ok' no campo CNPJ Operadora, mas não traz Registro ANS nem valor titular."
    )
    msg.fill = SOFT_FILL
    msg.alignment = Alignment(wrap_text=True, vertical="top")
    msg.border = TABLE_BORDER
    for row_index in range(4, 9):
        ws.row_dimensions[row_index].height = 26
        for col in range(1, 7):
            ws.cell(row_index, col).border = TABLE_BORDER
            ws.cell(row_index, col).fill = SOFT_FILL
    rows = [
        ["Plano de saúde", counts["plano"], "Jaque precisa reenviar CNPJ operadora numérico, Registro ANS e valor titular."],
    ]
    write_table(ws, 11, ["Tipo", "Qtd. trabalhadores", "Próxima ação"], rows, "JunhoResumo")
    ws["A16"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A16"].font = Font(italic=True, color="666666")
    autosize(ws, {1: 26, 2: 16, 3: 76})


def build_plan(ws, plan_rows: dict[str, dict[str, Any]], latest: dict[str, dict[str, Any]]) -> None:
    add_title(ws, "Plano de saúde - Junho/2025")
    headers = [
        "CPF",
        "CPF Normalizado",
        "Nome Trabalhador",
        "CNPJ Operadora",
        "Registro ANS",
        "Valor Titular Descontado em Folha",
        "O que falta corrigir",
        "Erro eSocial",
    ]
    rows = []
    for cpf in sorted(latest):
        data = plan_rows.get(cpf, {})
        item = latest[cpf]
        rows.append(
            [
                data.get("CPF") or cpf_mask(cpf),
                cpf,
                clean_text(data.get("Nome Trabalhador")),
                data.get("CNPJ Operadora"),
                data.get("Registro ANS"),
                data.get("Valor Titular Descontado em Folha"),
                "Informar CNPJ numérico válido da operadora, Registro ANS e valor titular descontado em folha.",
                short_error(str(item.get("erro_mensagem") or ""), 180),
            ]
        )
    write_table(ws, 3, headers, rows, "PlanoSaudeJunho")
    add_note(
        ws,
        "O eSocial recusou porque o grupo de plano de saúde coletivo não pôde ser montado. Na resposta de junho, estes CPFs aparecem como 'OK' no campo CNPJ Operadora, mas sem CNPJ numérico, sem Registro ANS e sem valor titular aproveitáveis.",
        10,
    )
    autosize(ws, {1: 18, 2: 17, 3: 34, 4: 20, 5: 18, 6: 24, 7: 52, 8: 52})


def build_workbook() -> dict[str, Any]:
    errors = latest_errors()
    plan_errors = {str(item["cpf"]): item for item in errors if is_plan_error(str(item.get("erro_mensagem") or ""))}
    other_errors = [item for item in errors if str(item["cpf"]) not in plan_errors]
    if other_errors:
        raise RuntimeError(f"Pendencias nao previstas para o relatorio limpo de junho: {other_errors[:3]}")

    plan_data = response_rows(set(plan_errors))
    counts = Counter({"plano": len(plan_errors)})
    workbook = Workbook()
    workbook.remove(workbook.active)
    build_summary(workbook.create_sheet("Junho"), counts)
    build_plan(workbook.create_sheet("Plano de saude"), plan_data, plan_errors)
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook.save(OUT_XLSX)
    return {
        "arquivo": str(OUT_XLSX),
        "abas": workbook.sheetnames,
        "contagens": dict(counts),
        "total_pendencias": len(errors),
        "plano_linhas_resposta": len(plan_data),
        "planilha_resposta": str(response_workbook_path()),
    }


def main() -> int:
    print(json.dumps(build_workbook(), ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())