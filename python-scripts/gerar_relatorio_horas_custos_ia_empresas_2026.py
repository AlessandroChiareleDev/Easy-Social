from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from html import escape
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from openpyxl.drawing.image import Image as ExcelImage
except ImportError:
    ExcelImage = None

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Image as PdfImage
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "relatorio_ana" / "RELATORIO_HORAS_CUSTOS_IA_EMPRESAS_2026"
EMPRESAS_DIR = OUT_DIR / "empresas"
DATA_PATH = OUT_DIR / "dados_relatorio_horas_custos_ia_empresas_2026.json"
LOGO_PATH = ROOT / "frontend" / "public" / "brand-logo.png"

NAVY = "0D1530"
BLUE = "0066FF"
GREEN = "137A3A"
AMBER = "B45309"
PURPLE = "6D28D9"
SLATE = "334155"
MUTED = "64748B"
GRID = "DBE5F3"
SOFT = "F5F8FC"
WHITE = "FFFFFF"

CAMBIO = {
    "abr_2026": {"data": "01/04/2026", "taxa": 5.1564},
    "mai_2026": {"data": "01/05/2026", "taxa": 4.9943},
}

PREMISSAS = [
    "Jornada-base: 8 horas de segunda a sexta-feira e 4 horas aos sábados, totalizando 44 horas por semana cheia.",
    "Domingos não são listados no corpo executivo; entram apenas no cálculo semanal como dias sem jornada.",
    "Datas de transição foram alocadas a uma única empresa no consolidado para evitar dupla contagem.",
    "Custos de IA são usados somente quando localizados em artefatos existentes do projeto.",
    "Para Soluções, o relatório de origem registra US$ 586,00 em tokens extras e 126 horas de processamento; a coluna de custo/hora desta v2 usa a base padronizada de jornada do recorte solicitado.",
    "Para Objetiva, não foi localizado custo de IA/tokens nos artefatos disponíveis; o campo permanece pendente, sem estimativa inventada.",
]

FONTES = [
    "APPA: dashboard/src/components/relatorio-pdf-inner.tsx, constantes GITHUB_TOTAL=468, CLAUDE_TOTAL=400 e TOTAL_USD=868.",
    "APPA: dashboard/src/components/relatorio-profissional.tsx, câmbio de abril/2026 e maio/2026 usado para conversão BRL.",
    "APPA: dashboard/src/components/relatorio-github.tsx, 12.340 requests estimados de Copilot e uso alocado majoritariamente/totalmente ao Easy e-Social APPA.",
    "Soluções: relatorio_ana/RELATORIO_CONCLUSAO_ATIVIDADE_SOLUCOES_2025/RELATORIO_CONCLUSAO_ATIVIDADE_EASY_SOCIAL_SOLUCOES_2025.html, linhas de horas totais e tokens extras.",
    "Objetiva: busca local em relatorio_ana, python-scripts e docs não localizou custo de IA/tokens documentado.",
]


@dataclass(frozen=True)
class EmpresaInput:
    ordem: int
    slug: str
    sigla: str
    nome: str
    cnpj: str
    frente: str
    inicio_informado: date
    fim_informado: date
    inicio_calculo: date
    fim_calculo: date
    cor: str
    observacao_horas: str
    custo_usd: float | None
    custo_brl: float | None
    custo_status: str
    custo_fonte: str
    tokens_resumo: str
    detalhe_ia: str
    horas_fonte_custo: int | None = None


def brl_from_usd(value: float, taxa: float) -> float:
    return round(value * taxa, 2)


APPA_BRL = round((434 * CAMBIO["abr_2026"]["taxa"]) + (434 * CAMBIO["mai_2026"]["taxa"]), 2)
SOLUCOES_BRL = brl_from_usd(586, CAMBIO["mai_2026"]["taxa"])

EMPRESAS = [
    EmpresaInput(
        ordem=1,
        slug="appa",
        sigla="APPA",
        nome="APPA - Administração dos Portos de Paranaguá e Antonina",
        cnpj="05.969.071/0001-10",
        frente="Operação Easy e-Social APPA, análise de rubricas, suporte técnico e processamento assistido.",
        inicio_informado=date(2026, 3, 24),
        fim_informado=date(2026, 5, 7),
        inicio_calculo=date(2026, 3, 24),
        fim_calculo=date(2026, 5, 6),
        cor=BLUE,
        observacao_horas="07/05/2026 foi tratado como transição para Soluções no consolidado.",
        custo_usd=868.00,
        custo_brl=APPA_BRL,
        custo_status="Documentado",
        custo_fonte="GitHub Copilot Pro: US$ 468,00; Claude Max 20x: US$ 400,00; total APPA: US$ 868,00.",
        tokens_resumo="12.340 requests Copilot estimados; Copilot alocado 98% a 100% ao APPA; Claude 100% APPA.",
        detalhe_ia="GitHub Copilot Pro, Claude Max 20x e uso de assistentes de desenvolvimento no Easy e-Social APPA.",
    ),
    EmpresaInput(
        ordem=2,
        slug="solucoes",
        sigla="SOLUÇÕES",
        nome="SOLUÇÕES SERVIÇOS TERCEIRIZADOS LTDA",
        cnpj="09.445.502/0001-09",
        frente="Reenvio massivo S-1210, saneamento de pendências e fechamento operacional dos períodos 2025.",
        inicio_informado=date(2026, 5, 7),
        fim_informado=date(2026, 5, 20),
        inicio_calculo=date(2026, 5, 7),
        fim_calculo=date(2026, 5, 19),
        cor=GREEN,
        observacao_horas="20/05/2026 foi tratado como transição para Objetiva no consolidado.",
        custo_usd=586.00,
        custo_brl=SOLUCOES_BRL,
        custo_status="Documentado",
        custo_fonte="Relatório de conclusão Soluções: US$ 586,00 em tokens extras para processamento da empresa.",
        tokens_resumo="Tokens extras documentados em valor financeiro: US$ 586,00; contagem técnica de tokens não informada no artefato.",
        detalhe_ia="Processamento assistido por IA para reenvio massivo S-1210 e fechamento operacional.",
        horas_fonte_custo=126,
    ),
    EmpresaInput(
        ordem=3,
        slug="objetiva",
        sigla="OBJETIVA",
        nome="OBJETIVA SERVIÇOS TERCEIRIZADOS LTDA",
        cnpj="10.874.523/0001-10",
        frente="Validação de base, preparação de execução e operação assistida da frente Objetiva.",
        inicio_informado=date(2026, 5, 20),
        fim_informado=date(2026, 5, 29),
        inicio_calculo=date(2026, 5, 20),
        fim_calculo=date(2026, 5, 29),
        cor=PURPLE,
        observacao_horas="Período final do recorte solicitado, sem transição posterior dentro deste relatório.",
        custo_usd=None,
        custo_brl=None,
        custo_status="Não localizado",
        custo_fonte="Não foi localizado relatório de custo IA/tokens para Objetiva nos artefatos pesquisados.",
        tokens_resumo="Sem valor de tokens ou custo de IA localizado; campo mantido pendente.",
        detalhe_ia="Sem fonte financeira localizada para atribuição de custo de IA.",
    ),
]


def br_date(value: date) -> str:
    return value.strftime("%d/%m/%Y")


def br_int(value: int | float) -> str:
    return f"{value:,.0f}".replace(",", ".")


def br_decimal(value: float) -> str:
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def br_percent(value: float) -> str:
    return f"{value * 100:,.1f}%".replace(",", "X").replace(".", ",").replace("X", ".")


def usd(value: float | None) -> str:
    if value is None:
        return "Não localizado"
    return f"US$ {br_decimal(value)}"


def brl(value: float | None) -> str:
    if value is None:
        return "Não localizado"
    return f"R$ {br_decimal(value)}"


def ratio(cost: float | None, hours: int) -> float | None:
    if cost is None or not hours:
        return None
    return round(cost / hours, 2)


def date_range(start: date, end: date) -> list[date]:
    days: list[date] = []
    current = start
    while current <= end:
        days.append(current)
        current += timedelta(days=1)
    return days


def hours_for_day(value: date) -> int:
    if value.weekday() <= 4:
        return 8
    if value.weekday() == 5:
        return 4
    return 0


def weekly_rows(start: date, end: date) -> list[dict[str, Any]]:
    weeks: dict[str, dict[str, Any]] = {}
    for current_date in date_range(start, end):
        iso_year, iso_week, _iso_day = current_date.isocalendar()
        key = f"{iso_year}-S{iso_week:02d}"
        if key not in weeks:
            weeks[key] = {
                "semana": key,
                "inicio": current_date,
                "fim": current_date,
                "dias_uteis": 0,
                "sabados": 0,
                "horas": 0,
            }

        week = weeks[key]
        week["fim"] = current_date
        hours = hours_for_day(current_date)
        week["horas"] += hours
        if current_date.weekday() <= 4:
            week["dias_uteis"] += 1
        elif current_date.weekday() == 5:
            week["sabados"] += 1

    return [
        {
            "semana": row["semana"],
            "periodo": f"{br_date(row['inicio'])} a {br_date(row['fim'])}",
            "dias_uteis": row["dias_uteis"],
            "sabados": row["sabados"],
            "horas": row["horas"],
        }
        for row in weeks.values()
    ]


def build_company(empresa: EmpresaInput) -> dict[str, Any]:
    weeks = weekly_rows(empresa.inicio_calculo, empresa.fim_calculo)
    total_hours = sum(row["horas"] for row in weeks)
    cost_hour_usd = ratio(empresa.custo_usd, total_hours)
    cost_hour_brl = ratio(empresa.custo_brl, total_hours)

    return {
        "ordem": empresa.ordem,
        "slug": empresa.slug,
        "sigla": empresa.sigla,
        "nome": empresa.nome,
        "cnpj": empresa.cnpj,
        "frente": empresa.frente,
        "cor": empresa.cor,
        "periodo_informado": f"{br_date(empresa.inicio_informado)} a {br_date(empresa.fim_informado)}",
        "periodo_calculo": f"{br_date(empresa.inicio_calculo)} a {br_date(empresa.fim_calculo)}",
        "observacao_horas": empresa.observacao_horas,
        "semanas": weeks,
        "horas": total_hours,
        "custo_usd": empresa.custo_usd,
        "custo_brl": empresa.custo_brl,
        "custo_status": empresa.custo_status,
        "custo_fonte": empresa.custo_fonte,
        "tokens_resumo": empresa.tokens_resumo,
        "detalhe_ia": empresa.detalhe_ia,
        "horas_fonte_custo": empresa.horas_fonte_custo,
        "custo_hora_usd": cost_hour_usd,
        "custo_hora_brl": cost_hour_brl,
    }


def build_report() -> dict[str, Any]:
    companies = [build_company(empresa) for empresa in EMPRESAS]
    total_hours = sum(company["horas"] for company in companies)
    known_cost_companies = [company for company in companies if company["custo_usd"] is not None]
    total_known_hours = sum(company["horas"] for company in known_cost_companies)
    total_known_usd = round(sum(company["custo_usd"] or 0 for company in companies), 2)
    total_known_brl = round(sum(company["custo_brl"] or 0 for company in companies), 2)
    first_date = min(empresa.inicio_calculo for empresa in EMPRESAS)
    last_date = max(empresa.fim_calculo for empresa in EMPRESAS)

    for company in companies:
        company["participacao_horas"] = round(company["horas"] / total_hours, 4) if total_hours else 0

    return {
        "documento": {
            "titulo": "Relatório Executivo de Horas e Custos de IA",
            "subtitulo": "APPA, Soluções e Objetiva - horas trabalhadas, tokens e ferramentas IA",
            "codigo": "ES-HC-IA-2026-002",
            "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "periodo_consolidado": f"{br_date(first_date)} a {br_date(last_date)}",
            "criterio_resumo": "8h de segunda a sexta + 4h aos sábados = 44h/semana",
        },
        "cambio": CAMBIO,
        "premissas": PREMISSAS,
        "fontes": FONTES,
        "empresas": companies,
        "totais": {
            "horas": total_hours,
            "horas_com_custo_documentado": total_known_hours,
            "custo_usd_documentado": total_known_usd,
            "custo_brl_referencia": total_known_brl,
            "custo_hora_usd_documentado": ratio(total_known_usd, total_known_hours),
            "custo_hora_brl_documentado": ratio(total_known_brl, total_known_hours),
            "empresas_com_custo": len(known_cost_companies),
            "empresas_sem_custo": len(companies) - len(known_cost_companies),
        },
    }


def html_logo(relative_path: str) -> str:
    if LOGO_PATH.exists():
        return f'<img src="{relative_path}" alt="Easy Social" class="logo-img">'
    return '<div class="logo-mark">ES</div><div class="logo-text">Easy <span>Social</span></div>'


def metric_card(label: str, value: str, sub: str = "") -> str:
    return f"""
      <div class="metric">
        <div class="metric-label">{escape(label)}</div>
        <div class="metric-value">{escape(value)}</div>
        <div class="metric-sub">{escape(sub)}</div>
      </div>
    """


def status_class(company: dict[str, Any]) -> str:
    return "ok" if company["custo_usd"] is not None else "pending"


def render_company_rows(companies: list[dict[str, Any]]) -> str:
    rows = []
    for company in companies:
        rows.append(
            f"""
            <tr>
              <td><strong>{escape(company['sigla'])}</strong><span>{escape(company['nome'])}<br>CNPJ {escape(company['cnpj'])}</span></td>
              <td>{escape(company['periodo_calculo'])}</td>
              <td class="num">{br_int(company['horas'])} h</td>
              <td class="num">{usd(company['custo_usd'])}</td>
              <td class="num">{brl(company['custo_brl'])}</td>
              <td class="num">{usd(company['custo_hora_usd'])}/h</td>
              <td><span class="badge {status_class(company)}">{escape(company['custo_status'])}</span></td>
            </tr>
            """
        )
    return "\n".join(rows)


def render_source_rows(companies: list[dict[str, Any]]) -> str:
    rows = []
    for company in companies:
        extra = ""
        if company["horas_fonte_custo"]:
            extra = f"<br><span>Fonte de custo também registra {company['horas_fonte_custo']} horas de processamento.</span>"
        rows.append(
            f"""
            <tr>
              <td><strong>{escape(company['sigla'])}</strong></td>
              <td>{escape(company['detalhe_ia'])}</td>
              <td>{escape(company['tokens_resumo'])}</td>
              <td>{escape(company['custo_fonte'])}{extra}</td>
            </tr>
            """
        )
    return "\n".join(rows)


def render_week_rows(company: dict[str, Any]) -> str:
    return "\n".join(
        f"""
        <tr>
          <td>{escape(row['semana'])}</td>
          <td>{escape(row['periodo'])}</td>
          <td class="num">{row['dias_uteis']}</td>
          <td class="num">{row['sabados']}</td>
          <td class="num">{row['horas']} h</td>
        </tr>
        """
        for row in company["semanas"]
    )


def base_css() -> str:
    return f"""
* {{ box-sizing: border-box; }}
body {{ margin: 0; background: #e8eef6; color: #172033; font-family: Arial, Helvetica, sans-serif; }}
.page {{ width: 210mm; min-height: 297mm; margin: 18px auto; background: #fff; box-shadow: 0 12px 32px rgba(15, 23, 42, .12); }}
.topbar {{ background: #{NAVY}; color: #fff; padding: 26px 30px 30px; }}
.brand-line {{ display: flex; align-items: center; justify-content: space-between; gap: 16px; margin-bottom: 20px; }}
.brand {{ display: flex; align-items: center; gap: 12px; font-weight: 800; letter-spacing: .2px; }}
.logo-img {{ width: 118px; max-height: 42px; object-fit: contain; display: block; }}
.logo-mark {{ width: 34px; height: 34px; border-radius: 9px; background: #{BLUE}; color: #fff; display: grid; place-items: center; font-weight: 900; }}
.logo-text span {{ color: #5db8ff; }}
.tag {{ border: 1px solid rgba(255,255,255,.24); border-radius: 999px; padding: 7px 11px; font-size: 11px; text-transform: uppercase; letter-spacing: .7px; color: #c9d7ea; }}
h1 {{ margin: 0; font-size: 27px; line-height: 1.12; letter-spacing: 0; }}
.subtitle {{ margin-top: 8px; color: #c9d7ea; font-size: 13px; }}
.content {{ padding: 24px 30px 30px; }}
.metrics {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 10px; margin-bottom: 18px; }}
.metric {{ border: 1px solid #{GRID}; border-radius: 8px; padding: 13px; background: #{SOFT}; min-height: 86px; }}
.metric-label {{ font-size: 10px; color: #{MUTED}; text-transform: uppercase; letter-spacing: .5px; margin-bottom: 7px; }}
.metric-value {{ font-size: 20px; line-height: 1.15; color: #{NAVY}; font-weight: 800; }}
.metric-sub {{ font-size: 10px; color: #{MUTED}; margin-top: 6px; line-height: 1.35; }}
.section-title {{ margin: 20px 0 10px; color: #{NAVY}; font-weight: 800; font-size: 15px; }}
table {{ width: 100%; border-collapse: collapse; font-size: 11px; }}
th {{ background: #{NAVY}; color: #fff; text-align: left; padding: 9px; font-size: 10px; text-transform: uppercase; letter-spacing: .45px; }}
td {{ border: 1px solid #{GRID}; padding: 9px; vertical-align: top; line-height: 1.35; }}
td span {{ display: block; color: #{MUTED}; font-size: 10px; margin-top: 3px; }}
.num {{ text-align: right; white-space: nowrap; font-variant-numeric: tabular-nums; }}
.badge {{ display: inline-block; border-radius: 999px; padding: 4px 8px; font-size: 10px; font-weight: 700; white-space: nowrap; }}
.badge.ok {{ color: #{GREEN}; background: #e9f8ef; border: 1px solid #b9e7c9; }}
.badge.pending {{ color: #{AMBER}; background: #fff7ed; border: 1px solid #fed7aa; }}
.note-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 12px; margin-top: 16px; }}
.note {{ border-left: 4px solid #{BLUE}; padding: 10px 12px; background: #f8fbff; color: #{SLATE}; font-size: 11px; line-height: 1.45; }}
.note strong {{ color: #{NAVY}; }}
.footer {{ display: flex; justify-content: space-between; gap: 12px; border-top: 1px solid #{GRID}; margin-top: 20px; padding-top: 12px; color: #{MUTED}; font-size: 10px; }}
.avoid-break {{ page-break-inside: avoid; }}
@media print {{ body {{ background: #fff; }} .page {{ margin: 0; box-shadow: none; width: auto; }} }}
"""


def render_consolidated_html(report: dict[str, Any]) -> str:
    doc = report["documento"]
    totals = report["totais"]
    companies = report["empresas"]
    return f"""<!doctype html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<title>{escape(doc['titulo'])}</title>
<style>{base_css()}</style>
</head>
<body>
<main class="page">
  <header class="topbar">
    <div class="brand-line">
      <div class="brand">{html_logo('../../frontend/public/brand-logo.png')}</div>
      <div class="tag">Relatório executivo</div>
    </div>
    <h1>{escape(doc['titulo'])}</h1>
    <div class="subtitle">{escape(doc['subtitulo'])} · {escape(doc['periodo_consolidado'])}</div>
  </header>
  <section class="content">
    <div class="metrics avoid-break">
      {metric_card('Horas trabalhadas', f"{br_int(totals['horas'])} h", doc['criterio_resumo'])}
      {metric_card('Custo IA documentado', usd(totals['custo_usd_documentado']), 'APPA + Soluções; Objetiva pendente')}
      {metric_card('Referência BRL', brl(totals['custo_brl_referencia']), 'Câmbio documentado em APPA e maio/2026 para Soluções')}
      {metric_card('Custo por hora', f"{usd(totals['custo_hora_usd_documentado'])}/h", f"{brl(totals['custo_hora_brl_documentado'])}/h em empresas com custo")}
    </div>

    <div class="section-title">Resumo por empresa</div>
    <table class="avoid-break">
      <thead>
        <tr>
          <th>Empresa</th>
          <th>Período usado</th>
          <th class="num">Horas</th>
          <th class="num">Custo IA USD</th>
          <th class="num">Referência BRL</th>
          <th class="num">USD/h</th>
          <th>Status</th>
        </tr>
      </thead>
      <tbody>{render_company_rows(companies)}</tbody>
    </table>

    <div class="section-title">Custo de IA e tokens</div>
    <table>
      <thead>
        <tr>
          <th>Empresa</th>
          <th>Uso relacionado</th>
          <th>Tokens / requests</th>
          <th>Fonte de custo</th>
        </tr>
      </thead>
      <tbody>{render_source_rows(companies)}</tbody>
    </table>

    <div class="note-grid avoid-break">
      <div class="note"><strong>Premissa de horas.</strong> O relatório usa uma base executiva de jornada, não memória diária: 8h em dias úteis e 4h aos sábados. Domingos zerados não são listados.</div>
      <div class="note"><strong>Objetiva.</strong> Há horas trabalhadas no recorte, mas custo de IA/tokens não foi localizado nos artefatos existentes. O relatório deixa o campo pendente para evitar estimativa sem fonte.</div>
    </div>

    <div class="section-title">Premissas e fontes</div>
    <table>
      <tbody>
        {''.join(f'<tr><td>{escape(item)}</td></tr>' for item in PREMISSAS[:4])}
        {''.join(f'<tr><td>{escape(item)}</td></tr>' for item in FONTES)}
      </tbody>
    </table>

    <div class="footer">
      <div>{escape(doc['codigo'])} · Gerado em {escape(doc['gerado_em'])}</div>
      <div>Easy Social · Relatório executivo de horas e custos IA</div>
    </div>
  </section>
</main>
</body>
</html>
"""


def render_company_html(report: dict[str, Any], company: dict[str, Any]) -> str:
    doc = report["documento"]
    return f"""<!doctype html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<title>{escape(company['sigla'])} - Horas e Custos IA</title>
<style>{base_css()}</style>
</head>
<body>
<main class="page">
  <header class="topbar">
    <div class="brand-line">
      <div class="brand">{html_logo('../../../frontend/public/brand-logo.png')}</div>
      <div class="tag">Relatório por empresa</div>
    </div>
    <h1>{escape(company['sigla'])} · Horas e Custos de IA</h1>
    <div class="subtitle">{escape(company['nome'])} · CNPJ {escape(company['cnpj'])}</div>
  </header>
  <section class="content">
    <div class="metrics avoid-break">
      {metric_card('Período usado', company['periodo_calculo'], company['observacao_horas'])}
      {metric_card('Horas trabalhadas', f"{br_int(company['horas'])} h", doc['criterio_resumo'])}
      {metric_card('Custo IA', usd(company['custo_usd']), company['custo_status'])}
      {metric_card('Custo por hora', f"{usd(company['custo_hora_usd'])}/h", f"{brl(company['custo_hora_brl'])}/h")}
    </div>

    <div class="section-title">Resumo executivo</div>
    <table class="avoid-break">
      <tbody>
        <tr><td><strong>Frente de trabalho</strong></td><td>{escape(company['frente'])}</td></tr>
        <tr><td><strong>Período informado</strong></td><td>{escape(company['periodo_informado'])}</td></tr>
        <tr><td><strong>Período usado para cálculo</strong></td><td>{escape(company['periodo_calculo'])}</td></tr>
        <tr><td><strong>Tokens / requests</strong></td><td>{escape(company['tokens_resumo'])}</td></tr>
        <tr><td><strong>Fonte de custo</strong></td><td>{escape(company['custo_fonte'])}</td></tr>
      </tbody>
    </table>

    <div class="section-title">Memória semanal de horas</div>
    <table>
      <thead>
        <tr>
          <th>Semana</th>
          <th>Período</th>
          <th class="num">Dias úteis</th>
          <th class="num">Sábados</th>
          <th class="num">Horas</th>
        </tr>
      </thead>
      <tbody>{render_week_rows(company)}</tbody>
    </table>

    <div class="section-title">Premissas aplicadas</div>
    <table>
      <tbody>{''.join(f'<tr><td>{escape(item)}</td></tr>' for item in PREMISSAS)}</tbody>
    </table>

    <div class="footer">
      <div>{escape(doc['codigo'])} · Gerado em {escape(doc['gerado_em'])}</div>
      <div>Easy Social · {escape(company['sigla'])}</div>
    </div>
  </section>
</main>
</body>
</html>
"""


def set_column_widths(sheet, widths: dict[str, float]) -> None:
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width


def style_header(sheet, title: str, subtitle: str) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:G1")
    sheet.merge_cells("A2:G2")
    sheet["A1"] = title
    sheet["A2"] = subtitle
    sheet["A1"].font = Font(bold=True, size=18, color=NAVY)
    sheet["A2"].font = Font(size=10, color=MUTED)
    sheet["A1"].alignment = Alignment(horizontal="left")
    sheet["A2"].alignment = Alignment(horizontal="left")
    if ExcelImage and LOGO_PATH.exists():
        try:
            image = ExcelImage(str(LOGO_PATH))
            image.width = 120
            image.height = 42
            sheet.add_image(image, "F1")
        except Exception:
            pass


def style_table(sheet, start_row: int, end_row: int, start_col: int, end_col: int) -> None:
    thin = Side(style="thin", color=GRID)
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in sheet[start_row]:
        if start_col <= cell.column <= end_col:
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(bold=True, color=WHITE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_rows(sheet, start_row: int, headers: list[str], rows: list[list[Any]]) -> int:
    for column_index, header in enumerate(headers, start=1):
        sheet.cell(start_row, column_index, header)
    for row_index, row_values in enumerate(rows, start=start_row + 1):
        for column_index, value in enumerate(row_values, start=1):
            sheet.cell(row_index, column_index, value)
    style_table(sheet, start_row, start_row + len(rows), 1, len(headers))
    return start_row + len(rows) + 2


def create_consolidated_workbook(report: dict[str, Any]) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "RESUMO_EXECUTIVO"
    style_header(summary, report["documento"]["titulo"], report["documento"]["subtitulo"])
    set_column_widths(summary, {"A": 22, "B": 38, "C": 15, "D": 18, "E": 18, "F": 18, "G": 18})

    totals = report["totais"]
    rows = [
        ["Horas trabalhadas", totals["horas"], "Critério", report["documento"]["criterio_resumo"]],
        ["Custo IA documentado USD", totals["custo_usd_documentado"], "Referência BRL", totals["custo_brl_referencia"]],
        ["USD/h em empresas com custo", totals["custo_hora_usd_documentado"], "BRL/h em empresas com custo", totals["custo_hora_brl_documentado"]],
        ["Empresas com custo", totals["empresas_com_custo"], "Empresas sem custo", totals["empresas_sem_custo"]],
    ]
    write_rows(summary, 4, ["Indicador", "Valor", "Indicador", "Valor"], rows)

    company_rows = [
        [
            company["sigla"],
            company["nome"],
            company["periodo_calculo"],
            company["horas"],
            company["participacao_horas"],
            company["custo_usd"],
            company["custo_brl"],
            company["custo_hora_usd"],
            company["custo_status"],
        ]
        for company in report["empresas"]
    ]
    write_rows(
        summary,
        11,
        ["Empresa", "Nome", "Período", "Horas", "% horas", "USD IA", "BRL ref.", "USD/h", "Status"],
        company_rows,
    )

    costs = workbook.create_sheet("CUSTOS_IA")
    style_header(costs, "Custos de IA e Tokens", "Fontes documentadas por empresa")
    set_column_widths(costs, {"A": 16, "B": 42, "C": 34, "D": 16, "E": 16, "F": 18, "G": 56})
    cost_rows = [
        [
            company["sigla"],
            company["detalhe_ia"],
            company["tokens_resumo"],
            company["custo_usd"],
            company["custo_brl"],
            company["custo_hora_usd"],
            company["custo_fonte"],
        ]
        for company in report["empresas"]
    ]
    write_rows(costs, 4, ["Empresa", "Uso relacionado", "Tokens / requests", "USD", "BRL ref.", "USD/h", "Fonte"], cost_rows)

    hours = workbook.create_sheet("HORAS")
    style_header(hours, "Horas por Empresa", "Base de jornada executiva")
    set_column_widths(hours, {"A": 16, "B": 44, "C": 24, "D": 24, "E": 14, "F": 50})
    hour_rows = [
        [company["sigla"], company["nome"], company["periodo_informado"], company["periodo_calculo"], company["horas"], company["observacao_horas"]]
        for company in report["empresas"]
    ]
    write_rows(hours, 4, ["Empresa", "Nome", "Período informado", "Período usado", "Horas", "Observação"], hour_rows)

    weekly = workbook.create_sheet("MEMORIA_SEMANAL")
    style_header(weekly, "Memória Semanal", "Sem detalhamento diário e sem domingos zerados")
    set_column_widths(weekly, {"A": 16, "B": 16, "C": 28, "D": 14, "E": 14, "F": 14})
    weekly_rows_flat = []
    for company in report["empresas"]:
        for row in company["semanas"]:
            weekly_rows_flat.append([company["sigla"], row["semana"], row["periodo"], row["dias_uteis"], row["sabados"], row["horas"]])
    write_rows(weekly, 4, ["Empresa", "Semana", "Período", "Dias úteis", "Sábados", "Horas"], weekly_rows_flat)

    sources = workbook.create_sheet("PREMISSAS_FONTES")
    style_header(sources, "Premissas e Fontes", "Rastreabilidade do relatório")
    set_column_widths(sources, {"A": 18, "B": 110})
    source_rows = [["Premissa", item] for item in report["premissas"]] + [["Fonte", item] for item in report["fontes"]]
    write_rows(sources, 4, ["Tipo", "Descrição"], source_rows)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A4"
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'
                if isinstance(cell.value, int):
                    cell.number_format = '#,##0'

    return workbook


def create_company_workbook(report: dict[str, Any], company: dict[str, Any]) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "RESUMO"
    style_header(summary, f"{company['sigla']} - Horas e Custos IA", company["nome"])
    set_column_widths(summary, {"A": 24, "B": 76, "C": 18, "D": 18})

    rows = [
        ["CNPJ", company["cnpj"]],
        ["Frente", company["frente"]],
        ["Período informado", company["periodo_informado"]],
        ["Período usado", company["periodo_calculo"]],
        ["Horas trabalhadas", company["horas"]],
        ["Custo IA USD", company["custo_usd"]],
        ["Referência BRL", company["custo_brl"]],
        ["Custo por hora USD", company["custo_hora_usd"]],
        ["Tokens / requests", company["tokens_resumo"]],
        ["Fonte", company["custo_fonte"]],
    ]
    write_rows(summary, 4, ["Campo", "Valor"], rows)

    weekly = workbook.create_sheet("MEMORIA_SEMANAL")
    style_header(weekly, "Memória Semanal", company["sigla"])
    set_column_widths(weekly, {"A": 18, "B": 28, "C": 16, "D": 16, "E": 16})
    weekly_rows_company = [[row["semana"], row["periodo"], row["dias_uteis"], row["sabados"], row["horas"]] for row in company["semanas"]]
    write_rows(weekly, 4, ["Semana", "Período", "Dias úteis", "Sábados", "Horas"], weekly_rows_company)

    sources = workbook.create_sheet("PREMISSAS_FONTES")
    style_header(sources, "Premissas e Fontes", company["sigla"])
    set_column_widths(sources, {"A": 18, "B": 110})
    rows_sources = [["Premissa", item] for item in report["premissas"]] + [["Fonte", item] for item in report["fontes"] if company["sigla"] in item or "Objetiva" in item and company["sigla"] == "OBJETIVA"]
    if not any(row[0] == "Fonte" for row in rows_sources):
        rows_sources.append(["Fonte", company["custo_fonte"]])
    write_rows(sources, 4, ["Tipo", "Descrição"], rows_sources)

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A4"
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'
                if isinstance(cell.value, int):
                    cell.number_format = '#,##0'
    return workbook


def pdf_styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("Title", parent=base["Title"], fontName="Helvetica-Bold", fontSize=20, leading=24, textColor=colors.HexColor(f"#{NAVY}"), alignment=TA_LEFT, spaceAfter=8),
        "subtitle": ParagraphStyle("Subtitle", parent=base["Normal"], fontSize=9, leading=12, textColor=colors.HexColor(f"#{MUTED}"), spaceAfter=14),
        "section": ParagraphStyle("Section", parent=base["Heading2"], fontName="Helvetica-Bold", fontSize=12, leading=15, textColor=colors.HexColor(f"#{NAVY}"), spaceBefore=12, spaceAfter=7),
        "header": ParagraphStyle("Header", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=7, leading=9, textColor=colors.white, alignment=TA_CENTER),
        "normal": ParagraphStyle("NormalCustom", parent=base["Normal"], fontSize=8, leading=10, textColor=colors.HexColor(f"#{SLATE}"), alignment=TA_LEFT),
        "small": ParagraphStyle("Small", parent=base["Normal"], fontSize=7, leading=9, textColor=colors.HexColor(f"#{MUTED}")),
        "center": ParagraphStyle("Center", parent=base["Normal"], fontSize=8, leading=10, textColor=colors.HexColor(f"#{SLATE}"), alignment=TA_CENTER),
    }


def p(text: Any, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape("" if text is None else str(text)), style)


def pdf_logo() -> Any | None:
    if not LOGO_PATH.exists():
        return None
    try:
        image = PdfImage(str(LOGO_PATH))
        max_width = 34 * mm
        max_height = 12 * mm
        scale = min(max_width / image.imageWidth, max_height / image.imageHeight)
        image.drawWidth = image.imageWidth * scale
        image.drawHeight = image.imageHeight * scale
        return image
    except Exception:
        return None


def make_pdf_table(data: list[list[Any]], col_widths: list[float], header: bool = True) -> Table:
    styles = pdf_styles()
    table_data = []
    for row_index, row in enumerate(data):
        row_style = styles["header"] if header and row_index == 0 else styles["normal"]
        table_data.append([value if hasattr(value, "wrap") else p(value, row_style) for value in row])
    table = Table(table_data, colWidths=col_widths, hAlign="LEFT", repeatRows=1 if header else 0)
    commands = [
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor(f"#{GRID}")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    if header:
        commands.extend([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{NAVY}")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ])
    table.setStyle(TableStyle(commands))
    return table


def pdf_header_story(title: str, subtitle: str) -> list[Any]:
    styles = pdf_styles()
    story: list[Any] = []
    logo = pdf_logo()
    if logo:
        story.append(logo)
        story.append(Spacer(1, 6))
    story.append(Paragraph(title, styles["title"]))
    story.append(Paragraph(subtitle, styles["subtitle"]))
    return story


def create_consolidated_pdf(report: dict[str, Any], path: Path) -> None:
    if not REPORTLAB_AVAILABLE:
        return
    styles = pdf_styles()
    doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    totals = report["totais"]
    story = pdf_header_story(report["documento"]["titulo"], f"{report['documento']['subtitulo']} - {report['documento']['periodo_consolidado']}")

    metrics = [
        ["Horas trabalhadas", "Custo IA USD", "Referência BRL", "USD/h documentado"],
        [f"{br_int(totals['horas'])} h", usd(totals["custo_usd_documentado"]), brl(totals["custo_brl_referencia"]), f"{usd(totals['custo_hora_usd_documentado'])}/h"],
    ]
    story.append(make_pdf_table(metrics, [42 * mm, 42 * mm, 42 * mm, 42 * mm]))
    story.append(Paragraph("Resumo por empresa", styles["section"]))
    summary_rows = [["Empresa", "Período", "Horas", "USD IA", "BRL ref.", "USD/h", "Status"]]
    for company in report["empresas"]:
        summary_rows.append([
            company["sigla"],
            company["periodo_calculo"],
            f"{company['horas']} h",
            usd(company["custo_usd"]),
            brl(company["custo_brl"]),
            f"{usd(company['custo_hora_usd'])}/h",
            company["custo_status"],
        ])
    story.append(make_pdf_table(summary_rows, [24 * mm, 32 * mm, 18 * mm, 25 * mm, 27 * mm, 23 * mm, 24 * mm]))

    story.append(Paragraph("Custo de IA e tokens", styles["section"]))
    source_rows = [["Empresa", "Uso relacionado", "Tokens / requests", "Fonte"]]
    for company in report["empresas"]:
        source_rows.append([company["sigla"], company["detalhe_ia"], company["tokens_resumo"], company["custo_fonte"]])
    story.append(make_pdf_table(source_rows, [20 * mm, 47 * mm, 50 * mm, 56 * mm]))

    story.append(Paragraph("Premissas e fontes", styles["section"]))
    for item in PREMISSAS[:4] + FONTES:
        story.append(Paragraph(f"- {escape(item)}", styles["small"]))
    doc.build(story)


def create_company_pdf(report: dict[str, Any], company: dict[str, Any], path: Path) -> None:
    if not REPORTLAB_AVAILABLE:
        return
    styles = pdf_styles()
    doc = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=15 * mm, leftMargin=15 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    story = pdf_header_story(f"{company['sigla']} - Horas e Custos de IA", f"{company['nome']} - CNPJ {company['cnpj']}")

    metrics = [
        ["Período usado", "Horas", "Custo IA", "USD/h"],
        [company["periodo_calculo"], f"{company['horas']} h", usd(company["custo_usd"]), f"{usd(company['custo_hora_usd'])}/h"],
    ]
    story.append(make_pdf_table(metrics, [45 * mm, 33 * mm, 45 * mm, 45 * mm]))
    story.append(Paragraph("Resumo executivo", styles["section"]))
    summary_rows = [
        ["Campo", "Valor"],
        ["Frente de trabalho", company["frente"]],
        ["Período informado", company["periodo_informado"]],
        ["Tokens / requests", company["tokens_resumo"]],
        ["Fonte de custo", company["custo_fonte"]],
    ]
    story.append(make_pdf_table(summary_rows, [38 * mm, 130 * mm]))

    story.append(Paragraph("Memória semanal de horas", styles["section"]))
    weekly_data = [["Semana", "Período", "Dias úteis", "Sábados", "Horas"]]
    for row in company["semanas"]:
        weekly_data.append([row["semana"], row["periodo"], row["dias_uteis"], row["sabados"], f"{row['horas']} h"])
    story.append(make_pdf_table(weekly_data, [28 * mm, 50 * mm, 28 * mm, 28 * mm, 28 * mm]))

    story.append(Paragraph("Premissas", styles["section"]))
    for item in PREMISSAS:
        story.append(Paragraph(f"- {escape(item)}", styles["small"]))
    doc.build(story)


def save_json(report: dict[str, Any]) -> None:
    DATA_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")


def save_outputs(report: dict[str, Any]) -> list[Path]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    EMPRESAS_DIR.mkdir(parents=True, exist_ok=True)
    generated: list[Path] = []

    save_json(report)
    generated.append(DATA_PATH)

    consolidated_html = OUT_DIR / "00_RELATORIO_HORAS_CUSTOS_IA_CONSOLIDADO_2026.html"
    consolidated_html.write_text(render_consolidated_html(report), encoding="utf-8")
    generated.append(consolidated_html)

    consolidated_xlsx = OUT_DIR / "00_RELATORIO_HORAS_CUSTOS_IA_CONSOLIDADO_2026.xlsx"
    create_consolidated_workbook(report).save(consolidated_xlsx)
    generated.append(consolidated_xlsx)

    consolidated_pdf = OUT_DIR / "00_RELATORIO_HORAS_CUSTOS_IA_CONSOLIDADO_2026.pdf"
    create_consolidated_pdf(report, consolidated_pdf)
    if consolidated_pdf.exists():
        generated.append(consolidated_pdf)

    for company in report["empresas"]:
        prefix = f"{company['ordem']:02d}_{company['slug'].upper()}_HORAS_CUSTOS_IA_2026"
        company_html = EMPRESAS_DIR / f"{prefix}.html"
        company_html.write_text(render_company_html(report, company), encoding="utf-8")
        generated.append(company_html)

        company_xlsx = EMPRESAS_DIR / f"{prefix}.xlsx"
        create_company_workbook(report, company).save(company_xlsx)
        generated.append(company_xlsx)

        company_pdf = EMPRESAS_DIR / f"{prefix}.pdf"
        create_company_pdf(report, company, company_pdf)
        if company_pdf.exists():
            generated.append(company_pdf)

    return generated


def main() -> None:
    report = build_report()
    generated = save_outputs(report)
    print("Relatório executivo de horas e custos de IA gerado.")
    for path in generated:
        print(path.relative_to(ROOT))


if __name__ == "__main__":
    main()