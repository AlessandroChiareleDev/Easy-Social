from __future__ import annotations

import json
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from html import escape
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "relatorio_ana" / "RELATORIO_HORAS_TRABALHADAS_EMPRESAS_2026"
EMPRESAS_DIR = OUT_DIR / "empresas"
DATA_PATH = OUT_DIR / "dados_relatorio_horas_trabalhadas.json"

NAVY = "0D1530"
BLUE = "0066FF"
GREEN = "137A3A"
AMBER = "B45309"
PURPLE = "7C3AED"
SLATE = "334155"
MUTED = "64748B"
GRID = "DBE5F3"
WHITE = "FFFFFF"

DIA_SEMANA = {
    0: "segunda-feira",
    1: "terça-feira",
    2: "quarta-feira",
    3: "quinta-feira",
    4: "sexta-feira",
    5: "sábado",
    6: "domingo",
}

CRITERIOS = [
    "Jornada-base: 8 horas de segunda a sexta-feira e 4 horas aos sábados, totalizando 44 horas por semana cheia.",
    "Domingos foram considerados sem jornada.",
    "Feriados não foram abatidos porque não foram informados como premissa específica.",
    "Datas de transição foram preservadas no documento e alocadas a uma única empresa no consolidado, evitando dupla contagem.",
    "Este relatório não atribui valor financeiro por hora; entrega a memória de horas para eventual precificação posterior.",
]


@dataclass(frozen=True)
class EmpresaInput:
    ordem: int
    slug: str
    sigla: str
    nome: str
    cnpj: str
    frente: str
    inicio_informado: date
    fim_informado: date
    inicio_calculo: date
    fim_calculo: date
    cor: str
    observacao: str


EMPRESAS = [
    EmpresaInput(
        ordem=1,
        slug="appa",
        sigla="APPA",
        nome="APPA - Administração dos Portos de Paranaguá e Antonina",
        cnpj="05.969.071/0001-10",
        frente="Operação Easy e-Social APPA, análise de rubricas, suporte técnico e processamento assistido",
        inicio_informado=date(2026, 3, 24),
        fim_informado=date(2026, 5, 7),
        inicio_calculo=date(2026, 3, 24),
        fim_calculo=date(2026, 5, 6),
        cor=BLUE,
        observacao="07/05/2026 foi tratado como marco de transição para Soluções no cálculo consolidado.",
    ),
    EmpresaInput(
        ordem=2,
        slug="solucoes",
        sigla="SOLUÇÕES",
        nome="SOLUÇÕES SERVIÇOS TERCEIRIZADOS LTDA",
        cnpj="09.445.502/0001-09",
        frente="Reenvio massivo S-1210, saneamento de pendências e fechamento operacional dos períodos 2025",
        inicio_informado=date(2026, 5, 7),
        fim_informado=date(2026, 5, 20),
        inicio_calculo=date(2026, 5, 7),
        fim_calculo=date(2026, 5, 19),
        cor=GREEN,
        observacao="20/05/2026 foi tratado como marco de transição para Objetiva no cálculo consolidado.",
    ),
    EmpresaInput(
        ordem=3,
        slug="objetiva",
        sigla="OBJETIVA",
        nome="OBJETIVA SERVIÇOS TERCEIRIZADOS LTDA",
        cnpj="10.874.523/0001-10",
        frente="Início da frente Objetiva, validação de base, preparação de execução e operação assistida",
        inicio_informado=date(2026, 5, 20),
        fim_informado=date(2026, 5, 29),
        inicio_calculo=date(2026, 5, 20),
        fim_calculo=date(2026, 5, 29),
        cor=PURPLE,
        observacao="Período final do recorte solicitado, sem transição posterior dentro deste relatório.",
    ),
]


def br_date(value: date) -> str:
    return value.strftime("%d/%m/%Y")


def br_int(value: int) -> str:
    return f"{value:,}".replace(",", ".")


def br_percent(value: float) -> str:
    return f"{value * 100:,.1f}%".replace(",", "X").replace(".", ",").replace("X", ".")


def date_range(start: date, end: date) -> list[date]:
    days: list[date] = []
    current = start
    while current <= end:
        days.append(current)
        current += timedelta(days=1)
    return days


def hours_for_day(value: date) -> int:
    if value.weekday() <= 4:
        return 8
    if value.weekday() == 5:
        return 4
    return 0


def day_type(value: date) -> str:
    if value.weekday() <= 4:
        return "Dia útil - 8h"
    if value.weekday() == 5:
        return "Sábado - 4h"
    return "Domingo - 0h"


def build_company(empresa: EmpresaInput) -> dict[str, Any]:
    daily_rows: list[dict[str, Any]] = []
    weekly_index: dict[str, dict[str, Any]] = {}

    for current_date in date_range(empresa.inicio_calculo, empresa.fim_calculo):
        iso_year, iso_week, _iso_day = current_date.isocalendar()
        week_key = f"{iso_year}-S{iso_week:02d}"
        hours = hours_for_day(current_date)
        daily_row = {
            "data": current_date.isoformat(),
            "data_br": br_date(current_date),
            "dia_semana": DIA_SEMANA[current_date.weekday()],
            "tipo": day_type(current_date),
            "semana": week_key,
            "horas": hours,
        }
        daily_rows.append(daily_row)

        if week_key not in weekly_index:
            weekly_index[week_key] = {
                "semana": week_key,
                "inicio": current_date,
                "fim": current_date,
                "dias_8h": 0,
                "sabados_4h": 0,
                "domingos_0h": 0,
                "horas": 0,
            }
        weekly_row = weekly_index[week_key]
        weekly_row["fim"] = current_date
        weekly_row["horas"] += hours
        if hours == 8:
            weekly_row["dias_8h"] += 1
        elif hours == 4:
            weekly_row["sabados_4h"] += 1
        else:
            weekly_row["domingos_0h"] += 1

    weekly_rows = []
    for weekly_row in weekly_index.values():
        weekly_rows.append(
            {
                "semana": weekly_row["semana"],
                "periodo": f"{br_date(weekly_row['inicio'])} a {br_date(weekly_row['fim'])}",
                "dias_8h": weekly_row["dias_8h"],
                "sabados_4h": weekly_row["sabados_4h"],
                "domingos_0h": weekly_row["domingos_0h"],
                "horas": weekly_row["horas"],
            }
        )

    return {
        "ordem": empresa.ordem,
        "slug": empresa.slug,
        "sigla": empresa.sigla,
        "nome": empresa.nome,
        "cnpj": empresa.cnpj,
        "frente": empresa.frente,
        "cor": empresa.cor,
        "periodo_informado": {
            "inicio": empresa.inicio_informado.isoformat(),
            "fim": empresa.fim_informado.isoformat(),
            "texto": f"{br_date(empresa.inicio_informado)} a {br_date(empresa.fim_informado)}",
        },
        "periodo_calculo": {
            "inicio": empresa.inicio_calculo.isoformat(),
            "fim": empresa.fim_calculo.isoformat(),
            "texto": f"{br_date(empresa.inicio_calculo)} a {br_date(empresa.fim_calculo)}",
        },
        "observacao": empresa.observacao,
        "dias": daily_rows,
        "semanas": weekly_rows,
        "totais": {
            "horas": sum(row["horas"] for row in daily_rows),
            "dias_calendario": len(daily_rows),
            "dias_8h": sum(1 for row in daily_rows if row["horas"] == 8),
            "sabados_4h": sum(1 for row in daily_rows if row["horas"] == 4),
            "domingos_0h": sum(1 for row in daily_rows if row["horas"] == 0),
            "semanas": len(weekly_rows),
        },
    }


def build_report() -> dict[str, Any]:
    companies = [build_company(empresa) for empresa in EMPRESAS]
    total_hours = sum(company["totais"]["horas"] for company in companies)
    first_date = min(date.fromisoformat(company["periodo_calculo"]["inicio"]) for company in companies)
    last_date = max(date.fromisoformat(company["periodo_calculo"]["fim"]) for company in companies)

    for company in companies:
        company["totais"]["participacao"] = company["totais"]["horas"] / total_hours if total_hours else 0

    return {
        "documento": {
            "titulo": "Relatório de Horas Trabalhadas por Empresa",
            "subtitulo": "Memória técnica de horas - APPA, Soluções e Objetiva",
            "codigo": "ES-HORAS-2026-001",
            "gerado_em": datetime.now().strftime("%d/%m/%Y %H:%M"),
            "periodo_consolidado": f"{br_date(first_date)} a {br_date(last_date)}",
            "criterio_resumo": "8h de segunda a sexta + 4h aos sábados = 44h/semana",
        },
        "criterios": CRITERIOS,
        "empresas": companies,
        "totais": {
            "empresas": len(companies),
            "horas": total_hours,
            "dias_8h": sum(company["totais"]["dias_8h"] for company in companies),
            "sabados_4h": sum(company["totais"]["sabados_4h"] for company in companies),
            "domingos_0h": sum(company["totais"]["domingos_0h"] for company in companies),
            "dias_calendario": sum(company["totais"]["dias_calendario"] for company in companies),
        },
    }


CSS = """
@page { size: A4; margin: 14mm; }
@media print { body { -webkit-print-color-adjust: exact; print-color-adjust: exact; } .page { box-shadow: none; } }
* { box-sizing: border-box; }
body { margin: 0; background: #edf2f7; color: #0f172a; font-family: "Segoe UI", Arial, sans-serif; font-size: 12px; line-height: 1.45; }
.page { width: 100%; max-width: 1060px; margin: 0 auto; background: #ffffff; border: 1px solid #dbe5f3; box-shadow: 0 18px 44px rgba(13, 21, 48, 0.12); }
.topbar { background: #0d1530; color: #ffffff; padding: 24px 28px 22px; border-bottom: 5px solid #0066ff; }
.brand-line { display: flex; align-items: center; justify-content: space-between; gap: 16px; margin-bottom: 18px; }
.brand { display: flex; align-items: center; gap: 10px; font-weight: 850; font-size: 18px; }
.mark { width: 36px; height: 36px; border-radius: 8px; background: linear-gradient(135deg, #0066ff, #37b7ff); display: inline-flex; align-items: center; justify-content: center; color: #ffffff; font-weight: 900; }
.tag { border: 1px solid rgba(255,255,255,0.24); color: #cbd5e1; border-radius: 999px; padding: 6px 11px; font-size: 10px; text-transform: uppercase; font-weight: 800; }
h1 { margin: 0; font-size: 29px; line-height: 1.12; letter-spacing: 0; }
.subtitle { margin-top: 9px; color: #cbd5e1; max-width: 780px; }
.content { padding: 24px 28px 28px; }
.hero-grid { display: grid; grid-template-columns: 1.2fr 0.8fr; gap: 14px; margin-bottom: 16px; }
.panel { border: 1px solid #e2e8f0; border-radius: 8px; background: #ffffff; padding: 14px; }
.panel.soft { background: #f8fbff; }
.panel.warning { background: #fff7ed; border-color: #fed7aa; }
.label { color: #64748b; font-size: 10px; font-weight: 800; text-transform: uppercase; margin-bottom: 4px; }
.value { font-size: 14px; font-weight: 800; color: #0f172a; }
.muted { color: #64748b; }
.metric-grid { display: grid; grid-template-columns: repeat(4, minmax(0, 1fr)); gap: 10px; margin: 16px 0; }
.metric { border: 1px solid #dbe7f6; background: #f8fbff; border-radius: 8px; padding: 13px 12px; }
.metric .k { color: #64748b; font-size: 10px; font-weight: 850; text-transform: uppercase; }
.metric .v { margin-top: 7px; color: #0d1530; font-size: 25px; font-weight: 900; line-height: 1; font-variant-numeric: tabular-nums; }
.metric.green { border-color: #bfe6d0; background: #f3fbf6; }
.metric.green .v { color: #137a3a; }
.metric.amber { border-color: #fed7aa; background: #fff7ed; }
.metric.amber .v { color: #b45309; }
.section-title { display: flex; align-items: center; gap: 8px; margin: 20px 0 10px; color: #0d1530; font-size: 15px; font-weight: 900; }
.section-title::before { content: ""; width: 4px; height: 17px; background: #0066ff; border-radius: 99px; }
table { width: 100%; border-collapse: collapse; font-size: 11px; }
th { background: #eef4ff; color: #334155; border: 1px solid #dbe5f3; padding: 8px 9px; text-align: left; font-size: 10px; text-transform: uppercase; }
td { border: 1px solid #e2e8f0; padding: 8px 9px; vertical-align: top; }
tr:nth-child(even) td { background: #fbfdff; }
.num { text-align: right; font-variant-numeric: tabular-nums; }
.ok { color: #137a3a; font-weight: 900; }
.pill { display: inline-block; border-radius: 999px; padding: 3px 8px; font-size: 10px; font-weight: 850; white-space: nowrap; }
.pill.blue { color: #0757d8; background: #eaf2ff; border: 1px solid #bcd6ff; }
.pill.green { color: #137a3a; background: #e9f8ef; border: 1px solid #bfe6d0; }
.pill.amber { color: #9a3412; background: #fff7ed; border: 1px solid #fed7aa; }
.company-card { border: 1px solid #dbe5f3; border-radius: 8px; background: #ffffff; padding: 14px; margin-bottom: 12px; page-break-inside: avoid; }
.company-head { display: flex; justify-content: space-between; align-items: flex-start; gap: 16px; margin-bottom: 10px; }
.company-title { font-size: 16px; font-weight: 900; color: #0d1530; }
.accent-line { height: 4px; border-radius: 99px; margin: 11px 0 12px; }
.bar-track { height: 8px; border-radius: 999px; background: #e2e8f0; overflow: hidden; }
.bar-fill { height: 8px; border-radius: 999px; }
.two-col { display: grid; grid-template-columns: 1fr 1fr; gap: 14px; }
.note-list { margin: 0; padding-left: 18px; color: #334155; }
.note-list li { margin-bottom: 5px; }
.conclusion { margin-top: 16px; border-radius: 8px; background: #0d1530; color: #ffffff; padding: 15px 16px; }
.conclusion b { color: #7cc7ff; }
.footer { display: flex; justify-content: space-between; gap: 12px; border-top: 1px solid #dbe5f3; margin-top: 18px; padding-top: 11px; color: #64748b; font-size: 10px; }
.avoid-break { page-break-inside: avoid; }
@media (max-width: 760px) { .hero-grid, .metric-grid, .two-col { grid-template-columns: 1fr; } .brand-line, .company-head, .footer { flex-direction: column; align-items: flex-start; } h1 { font-size: 24px; } }
"""


def criteria_html() -> str:
    items = "".join(f"<li>{escape(item)}</li>" for item in CRITERIOS)
    return f"<ul class=\"note-list\">{items}</ul>"


def topbar(title: str, subtitle: str, tag: str, color: str = BLUE) -> str:
    return f"""
  <header class="topbar" style="border-bottom-color:#{escape(color)}">
    <div class="brand-line">
      <div class="brand"><span class="mark">ES</span><span>Easy <span style="color:#5db8ff">Social</span></span></div>
      <div class="tag">{escape(tag)}</div>
    </div>
    <h1>{escape(title)}</h1>
    <div class="subtitle">{escape(subtitle)}</div>
  </header>
"""


def summary_rows(companies: list[dict[str, Any]]) -> str:
    rows = []
    for company in companies:
        rows.append(
            f"""
            <tr>
              <td><strong>{escape(company['sigla'])}</strong><br><span class="muted">{escape(company['nome'])}</span></td>
              <td>{escape(company['periodo_informado']['texto'])}</td>
              <td>{escape(company['periodo_calculo']['texto'])}</td>
              <td class="num">{br_int(company['totais']['dias_8h'])}</td>
              <td class="num">{br_int(company['totais']['sabados_4h'])}</td>
              <td class="num">{br_int(company['totais']['domingos_0h'])}</td>
              <td class="num ok">{br_int(company['totais']['horas'])} h</td>
              <td class="num">{br_percent(company['totais']['participacao'])}</td>
            </tr>
            """
        )
    return "\n".join(rows)


def weekly_rows(company: dict[str, Any]) -> str:
    rows = []
    for weekly_row in company["semanas"]:
        rows.append(
            f"""
            <tr>
              <td>{escape(weekly_row['semana'])}</td>
              <td>{escape(weekly_row['periodo'])}</td>
              <td class="num">{br_int(weekly_row['dias_8h'])}</td>
              <td class="num">{br_int(weekly_row['sabados_4h'])}</td>
              <td class="num">{br_int(weekly_row['domingos_0h'])}</td>
              <td class="num ok">{br_int(weekly_row['horas'])} h</td>
            </tr>
            """
        )
    return "\n".join(rows)


def daily_rows(company: dict[str, Any]) -> str:
    rows = []
    for daily_row in company["dias"]:
        pill_class = "green" if daily_row["horas"] == 8 else "amber" if daily_row["horas"] == 4 else "blue"
        rows.append(
            f"""
            <tr>
              <td>{escape(daily_row['data_br'])}</td>
              <td>{escape(daily_row['dia_semana'])}</td>
              <td><span class="pill {pill_class}">{escape(daily_row['tipo'])}</span></td>
              <td>{escape(daily_row['semana'])}</td>
              <td class="num ok">{br_int(daily_row['horas'])} h</td>
            </tr>
            """
        )
    return "\n".join(rows)


def company_cards(companies: list[dict[str, Any]]) -> str:
    cards = []
    for company in companies:
        participation = company["totais"]["participacao"]
        cards.append(
            f"""
            <article class="company-card">
              <div class="company-head">
                <div>
                  <div class="label">Empresa</div>
                  <div class="company-title">{escape(company['sigla'])}</div>
                  <div class="muted">{escape(company['nome'])} - CNPJ {escape(company['cnpj'])}</div>
                </div>
                <div><span class="pill blue">{escape(company['periodo_informado']['texto'])}</span></div>
              </div>
              <div class="accent-line" style="background:#{escape(company['cor'])}"></div>
              <div class="two-col">
                <div><div class="label">Frente de trabalho</div><div>{escape(company['frente'])}</div></div>
                <div><div class="label">Horas apuradas</div><div class="value">{br_int(company['totais']['horas'])} horas</div><div class="muted">{br_int(company['totais']['dias_8h'])} dias de 8h, {br_int(company['totais']['sabados_4h'])} sábados de 4h e {br_int(company['totais']['domingos_0h'])} domingos sem jornada.</div></div>
              </div>
              <div style="margin-top:12px"><div class="bar-track"><div class="bar-fill" style="width:{participation * 100:.2f}%; background:#{escape(company['cor'])}"></div></div><div class="muted" style="margin-top:5px">Participação no consolidado: {br_percent(participation)}</div></div>
              <div class="panel warning" style="margin-top:12px"><div class="label">Regra de transição</div><div>{escape(company['observacao'])}</div></div>
            </article>
            """
        )
    return "\n".join(cards)


def weekly_block(company: dict[str, Any]) -> str:
    return f"""
    <div class="company-card">
      <div class="company-head"><div><div class="label">Memória semanal</div><div class="company-title">{escape(company['sigla'])}</div></div><div><span class="pill green">{br_int(company['totais']['horas'])} h</span></div></div>
      <table>
        <thead><tr><th>Semana</th><th>Período</th><th class="num">Dias 8h</th><th class="num">Sáb. 4h</th><th class="num">Dom. 0h</th><th class="num">Horas</th></tr></thead>
        <tbody>{weekly_rows(company)}</tbody>
      </table>
    </div>
    """


def render_consolidated_html(report: dict[str, Any]) -> str:
    document = report["documento"]
    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head><meta charset="UTF-8" /><meta name="viewport" content="width=device-width, initial-scale=1.0" /><title>{escape(document['titulo'])} - Consolidado</title><style>{CSS}</style></head>
<body><main class="page">
{topbar(document['titulo'], f"{document['subtitulo']}. Documento executivo com memória de cálculo, separação por empresa e critério auditável de jornada.", document['codigo'])}
  <section class="content">
    <div class="hero-grid avoid-break"><div class="panel soft"><div class="label">Período consolidado</div><div class="value">{escape(document['periodo_consolidado'])}</div><div class="muted">Recorte solicitado: APPA, Soluções e Objetiva, com transições controladas.</div></div><div class="panel"><div class="label">Critério de jornada</div><div class="value">{escape(document['criterio_resumo'])}</div><div class="muted">Gerado em {escape(document['gerado_em'])}</div></div></div>
    <div class="metric-grid avoid-break"><div class="metric"><div class="k">Empresas</div><div class="v">{br_int(report['totais']['empresas'])}</div></div><div class="metric green"><div class="k">Horas consolidadas</div><div class="v">{br_int(report['totais']['horas'])}</div></div><div class="metric"><div class="k">Dias de 8h</div><div class="v">{br_int(report['totais']['dias_8h'])}</div></div><div class="metric amber"><div class="k">Sábados de 4h</div><div class="v">{br_int(report['totais']['sabados_4h'])}</div></div></div>
    <div class="conclusion avoid-break">Total consolidado apurado: <b>{br_int(report['totais']['horas'])} horas trabalhadas</b>, separado por empresa e sem dupla contagem nas datas de transição entre frentes.</div>
    <div class="section-title">Resumo executivo por empresa</div>
    <table><thead><tr><th>Empresa</th><th>Período informado</th><th>Período apurado</th><th class="num">Dias 8h</th><th class="num">Sáb. 4h</th><th class="num">Dom. 0h</th><th class="num">Horas</th><th class="num">Part.</th></tr></thead><tbody>{summary_rows(report['empresas'])}</tbody></table>
    <div class="section-title">Leitura por empresa</div>{company_cards(report['empresas'])}
    <div class="section-title">Critérios de cálculo</div><div class="panel soft">{criteria_html()}</div>
    <div class="section-title">Memória semanal</div>{''.join(weekly_block(company) for company in report['empresas'])}
    <div class="footer"><div>Easy Social - Relatório de horas trabalhadas</div><div>Fonte: premissas informadas pelo solicitante; cálculo estático sem consulta ao eSocial.</div></div>
  </section>
</main></body></html>
"""


def render_company_html(report: dict[str, Any], company: dict[str, Any]) -> str:
    document = report["documento"]
    return f"""<!DOCTYPE html>
<html lang="pt-BR">
<head><meta charset="UTF-8" /><meta name="viewport" content="width=device-width, initial-scale=1.0" /><title>Relatório de Horas Trabalhadas - {escape(company['sigla'])}</title><style>{CSS}</style></head>
<body><main class="page">
{topbar(f"Relatório de Horas Trabalhadas - {company['sigla']}", "Documento individual com memória diária, resumo semanal e critério de jornada.", f"{document['codigo']} / {company['sigla']}", company['cor'])}
  <section class="content">
    <div class="hero-grid avoid-break"><div class="panel soft"><div class="label">Empresa</div><div class="value">{escape(company['nome'])}</div><div class="muted">CNPJ {escape(company['cnpj'])}</div></div><div class="panel"><div class="label">Horas apuradas</div><div class="value">{br_int(company['totais']['horas'])} horas</div><div class="muted">{escape(company['periodo_calculo']['texto'])}</div></div></div>
    <div class="panel avoid-break"><div class="two-col"><div><div class="label">Período informado</div><div class="value">{escape(company['periodo_informado']['texto'])}</div></div><div><div class="label">Período apurado sem dupla contagem</div><div class="value">{escape(company['periodo_calculo']['texto'])}</div></div><div><div class="label">Frente de trabalho</div><div>{escape(company['frente'])}</div></div><div><div class="label">Regra de transição</div><div>{escape(company['observacao'])}</div></div></div></div>
    <div class="metric-grid avoid-break"><div class="metric green"><div class="k">Horas</div><div class="v">{br_int(company['totais']['horas'])}</div></div><div class="metric"><div class="k">Dias de 8h</div><div class="v">{br_int(company['totais']['dias_8h'])}</div></div><div class="metric amber"><div class="k">Sábados de 4h</div><div class="v">{br_int(company['totais']['sabados_4h'])}</div></div><div class="metric"><div class="k">Semanas</div><div class="v">{br_int(company['totais']['semanas'])}</div></div></div>
    <div class="section-title">Resumo semanal</div><table><thead><tr><th>Semana</th><th>Período</th><th class="num">Dias 8h</th><th class="num">Sáb. 4h</th><th class="num">Dom. 0h</th><th class="num">Horas</th></tr></thead><tbody>{weekly_rows(company)}</tbody></table>
    <div class="section-title">Memória diária</div><table><thead><tr><th>Data</th><th>Dia</th><th>Tipo</th><th>Semana</th><th class="num">Horas</th></tr></thead><tbody>{daily_rows(company)}</tbody></table>
    <div class="section-title">Critérios de cálculo</div><div class="panel soft">{criteria_html()}</div>
    <div class="footer"><div>Easy Social - Relatório individual de horas trabalhadas</div><div>Gerado em {escape(document['gerado_em'])}</div></div>
  </section>
</main></body></html>
"""


def excel_fill(color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=color.replace("#", ""))


def excel_border() -> Border:
    side = Side(style="thin", color=GRID)
    return Border(left=side, right=side, top=side, bottom=side)


def write_title(sheet: Any, title: str, subtitle: str, width: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Calibri", bold=True, size=16, color=NAVY)
    title_cell.alignment = Alignment(horizontal="center")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
    subtitle_cell = sheet.cell(row=2, column=1, value=subtitle)
    subtitle_cell.font = Font(name="Calibri", bold=True, size=11, color=SLATE)
    subtitle_cell.alignment = Alignment(horizontal="center")


def style_header(sheet: Any, row_number: int, column_count: int, color: str = NAVY) -> None:
    for column_number in range(1, column_count + 1):
        cell = sheet.cell(row=row_number, column=column_number)
        cell.font = Font(name="Calibri", bold=True, color=WHITE)
        cell.fill = excel_fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = excel_border()


def style_cells(sheet: Any, start_row: int, end_row: int, column_count: int) -> None:
    for row_number in range(start_row, end_row + 1):
        for column_number in range(1, column_count + 1):
            cell = sheet.cell(row=row_number, column=column_number)
            cell.border = excel_border()
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(sheet: Any, widths: list[int]) -> None:
    for column_number, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_number)].width = width


def populate_company_sheet(sheet: Any, company: dict[str, Any]) -> None:
    write_title(sheet, f"Relatório de Horas - {company['sigla']}", company["nome"], 7)
    sheet.append([])
    sheet.append(["CNPJ", company["cnpj"], "Frente", company["frente"], "Horas", company["totais"]["horas"], ""])
    sheet.append(["Período informado", company["periodo_informado"]["texto"], "Período apurado", company["periodo_calculo"]["texto"], "Observação", company["observacao"], ""])
    sheet.append([])
    sheet.append(["Data", "Dia da semana", "Tipo", "Semana", "Horas", "Período apurado", "Empresa"])
    style_header(sheet, 7, 7, company["cor"])
    for daily_row in company["dias"]:
        sheet.append([daily_row["data_br"], daily_row["dia_semana"], daily_row["tipo"], daily_row["semana"], daily_row["horas"], company["periodo_calculo"]["texto"], company["sigla"]])
    style_cells(sheet, 4, sheet.max_row, 7)
    sheet.freeze_panes = "A8"
    sheet.auto_filter.ref = f"A7:G{sheet.max_row}"
    set_widths(sheet, [13, 18, 22, 12, 10, 24, 16])


def create_consolidated_workbook(report: dict[str, Any], path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "RESUMO"
    summary.sheet_properties.tabColor = BLUE
    write_title(summary, report["documento"]["titulo"], report["documento"]["subtitulo"], 8)
    summary.append([])
    summary.append(["Documento", report["documento"]["codigo"], "Gerado em", report["documento"]["gerado_em"]])
    summary.append(["Período consolidado", report["documento"]["periodo_consolidado"], "Critério", report["documento"]["criterio_resumo"]])
    summary.append([])
    summary.append(["Empresa", "Nome", "CNPJ", "Período informado", "Período apurado", "Horas", "Participação", "Observação"])
    style_header(summary, 7, 8)
    for company in report["empresas"]:
        summary.append([company["sigla"], company["nome"], company["cnpj"], company["periodo_informado"]["texto"], company["periodo_calculo"]["texto"], company["totais"]["horas"], company["totais"]["participacao"], company["observacao"]])
    summary.append(["TOTAL", "", "", "", "", report["totais"]["horas"], 1, "Sem dupla contagem nas transições"])
    style_cells(summary, 4, summary.max_row, 8)
    for row_number in range(8, summary.max_row + 1):
        summary.cell(row=row_number, column=7).number_format = "0.0%"
    summary.freeze_panes = "A8"
    summary.auto_filter.ref = f"A7:H{summary.max_row}"
    set_widths(summary, [15, 42, 20, 22, 22, 12, 14, 58])

    memory = workbook.create_sheet("MEMORIA_DIARIA")
    memory.sheet_properties.tabColor = GREEN
    memory.append(["Empresa", "Data", "Dia da semana", "Tipo", "Semana", "Horas", "Período apurado"])
    style_header(memory, 1, 7, GREEN)
    for company in report["empresas"]:
        for daily_row in company["dias"]:
            memory.append([company["sigla"], daily_row["data_br"], daily_row["dia_semana"], daily_row["tipo"], daily_row["semana"], daily_row["horas"], company["periodo_calculo"]["texto"]])
    style_cells(memory, 2, memory.max_row, 7)
    memory.freeze_panes = "A2"
    memory.auto_filter.ref = f"A1:G{memory.max_row}"
    set_widths(memory, [14, 13, 18, 22, 12, 10, 24])

    criteria = workbook.create_sheet("CRITERIOS")
    criteria.sheet_properties.tabColor = AMBER
    write_title(criteria, "Critérios de cálculo", "Premissas aplicadas no relatório", 3)
    criteria.append([])
    criteria.append(["#", "Critério", "Status"])
    style_header(criteria, 4, 3, AMBER)
    for item_number, item in enumerate(report["criterios"], start=1):
        criteria.append([item_number, item, "Aplicado"])
    style_cells(criteria, 5, criteria.max_row, 3)
    set_widths(criteria, [8, 95, 18])

    for company in report["empresas"]:
        company_sheet = workbook.create_sheet(company["slug"].upper())
        company_sheet.sheet_properties.tabColor = company["cor"]
        populate_company_sheet(company_sheet, company)

    workbook.save(path)


def create_company_workbook(company: dict[str, Any], path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "HORAS"
    sheet.sheet_properties.tabColor = company["cor"]
    populate_company_sheet(sheet, company)

    weekly = workbook.create_sheet("SEMANAS")
    weekly.sheet_properties.tabColor = NAVY
    write_title(weekly, f"Resumo semanal - {company['sigla']}", company["periodo_calculo"]["texto"], 6)
    weekly.append([])
    weekly.append(["Semana", "Período", "Dias 8h", "Sábados 4h", "Domingos 0h", "Horas"])
    style_header(weekly, 4, 6)
    for weekly_row in company["semanas"]:
        weekly.append([weekly_row["semana"], weekly_row["periodo"], weekly_row["dias_8h"], weekly_row["sabados_4h"], weekly_row["domingos_0h"], weekly_row["horas"]])
    style_cells(weekly, 5, weekly.max_row, 6)
    set_widths(weekly, [14, 24, 12, 14, 14, 10])
    workbook.save(path)


def paragraph(value: Any, style: ParagraphStyle) -> Paragraph:
    return Paragraph(escape(str(value)), style)


def pdf_styles() -> dict[str, ParagraphStyle]:
    stylesheet = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("ReportTitle", parent=stylesheet["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22, textColor=colors.HexColor(f"#{NAVY}"), alignment=TA_LEFT, spaceAfter=8),
        "subtitle": ParagraphStyle("ReportSubtitle", parent=stylesheet["BodyText"], fontName="Helvetica", fontSize=9, leading=13, textColor=colors.HexColor(f"#{SLATE}"), spaceAfter=10),
        "section": ParagraphStyle("SectionTitle", parent=stylesheet["Heading2"], fontName="Helvetica-Bold", fontSize=12, leading=15, textColor=colors.HexColor(f"#{NAVY}"), spaceBefore=10, spaceAfter=6),
        "body": ParagraphStyle("BodySmall", parent=stylesheet["BodyText"], fontName="Helvetica", fontSize=8, leading=10, textColor=colors.HexColor(f"#{SLATE}")),
        "body_bold": ParagraphStyle("BodySmallBold", parent=stylesheet["BodyText"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.HexColor(f"#{NAVY}")),
        "center": ParagraphStyle("CenterSmall", parent=stylesheet["BodyText"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.HexColor(f"#{NAVY}"), alignment=TA_CENTER),
    }


def pdf_table_style(header_fill: str = NAVY) -> TableStyle:
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{header_fill.replace('#', '')}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("LEADING", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{GRID}")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FBFDFF")]),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ])


def add_pdf_header(story: list[Any], title: str, subtitle: str, styles: dict[str, ParagraphStyle]) -> None:
    logo = Table([[Paragraph("<font color='white'><b>ES</b></font>", styles["center"])]], colWidths=[18 * mm], rowHeights=[12 * mm])
    logo.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(f"#{BLUE}")), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (0, 0), (-1, -1), "CENTER")]))
    header_table = Table([[logo, paragraph(title, styles["title"])]], colWidths=[22 * mm, 158 * mm])
    header_table.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "MIDDLE")]))
    story.append(header_table)
    story.append(paragraph(subtitle, styles["subtitle"]))


def footer(canvas_obj: Any, document: Any) -> None:
    canvas_obj.saveState()
    canvas_obj.setStrokeColor(colors.HexColor(f"#{GRID}"))
    canvas_obj.line(document.leftMargin, 16 * mm, A4[0] - document.rightMargin, 16 * mm)
    canvas_obj.setFont("Helvetica", 7)
    canvas_obj.setFillColor(colors.HexColor(f"#{MUTED}"))
    canvas_obj.drawString(document.leftMargin, 10 * mm, "Easy Social - Relatório de horas trabalhadas")
    canvas_obj.drawRightString(A4[0] - document.rightMargin, 10 * mm, f"Página {document.page}")
    canvas_obj.restoreState()


def metric_table(metrics: list[tuple[str, str]], styles: dict[str, ParagraphStyle]) -> Table:
    table_data = [[paragraph(label, styles["body_bold"]), paragraph(value, styles["center"])] for label, value in metrics]
    table = Table(table_data, colWidths=[50 * mm, 35 * mm])
    table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F8FBFF")), ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor(f"#{GRID}")), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("LEFTPADDING", (0, 0), (-1, -1), 6), ("RIGHTPADDING", (0, 0), (-1, -1), 6), ("TOPPADDING", (0, 0), (-1, -1), 6), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    return table


def add_company_pdf_section(story: list[Any], company: dict[str, Any], styles: dict[str, ParagraphStyle], include_daily: bool) -> None:
    story.append(paragraph(f"{company['sigla']} - {company['nome']}", styles["section"]))
    story.append(metric_table([("CNPJ", company["cnpj"]), ("Período informado", company["periodo_informado"]["texto"]), ("Período apurado", company["periodo_calculo"]["texto"]), ("Horas apuradas", f"{br_int(company['totais']['horas'])} h")], styles))
    story.append(paragraph(f"Frente: {company['frente']}", styles["body"]))
    story.append(paragraph(f"Transição: {company['observacao']}", styles["body"]))
    story.append(Spacer(1, 8))
    story.append(paragraph("Resumo semanal", styles["section"]))
    weekly_table_data = [["Semana", "Período", "Dias 8h", "Sáb. 4h", "Dom. 0h", "Horas"]]
    for weekly_row in company["semanas"]:
        weekly_table_data.append([weekly_row["semana"], weekly_row["periodo"], br_int(weekly_row["dias_8h"]), br_int(weekly_row["sabados_4h"]), br_int(weekly_row["domingos_0h"]), f"{br_int(weekly_row['horas'])} h"])
    weekly_table = Table(weekly_table_data, colWidths=[24 * mm, 42 * mm, 22 * mm, 22 * mm, 22 * mm, 24 * mm], repeatRows=1)
    weekly_table.setStyle(pdf_table_style(company["cor"]))
    story.append(weekly_table)

    if include_daily:
        story.append(paragraph("Memória diária", styles["section"]))
        daily_table_data = [["Data", "Dia", "Tipo", "Semana", "Horas"]]
        for daily_row in company["dias"]:
            daily_table_data.append([daily_row["data_br"], daily_row["dia_semana"], daily_row["tipo"], daily_row["semana"], f"{br_int(daily_row['horas'])} h"])
        daily_table = Table(daily_table_data, colWidths=[24 * mm, 35 * mm, 43 * mm, 27 * mm, 22 * mm], repeatRows=1)
        daily_table.setStyle(pdf_table_style(company["cor"]))
        story.append(daily_table)


def create_consolidated_pdf(report: dict[str, Any], path: Path) -> bool:
    if not REPORTLAB_AVAILABLE:
        return False
    styles = pdf_styles()
    document = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=14 * mm, bottomMargin=20 * mm)
    story: list[Any] = []
    add_pdf_header(story, report["documento"]["titulo"], report["documento"]["subtitulo"], styles)
    story.append(metric_table([("Documento", report["documento"]["codigo"]), ("Período", report["documento"]["periodo_consolidado"]), ("Critério", report["documento"]["criterio_resumo"]), ("Horas consolidadas", f"{br_int(report['totais']['horas'])} h")], styles))
    story.append(Spacer(1, 8))
    story.append(paragraph("Resumo executivo por empresa", styles["section"]))
    summary_table_data = [["Empresa", "Período informado", "Período apurado", "Dias 8h", "Sáb. 4h", "Dom. 0h", "Horas", "Part."]]
    for company in report["empresas"]:
        summary_table_data.append([company["sigla"], company["periodo_informado"]["texto"], company["periodo_calculo"]["texto"], br_int(company["totais"]["dias_8h"]), br_int(company["totais"]["sabados_4h"]), br_int(company["totais"]["domingos_0h"]), f"{br_int(company['totais']['horas'])} h", br_percent(company["totais"]["participacao"])])
    summary_table = Table(summary_table_data, colWidths=[24 * mm, 28 * mm, 28 * mm, 17 * mm, 17 * mm, 17 * mm, 19 * mm, 18 * mm], repeatRows=1)
    summary_table.setStyle(pdf_table_style())
    story.append(summary_table)
    story.append(paragraph("Critérios de cálculo", styles["section"]))
    for criteria in report["criterios"]:
        story.append(paragraph(f"- {criteria}", styles["body"]))
    for company in report["empresas"]:
        story.append(PageBreak())
        add_company_pdf_section(story, company, styles, include_daily=False)
    document.build(story, onFirstPage=footer, onLaterPages=footer)
    return True


def create_company_pdf(company: dict[str, Any], path: Path) -> bool:
    if not REPORTLAB_AVAILABLE:
        return False
    styles = pdf_styles()
    document = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=14 * mm, leftMargin=14 * mm, topMargin=14 * mm, bottomMargin=20 * mm)
    story: list[Any] = []
    add_pdf_header(story, f"Relatório de Horas Trabalhadas - {company['sigla']}", company["nome"], styles)
    add_company_pdf_section(story, company, styles, include_daily=True)
    document.build(story, onFirstPage=footer, onLaterPages=footer)
    return True


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    EMPRESAS_DIR.mkdir(parents=True, exist_ok=True)
    report = build_report()
    DATA_PATH.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    consolidated_html = OUT_DIR / "00_RELATORIO_HORAS_TRABALHADAS_CONSOLIDADO_2026.html"
    consolidated_pdf = OUT_DIR / "00_RELATORIO_HORAS_TRABALHADAS_CONSOLIDADO_2026.pdf"
    consolidated_xlsx = OUT_DIR / "00_RELATORIO_HORAS_TRABALHADAS_CONSOLIDADO_2026.xlsx"
    consolidated_html.write_text(render_consolidated_html(report), encoding="utf-8")
    create_consolidated_workbook(report, consolidated_xlsx)
    generated_paths = [DATA_PATH, consolidated_html, consolidated_xlsx]
    if create_consolidated_pdf(report, consolidated_pdf):
        generated_paths.append(consolidated_pdf)

    for company in report["empresas"]:
        prefix = f"{company['ordem']:02d}_{company['slug'].upper()}_RELATORIO_HORAS_TRABALHADAS_2026"
        html_path = EMPRESAS_DIR / f"{prefix}.html"
        pdf_path = EMPRESAS_DIR / f"{prefix}.pdf"
        xlsx_path = EMPRESAS_DIR / f"{prefix}.xlsx"
        html_path.write_text(render_company_html(report, company), encoding="utf-8")
        create_company_workbook(company, xlsx_path)
        generated_paths.extend([html_path, xlsx_path])
        if create_company_pdf(company, pdf_path):
            generated_paths.append(pdf_path)

    print("Relatório de horas trabalhadas gerado com sucesso.")
    print(f"Total consolidado: {br_int(report['totais']['horas'])} horas")
    for company in report["empresas"]:
        print(f"- {company['sigla']}: {br_int(company['totais']['horas'])} horas ({company['periodo_calculo']['texto']})")
    print("\nArquivos:")
    for path in generated_paths:
        print(path)


if __name__ == "__main__":
    main()
