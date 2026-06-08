from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

try:
    from openpyxl.drawing.image import Image as ExcelImage
except ImportError:
    ExcelImage = None

try:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Image as PdfImage
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "relatorio_ana" / "RELATORIOS_FINAIS_HORAS_GASTO_IA_EMPRESAS_2026"
LOGO_PATH = ROOT / "frontend" / "public" / "brand-logo.png"

NAVY = "0D1530"
BLUE = "0066FF"
GRID = "DBE5F3"
SOFT = "F5F8FC"
TEXT = "172033"
MUTED = "64748B"
WHITE = "FFFFFF"


@dataclass(frozen=True)
class Empresa:
    ordem: int
    slug: str
    nome: str
    cnpj: str
    periodo: str
    horas: int
    gasto_ia_usd: float


EMPRESAS = [
    Empresa(
        ordem=1,
        slug="APPA",
        nome="APPA - Administração dos Portos de Paranaguá e Antonina",
        cnpj="05.969.071/0001-10",
        periodo="24/03/2026 a 06/05/2026",
        horas=280,
        gasto_ia_usd=868.00,
    ),
    Empresa(
        ordem=2,
        slug="SOLUCOES",
        nome="SOLUÇÕES SERVIÇOS TERCEIRIZADOS LTDA",
        cnpj="09.445.502/0001-09",
        periodo="07/05/2026 a 19/05/2026",
        horas=80,
        gasto_ia_usd=586.00,
    ),
    Empresa(
        ordem=3,
        slug="OBJETIVA",
        nome="OBJETIVA SERVIÇOS TERCEIRIZADOS LTDA",
        cnpj="10.874.523/0001-10",
        periodo="20/05/2026 a 29/05/2026",
        horas=68,
        gasto_ia_usd=195.33,
    ),
]


def br_int(value: int) -> str:
    return f"{value:,}".replace(",", ".")


def usd(value: float) -> str:
    formatted = f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"US$ {formatted}"


def html_logo() -> str:
    if LOGO_PATH.exists():
        return '<img src="../../frontend/public/brand-logo.png" alt="Easy Social" class="logo">'
    return '<div class="logo-fallback">Easy Social</div>'


def css() -> str:
    return f"""
* {{ box-sizing: border-box; }}
body {{ margin: 0; background: #edf2f7; color: #{TEXT}; font-family: Arial, Helvetica, sans-serif; }}
.page {{ width: 210mm; min-height: 297mm; margin: 18px auto; background: #fff; box-shadow: 0 12px 30px rgba(15,23,42,.12); }}
.topbar {{ background: #{NAVY}; color: #fff; padding: 28px 32px; }}
.logo {{ width: 128px; max-height: 46px; object-fit: contain; display: block; margin-bottom: 24px; }}
.logo-fallback {{ font-size: 18px; font-weight: 800; margin-bottom: 24px; }}
h1 {{ margin: 0; font-size: 25px; line-height: 1.18; letter-spacing: 0; }}
.company {{ margin-top: 9px; color: #c9d7ea; font-size: 13px; line-height: 1.4; }}
.content {{ padding: 30px 32px; }}
.summary {{ display: grid; grid-template-columns: 1fr 1fr; gap: 14px; margin-bottom: 18px; }}
.box {{ border: 1px solid #{GRID}; border-radius: 8px; background: #{SOFT}; padding: 18px; min-height: 132px; }}
.label {{ color: #{MUTED}; font-size: 11px; text-transform: uppercase; letter-spacing: .45px; font-weight: 700; margin-bottom: 10px; }}
.value {{ color: #{NAVY}; font-size: 31px; font-weight: 850; line-height: 1.05; }}
.sub {{ margin-top: 10px; color: #{MUTED}; font-size: 12px; line-height: 1.35; }}
table {{ width: 100%; border-collapse: collapse; margin-top: 18px; font-size: 12px; }}
th {{ width: 34%; background: #{NAVY}; color: #fff; text-align: left; padding: 11px; font-size: 11px; text-transform: uppercase; letter-spacing: .45px; }}
td {{ border: 1px solid #{GRID}; padding: 12px; vertical-align: top; }}
.footer-line {{ margin-top: 22px; height: 3px; width: 90px; background: #{BLUE}; border-radius: 999px; }}
@media print {{ body {{ background: #fff; }} .page {{ margin: 0; box-shadow: none; }} }}
"""


def render_html(empresa: Empresa) -> str:
    return f"""<!doctype html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<title>Relatório - {empresa.slug}</title>
<style>{css()}</style>
</head>
<body>
<main class="page">
  <header class="topbar">
    {html_logo()}
    <h1>Relatório de Horas Trabalhadas e Gasto com IA</h1>
    <div class="company">{empresa.nome}<br>CNPJ {empresa.cnpj}</div>
  </header>
  <section class="content">
    <div class="summary">
      <div class="box">
        <div class="label">Horas trabalhadas</div>
        <div class="value">{br_int(empresa.horas)} h</div>
        <div class="sub">Período: {empresa.periodo}</div>
      </div>
      <div class="box">
        <div class="label">Valor gasto com IA</div>
        <div class="value">{usd(empresa.gasto_ia_usd)}</div>
        <div class="sub">Gasto de IA/tokens vinculado à empresa.</div>
      </div>
    </div>
    <table>
      <tbody>
        <tr><th>Empresa</th><td>{empresa.nome}</td></tr>
        <tr><th>CNPJ</th><td>{empresa.cnpj}</td></tr>
        <tr><th>Período</th><td>{empresa.periodo}</td></tr>
        <tr><th>Total de horas trabalhadas</th><td>{br_int(empresa.horas)} horas</td></tr>
        <tr><th>Valor gasto com IA</th><td>{usd(empresa.gasto_ia_usd)}</td></tr>
      </tbody>
    </table>
    <div class="footer-line"></div>
  </section>
</main>
</body>
</html>
"""


def create_excel(empresa: Empresa, path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "RELATORIO"
    sheet.sheet_view.showGridLines = False
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 62
    sheet.row_dimensions[1].height = 44

    if ExcelImage and LOGO_PATH.exists():
        try:
            image = ExcelImage(str(LOGO_PATH))
            image.width = 126
            image.height = 44
            sheet.add_image(image, "A1")
        except Exception:
            sheet["A1"] = "Easy Social"
    else:
        sheet["A1"] = "Easy Social"

    sheet.merge_cells("A3:B3")
    sheet["A3"] = "Relatório de Horas Trabalhadas e Gasto com IA"
    sheet["A3"].font = Font(bold=True, size=16, color=NAVY)

    rows = [
        ("Empresa", empresa.nome),
        ("CNPJ", empresa.cnpj),
        ("Período", empresa.periodo),
        ("Total de horas trabalhadas", f"{empresa.horas} horas"),
        ("Valor gasto com IA", usd(empresa.gasto_ia_usd)),
    ]

    start_row = 5
    thin = Side(style="thin", color=GRID)
    for row_index, (label, value) in enumerate(rows, start=start_row):
        sheet.cell(row_index, 1, label)
        sheet.cell(row_index, 2, value)
        for column_index in (1, 2):
            cell = sheet.cell(row_index, column_index)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        sheet.cell(row_index, 1).fill = PatternFill("solid", fgColor=NAVY)
        sheet.cell(row_index, 1).font = Font(bold=True, color=WHITE)
        sheet.cell(row_index, 2).font = Font(color=TEXT)

    workbook.save(path)


def pdf_styles() -> dict[str, ParagraphStyle]:
    base = getSampleStyleSheet()
    return {
        "title": ParagraphStyle("Title", parent=base["Title"], fontName="Helvetica-Bold", fontSize=18, leading=22, textColor=colors.HexColor(f"#{NAVY}"), alignment=TA_LEFT, spaceAfter=8),
        "company": ParagraphStyle("Company", parent=base["Normal"], fontSize=9, leading=12, textColor=colors.HexColor(f"#{MUTED}"), spaceAfter=16),
        "normal": ParagraphStyle("NormalCustom", parent=base["Normal"], fontSize=9, leading=12, textColor=colors.HexColor(f"#{TEXT}"), alignment=TA_LEFT),
        "header": ParagraphStyle("Header", parent=base["Normal"], fontName="Helvetica-Bold", fontSize=8, leading=10, textColor=colors.white, alignment=TA_CENTER),
    }


def pdf_logo() -> PdfImage | None:
    if not LOGO_PATH.exists():
        return None
    try:
        image = PdfImage(str(LOGO_PATH))
        max_width = 38 * mm
        max_height = 14 * mm
        scale = min(max_width / image.imageWidth, max_height / image.imageHeight)
        image.drawWidth = image.imageWidth * scale
        image.drawHeight = image.imageHeight * scale
        return image
    except Exception:
        return None


def create_pdf(empresa: Empresa, path: Path) -> None:
    if not REPORTLAB_AVAILABLE:
        return

    styles = pdf_styles()
    document = SimpleDocTemplate(str(path), pagesize=A4, rightMargin=18 * mm, leftMargin=18 * mm, topMargin=18 * mm, bottomMargin=18 * mm)
    story = []
    logo = pdf_logo()
    if logo:
        story.append(logo)
        story.append(Spacer(1, 10))

    story.append(Paragraph("Relatório de Horas Trabalhadas e Gasto com IA", styles["title"]))
    story.append(Paragraph(f"{empresa.nome}<br/>CNPJ {empresa.cnpj}", styles["company"]))

    table_data = [
        [Paragraph("Campo", styles["header"]), Paragraph("Informação", styles["header"])],
        [Paragraph("Empresa", styles["normal"]), Paragraph(empresa.nome, styles["normal"])],
        [Paragraph("CNPJ", styles["normal"]), Paragraph(empresa.cnpj, styles["normal"])],
        [Paragraph("Período", styles["normal"]), Paragraph(empresa.periodo, styles["normal"])],
        [Paragraph("Total de horas trabalhadas", styles["normal"]), Paragraph(f"{br_int(empresa.horas)} horas", styles["normal"])],
        [Paragraph("Valor gasto com IA", styles["normal"]), Paragraph(usd(empresa.gasto_ia_usd), styles["normal"])],
    ]
    table = Table(table_data, colWidths=[55 * mm, 115 * mm], hAlign="LEFT", repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{NAVY}")),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor(f"#{GRID}")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 7),
        ("RIGHTPADDING", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 7),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
    ]))
    story.append(table)
    document.build(story)


def main() -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    generated: list[Path] = []

    for empresa in EMPRESAS:
        prefix = f"{empresa.ordem:02d}_{empresa.slug}_RELATORIO_HORAS_GASTO_IA_2026"
        html_path = OUT_DIR / f"{prefix}.html"
        pdf_path = OUT_DIR / f"{prefix}.pdf"
        xlsx_path = OUT_DIR / f"{prefix}.xlsx"

        html_path.write_text(render_html(empresa), encoding="utf-8")
        create_pdf(empresa, pdf_path)
        create_excel(empresa, xlsx_path)

        generated.extend([html_path, pdf_path, xlsx_path])

    print("Relatórios finais gerados:")
    for path in generated:
        print(path.relative_to(ROOT))


if __name__ == "__main__":
    main()