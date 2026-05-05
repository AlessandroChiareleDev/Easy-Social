"""
Ingestao Lote 2 Dezembro/2025 — popula s1210_cpf_scope + s1210_operadoras.

XLSX: 12 Dezembro_lote 002_APPA (2).xlsx
  1001 CPFs unicos, todos '2 Lote'

Aba "Lote para Envio " (trailing space intencional):
  col 0  = CPF formatado
  col 8  = Lotes ("2 Lote")
  col 9  = CPF (duplicado — mais confiavel, usa esse)

Aba "Assistencia Medica":
  col 0  = Lotes
  col 8  = CPF
  col 10 = CodigoEvento
  col 14 = CNPJ OPERADORA
  col 15 = ANS
  col 18 = ValorEvento (centavos)

Regras planSaude:
  - Ignorar CodigoEvento 9279 e 9281 (informativos)
  - Ignorar CNPJ 'Informativo', '-', vazio
  - Exige CNPJ com 14 digitos e ANS numerico
  - Soma valores por (cpf, cnpj_operadora)
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import psycopg2
import psycopg2.extras
from db_config import DB_CONFIG
from openpyxl import load_workbook

XLSX_PATH  = Path(r"C:\Users\xandao\Downloads\12 Dezembro_lote 002_APPA (2).xlsx")
ABA_GERAL  = "Lote para Envio "   # trailing space intencional — nome real da aba
ABA_OPER   = "Assistencia Médica"
PER_APUR   = "2025-12"
EMPRESA_ID = 1
LOTE_ALVO  = 2

IGNORAR_CNPJ = {"Informativo", "-", "", "None", "nan"}
IGNORAR_COD = {"9279", "9281", "774"}  # 774 = DESC.ASSIST.MEDICA (informativa, não gera planSaude)


def _cpf(v) -> str:
    return re.sub(r"\D", "", str(v or "")).zfill(11)


def _cnpj_limpo(v) -> str:
    return re.sub(r"\D", "", str(v or ""))


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"[ERR] XLSX nao encontrado: {XLSX_PATH}")
        return 2

    data = XLSX_PATH.read_bytes()
    sha  = hashlib.sha256(data).hexdigest()
    print(f"[info] sha256={sha[:12]}... size={len(data)/1024:.1f}KB")

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    print(f"[info] abas: {wb.sheetnames}")

    # ── Aba principal: CPFs do Lote 2 ────────────────────────────────
    if ABA_GERAL not in wb.sheetnames:
        print(f"[ERR] aba '{ABA_GERAL}' nao encontrada")
        return 3

    ws = wb[ABA_GERAL]
    it = ws.iter_rows(values_only=True)
    next(it)  # pula header

    scope_rows: list[tuple] = []
    vistos: set[str] = set()
    totais: dict[int, int] = {}

    for idx, row in enumerate(it, start=2):
        if not row or row[0] is None:
            continue
        if len(row) < 10:
            continue
        lote_str = str(row[8] or "").strip()
        # Normaliza: "2 Lote", "2o Lote", "2º Lote" -> extrai o digito
        lote_num = None
        for ch in lote_str:
            if ch.isdigit():
                lote_num = int(ch)
                break
        if lote_num is None:
            continue

        # col 9 = CPF duplicado (mais limpo), fallback col 0
        cpf = _cpf(row[9] if row[9] is not None else row[0])
        if len(cpf) != 11:
            continue
        if cpf in vistos:
            continue
        vistos.add(cpf)
        totais[lote_num] = totais.get(lote_num, 0) + 1

        raw = {f"c{i}": (str(v) if v is not None else None) for i, v in enumerate(row)}
        scope_rows.append((EMPRESA_ID, PER_APUR, cpf, None, None, lote_num, idx, json.dumps(raw)))

    print(f"[info] parse scope: {totais} (total={sum(totais.values())})")
    if totais.get(LOTE_ALVO, 0) == 0:
        print(f"[ERR] nenhuma linha com Lote {LOTE_ALVO} encontrada")
        return 4

    # ── Aba Assistencia Medica: planSaude por CPF ─────────────────────
    if ABA_OPER not in wb.sheetnames:
        print(f"[WARN] aba '{ABA_OPER}' nao encontrada — operadoras nao populadas")
        agg: dict = {}
    else:
        ws2 = wb[ABA_OPER]
        it2 = ws2.iter_rows(values_only=True)
        next(it2)  # pula header

        # agg[cpf][(cnpj, rubrica)] = {"reg_ans": str, "vlr_cent": int}
        # Agrupamento por (cnpj, rubrica) para nao somar verbas de naturezas diferentes
        agg = defaultdict(lambda: defaultdict(lambda: {"reg_ans": "", "vlr_cent": 0}))

        linhas_processadas = 0
        linhas_ignoradas   = 0
        for row2 in it2:
            if not row2 or row2[0] is None:
                continue
            lote_str = str(row2[0] or "").strip()
            if "2" not in lote_str:
                continue

            cpf = _cpf(row2[8])
            if len(cpf) != 11:
                linhas_ignoradas += 1
                continue

            cod_evento = str(row2[10] or "").strip()
            if cod_evento in IGNORAR_COD:
                linhas_ignoradas += 1
                continue

            cnpj_raw = str(row2[14] or "").strip()
            if cnpj_raw in IGNORAR_CNPJ:
                linhas_ignoradas += 1
                continue

            cnpj = _cnpj_limpo(cnpj_raw)
            if len(cnpj) != 14:
                linhas_ignoradas += 1
                continue

            ans = re.sub(r"\D", "", str(row2[15] or "")).strip()
            if not ans or ans == "0":
                linhas_ignoradas += 1
                continue

            valor = row2[18]
            try:
                centavos = int(valor or 0)
            except (TypeError, ValueError):
                centavos = 0

            agg[cpf][(cnpj, cod_evento)]["reg_ans"] = ans
            agg[cpf][(cnpj, cod_evento)]["vlr_cent"] += centavos
            linhas_processadas += 1

        print(f"[info] operadoras: {len(agg)} CPFs, "
              f"{linhas_processadas} linhas incluidas, {linhas_ignoradas} ignoradas")

    wb.close()

    # ── DB ────────────────────────────────────────────────────────────
    conn = psycopg2.connect(**DB_CONFIG)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Upsert s1210_xlsx
            cur.execute(
                "SELECT id FROM s1210_xlsx WHERE empresa_id=%s AND per_apur=%s AND sha256=%s",
                (EMPRESA_ID, PER_APUR, sha),
            )
            row_ex = cur.fetchone()
            if row_ex:
                xlsx_id = row_ex["id"]
                print(f"[info] s1210_xlsx ja existia id={xlsx_id}")
            else:
                cur.execute(
                    """INSERT INTO s1210_xlsx
                         (empresa_id, per_apur, nome_arquivo, tamanho_bytes,
                          sha256, storage_path, aba_geral, aba_operadoras,
                          parse_ok, totais_json)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,TRUE,%s)
                       RETURNING id""",
                    (EMPRESA_ID, PER_APUR, XLSX_PATH.name, len(data),
                     sha, f"local/{PER_APUR}/{sha[:12]}.xlsx",
                     ABA_GERAL, ABA_OPER,
                     json.dumps({f"lote{k}": v for k, v in totais.items()})),
                )
                xlsx_id = cur.fetchone()["id"]
                print(f"[ok] s1210_xlsx inserido id={xlsx_id}")

            # Limpa scope lote 2 anterior (preserva lote 1/3)
            cur.execute(
                "DELETE FROM s1210_cpf_scope "
                "WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s",
                (EMPRESA_ID, PER_APUR, LOTE_ALVO),
            )
            print(f"[info] scope anterior deletado ({cur.rowcount} linhas)")

            # Insere scope (apenas lote 2)
            scope_lote2 = [r for r in scope_rows if r[5] == LOTE_ALVO]
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO s1210_cpf_scope
                     (xlsx_id, empresa_id, per_apur, cpf, nome, matricula,
                      lote_num, row_number, raw_row)
                   VALUES %s""",
                [(xlsx_id, *r) for r in scope_lote2],
                template="(%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb)",
            )
            print(f"[ok] {len(scope_lote2)} CPFs inseridos em s1210_cpf_scope (lote={LOTE_ALVO})")

            # Limpa operadoras lote 2 anterior
            cur.execute(
                "DELETE FROM s1210_operadoras "
                "WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s",
                (EMPRESA_ID, PER_APUR, LOTE_ALVO),
            )
            print(f"[info] operadoras anterior deletado ({cur.rowcount} linhas)")

            # Insere operadoras (1 linha por (cpf, cnpj, rubrica) unico)
            oper_rows: list[tuple] = []
            for cpf, combos in agg.items():
                for (cnpj, rubrica), d in combos.items():
                    if d["vlr_cent"] <= 0:
                        continue
                    oper_rows.append((
                        xlsx_id, EMPRESA_ID, PER_APUR, cpf, rubrica,
                        cnpj, d["reg_ans"], None,
                        d["vlr_cent"], LOTE_ALVO,
                    ))

            if oper_rows:
                psycopg2.extras.execute_values(
                    cur,
                    """INSERT INTO s1210_operadoras
                         (xlsx_id, empresa_id, per_apur, cpf, rubrica_origem,
                          cnpj_operadora, reg_ans, nome_operadora,
                          valor, lote_num)
                       VALUES %s""",
                    oper_rows,
                )
                print(f"[ok] {len(oper_rows)} linhas inseridas em s1210_operadoras (lote={LOTE_ALVO})")
            else:
                print("[warn] nenhuma linha para s1210_operadoras")

        conn.commit()
    finally:
        conn.close()

    # ── Confirmacao ───────────────────────────────────────────────────
    conn = psycopg2.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT lote_num, COUNT(*) FROM s1210_cpf_scope
                    WHERE empresa_id=%s AND per_apur=%s
                    GROUP BY lote_num ORDER BY lote_num""",
                (EMPRESA_ID, PER_APUR),
            )
            print("[check scope 2025-12]", cur.fetchall())

            cur.execute(
                """SELECT lote_num, COUNT(*) FROM s1210_operadoras
                    WHERE empresa_id=%s AND per_apur=%s
                    GROUP BY lote_num ORDER BY lote_num""",
                (EMPRESA_ID, PER_APUR),
            )
            print("[check operadoras 2025-12]", cur.fetchall())
    finally:
        conn.close()

    return 0


if __name__ == "__main__":
    sys.exit(main())
