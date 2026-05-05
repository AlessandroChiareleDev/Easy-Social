"""
Ingestao COMPLEMENTAR Lote 2 Maio/2025 — 411 CPFs.
XLSX: 05_Maio_lote 002_APPA_Complementar.xlsx
Recria s1210_cpf_scope e s1210_operadoras de 2025-05 lote=2.
Envios anteriores ('ok','na') sao preservados (tabela s1210_cpf_envios).
"""
from __future__ import annotations

import hashlib
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import psycopg2
import psycopg2.extras
from db_config import DB_CONFIG
from openpyxl import load_workbook

XLSX_PATH  = Path(r"C:\Users\xandao\Downloads\05_Maio_lote 002_APPA_Complementar.xlsx")
ABA_GERAL  = "Lote para envio"
ABA_OPER   = "Assistencia Médica"
PER_APUR   = "2025-05"
EMPRESA_ID = 1
LOTE_ALVO  = 2

IGNORAR_CNPJ = {"Informativo", "-", "", "None", "nan"}
IGNORAR_COD  = {"9279", "9281", "774"}


def _cpf(v) -> str:
    return re.sub(r"\D", "", str(v or "")).zfill(11)


def _cnpj_limpo(v) -> str:
    return re.sub(r"\D", "", str(v or ""))


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"[ERR] XLSX nao encontrado: {XLSX_PATH}")
        return 2

    data = XLSX_PATH.read_bytes()
    sha  = hashlib.sha256(data).hexdigest()
    print(f"[info] sha256={sha[:12]}... size={len(data)/1024:.1f}KB")

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    print(f"[info] abas: {wb.sheetnames}")

    if ABA_GERAL not in wb.sheetnames:
        print(f"[ERR] aba '{ABA_GERAL}' nao encontrada")
        return 3

    ws = wb[ABA_GERAL]
    it = ws.iter_rows(values_only=True)
    next(it)

    scope_rows: list[tuple] = []
    vistos: set[str] = set()
    totais: dict[int, int] = {}

    for idx, row in enumerate(it, start=2):
        if not row or row[0] is None:
            continue
        if len(row) < 9:
            continue
        lote_str = str(row[0] or "").strip()
        lote_num = None
        for ch in lote_str:
            if ch.isdigit():
                lote_num = int(ch)
                break
        if lote_num is None:
            continue
        cpf = _cpf(row[8])
        if len(cpf) != 11:
            continue
        if cpf in vistos:
            continue
        vistos.add(cpf)
        totais[lote_num] = totais.get(lote_num, 0) + 1
        raw = {f"c{i}": (str(v) if v is not None else None) for i, v in enumerate(row)}
        scope_rows.append((EMPRESA_ID, PER_APUR, cpf, None, None, lote_num, idx, json.dumps(raw)))

    print(f"[info] parse scope: {totais} (total={sum(totais.values())})")
    if totais.get(LOTE_ALVO, 0) == 0:
        print(f"[ERR] nenhuma linha com Lote {LOTE_ALVO}")
        return 4

    if ABA_OPER not in wb.sheetnames:
        print(f"[WARN] aba '{ABA_OPER}' nao encontrada")
        agg = {}
    else:
        ws2 = wb[ABA_OPER]
        it2 = ws2.iter_rows(values_only=True)
        next(it2)
        agg = defaultdict(lambda: defaultdict(lambda: {"reg_ans": "", "vlr_cent": 0}))

        proc = ign = 0
        for row2 in it2:
            if not row2 or row2[0] is None:
                continue
            lote_str = str(row2[0] or "").strip()
            if "2" not in lote_str:
                continue
            cpf = _cpf(row2[8])
            if len(cpf) != 11:
                ign += 1; continue
            cod = str(row2[10] or "").strip()
            if cod in IGNORAR_COD:
                ign += 1; continue
            cnpj_raw = str(row2[14] or "").strip()
            if cnpj_raw in IGNORAR_CNPJ:
                ign += 1; continue
            cnpj = _cnpj_limpo(cnpj_raw)
            if len(cnpj) != 14:
                ign += 1; continue
            ans = re.sub(r"\D", "", str(row2[15] or "")).strip()
            if not ans or ans == "0":
                ign += 1; continue
            try:
                cents = int(row2[18] or 0)
            except (TypeError, ValueError):
                cents = 0
            agg[cpf][(cnpj, cod)]["reg_ans"] = ans
            agg[cpf][(cnpj, cod)]["vlr_cent"] += cents
            proc += 1
        print(f"[info] operadoras: {len(agg)} CPFs, {proc} linhas inc, {ign} ignoradas")

    wb.close()

    conn = psycopg2.connect(**DB_CONFIG)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT id FROM s1210_xlsx WHERE empresa_id=%s AND per_apur=%s AND sha256=%s",
                (EMPRESA_ID, PER_APUR, sha),
            )
            row_ex = cur.fetchone()
            if row_ex:
                xlsx_id = row_ex["id"]
                print(f"[info] s1210_xlsx ja existia id={xlsx_id}")
            else:
                cur.execute(
                    """INSERT INTO s1210_xlsx
                         (empresa_id, per_apur, nome_arquivo, tamanho_bytes,
                          sha256, storage_path, aba_geral, aba_operadoras,
                          parse_ok, totais_json)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,TRUE,%s)
                       RETURNING id""",
                    (EMPRESA_ID, PER_APUR, XLSX_PATH.name, len(data),
                     sha, f"local/{PER_APUR}/{sha[:12]}.xlsx",
                     ABA_GERAL, ABA_OPER,
                     json.dumps({f"lote{k}": v for k, v in totais.items()})),
                )
                xlsx_id = cur.fetchone()["id"]
                print(f"[ok] s1210_xlsx inserido id={xlsx_id}")

            # UPSERT: move CPFs para lote 2 mesmo se ja existirem em outro lote do mesmo per_apur.
            scope_lote2 = [r for r in scope_rows if r[5] == LOTE_ALVO]
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO s1210_cpf_scope
                     (xlsx_id, empresa_id, per_apur, cpf, nome, matricula,
                      lote_num, row_number, raw_row)
                   VALUES %s
                   ON CONFLICT (empresa_id, per_apur, cpf) DO UPDATE SET
                     xlsx_id=EXCLUDED.xlsx_id,
                     lote_num=EXCLUDED.lote_num,
                     row_number=EXCLUDED.row_number,
                     raw_row=EXCLUDED.raw_row""",
                [(xlsx_id, *r) for r in scope_lote2],
                template="(%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb)",
            )
            print(f"[ok] {len(scope_lote2)} CPFs upserted em s1210_cpf_scope (lote=2)")

            # Operadoras: limpa apenas lote 2 (operadoras nao tem unique global por cpf)
            cpfs_xlsx = tuple({r[2] for r in scope_lote2})
            if cpfs_xlsx:
                cur.execute(
                    "DELETE FROM s1210_operadoras WHERE empresa_id=%s AND per_apur=%s AND cpf IN %s",
                    (EMPRESA_ID, PER_APUR, cpfs_xlsx),
                )
                print(f"[info] operadoras deletadas dos {len(cpfs_xlsx)} CPFs ({cur.rowcount} linhas)")

            oper_rows: list[tuple] = []
            for cpf, combos in agg.items():
                for (cnpj, rub), d in combos.items():
                    if d["vlr_cent"] <= 0:
                        continue
                    oper_rows.append((
                        xlsx_id, EMPRESA_ID, PER_APUR, cpf, rub,
                        cnpj, d["reg_ans"], None,
                        d["vlr_cent"], LOTE_ALVO,
                    ))
            if oper_rows:
                psycopg2.extras.execute_values(
                    cur,
                    """INSERT INTO s1210_operadoras
                         (xlsx_id, empresa_id, per_apur, cpf, rubrica_origem,
                          cnpj_operadora, reg_ans, nome_operadora,
                          valor, lote_num)
                       VALUES %s""",
                    oper_rows,
                )
                print(f"[ok] {len(oper_rows)} em s1210_operadoras")

        conn.commit()
        with conn.cursor() as cur:
            cur.execute(
                "SELECT lote_num, COUNT(*) FROM s1210_cpf_scope "
                "WHERE empresa_id=%s AND per_apur=%s GROUP BY lote_num ORDER BY lote_num",
                (EMPRESA_ID, PER_APUR),
            )
            print("[scope 2025-05]", cur.fetchall())
            cur.execute("""
                WITH lv AS (
                  SELECT DISTINCT ON (cpf) cpf, status
                  FROM s1210_cpf_envios
                  WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s
                  ORDER BY cpf, enviado_em DESC NULLS LAST
                )
                SELECT
                  COUNT(*) AS escopo,
                  COUNT(*) FILTER (WHERE lv.status='ok') AS ja_ok,
                  COUNT(*) FILTER (WHERE lv.status='na') AS ja_na,
                  COUNT(*) FILTER (WHERE lv.status IS NULL OR lv.status NOT IN ('ok','na')) AS pendentes
                FROM s1210_cpf_scope s
                LEFT JOIN lv ON lv.cpf=s.cpf
                WHERE s.empresa_id=%s AND s.per_apur=%s AND s.lote_num=%s
            """, (EMPRESA_ID, PER_APUR, LOTE_ALVO, EMPRESA_ID, PER_APUR, LOTE_ALVO))
            print("[pendentes]", cur.fetchone())
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
