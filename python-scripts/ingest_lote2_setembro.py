"""Ingestao Lote 2 Setembro/2025 APPA."""
from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

import psycopg2
import psycopg2.extras
from db_config import DB_CONFIG
from openpyxl import load_workbook

XLSX_PATH = Path(r"C:\Users\xandao\Downloads\09 Setembro_lote 002_APPA.xlsx")
ABA_GERAL = "Lote para Envio"
ABA_OPER = "Assistencia Médica"
PER_APUR = "2025-09"
EMPRESA_ID = 1
LOTE_ALVO = 2

COL_GERAL_LOTE = 0
COL_GERAL_CPF = 8
COL_OPER_LOTE = 0
COL_OPER_CPF = 8
COL_OPER_COD = 10
COL_OPER_CNPJ = 14
COL_OPER_ANS = 15
COL_OPER_VALOR = 18

IGNORAR_CNPJ = {"Informativo", "-", "", "None", "nan"}
IGNORAR_COD = {"9279", "9281", "774"}  # 774 = DESC.ASSIST.MEDICA (informativa, não gera planSaude)


def _cpf(valor) -> str:
    return re.sub(r"\D", "", str(valor or "")).zfill(11)


def _cnpj_limpo(valor) -> str:
    return re.sub(r"\D", "", str(valor or ""))


def _lote_num(valor) -> int | None:
    texto = str(valor or "").strip()
    for ch in texto:
        if ch.isdigit():
            return int(ch)
    return None


def parse_xlsx():
    if not XLSX_PATH.exists():
        raise FileNotFoundError(f"XLSX nao encontrado: {XLSX_PATH}")

    data = XLSX_PATH.read_bytes()
    sha = hashlib.sha256(data).hexdigest()
    print(f"[info] arquivo={XLSX_PATH.name} sha256={sha[:12]} size={len(data)/1024:.1f}KB")

    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    try:
        print(f"[info] abas={wb.sheetnames}")
        if ABA_GERAL not in wb.sheetnames:
            raise RuntimeError(f"aba '{ABA_GERAL}' nao encontrada")
        if ABA_OPER not in wb.sheetnames:
            raise RuntimeError(f"aba '{ABA_OPER}' nao encontrada")

        scope_rows: list[tuple] = []
        vistos: set[str] = set()
        totais: dict[int, int] = {}
        ws = wb[ABA_GERAL]
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) <= COL_GERAL_CPF:
                continue
            lote_num = _lote_num(row[COL_GERAL_LOTE])
            if lote_num is None:
                continue
            cpf = _cpf(row[COL_GERAL_CPF])
            if len(cpf) != 11 or cpf in vistos:
                continue
            vistos.add(cpf)
            totais[lote_num] = totais.get(lote_num, 0) + 1
            raw = {f"c{i}": (str(v) if v is not None else None) for i, v in enumerate(row)}
            scope_rows.append((EMPRESA_ID, PER_APUR, cpf, None, None, lote_num, idx, json.dumps(raw)))

        print(f"[info] parse scope: {totais} total={sum(totais.values())}")
        if totais.get(LOTE_ALVO, 0) == 0:
            raise RuntimeError(f"nenhuma linha do lote {LOTE_ALVO} encontrada")

        agg = defaultdict(lambda: defaultdict(lambda: {"reg_ans": "", "vlr_cent": 0}))
        filtros = {"incluidas": 0, "cpf": 0, "cod_info": 0, "cnpj": 0, "ans": 0, "valor": 0}
        ws2 = wb[ABA_OPER]
        for row in ws2.iter_rows(min_row=2, values_only=True):
            if not row or len(row) <= COL_OPER_VALOR:
                continue
            if _lote_num(row[COL_OPER_LOTE]) != LOTE_ALVO:
                continue
            cpf = _cpf(row[COL_OPER_CPF])
            if len(cpf) != 11:
                filtros["cpf"] += 1
                continue
            cod_evento = str(row[COL_OPER_COD] or "").strip()
            if cod_evento in IGNORAR_COD:
                filtros["cod_info"] += 1
                continue
            cnpj_raw = str(row[COL_OPER_CNPJ] or "").strip()
            if cnpj_raw in IGNORAR_CNPJ:
                filtros["cnpj"] += 1
                continue
            cnpj = _cnpj_limpo(cnpj_raw)
            if len(cnpj) != 14:
                filtros["cnpj"] += 1
                continue
            ans = re.sub(r"\D", "", str(row[COL_OPER_ANS] or "")).strip()
            if not ans or ans == "0":
                filtros["ans"] += 1
                continue
            try:
                centavos = int(row[COL_OPER_VALOR] or 0)
            except (TypeError, ValueError):
                centavos = 0
            if centavos <= 0:
                filtros["valor"] += 1
                continue
            agg[cpf][(cnpj, cod_evento)]["reg_ans"] = ans
            agg[cpf][(cnpj, cod_evento)]["vlr_cent"] += centavos
            filtros["incluidas"] += 1

        print(f"[info] operadoras: {len(agg)} CPFs filtros={filtros}")
        return data, sha, scope_rows, totais, agg
    finally:
        wb.close()


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--executar", action="store_true", help="grava no banco")
    args = ap.parse_args()

    data, sha, scope_rows, totais, agg = parse_xlsx()
    scope_lote2 = [row for row in scope_rows if row[5] == LOTE_ALVO]
    oper_rows: list[tuple] = []
    for cpf, combos in agg.items():
        for (cnpj, rubrica), item in combos.items():
            if item["vlr_cent"] <= 0:
                continue
            oper_rows.append((EMPRESA_ID, PER_APUR, cpf, LOTE_ALVO, rubrica, cnpj, item["reg_ans"], item["vlr_cent"]))

    print(f"[resumo] scope_lote2={len(scope_lote2)} oper_rows={len(oper_rows)}")
    print(f"[resumo] CPFs com operadora={len({r[2] for r in oper_rows})}")

    conn = psycopg2.connect(**DB_CONFIG)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                "SELECT lote_num, COUNT(*) FROM s1210_cpf_scope "
                "WHERE empresa_id=%s AND per_apur=%s AND cpf=ANY(%s) "
                "GROUP BY lote_num ORDER BY lote_num",
                (EMPRESA_ID, PER_APUR, [row[2] for row in scope_lote2]),
            )
            conflitos = cur.fetchall()
            print(f"[check] conflitos com scope existente={conflitos}")
            if conflitos:
                raise RuntimeError(f"CPFs do lote 2 ja existem em outro scope de {PER_APUR}: {conflitos}")

            if not args.executar:
                print("[dry-run] use --executar para gravar no banco")
                return 0

            cur.execute(
                "SELECT id FROM s1210_xlsx WHERE empresa_id=%s AND per_apur=%s AND sha256=%s",
                (EMPRESA_ID, PER_APUR, sha),
            )
            row_ex = cur.fetchone()
            if row_ex:
                xlsx_id = row_ex["id"]
                print(f"[info] s1210_xlsx ja existia id={xlsx_id}")
            else:
                cur.execute(
                    """INSERT INTO s1210_xlsx
                         (empresa_id, per_apur, nome_arquivo, tamanho_bytes,
                          sha256, storage_path, aba_geral, aba_operadoras,
                          parse_ok, totais_json)
                       VALUES (%s,%s,%s,%s,%s,%s,%s,%s,TRUE,%s)
                       RETURNING id""",
                    (EMPRESA_ID, PER_APUR, XLSX_PATH.name, len(data), sha,
                     f"local/{PER_APUR}/{sha[:12]}.xlsx", ABA_GERAL, ABA_OPER,
                     json.dumps({f"{k}_LOTE": v for k, v in totais.items()})),
                )
                xlsx_id = cur.fetchone()["id"]
                print(f"[ok] s1210_xlsx inserido id={xlsx_id}")

            cur.execute(
                "DELETE FROM s1210_cpf_scope WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s",
                (EMPRESA_ID, PER_APUR, LOTE_ALVO),
            )
            print(f"[info] scope anterior deletado={cur.rowcount}")

            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO s1210_cpf_scope
                     (xlsx_id, empresa_id, per_apur, cpf, nome, matricula,
                      lote_num, row_number, raw_row)
                   VALUES %s""",
                [(xlsx_id, *row) for row in scope_lote2],
                template="(%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb)",
                page_size=500,
            )
            print(f"[ok] scope inserido={len(scope_lote2)}")

            cur.execute(
                "DELETE FROM s1210_operadoras WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s",
                (EMPRESA_ID, PER_APUR, LOTE_ALVO),
            )
            print(f"[info] operadoras anteriores deletadas={cur.rowcount}")

            if oper_rows:
                psycopg2.extras.execute_values(
                    cur,
                    """INSERT INTO s1210_operadoras
                         (empresa_id, per_apur, cpf, lote_num,
                          rubrica_origem, cnpj_operadora, reg_ans, valor)
                       VALUES %s""",
                    oper_rows,
                    page_size=500,
                )
                print(f"[ok] operadoras inseridas={len(oper_rows)}")
        conn.commit()
    finally:
        conn.close()

    with psycopg2.connect(**DB_CONFIG) as conn, conn.cursor() as cur:
        cur.execute(
            "SELECT lote_num, COUNT(*) FROM s1210_cpf_scope "
            "WHERE empresa_id=%s AND per_apur=%s GROUP BY lote_num ORDER BY lote_num",
            (EMPRESA_ID, PER_APUR),
        )
        print(f"[check scope {PER_APUR}]", cur.fetchall())
        cur.execute(
            "SELECT lote_num, COUNT(*), COUNT(DISTINCT cpf) FROM s1210_operadoras "
            "WHERE empresa_id=%s AND per_apur=%s GROUP BY lote_num ORDER BY lote_num",
            (EMPRESA_ID, PER_APUR),
        )
        print(f"[check operadoras {PER_APUR}]", cur.fetchall())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())