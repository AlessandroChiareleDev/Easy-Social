"""
Ingestao direta (SQL) do XLSX 05 Maio_lote 003_APPA.xlsx → popula
s1210_xlsx + s1210_cpf_scope para (empresa_id=1, per_apur=2025-05).

Por que nao usar a rota /xlsx/ingest?
- A rota tem whitelist hardcoded ("2025-02","2025-03","2025-04") e nao
  parseia a aba de operadoras (codigo ANS + CNPJ) — o XLSX do Maio
  nem veio com essa aba ainda. Entao fazemos ingestao manual do scope
  aqui, do mesmo jeito que Fev/Mar/Abr tiveram as operadoras populadas
  por outro caminho (xlsx_id=NULL naquelas linhas).

Layout aba "Lote Para Envio":
  c0=CodigoEmpresa c1=Competencia c2=CodigoLote c3=CodigoFilial
  c4=CodigoFuncionario c5=Concatenar c6='3° Lote' c7=CPF formatado
"""
from __future__ import annotations
import hashlib, json, sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
import psycopg2, psycopg2.extras
from db_config import DB_CONFIG
from openpyxl import load_workbook

XLSX_PATH = Path(r"C:\Users\NITRO\Downloads\08 Agosto_lote 003_APPA  com cpf.xlsx")
ABA_GERAL = "Lote para Envio"
ABA_OPER  = "Assistencia Médica"   # nao eh usada agora (sem CNPJ/ANS)
PER_APUR  = "2025-08"
COL_LOTE  = 0   # 'Lotes' = '3º Lote'
COL_CPF   = 9   # 'CPF' = '247.672.218-97'
EMPRESA_ID = 1
LOTE_ALVO = 3  # ingere so o Lote 3 (o XLSX jah veio filtrado para o 3)


def _norm_lote(v) -> int | None:
    s = str(v or "").strip()
    for ch in s:
        if ch.isdigit() and ch in "1234":
            return int(ch)
    return None


def main() -> int:
    if not XLSX_PATH.exists():
        print(f"[ERR] XLSX nao encontrado: {XLSX_PATH}")
        return 2

    data = XLSX_PATH.read_bytes()
    sha = hashlib.sha256(data).hexdigest()
    print(f"[info] XLSX sha256={sha[:12]}... size={len(data)/1024:.1f}KB")

    # Parse
    wb = load_workbook(XLSX_PATH, read_only=True, data_only=True)
    try:
        if ABA_GERAL not in wb.sheetnames:
            print(f"[ERR] aba '{ABA_GERAL}' nao encontrada; abas={wb.sheetnames}")
            return 3
        ws = wb[ABA_GERAL]
        rows: list[tuple] = []
        vistos: set[str] = set()
        totais = {1: 0, 2: 0, 3: 0, 4: 0}
        it = ws.iter_rows(values_only=True)
        next(it, None)  # header
        for idx, row in enumerate(it, start=2):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            if len(row) < 10:
                continue
            lote = _norm_lote(row[COL_LOTE])
            if lote is None:
                continue
            cpf_raw = str(row[COL_CPF] or "")
            cpf = "".join(ch for ch in cpf_raw if ch.isdigit())
            if len(cpf) != 11:
                continue
            if cpf in vistos:
                continue
            vistos.add(cpf)
            totais[lote] = totais.get(lote, 0) + 1
            raw = {f"c{i}": (str(v) if v is not None else None) for i, v in enumerate(row)}
            rows.append((
                EMPRESA_ID, PER_APUR, cpf, None, None,
                lote, idx, json.dumps(raw),
            ))
    finally:
        wb.close()

    print(f"[info] parse: {totais} (total unicos={sum(totais.values())})")
    if totais.get(LOTE_ALVO, 0) == 0:
        print(f"[ERR] nenhuma linha com Lote {LOTE_ALVO} encontrada")
        return 4

    conn = psycopg2.connect(**DB_CONFIG)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            # Upsert s1210_xlsx
            cur.execute(
                """SELECT id FROM s1210_xlsx
                    WHERE empresa_id=%s AND per_apur=%s AND sha256=%s""",
                (EMPRESA_ID, PER_APUR, sha),
            )
            existing = cur.fetchone()
            if existing:
                xlsx_id = existing["id"]
                print(f"[info] s1210_xlsx ja existia id={xlsx_id}")
            else:
                cur.execute(
                    """INSERT INTO s1210_xlsx
                      (empresa_id, per_apur, nome_arquivo, tamanho_bytes,
                       sha256, storage_path, aba_geral, aba_operadoras,
                       parse_ok, totais_json)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,TRUE,%s)
                    RETURNING id""",
                    (EMPRESA_ID, PER_APUR, XLSX_PATH.name, len(data),
                     sha, f"local/{PER_APUR}/{sha[:12]}.xlsx",
                     ABA_GERAL, ABA_OPER,
                     json.dumps({f"{k}_LOTE": v for k, v in totais.items()})),
                )
                xlsx_id = cur.fetchone()["id"]
                print(f"[ok] s1210_xlsx inserido id={xlsx_id}")

            # Limpa SO o lote alvo (preserva L1/L2/L4 ja existentes)
            cur.execute(
                "DELETE FROM s1210_cpf_scope WHERE empresa_id=%s AND per_apur=%s AND lote_num=%s",
                (EMPRESA_ID, PER_APUR, LOTE_ALVO),
            )
            print(f"[info] scope anterior de {PER_APUR} lote={LOTE_ALVO} deletado ({cur.rowcount} linhas)")

            # Insere linhas
            psycopg2.extras.execute_values(
                cur,
                """INSERT INTO s1210_cpf_scope
                    (xlsx_id, empresa_id, per_apur, cpf, nome, matricula,
                     lote_num, row_number, raw_row)
                   VALUES %s""",
                [(xlsx_id, *r) for r in rows],
                template="(%s,%s,%s,%s,%s,%s,%s,%s,%s::jsonb)",
            )
            print(f"[ok] {len(rows)} linhas inseridas em s1210_cpf_scope")
        conn.commit()
    finally:
        conn.close()

    # Confirmacao
    conn = psycopg2.connect(**DB_CONFIG)
    try:
        with conn.cursor() as cur:
            cur.execute(
                """SELECT lote_num, COUNT(*) FROM s1210_cpf_scope
                    WHERE empresa_id=%s AND per_apur=%s
                    GROUP BY lote_num ORDER BY lote_num""",
                (EMPRESA_ID, PER_APUR),
            )
            print("[check]", cur.fetchall())
    finally:
        conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
