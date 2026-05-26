from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2.extras
from lxml import etree


ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
BACKEND_V2 = Path(r"C:\Users\xandao\Documents\GitHub\Easy-eSocial-v2\backend")
if str(ROOT / "python-scripts") not in sys.path:
    sys.path.insert(0, str(ROOT / "python-scripts"))
if str(BACKEND_V2) not in sys.path:
    sys.path.insert(0, str(BACKEND_V2))

import corrigir_mes_respostas_jaque_plano_pensao as correcao_base  # noqa: E402
import enviar_fechamento_agosto_solucoes_s1299 as fechamento_base  # noqa: E402
from app import db, tenant  # noqa: E402
from app.xml_diff import eventos_iguais  # noqa: E402
from app.xml_extractor import extrair_s1210  # noqa: E402
from app.xml_s1210 import NS as S1210_NS, S1210XMLGenerator  # noqa: E402


EMPRESA_ID = 2
CNPJ = "09445502000109"
MONTHS = ["2025-02", "2025-03", "2025-04", "2025-05", "2025-06", "2025-07", "2025-09", "2025-10"]
REPORT_MONTHS = ["2025-11", "2025-12"]
CONFIRM_TOKEN = "EXECUTAR_20_05_SOLUCOES"
INPUT_DIR = Path(r"C:\Users\xandao\Downloads\resposta final")
OUT_BASE = ROOT / "relatorio_ana" / "OPERACAO_20_05_SOLUCOES"
INPUTS_CSV = OUT_BASE / "inputs_jaque_resposta_final.csv"
EMPTY_MISSING_CSV = OUT_BASE / "faltantes_vazio.csv"
INPUTS_JSON = OUT_BASE / "inputs_jaque_resposta_final.json"
FINAL_REPORT = OUT_BASE / "resultado_operacao_20_05.json"
SUMMARY_MD = OUT_BASE / "RESULTADO_OPERACAO_20_05.md"


def text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def digits(value: Any) -> str:
    return re.sub(r"\D", "", text(value))


def cpf11(value: Any) -> str:
    only_digits = digits(value)
    return only_digits.zfill(11)[-11:] if only_digits else ""


def cnpj14(value: Any) -> str:
    only_digits = digits(value)
    return only_digits.zfill(14)[-14:] if only_digits else ""


def money(value: Any) -> Decimal:
    if value is None or text(value) == "":
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, int):
        return Decimal(value).quantize(Decimal("0.01"))
    if isinstance(value, float):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    raw = text(value).replace("R$", "").replace(" ", "")
    raw = re.sub(r"[^0-9,.-]", "", raw)
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif "," in raw:
        raw = raw.replace(",", ".")
    try:
        return Decimal(raw).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


def money_str(value: Any) -> str:
    return f"{money(value):.2f}"


def normalize_tp_rend(value: Any) -> str:
    raw = text(value).lower()
    raw_ascii = raw.translate(str.maketrans("áàãâéêíóôõúüç", "aaaaeeiooouuc"))
    only_digits = digits(value)
    if only_digits in {"11", "12", "13", "14", "18"}:
        return only_digits
    if "mensal" in raw_ascii:
        return "11"
    if "ferias" in raw_ascii:
        return "13"
    if "13" in raw_ascii or "decimo" in raw_ascii:
        return "12"
    return raw_ascii


def qname(tag: str) -> str:
    return f"{{{S1210_NS}}}{tag}"


def direct_children(parent: etree._Element, tag: str) -> list[etree._Element]:
    return parent.xpath(f'./*[local-name()="{tag}"]')


def sub(parent: etree._Element, tag: str, value: str | None = None) -> etree._Element:
    child = etree.SubElement(parent, qname(tag))
    if value is not None:
        child.text = value
    return child


def sheet_kind(name: str) -> str:
    lowered = name.lower()
    folded = lowered.translate(str.maketrans("áàãâéêíóôõúüç", "aaaaeeiooouuc"))
    if "plano" in folded:
        return "PLANO"
    if "pensao" in folded:
        return "PENSAO"
    if "dependente" in folded:
        return "DEPENDENTE"
    return ""


def month_file(per_apur: str) -> Path:
    return INPUT_DIR / f"{per_apur}_relatorio_final_jaque.xlsx"


def parse_workbooks() -> dict[str, Any]:
    OUT_BASE.mkdir(parents=True, exist_ok=True)
    parsed: dict[str, Any] = {month: {"plano": [], "pensao": [], "dependente": [], "invalid": []} for month in MONTHS}
    for per_apur in MONTHS:
        path = month_file(per_apur)
        if not path.exists():
            parsed[per_apur]["invalid"].append({"reason": f"arquivo nao encontrado: {path}"})
            continue
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        for sheet_name in workbook.sheetnames:
            kind = sheet_kind(sheet_name)
            if not kind:
                continue
            sheet = workbook[sheet_name]
            for row_number, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                cpf = cpf11(row[0] if len(row) > 0 else None)
                if not cpf or cpf == "00000000000":
                    continue
                if kind == "PLANO":
                    cpf_norm = cpf11(row[1] if len(row) > 1 else cpf) or cpf
                    cnpj_oper = cnpj14(row[3] if len(row) > 3 else None)
                    reg_ans = digits(row[4] if len(row) > 4 else None)
                    valor = money(row[5] if len(row) > 5 else None)
                    item = {
                        "per_apur": per_apur,
                        "cpf": cpf_norm,
                        "tipo": "PLANO",
                        "source_file": str(path),
                        "sheet": sheet_name,
                        "row_number": row_number,
                        "cnpj_operadora": cnpj_oper,
                        "registro_ans": reg_ans,
                        "valor_titular": money_str(valor),
                    }
                    if len(cnpj_oper) != 14 or not reg_ans or valor <= 0:
                        item["reason"] = "dados incompletos de plano de saude"
                        parsed[per_apur]["invalid"].append(item)
                    else:
                        parsed[per_apur]["plano"].append(item)
                elif kind == "PENSAO":
                    cpf_norm = cpf11(row[1] if len(row) > 1 else cpf) or cpf
                    cpf_dep = cpf11(row[3] if len(row) > 3 else None)
                    tp_rend = normalize_tp_rend(row[4] if len(row) > 4 else None)
                    percentual = text(row[5] if len(row) > 5 else None)
                    valor = money(row[6] if len(row) > 6 else None)
                    item = {
                        "per_apur": per_apur,
                        "cpf": cpf_norm,
                        "tipo": "PENSAO",
                        "source_file": str(path),
                        "sheet": sheet_name,
                        "row_number": row_number,
                        "cpf_beneficiario": cpf_dep,
                        "tipo_rendimento": tp_rend,
                        "percentual": percentual,
                        "valor_deduzido": money_str(valor),
                    }
                    if len(cpf_dep) != 11 or tp_rend not in {"11", "12", "13", "14", "18"} or valor <= 0:
                        item["reason"] = "dados incompletos de pensao alimenticia"
                        parsed[per_apur]["invalid"].append(item)
                    else:
                        parsed[per_apur]["pensao"].append(item)
                elif kind == "DEPENDENTE":
                    cpf_dep = cpf11(row[2] if len(row) > 2 else None)
                    item = {
                        "per_apur": per_apur,
                        "cpf": cpf,
                        "tipo": "DEPENDENTE",
                        "source_file": str(path),
                        "sheet": sheet_name,
                        "row_number": row_number,
                        "cpf_dependente": cpf_dep,
                        "erro_original": text(row[4] if len(row) > 4 else None),
                    }
                    if len(cpf_dep) != 11:
                        item["reason"] = "cpf dependente ausente/invalido"
                        parsed[per_apur]["invalid"].append(item)
                    else:
                        parsed[per_apur]["dependente"].append(item)
    INPUTS_JSON.write_text(json.dumps(parsed, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return parsed


def configure_base() -> None:
    EMPTY_MISSING_CSV.parent.mkdir(parents=True, exist_ok=True)
    if not EMPTY_MISSING_CSV.exists():
        EMPTY_MISSING_CSV.write_text("per_apur,cpf,tipo,motivo\n", encoding="utf-8")
    correcao_base.VALID_RESPONSES_CSV = INPUTS_CSV
    correcao_base.MISSING_CSV = EMPTY_MISSING_CSV
    correcao_base.OUT_BASE = OUT_BASE / "S1210_CORRECOES"
    correcao_base.POLL_TENTATIVAS = 12
    correcao_base.POLL_INTERVALO_S = 8


def final_error_rows(months: list[str]) -> dict[str, list[dict[str, Any]]]:
    internal_empresa_id = tenant.internal_empresa_id(EMPRESA_ID)
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(
                """
                WITH scoped AS (
                    SELECT tm.per_apur, it.cpf, it.status, it.erro_codigo, it.erro_mensagem,
                           it.nr_recibo_anterior, it.nr_recibo_novo,
                           it.id AS item_id, te.id AS envio_id, it.criado_em
                      FROM timeline_envio_item it
                      JOIN timeline_envio te ON te.id = it.timeline_envio_id
                      JOIN timeline_mes tm ON tm.id = te.timeline_mes_id
                     WHERE tm.empresa_id = %s
                       AND tm.per_apur = ANY(%s)
                       AND it.tipo_evento = 'S-1210'
                       AND it.cpf IS NOT NULL
                ), per_cpf AS (
                    SELECT per_apur, cpf,
                           BOOL_OR(status = 'sucesso') AS has_success
                      FROM scoped
                     GROUP BY per_apur, cpf
                ), ranked_errors AS (
                    SELECT scoped.*,
                           ROW_NUMBER() OVER (
                               PARTITION BY per_apur, cpf
                               ORDER BY criado_em DESC NULLS LAST, item_id DESC
                           ) AS rn
                      FROM scoped
                     WHERE status <> 'sucesso'
                        OR COALESCE(erro_codigo, '') NOT IN ('', '202')
                )
                SELECT r.per_apur, r.cpf, r.status, r.erro_codigo, r.erro_mensagem,
                       r.nr_recibo_anterior, r.nr_recibo_novo,
                       r.item_id, r.envio_id, r.criado_em
                  FROM per_cpf p
                  JOIN ranked_errors r
                    ON r.per_apur = p.per_apur
                   AND r.cpf = p.cpf
                   AND r.rn = 1
                 WHERE NOT p.has_success
                 ORDER BY r.per_apur, r.cpf
                """,
                (internal_empresa_id, months),
            )
            rows = [dict(row) for row in cursor.fetchall()]
    finally:
        conn.close()
    grouped: dict[str, list[dict[str, Any]]] = {month: [] for month in months}
    for row in rows:
        grouped.setdefault(row["per_apur"], []).append(row)
    return grouped


def write_operational_csv(parsed: dict[str, Any], errors_by_month: dict[str, list[dict[str, Any]]]) -> dict[str, Any]:
    rows: list[dict[str, str]] = []
    included = {month: {"plano": 0, "pensao": 0, "dependente": 0, "error_cpfs": len(errors_by_month.get(month, []))} for month in MONTHS}
    for per_apur in MONTHS:
        error_cpfs = {row["cpf"] for row in errors_by_month.get(per_apur, [])}
        for item in parsed[per_apur]["plano"]:
            if item["cpf"] not in error_cpfs:
                continue
            rows.append(
                {
                    "per_apur": per_apur,
                    "cpf": item["cpf"],
                    "tipo": "PLANO",
                    "source_file": item["source_file"],
                    "sheet": item["sheet"],
                    "row_number": str(item["row_number"]),
                    "data_json": json.dumps(
                        {
                            "cnpj_operadora": item["cnpj_operadora"],
                            "registro_ans": item["registro_ans"],
                            "valor_titular": item["valor_titular"],
                        },
                        ensure_ascii=False,
                    ),
                }
            )
            included[per_apur]["plano"] += 1
        for item in parsed[per_apur]["pensao"]:
            if item["cpf"] not in error_cpfs:
                continue
            rows.append(
                {
                    "per_apur": per_apur,
                    "cpf": item["cpf"],
                    "tipo": "PENSAO",
                    "source_file": item["source_file"],
                    "sheet": item["sheet"],
                    "row_number": str(item["row_number"]),
                    "data_json": json.dumps(
                        {
                            "beneficiarios": [
                                {
                                    "cpf_beneficiario": item["cpf_beneficiario"],
                                    "tipo_rendimento": item["tipo_rendimento"],
                                    "percentual": item["percentual"],
                                    "valor_deduzido": item["valor_deduzido"],
                                }
                            ]
                        },
                        ensure_ascii=False,
                    ),
                }
            )
            included[per_apur]["pensao"] += 1
        included[per_apur]["dependente"] = sum(1 for item in parsed[per_apur]["dependente"] if item["cpf"] in error_cpfs)
    with INPUTS_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        fieldnames = ["per_apur", "cpf", "tipo", "source_file", "sheet", "row_number", "data_json"]
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    return {"csv": str(INPUTS_CSV), "rows": len(rows), "included": included}


def dep_map_for_month(parsed: dict[str, Any], per_apur: str, error_cpfs: set[str]) -> dict[str, list[str]]:
    out: dict[str, list[str]] = defaultdict(list)
    for item in parsed[per_apur]["dependente"]:
        if item["cpf"] not in error_cpfs:
            continue
        cpf_dep = item["cpf_dependente"]
        if cpf_dep and cpf_dep not in out[item["cpf"]]:
            out[item["cpf"]].append(cpf_dep)
    return dict(out)


def pensao_dep_map_for_month(parsed: dict[str, Any], per_apur: str) -> dict[str, list[str]]:
    out: dict[str, list[str]] = defaultdict(list)
    for item in parsed[per_apur]["pensao"]:
        cpf_dep = item["cpf_beneficiario"]
        if cpf_dep and cpf_dep not in out[item["cpf"]]:
            out[item["cpf"]].append(cpf_dep)
    return dict(out)


def copy_info_dep_nodes(xml_old: bytes) -> list[etree._Element]:
    root = etree.fromstring(xml_old)
    nodes: list[etree._Element] = []
    for node in root.xpath('//*[local-name()="infoIRComplem"]/*[local-name()="infoDep"]'):
        nodes.append(etree.fromstring(etree.tostring(node)))
    return nodes


def insert_info_deps(xml_new: bytes, *, xml_old: bytes | None = None, extra_cpfs: list[str] | None = None) -> bytes:
    root = etree.fromstring(xml_new)
    for signature in root.xpath('.//*[local-name()="Signature"]'):
        parent = signature.getparent()
        if parent is not None:
            parent.remove(signature)
    info_ir_nodes = root.xpath('//*[local-name()="infoIRComplem"]')
    if not info_ir_nodes:
        ide_benef_nodes = root.xpath('//*[local-name()="ideBenef"]')
        if not ide_benef_nodes:
            raise RuntimeError("ideBenef ausente")
        info_ir = etree.SubElement(ide_benef_nodes[0], qname("infoIRComplem"))
    else:
        info_ir = info_ir_nodes[0]

    existing = {node.xpath('string(./*[local-name()="cpfDep"])').strip() for node in direct_children(info_ir, "infoDep")}
    to_insert: list[etree._Element] = []
    if xml_old:
        for node in copy_info_dep_nodes(xml_old):
            cpf_dep = node.xpath('string(./*[local-name()="cpfDep"])').strip()
            if cpf_dep and cpf_dep not in existing:
                to_insert.append(node)
                existing.add(cpf_dep)
    for cpf_dep in extra_cpfs or []:
        cpf_dep = cpf11(cpf_dep)
        if not cpf_dep or cpf_dep in existing:
            continue
        info_dep = etree.Element(qname("infoDep"))
        sub(info_dep, "cpfDep", cpf_dep)
        to_insert.append(info_dep)
        existing.add(cpf_dep)
    if not to_insert:
        return etree.tostring(root, xml_declaration=True, encoding="UTF-8")
    ircr_nodes = direct_children(info_ir, "infoIRCR")
    plan_nodes = direct_children(info_ir, "planSaude")
    anchor = ircr_nodes[0] if ircr_nodes else (plan_nodes[0] if plan_nodes else None)
    insert_at = info_ir.index(anchor) if anchor is not None else 0
    for offset, node in enumerate(to_insert):
        info_ir.insert(insert_at + offset, node)
    return etree.tostring(root, xml_declaration=True, encoding="UTF-8")


def postprocess_pensao_info_deps(manifest: dict[str, Any], parsed: dict[str, Any]) -> dict[str, Any]:
    per_apur = manifest["per_apur"]
    dep_map = pensao_dep_map_for_month(parsed, per_apur)
    targets = [item for item in manifest.get("targets") or [] if item.get("generated") and item.get("has_pensao")]
    if not targets:
        return {"updated": 0, "cpfs": []}
    current_rows = correcao_base.load_current_rows(per_apur, [item["cpf"] for item in targets])
    updated = []
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        for item in targets:
            xml_path = Path(item["xml"])
            row = current_rows.get(item["cpf"])
            xml_old = correcao_base.read_xml_event(conn, row) if row else None
            xml_new = insert_info_deps(xml_path.read_bytes(), xml_old=xml_old, extra_cpfs=dep_map.get(item["cpf"], []))
            xml_path.write_bytes(xml_new)
            item["info_dep_auto"] = dep_map.get(item["cpf"], [])
            updated.append(item["cpf"])
    finally:
        conn.close()
    correcao_base.manifest_path(per_apur).write_text(json.dumps(manifest, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return {"updated": len(updated), "cpfs": updated}


def generate_dependente_manifest(parsed: dict[str, Any], per_apur: str, error_cpfs: set[str]) -> dict[str, Any]:
    out_dir = OUT_BASE / "S1210_CORRECOES" / per_apur / "dependente_infodep"
    xml_dir = out_dir / "xml_unsigned"
    manifest_path = out_dir / f"manifest_dependente_infodep_{per_apur}.json"
    xml_dir.mkdir(parents=True, exist_ok=True)
    for old_xml in xml_dir.glob("*.xml"):
        old_xml.unlink()
    dep_map = dep_map_for_month(parsed, per_apur, error_cpfs)
    current_rows = correcao_base.load_current_rows(per_apur, sorted(dep_map))
    targets: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        for seq, cpf in enumerate(sorted(dep_map), start=1):
            row = current_rows.get(cpf)
            record: dict[str, Any] = {"cpf": cpf, "cpf_dependentes": dep_map[cpf], "generated": False}
            if not row:
                record["reason"] = "CPF sem S-1210 ativo local com XML"
                skipped.append(record)
                continue
            if row.get("item_status") == "sucesso":
                record["reason"] = "Ultimo status local ja e sucesso"
                skipped.append(record)
                continue
            try:
                xml_old = correcao_base.read_xml_event(conn, row)
                campos = extrair_s1210(xml_old)
                nr_recibo = row.get("nr_recibo") or row.get("nr_recibo_anterior") or campos.get("nr_recibo_atual")
                if not nr_recibo:
                    record["reason"] = "Sem nrRecibo ativo para retificar"
                    skipped.append(record)
                    continue
                xml_generated = S1210XMLGenerator.gerar(
                    empregador=campos["empregador"],
                    beneficiario=campos["beneficiario"],
                    info_pgtos=campos["info_pgtos"],
                    per_apur=per_apur,
                    ind_retif="2",
                    nr_recibo=nr_recibo,
                    info_ir_complem=campos.get("info_ir_complem"),
                    plan_saude=campos.get("plan_saude"),
                    seq=seq,
                    tp_amb=correcao_base.TP_AMB,
                )
                xml_new = insert_info_deps(xml_generated, xml_old=xml_old, extra_cpfs=dep_map[cpf])
                if eventos_iguais(xml_old, xml_new):
                    record["reason"] = "XML novo ficou identico ao atual"
                    skipped.append(record)
                    continue
                out_xml = xml_dir / f"S1210_{per_apur}_{cpf}_dependente_infodep_unsigned.xml"
                out_xml.write_bytes(xml_new)
                record.update(
                    {
                        "generated": True,
                        "xml": str(out_xml),
                        "evento_id": row.get("id"),
                        "nr_recibo": nr_recibo,
                        "status": row.get("item_status"),
                        "erro_codigo": row.get("erro_codigo"),
                        "erro_mensagem": row.get("erro_mensagem"),
                        "has_plano": False,
                        "has_pensao": False,
                        "has_dependente": True,
                    }
                )
                targets.append(record)
            except Exception as exc:
                record["reason"] = f"{type(exc).__name__}: {exc}"
                skipped.append(record)
    finally:
        conn.close()
    manifest = {
        "empresa_id": EMPRESA_ID,
        "per_apur": per_apur,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "out_dir": str(out_dir),
        "target_cpfs": len(dep_map),
        "xmls_generated": len(targets),
        "blocked_count": len(skipped),
        "xml_type_counts": {"dependente": len(targets)},
        "targets": targets + skipped,
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return {"manifest": manifest, "path": manifest_path}


def validate_infodep_manifest(manifest: dict[str, Any]) -> dict[str, Any]:
    rows = []
    for item in manifest.get("targets") or []:
        if not item.get("generated"):
            continue
        root = etree.fromstring(Path(item["xml"]).read_bytes())
        rows.append(
            {
                "cpf": item["cpf"],
                "perApur": root.xpath('string(//*[local-name()="perApur"])'),
                "indRetif": root.xpath('string(//*[local-name()="indRetif"])'),
                "infoDep": root.xpath('//*[local-name()="infoDep"]/*[local-name()="cpfDep"]/text()'),
                "signature": bool(root.xpath('//*[local-name()="Signature"]')),
            }
        )
    wrong = [row for row in rows if row["perApur"] != manifest["per_apur"] or row["indRetif"] != "2" or row["signature"]]
    return {"total": len(rows), "wrong": wrong, "sample": rows[:10]}


def execute_s1210_manifest(manifest: dict[str, Any], manifest_path: Path) -> dict[str, Any]:
    if not any(item.get("generated") for item in manifest.get("targets") or []):
        return {"ok": True, "skipped": True, "reason": "nenhum XML gerado"}
    if manifest.get("blocked_count"):
        return {"ok": False, "skipped": True, "reason": f"preflight bloqueado: {manifest['blocked_count']}"}
    fake_reabertura = {"persisted": {"codigo": "LOCAL_OPEN", "nr_recibo": None, "origem": "s1299_status_aberto"}}
    return correcao_base.execute_s1210(manifest, correcao_base.DEFAULT_CERT, correcao_base.read_password(), fake_reabertura)


def audit_counts(months: list[str]) -> dict[str, Any]:
    errors = final_error_rows(months)
    return {
        month: {
            "erros": len(rows),
            "codigos": dict(Counter(str(row.get("erro_codigo") or "") for row in rows)),
            "cpfs": [row["cpf"] for row in rows],
            "rows": rows,
        }
        for month, rows in errors.items()
    }


def configure_closing(per_apur: str) -> None:
    out_dir = OUT_BASE / "S1299_FECHAMENTOS" / per_apur
    fechamento_base.PER_APUR = per_apur
    fechamento_base.CONFIRM_TOKEN = f"FECHAR_{per_apur.replace('-', '_')}_SOLUCOES_S1299"
    fechamento_base.OUT_DIR = out_dir
    fechamento_base.XML_UNSIGNED = out_dir / f"S1299_{per_apur}_SOLUCOES_unsigned.xml"
    fechamento_base.XML_SIGNED = out_dir / f"S1299_{per_apur}_SOLUCOES_signed.xml"
    fechamento_base.MANIFEST = out_dir / f"manifest_fechamento_s1299_{per_apur}_solucoes.json"
    fechamento_base.POLL_TENTATIVAS = 30
    fechamento_base.POLL_INTERVALO_S = 8


def close_month(per_apur: str, execute: bool) -> dict[str, Any]:
    configure_closing(per_apur)
    if not execute:
        return fechamento_base.dry_run()
    return fechamento_base.execute()


def classify_error(row: dict[str, Any]) -> str:
    msg = text(row.get("erro_mensagem")).lower()
    folded = msg.translate(str.maketrans("áàãâéêíóôõúüç", "aaaaeeiooouuc"))
    if "plano" in folded or "plansaude" in folded or "saude coletivo" in folded:
        return "Jaque - plano de saude"
    if "pensao" in folded or "penalim" in folded or "vlrdedpenalim" in folded:
        return "Jaque - pensao alimenticia"
    if "depend" in folded or "cpfdep" in folded or "1861" in folded:
        return "Jaque - cpf dependente"
    return "Dev"


def generate_reports_nov_dec() -> dict[str, Any]:
    report_dir = OUT_BASE / "RELATORIOS_NOV_DEZ"
    report_dir.mkdir(parents=True, exist_ok=True)
    errors = final_error_rows(REPORT_MONTHS)
    all_rows = []
    for per_apur, rows in errors.items():
        for row in rows:
            cls = classify_error(row)
            all_rows.append({**row, "classe_operacional": cls, "responsavel": "Jaque" if cls.startswith("Jaque") else "Dev"})

    jaque_rows = [row for row in all_rows if row["responsavel"] == "Jaque"]
    dev_rows = [row for row in all_rows if row["responsavel"] == "Dev"]
    for filename, rows in [("relatorio_jaque_nov_dez.csv", jaque_rows), ("relatorio_dev_nov_dez.csv", dev_rows), ("relatorio_todos_nov_dez.csv", all_rows)]:
        with (report_dir / filename).open("w", encoding="utf-8-sig", newline="") as handle:
            fieldnames = [
                "per_apur",
                "cpf",
                "responsavel",
                "classe_operacional",
                "status",
                "erro_codigo",
                "erro_mensagem",
                "item_id",
                "envio_id",
            ]
            writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(rows)

    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "months": REPORT_MONTHS,
        "total": len(all_rows),
        "by_month": {month: len(errors.get(month, [])) for month in REPORT_MONTHS},
        "jaque": len(jaque_rows),
        "dev": len(dev_rows),
        "jaque_by_class": dict(Counter(row["classe_operacional"] for row in jaque_rows)),
        "dev_by_code": dict(Counter(str(row.get("erro_codigo") or "") for row in dev_rows)),
        "files": {
            "jaque_csv": str(report_dir / "relatorio_jaque_nov_dez.csv"),
            "dev_csv": str(report_dir / "relatorio_dev_nov_dez.csv"),
            "todos_csv": str(report_dir / "relatorio_todos_nov_dez.csv"),
        },
    }
    md_lines = [
        "# Relatorios Novembro/Dezembro - SOLUCOES",
        "",
        f"Gerado em: {summary['generated_at']}",
        "",
        "## Totais",
        "",
        f"- Total final auditado: {summary['total']}",
        f"- Novembro: {summary['by_month'].get('2025-11', 0)}",
        f"- Dezembro: {summary['by_month'].get('2025-12', 0)}",
        f"- Jaque: {summary['jaque']}",
        f"- Dev: {summary['dev']}",
        "",
        "## Jaque",
        "",
    ]
    for cls, total in summary["jaque_by_class"].items():
        md_lines.append(f"- {cls}: {total}")
    md_lines.extend(["", "## Dev", ""])
    for code, total in summary["dev_by_code"].items():
        md_lines.append(f"- Codigo {code or 'sem codigo'}: {total}")
    md_lines.extend(["", "## Arquivos", ""])
    for label, path in summary["files"].items():
        md_lines.append(f"- {label}: `{path}`")
    (report_dir / "RELATORIO_NOV_DEZ.md").write_text("\n".join(md_lines) + "\n", encoding="utf-8")
    summary["files"]["md"] = str(report_dir / "RELATORIO_NOV_DEZ.md")
    (report_dir / "summary_nov_dez.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    summary["files"]["json"] = str(report_dir / "summary_nov_dez.json")
    return summary


def run_month(per_apur: str, parsed: dict[str, Any], execute: bool, close_zeroed: bool) -> dict[str, Any]:
    configure_base()
    before = audit_counts([per_apur])[per_apur]
    error_cpfs = set(before["cpfs"])
    result: dict[str, Any] = {"per_apur": per_apur, "before": before, "plan_pensao": None, "dependente": None, "after": None, "closing": None}
    print(f"\n===== {per_apur}: erros antes={before['erros']} =====")
    if not error_cpfs:
        result["after"] = before
        if close_zeroed:
            result["closing"] = close_month(per_apur, execute)
        return result

    manifest = correcao_base.generate_manifest(per_apur)
    info_dep_patch = postprocess_pensao_info_deps(manifest, parsed)
    validation = correcao_base.validate_manifest(manifest)
    result["plan_pensao"] = {
        "manifest": str(correcao_base.manifest_path(per_apur)),
        "xmls_generated": manifest.get("xmls_generated"),
        "blocked_count": manifest.get("blocked_count"),
        "info_dep_patch": info_dep_patch,
        "validation": validation,
    }
    if validation.get("wrong"):
        result["plan_pensao"]["execute"] = {"ok": False, "reason": "validacao falhou"}
    elif manifest.get("xmls_generated"):
        if execute:
            try:
                result["plan_pensao"]["execute"] = execute_s1210_manifest(manifest, correcao_base.manifest_path(per_apur))
            except Exception as exc:
                result["plan_pensao"]["execute"] = {"ok": False, "exception": f"{type(exc).__name__}: {exc}"}
        else:
            result["plan_pensao"]["execute"] = {"ok": True, "dry_run": True}
    else:
        result["plan_pensao"]["execute"] = {"ok": True, "skipped": True, "reason": "sem plano/pensao"}

    if execute:
        after_plan = audit_counts([per_apur])[per_apur]
        error_cpfs = set(after_plan["cpfs"])
    dep_manifest_info = generate_dependente_manifest(parsed, per_apur, error_cpfs)
    dep_manifest = dep_manifest_info["manifest"]
    dep_validation = validate_infodep_manifest(dep_manifest)
    result["dependente"] = {
        "manifest": str(dep_manifest_info["path"]),
        "xmls_generated": dep_manifest.get("xmls_generated"),
        "blocked_count": dep_manifest.get("blocked_count"),
        "validation": dep_validation,
    }
    if dep_validation.get("wrong"):
        result["dependente"]["execute"] = {"ok": False, "reason": "validacao falhou"}
    elif dep_manifest.get("xmls_generated"):
        if execute:
            try:
                result["dependente"]["execute"] = execute_s1210_manifest(dep_manifest, dep_manifest_info["path"])
            except Exception as exc:
                result["dependente"]["execute"] = {"ok": False, "exception": f"{type(exc).__name__}: {exc}"}
        else:
            result["dependente"]["execute"] = {"ok": True, "dry_run": True}
    else:
        result["dependente"]["execute"] = {"ok": True, "skipped": True, "reason": "sem dependente"}

    after = audit_counts([per_apur])[per_apur] if execute else before
    result["after"] = after
    print(f"===== {per_apur}: erros depois={after['erros']} =====")
    if after["erros"] == 0 and close_zeroed:
        try:
            result["closing"] = close_month(per_apur, execute)
        except Exception as exc:
            result["closing"] = {"ok": False, "exception": f"{type(exc).__name__}: {exc}"}
    else:
        result["closing"] = {"skipped": True, "reason": "mes ainda tem erros" if after["erros"] else "fechamento nao solicitado"}
    return result


def write_summary(result: dict[str, Any]) -> None:
    lines = [
        "# Resultado Operacao 20/05 - SOLUCOES",
        "",
        f"Gerado em: {result.get('generated_at')}",
        "",
        "## Meses",
        "",
        "| Mes | Erros antes | Erros depois | Fechamento |",
        "|---|---:|---:|---|",
    ]
    for item in result.get("months", []):
        closing = item.get("closing") or {}
        if closing.get("ok"):
            close_status = f"OK {closing.get('nr_recibo') or ''}".strip()
        elif closing.get("skipped"):
            close_status = f"Pulou: {closing.get('reason')}"
        else:
            close_status = f"Erro: {closing.get('descricao') or closing.get('reason') or ''}".strip()
        lines.append(f"| {item['per_apur']} | {item['before']['erros']} | {item['after']['erros']} | {close_status} |")
    reports = result.get("reports_nov_dec") or {}
    lines.extend([
        "",
        "## Novembro/Dezembro",
        "",
        f"- Total: {reports.get('total')}",
        f"- Jaque: {reports.get('jaque')}",
        f"- Dev: {reports.get('dev')}",
        "",
    ])
    for label, path in (reports.get("files") or {}).items():
        lines.append(f"- {label}: `{path}`")
    SUMMARY_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def run_all(*, execute: bool, close_zeroed: bool, only_months: list[str] | None = None) -> dict[str, Any]:
    configure_base()
    parsed = parse_workbooks()
    initial_errors = final_error_rows(MONTHS)
    inputs = write_operational_csv(parsed, initial_errors)
    months = only_months or MONTHS
    results = []
    for per_apur in months:
        results.append(run_month(per_apur, parsed, execute=execute, close_zeroed=close_zeroed))
        FINAL_REPORT.write_text(
            json.dumps(
                {
                    "generated_at": datetime.now().isoformat(timespec="seconds"),
                    "execute": execute,
                    "close_zeroed": close_zeroed,
                    "inputs": inputs,
                    "months": results,
                },
                ensure_ascii=False,
                indent=2,
                default=str,
            ),
            encoding="utf-8",
        )
    reports = generate_reports_nov_dec()
    result = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "execute": execute,
        "close_zeroed": close_zeroed,
        "inputs": inputs,
        "months": results,
        "reports_nov_dec": reports,
    }
    FINAL_REPORT.write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    write_summary(result)
    return result


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--close-zeroed", action="store_true")
    parser.add_argument("--months", default=",".join(MONTHS))
    parser.add_argument("--reports-only", action="store_true")
    parser.add_argument("--confirmar", default="")
    args = parser.parse_args()
    if args.reports_only:
        print(json.dumps(generate_reports_nov_dec(), ensure_ascii=False, indent=2, default=str))
        return 0
    months = [month.strip() for month in args.months.split(",") if month.strip()]
    invalid_months = [month for month in months if month not in MONTHS]
    if invalid_months:
        raise SystemExit(f"Meses fora do escopo: {invalid_months}")
    if args.execute and args.confirmar != CONFIRM_TOKEN:
        raise SystemExit(f"Para executar, use --confirmar {CONFIRM_TOKEN}")
    result = run_all(execute=args.execute, close_zeroed=args.close_zeroed, only_months=months)
    print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())