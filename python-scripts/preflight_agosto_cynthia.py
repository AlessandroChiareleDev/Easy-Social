from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2.extras


BACKEND = Path(r"C:\Users\xandao\Documents\GitHub\Easy-eSocial-v2\backend")
if str(BACKEND) not in sys.path:
    sys.path.insert(0, str(BACKEND))

from app import db, tenant  # noqa: E402
from app.envio_teste_100 import _ler_xml_evento  # noqa: E402
from app.xml_extractor import extrair_s1210  # noqa: E402


EMPRESA_ID = 2
PER_APUR = "2025-08"
DOWNLOADS = Path.home() / "Downloads"
ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
OUT_DIR = ROOT / "relatorio_ana"

OPERADORAS_XLSX = DOWNLOADS / "OPERADORAS.xlsx"
PENSAO_XLSX = DOWNLOADS / "pensao.xlsx"
RECIBOS_XLSX = DOWNLOADS / "RELATORIO_SOLUCOES_AGOSTO_2025_ERROS_S1210 resposta erro 456.xlsx"


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def digits(value: Any) -> str:
    return re.sub(r"\D", "", norm(value))


def money_text(value: Any) -> str:
    if value is None or norm(value) == "":
        return ""
    if isinstance(value, (int, float)):
        return f"{value:.2f}"
    text = norm(value).replace("R$", "").strip()
    text = text.replace(".", "").replace(",", ".") if "," in text else text
    try:
        return f"{float(text):.2f}"
    except ValueError:
        return norm(value)


def is_receipt(value: Any) -> bool:
    return norm(value).startswith("1.1.")


def categorize(status: str | None, code: str | None, message: str | None) -> set[str]:
    out = {"nao_ok"}
    status_l = (status or "").lower()
    code_s = norm(code)
    msg = norm(message).lower()
    if code_s == "202":
        out.add("aviso_202")
    if "459:" in msg or "recibo de entrega informado" in msg:
        out.add("recibo_459")
    if "plano de saude coletivo" in msg or "plano de saúde coletivo" in msg:
        out.add("plano_saude_codigo8")
    if "beneficiarios da pensao" in msg or "beneficiários da pensão" in msg:
        out.add("pensao_codigo8")
    if "1089:" in msg:
        out.add("overlap_1089")
    if "543:" in msg or code_s == "SEM_MUDANCA" or status_l == "sem_mudanca":
        out.add("duplicidade_543_sem_mudanca")
    return out


def load_current_errors() -> list[dict[str, Any]]:
    internal_id = tenant.internal_empresa_id(EMPRESA_ID)
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                WITH scope AS (
                    SELECT DISTINCT ON (ev.cpf)
                           ev.id, ev.cpf, ev.nr_recibo, ev.id_evento,
                           ev.xml_oid, ev.xml_bytes, ev.xml_size_bytes,
                           ev.dt_processamento
                      FROM explorador_eventos ev
                      JOIN empresa_zips_brutos z ON z.id=ev.zip_id
                     WHERE z.empresa_id=%s
                       AND ev.tipo_evento='S-1210'
                       AND ev.per_apur=%s
                       AND ev.retificado_por_id IS NULL
                       AND ev.cpf IS NOT NULL
                       AND (ev.xml_oid IS NOT NULL OR ev.xml_bytes IS NOT NULL)
                     ORDER BY ev.cpf ASC, ev.dt_processamento DESC NULLS LAST, ev.id DESC
                ), ult AS (
                    SELECT DISTINCT ON (it.cpf)
                           it.cpf, it.status, it.erro_codigo, it.erro_mensagem,
                           it.nr_recibo_anterior, it.nr_recibo_novo,
                           it.criado_em, it.id AS item_id, te.id AS envio_id
                      FROM timeline_envio_item it
                      JOIN timeline_envio te ON te.id=it.timeline_envio_id
                      JOIN timeline_mes tm ON tm.id=te.timeline_mes_id
                     WHERE tm.empresa_id=%s
                       AND tm.per_apur=%s
                       AND it.tipo_evento='S-1210'
                       AND it.cpf IS NOT NULL
                     ORDER BY it.cpf, it.criado_em DESC NULLS LAST, it.id DESC
                )
                SELECT s.*, u.status AS item_status, u.erro_codigo, u.erro_mensagem,
                       u.nr_recibo_anterior, u.nr_recibo_novo,
                       u.criado_em AS ultimo_item_em, u.item_id, u.envio_id
                  FROM scope s
                  LEFT JOIN ult u ON u.cpf=s.cpf
                 WHERE u.status IS DISTINCT FROM 'sucesso'
                 ORDER BY s.cpf
                """,
                (internal_id, PER_APUR, internal_id, PER_APUR),
            )
            return [dict(row) for row in cur.fetchall()]
    finally:
        conn.close()


def error_sets(rows: list[dict[str, Any]]) -> dict[str, set[str]]:
    grouped: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        cpf = digits(row.get("cpf"))
        for cat in categorize(row.get("item_status"), row.get("erro_codigo"), row.get("erro_mensagem")):
            grouped[cat].add(cpf)
    return grouped


def load_receipt_overrides(path: Path) -> dict[str, dict[str, str]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["ERROS - CORRIGIDO"]
    rows = ws.iter_rows(values_only=True)
    headers = [norm(cell).upper() for cell in next(rows)]
    idx = {name: i for i, name in enumerate(headers)}
    out: dict[str, dict[str, str]] = {}
    for row in rows:
        cpf = digits(row[idx["CPF NORMALIZADO"]]) if "CPF NORMALIZADO" in idx else ""
        if len(cpf) != 11:
            continue
        used = norm(row[idx["RECIBO USADO"]]) if "RECIBO USADO" in idx else ""
        correct = norm(row[idx["RECIBO CORRETO"]]) if "RECIBO CORRETO" in idx else ""
        erro = norm(row[idx["ERRO"]]) if "ERRO" in idx else ""
        if is_receipt(correct):
            out[cpf] = {"recibo_usado": used, "recibo_correto": correct, "erro": erro}
    return out


def load_pensao(path: Path) -> dict[str, list[dict[str, str]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [norm(cell).upper() for cell in next(rows)]
    idx = {name: i for i, name in enumerate(headers)}
    out: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        cpf = digits(row[idx["CPF"]]) if "CPF" in idx else ""
        if len(cpf) != 11:
            continue
        out[cpf].append(
            {
                "vinculo": norm(row[idx.get("VÍNCULO", -1)]) if "VÍNCULO" in idx else "",
                "evento": norm(row[idx.get("EVENTO", -1)]) if "EVENTO" in idx else "",
                "nr_evento": norm(row[idx.get("NR. EVENTO", -1)]) if "NR. EVENTO" in idx else "",
                "percentual": norm(row[idx.get("VALOR DE REFERÊNCIA %", -1)]) if "VALOR DE REFERÊNCIA %" in idx else "",
                "valor": money_text(row[idx.get("VALOR", -1)]) if "VALOR" in idx else "",
                "data": norm(row[idx.get("DATA OCORRÊNCIA", -1)]) if "DATA OCORRÊNCIA" in idx else "",
            }
        )
    return dict(out)


def load_operadoras(path: Path) -> dict[str, Any]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    by_event: dict[str, dict[str, str]] = {}
    by_name: dict[str, dict[str, str]] = {}
    sheet_rows: dict[str, int] = {}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        headers = [norm(cell).upper() for cell in rows[0]]
        idx = {name: i for i, name in enumerate(headers)}
        sheet_rows[ws.title] = max(len(rows) - 1, 0)
        for row in rows[1:]:
            cnpj = digits(row[idx.get("CNPJ", -1)]) if "CNPJ" in idx else ""
            ans = digits(row[idx.get("ANS", -1)]) if "ANS" in idx else ""
            if len(cnpj) != 14 or not ans:
                continue
            name = norm(row[idx.get("OPERADORA", -1)]) if "OPERADORA" in idx else norm(row[idx.get("NOME EMPRESARIAL DA OPERADORA", -1)])
            event_code = norm(row[idx.get("COD. EVENTO", -1)]) if "COD. EVENTO" in idx else ""
            event_name = norm(row[idx.get("EVENTO", -1)]) if "EVENTO" in idx else ""
            item = {"cnpjOper": cnpj, "regANS": ans, "operadora": name, "evento": event_name}
            if event_code:
                by_event[event_code] = item
            if name:
                by_name[name.upper()] = item
    return {"by_event": by_event, "by_name": by_name, "sheet_rows": sheet_rows}


def check_s1210_operadoras() -> dict[str, Any]:
    internal_id = tenant.internal_empresa_id(EMPRESA_ID)
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            try:
                cur.execute(
                    """
                    SELECT empresa_id, COUNT(*) AS linhas, COUNT(DISTINCT cpf) AS cpfs
                      FROM s1210_operadoras
                     WHERE per_apur=%s AND empresa_id IN (%s, %s)
                     GROUP BY empresa_id
                     ORDER BY empresa_id
                    """,
                    (PER_APUR, internal_id, EMPRESA_ID),
                )
                rows = [dict(row) for row in cur.fetchall()]
            except Exception as exc:  # noqa: BLE001
                conn.rollback()
                return {"ok": False, "erro": f"{type(exc).__name__}: {exc}"}
    finally:
        conn.close()
    return {"ok": True, "rows": rows}


def count_ir_entries(info_ir: dict | None) -> dict[str, int]:
    out = {"infoIRCR": 0, "dedDepen": 0, "penAlim": 0}
    if not info_ir:
        return out
    for item in info_ir.get("infoIRCR") or []:
        out["infoIRCR"] += 1
        out["dedDepen"] += len(item.get("dedDepen") or [])
        out["penAlim"] += len(item.get("penAlim") or [])
    return out


def inspect_local_xml(rows_by_cpf: dict[str, dict[str, Any]], target_cpfs: set[str]) -> dict[str, Any]:
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        details = []
        counters = {"total": 0, "plan_saude_presente": 0, "penAlim_presente": 0, "dedDepen_presente": 0}
        for cpf in sorted(target_cpfs):
            event = rows_by_cpf.get(cpf)
            if not event:
                continue
            counters["total"] += 1
            try:
                xml = _ler_xml_evento(conn, event)
                campos = extrair_s1210(xml)
                ir_counts = count_ir_entries(campos.get("info_ir_complem"))
                has_plan = bool(campos.get("plan_saude"))
                if has_plan:
                    counters["plan_saude_presente"] += 1
                if ir_counts["penAlim"]:
                    counters["penAlim_presente"] += 1
                if ir_counts["dedDepen"]:
                    counters["dedDepen_presente"] += 1
                details.append(
                    {
                        "cpf": cpf,
                        "recibo_xml_local": event.get("nr_recibo"),
                        "plan_saude": has_plan,
                        **ir_counts,
                    }
                )
            except Exception as exc:  # noqa: BLE001
                details.append({"cpf": cpf, "erro_xml": f"{type(exc).__name__}: {exc}"})
    finally:
        conn.close()
    return {"counters": counters, "details": details}


def build_summary() -> dict[str, Any]:
    current_rows = load_current_errors()
    rows_by_cpf = {digits(row.get("cpf")): row for row in current_rows}
    sets = error_sets(current_rows)
    receipts = load_receipt_overrides(RECIBOS_XLSX)
    pensao = load_pensao(PENSAO_XLSX)
    operadoras = load_operadoras(OPERADORAS_XLSX)
    operadoras_db = check_s1210_operadoras()

    receipt_ready = sorted(sets["recibo_459"] & set(receipts))
    receipt_missing = sorted(sets["recibo_459"] - set(receipts))
    pensao_cover = sorted(sets["pensao_codigo8"] & set(pensao))
    pensao_missing = sorted(sets["pensao_codigo8"] - set(pensao))

    xml_plan = inspect_local_xml(rows_by_cpf, sets["plano_saude_codigo8"])
    xml_pensao = inspect_local_xml(rows_by_cpf, sets["pensao_codigo8"])
    xml_aviso = inspect_local_xml(rows_by_cpf, sets["aviso_202"])

    receipt_override_current = {
        cpf: receipts[cpf]["recibo_correto"]
        for cpf in receipt_ready
    }

    return {
        "empresa_id": EMPRESA_ID,
        "per_apur": PER_APUR,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "files": {
            "operadoras": str(OPERADORAS_XLSX),
            "pensao": str(PENSAO_XLSX),
            "recibos": str(RECIBOS_XLSX),
        },
        "current_counts": {key: len(value) for key, value in sorted(sets.items())},
        "receipt_overrides": {
            "rows_total": len(receipts),
            "current_459_total": len(sets["recibo_459"]),
            "ready": len(receipt_ready),
            "missing": receipt_missing,
            "sample": {cpf: receipts[cpf] for cpf in receipt_ready[:10]},
            "override_por_cpf": receipt_override_current,
        },
        "pensao": {
            "rows_cpfs_total": len(pensao),
            "current_pensao_total": len(sets["pensao_codigo8"]),
            "covered": len(pensao_cover),
            "missing": pensao_missing,
            "sample": {cpf: pensao[cpf] for cpf in pensao_cover[:10]},
            "xml_inspection": xml_pensao,
        },
        "plano_saude": {
            "current_plano_total": len(sets["plano_saude_codigo8"]),
            "operadoras_event_codes": len(operadoras["by_event"]),
            "operadoras_names": len(operadoras["by_name"]),
            "operadoras_sheets": operadoras["sheet_rows"],
            "s1210_operadoras_db": operadoras_db,
            "xml_inspection": xml_plan,
        },
        "aviso_202": {
            "total": len(sets["aviso_202"]),
            "xml_inspection": xml_aviso,
        },
        "other_groups": {
            "overlap_1089": len(sets["overlap_1089"]),
            "duplicidade_543_sem_mudanca": len(sets["duplicidade_543_sem_mudanca"]),
        },
    }


def write_outputs(summary: dict[str, Any]) -> dict[str, str]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_path = OUT_DIR / f"PREFLIGHT_AGOSTO_CYNTHIA_{stamp}.json"
    receipt_path = OUT_DIR / f"AGOSTO_RECIBO_OVERRIDE_CYNTHIA_{stamp}.json"
    md_path = OUT_DIR / f"PREFLIGHT_AGOSTO_CYNTHIA_{stamp}.md"

    receipt_override = summary["receipt_overrides"]["override_por_cpf"]
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    receipt_path.write_text(json.dumps(receipt_override, ensure_ascii=False, indent=2), encoding="utf-8")

    counts = summary["current_counts"]
    plan = summary["plano_saude"]
    pensao = summary["pensao"]
    recibos = summary["receipt_overrides"]
    aviso = summary["aviso_202"]
    other = summary["other_groups"]

    lines = [
        "# Preflight Agosto/2025 - Cynthia",
        "",
        f"Gerado em: {summary['generated_at']}",
        "",
        "## Erros atuais locais",
        f"- Nao OK total: {counts.get('nao_ok', 0)}",
        f"- Recibo 459: {counts.get('recibo_459', 0)}",
        f"- Plano saude codigo 8: {counts.get('plano_saude_codigo8', 0)}",
        f"- Pensao codigo 8: {counts.get('pensao_codigo8', 0)}",
        f"- Aviso 202: {counts.get('aviso_202', 0)}",
        f"- Overlap 1089: {other['overlap_1089']}",
        f"- Duplicidade/sem mudanca 543: {other['duplicidade_543_sem_mudanca']}",
        "",
        "## Recibos 459",
        f"- Planilha tem {recibos['rows_total']} CPFs com recibo correto.",
        f"- Cobre {recibos['ready']} de {recibos['current_459_total']} CPFs que ainda estao em 459.",
        f"- Faltantes: {len(recibos['missing'])}.",
        f"- Arquivo JSON pronto para override: {receipt_path.name}",
        "- Conclusao: grupo pronto para dry-run/reenvio controlado com recibo_override_por_cpf, sem consultar eSocial.",
        "",
        "## Plano de saude codigo 8",
        f"- Erros atuais: {plan['current_plano_total']} CPFs.",
        f"- OPERADORAS.xlsx traz {plan['operadoras_event_codes']} codigos de evento e {plan['operadoras_names']} operadoras com CNPJ/ANS.",
        f"- s1210_operadoras no banco para Agosto: {plan['s1210_operadoras_db']}.",
        f"- XML local com planSaude presente: {plan['xml_inspection']['counters']['plan_saude_presente']} de {plan['xml_inspection']['counters']['total']}.",
        "- Conclusao: util como mapa CNPJ/ANS, mas ainda falta CPF+valor por trabalhador para montar plan_saude_por_cpf.",
        "",
        "## Pensao codigo 8",
        f"- Erros atuais: {pensao['current_pensao_total']} CPFs.",
        f"- pensao.xlsx cobre {pensao['covered']} desses CPFs.",
        f"- XML local com penAlim presente: {pensao['xml_inspection']['counters']['penAlim_presente']} de {pensao['xml_inspection']['counters']['total']}.",
        "- Conclusao: arquivo ajuda com valor/percentual, mas falta CPF do beneficiario da pensao para gerar penAlim com seguranca.",
        "",
        "## Avisos 202",
        f"- Total: {aviso['total']}.",
        f"- XML local com dedDepen presente: {aviso['xml_inspection']['counters']['dedDepen_presente']} de {aviso['xml_inspection']['counters']['total']}.",
        "- Conclusao: ja foram aceitos com advertencia; nao sao o primeiro bloqueio para limpar Agosto.",
        "",
        "## Ordem segura sugerida",
        "1. Primeiro atacar 459 com recibo correto, em lote pequeno 1-10 antes de ampliar.",
        "2. Depois preparar plano de saude quando houver CPF+valor por trabalhador.",
        "3. Depois preparar pensao quando houver beneficiario/cpfDep por trabalhador.",
        "4. Deixar 1089/543 para reavaliacao depois das correcoes principais.",
        "",
        "Nenhum envio eSocial, Download ou ConsultarIdentificadores foi executado neste preflight.",
    ]
    md_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    return {"summary_json": str(json_path), "receipt_override_json": str(receipt_path), "markdown": str(md_path)}


def main() -> int:
    summary = build_summary()
    outputs = write_outputs(summary)
    print("PREFLIGHT_AGOSTO_CYNTHIA_OK")
    print(f"nao_ok={summary['current_counts'].get('nao_ok', 0)}")
    print(
        "recibos_459_ready="
        f"{summary['receipt_overrides']['ready']}/{summary['receipt_overrides']['current_459_total']}"
    )
    print(
        "plano_saude_xml_com_planSaude="
        f"{summary['plano_saude']['xml_inspection']['counters']['plan_saude_presente']}/"
        f"{summary['plano_saude']['xml_inspection']['counters']['total']}"
    )
    print(
        "pensao_cobertura="
        f"{summary['pensao']['covered']}/{summary['pensao']['current_pensao_total']} "
        f"penAlim_xml={summary['pensao']['xml_inspection']['counters']['penAlim_presente']}/"
        f"{summary['pensao']['xml_inspection']['counters']['total']}"
    )
    print(f"markdown={outputs['markdown']}")
    print(f"receipt_override_json={outputs['receipt_override_json']}")
    print(f"summary_json={outputs['summary_json']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())