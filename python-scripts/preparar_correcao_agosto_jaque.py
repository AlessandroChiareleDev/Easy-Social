from __future__ import annotations

import copy
import csv
import json
import re
import sys
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import openpyxl
import psycopg2.extras


BACKEND_V2 = Path(r"C:\Users\xandao\Documents\GitHub\Easy-eSocial-v2\backend")
if str(BACKEND_V2) not in sys.path:
    sys.path.insert(0, str(BACKEND_V2))

from app import db, storage, tenant  # noqa: E402
from app.envio_teste_100 import _ler_xml_evento  # noqa: E402
from app.xml_diff import eventos_iguais  # noqa: E402
from app.xml_extractor import extrair_s1210  # noqa: E402
from app.xml_s1210 import S1210XMLGenerator  # noqa: E402


EMPRESA_ID = 2
PER_APUR = "2025-08"
ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
JAQUE_XLSX = Path(r"C:\Users\xandao\Downloads\SOLUCOES_AGOSTO_2025_JAQUE_PLANO_PENSAO_PREENCHida pela jaque.xlsx")
MODELO_XLSX = ROOT / "relatorio_ana" / "SOLUCOES_AGOSTO_2025_JAQUE_PLANO_PENSAO_PREENCHER_CORRIGIDO.xlsx"
OUT_DIR = ROOT / "relatorio_ana" / "CORRECAO_AGOSTO_JAQUE"
XML_DIR = OUT_DIR / "xml_unsigned"
LOCAL_ZIP_ROOTS = [
    Path.home() / "Downloads" / "todos os meses 2025 SOLUCOES",
    Path.home() / "Downloads",
]


def text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value).strip())


def ascii_fold(value: Any) -> str:
    raw = text(value).lower()
    return "".join(ch for ch in unicodedata.normalize("NFKD", raw) if not unicodedata.combining(ch))


def digits(value: Any) -> str:
    return re.sub(r"\D", "", text(value))


def cpf11(value: Any) -> str:
    d = digits(value)
    return d.zfill(11) if d else ""


def cnpj14(value: Any) -> str:
    d = digits(value)
    return d.zfill(14) if d else ""


def money(value: Any) -> Decimal:
    if value is None or text(value) == "":
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, int):
        return Decimal(value).quantize(Decimal("0.01"))
    if isinstance(value, float):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    raw = text(value).replace("R$", "").strip()
    if "," in raw:
        raw = raw.replace(".", "").replace(",", ".")
    try:
        return Decimal(raw).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


def money_str(value: Decimal) -> str:
    return f"{value.quantize(Decimal('0.01')):.2f}"


def load_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Aba '{sheet_name}' nao encontrada em {path}")
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        headers = [text(cell) for cell in next(rows)]
        data: list[dict[str, Any]] = []
        for row_number, row in enumerate(rows, start=2):
            if not any(cell is not None and text(cell) != "" for cell in row):
                continue
            item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
            item["__row__"] = row_number
            data.append(item)
        return headers, data
    finally:
        wb.close()


def find_col(headers: list[str], *needles: str) -> str:
    folded = [(h, ascii_fold(h)) for h in headers]
    for needle in needles:
        n = ascii_fold(needle)
        for original, folded_header in folded:
            if n in folded_header:
                return original
    raise RuntimeError(f"Coluna nao encontrada. Procurado={needles}; headers={headers}")


def load_original_cpfs() -> dict[str, dict[str, str]]:
    out: dict[str, dict[str, str]] = {}
    for sheet in ("Plano de saude", "Pensao alimenticia"):
        headers, rows = load_sheet(MODELO_XLSX, sheet)
        cpf_col = find_col(headers, "CPF Normalizado", "CPF")
        nome_col = find_col(headers, "Nome Trabalhador", "Nome")
        for row in rows:
            cpf = cpf11(row.get(cpf_col))
            if len(cpf) == 11:
                out[cpf] = {"nome": text(row.get(nome_col)), "sheet": sheet}
    return out


def parse_plano() -> dict[str, Any]:
    headers, rows = load_sheet(JAQUE_XLSX, "PLANO SAUDE PREENCHIDO")
    cpf_col = find_col(headers, "CPF")
    nome_col = find_col(headers, "Vinculo", "Nome")
    cnpj_col = find_col(headers, "CNPJ Operadora", "CNPJ")
    ans_col = find_col(headers, "Registro ANS", "ANS")
    evento_col = find_col(headers, "Evento")
    valor_col = find_col(headers, "Valor")

    by_key: dict[tuple[str, str, str], Decimal] = defaultdict(lambda: Decimal("0.00"))
    names: dict[str, str] = {}
    invalid: list[dict[str, Any]] = []
    raw_counter: Counter[tuple[str, ...]] = Counter()
    rows_by_cpf: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for row in rows:
        cpf = cpf11(row.get(cpf_col))
        cnpj = cnpj14(row.get(cnpj_col))
        ans = digits(row.get(ans_col))
        valor = money(row.get(valor_col))
        nome = text(row.get(nome_col))
        evento = text(row.get(evento_col))
        raw_key = tuple(text(row.get(h)) for h in headers)
        raw_counter[raw_key] += 1

        if len(cpf) != 11 or len(cnpj) != 14 or not ans or valor <= 0:
            invalid.append({
                "row": row.get("__row__"),
                "cpf": cpf,
                "nome": nome,
                "cnpj": cnpj,
                "ans": ans,
                "valor": money_str(valor),
                "evento": evento,
            })
            continue

        names[cpf] = nome
        by_key[(cpf, cnpj, ans)] += valor
        rows_by_cpf[cpf].append({
            "row": row.get("__row__"),
            "nome": nome,
            "evento": evento,
            "cnpj": cnpj,
            "ans": ans,
            "valor": money_str(valor),
        })

    plan_by_cpf: dict[str, list[dict[str, str]]] = defaultdict(list)
    for (cpf, cnpj, ans), total in sorted(by_key.items()):
        plan_by_cpf[cpf].append({
            "cnpjOper": cnpj,
            "regANS": ans,
            "vlrSaudeTit": money_str(total),
        })

    exact_duplicate_rows = sum(count - 1 for count in raw_counter.values() if count > 1)
    return {
        "headers": headers,
        "row_count": len(rows),
        "cpf_count": len(plan_by_cpf),
        "entry_count": sum(len(items) for items in plan_by_cpf.values()),
        "exact_duplicate_rows": exact_duplicate_rows,
        "invalid": invalid,
        "plan_by_cpf": dict(plan_by_cpf),
        "names": names,
        "rows_by_cpf": dict(rows_by_cpf),
    }


def normalize_tp_rend(value: Any) -> tuple[str, str | None]:
    raw = ascii_fold(value)
    d = digits(value)
    if d in {"11", "12", "13", "14", "18"}:
        return d, None
    if "mensal" in raw or raw in {"mes", "m"}:
        return "11", None
    if "13" in raw or "decimo" in raw or "decimo terceiro" in raw:
        return "12", None
    return raw or "", f"Tipo Rendimento nao mapeado: {text(value)}"


def parse_pensao() -> dict[str, Any]:
    headers, rows = load_sheet(JAQUE_XLSX, "Pensao alimenticia Preenchido")
    cpf_col = find_col(headers, "CPF Normalizado", "CPF")
    nome_col = find_col(headers, "Nome Trabalhador", "Nome")
    pensao_by_cpf: dict[str, list[dict[str, str]]] = defaultdict(list)
    names: dict[str, str] = {}
    warnings: list[dict[str, Any]] = []

    for row in rows:
        cpf = cpf11(row.get(cpf_col))
        if len(cpf) != 11:
            continue
        names[cpf] = text(row.get(nome_col))
        for idx in range(1, 5):
            ben_col = f"CPF Beneficiario {idx}"
            tipo_col = f"Tipo Rendimento {idx}"
            valor_col = f"Valor Deduzido {idx}"
            if ben_col not in row:
                continue
            cpf_dep = cpf11(row.get(ben_col))
            if len(cpf_dep) != 11:
                continue
            tp_rend, warning = normalize_tp_rend(row.get(tipo_col))
            valor = money(row.get(valor_col))
            if warning:
                warnings.append({"row": row.get("__row__"), "cpf": cpf, "warning": warning})
            if not tp_rend or valor <= 0:
                warnings.append({
                    "row": row.get("__row__"),
                    "cpf": cpf,
                    "cpfDep": cpf_dep,
                    "warning": "Pensao sem tpRend ou valor positivo",
                })
                continue
            pensao_by_cpf[cpf].append({
                "tpRend": tp_rend,
                "cpfDep": cpf_dep,
                "vlrDedPenAlim": money_str(valor),
            })

    return {
        "headers": headers,
        "row_count": len(rows),
        "cpf_count": len(pensao_by_cpf),
        "entry_count": sum(len(items) for items in pensao_by_cpf.values()),
        "warnings": warnings,
        "pensao_by_cpf": dict(pensao_by_cpf),
        "names": names,
    }


def categorize(status: str | None, code: str | None, message: str | None) -> set[str]:
    out = {"nao_ok"}
    status_l = ascii_fold(status)
    code_s = text(code)
    msg = ascii_fold(message)
    if code_s == "202":
        out.add("aviso_202")
    if "459:" in msg or "recibo de entrega informado" in msg:
        out.add("recibo_459")
    if "plano de saude coletivo" in msg:
        out.add("plano_saude_codigo8")
    if "beneficiarios da pensao" in msg or "beneficiarios da pensao alimenticia" in msg:
        out.add("pensao_codigo8")
    if "1089:" in msg:
        out.add("overlap_1089")
    if "543:" in msg or code_s == "SEM_MUDANCA" or status_l == "sem_mudanca":
        out.add("duplicidade_543_sem_mudanca")
    return out


def load_current_errors() -> list[dict[str, Any]]:
    internal_id = tenant.internal_empresa_id(EMPRESA_ID)
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                WITH scope AS (
                    SELECT DISTINCT ON (ev.cpf)
                           ev.id, ev.cpf, ev.nr_recibo, ev.id_evento,
                           ev.xml_oid, ev.xml_bytes, ev.xml_size_bytes,
                           ev.xml_entry_name, ev.zip_id,
                           z.conteudo_oid AS zip_conteudo_oid,
                           z.tamanho_bytes AS zip_tamanho_bytes,
                           z.nome_arquivo_original AS zip_nome,
                           ev.dt_processamento
                      FROM explorador_eventos ev
                      JOIN empresa_zips_brutos z ON z.id=ev.zip_id
                     WHERE z.empresa_id=%s
                       AND ev.tipo_evento='S-1210'
                       AND ev.per_apur=%s
                       AND ev.retificado_por_id IS NULL
                       AND ev.cpf IS NOT NULL
                       AND (ev.xml_oid IS NOT NULL OR ev.xml_bytes IS NOT NULL)
                     ORDER BY ev.cpf ASC, ev.dt_processamento DESC NULLS LAST, ev.id DESC
                ), ult AS (
                    SELECT DISTINCT ON (it.cpf)
                           it.cpf, it.status, it.erro_codigo, it.erro_mensagem,
                           it.nr_recibo_anterior, it.nr_recibo_novo,
                           it.criado_em, it.id AS item_id, te.id AS envio_id
                      FROM timeline_envio_item it
                      JOIN timeline_envio te ON te.id=it.timeline_envio_id
                      JOIN timeline_mes tm ON tm.id=te.timeline_mes_id
                     WHERE tm.empresa_id=%s
                       AND tm.per_apur=%s
                       AND it.tipo_evento='S-1210'
                       AND it.cpf IS NOT NULL
                     ORDER BY it.cpf, it.criado_em DESC NULLS LAST, it.id DESC
                )
                SELECT s.*, u.status AS item_status, u.erro_codigo, u.erro_mensagem,
                       u.nr_recibo_anterior, u.nr_recibo_novo,
                       u.criado_em AS ultimo_item_em, u.item_id, u.envio_id
                  FROM scope s
                  LEFT JOIN ult u ON u.cpf=s.cpf
                 WHERE u.status IS DISTINCT FROM 'sucesso'
                 ORDER BY s.cpf
                """,
                (internal_id, PER_APUR, internal_id, PER_APUR),
            )
            return [dict(row) for row in cur.fetchall()]
    finally:
        conn.close()


def error_sets(rows: list[dict[str, Any]]) -> dict[str, set[str]]:
    grouped: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        cpf = cpf11(row.get("cpf"))
        for cat in categorize(row.get("item_status"), row.get("erro_codigo"), row.get("erro_mensagem")):
            grouped[cat].add(cpf)
    return grouped


def find_local_zip(zip_name: str | None) -> Path | None:
    if not zip_name:
        return None
    for root in LOCAL_ZIP_ROOTS:
        candidate = root / zip_name
        if candidate.exists():
            return candidate
    for root in LOCAL_ZIP_ROOTS:
        if not root.exists():
            continue
        for candidate in root.glob("*.zip"):
            if candidate.name == zip_name:
                return candidate
    return None


def safe_rollback(conn) -> None:
    try:
        conn.rollback()
    except Exception:
        pass


def read_xml_event(conn, row: dict[str, Any]) -> bytes:
    """Le XML do evento com fallback se o LO individual estiver ausente."""
    xml_bytes = row.get("xml_bytes")
    if xml_bytes is not None:
        return bytes(xml_bytes)

    entry = row.get("xml_entry_name")
    local_zip = find_local_zip(row.get("zip_nome"))
    if local_zip and entry:
        try:
            with zipfile.ZipFile(local_zip, mode="r") as zf:
                return zf.read(entry)
        except Exception:
            pass

    if row.get("xml_oid") is not None:
        try:
            return _ler_xml_evento(conn, row)
        except Exception:
            safe_rollback(conn)

    if entry and row.get("zip_conteudo_oid") and row.get("zip_tamanho_bytes"):
        try:
            reader = storage.LargeObjectReader(
                conn,
                int(row["zip_conteudo_oid"]),
                int(row["zip_tamanho_bytes"]),
            )
            try:
                with zipfile.ZipFile(reader, mode="r") as zf:
                    return zf.read(entry)
            finally:
                try:
                    reader.close()
                except Exception:
                    pass
        except Exception:
            safe_rollback(conn)

    raise RuntimeError(
        f"XML indisponivel para evento={row.get('id')} cpf={row.get('cpf')} "
        f"zip={row.get('zip_nome')} entry={entry}"
    )


def count_ir(info_ir: dict | None) -> dict[str, int]:
    out = {"infoIRCR": 0, "dedDepen": 0, "penAlim": 0}
    if not info_ir:
        return out
    for item in info_ir.get("infoIRCR") or []:
        out["infoIRCR"] += 1
        out["dedDepen"] += len(item.get("dedDepen") or [])
        out["penAlim"] += len(item.get("penAlim") or [])
    return out


def merge_pensao(info_ir: dict | None, pensoes: list[dict[str, str]]) -> tuple[dict | None, str | None]:
    if not pensoes:
        return info_ir, None
    if not info_ir or not info_ir.get("infoIRCR"):
        return None, "Sem infoIRCR no S-1210 atual; nao ha tpCR para inserir penAlim com seguranca"
    out = copy.deepcopy(info_ir)
    irs = out.get("infoIRCR") or []
    if len(irs) != 1:
        return None, f"S-1210 tem {len(irs)} infoIRCR; precisa decisao manual de qual tpCR recebe penAlim"
    irs[0]["penAlim"] = pensoes
    return out, None


def generate_artifacts() -> dict[str, Any]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    XML_DIR.mkdir(parents=True, exist_ok=True)

    original_cpfs = load_original_cpfs()
    plano = parse_plano()
    pensao = parse_pensao()
    current_rows = load_current_errors()
    rows_by_cpf = {cpf11(row.get("cpf")): row for row in current_rows}
    sets = error_sets(current_rows)

    plan_map: dict[str, list[dict[str, str]]] = plano["plan_by_cpf"]
    pensao_map: dict[str, list[dict[str, str]]] = pensao["pensao_by_cpf"]
    original_plano = {
        cpf for cpf, item in original_cpfs.items()
        if item.get("sheet") == "Plano de saude"
    }
    missing_plano = sorted(original_plano - set(plan_map))

    plano_error_cpfs = sets.get("plano_saude_codigo8", set())
    pensao_error_cpfs = sets.get("pensao_codigo8", set())
    corrigiveis_plano = sorted((plano_error_cpfs & set(plan_map)) - set(missing_plano))
    corrigiveis_pensao = sorted(pensao_error_cpfs & set(pensao_map))
    target_cpfs = sorted(set(corrigiveis_plano) | set(corrigiveis_pensao))

    generated: list[dict[str, Any]] = []
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        for seq, cpf in enumerate(target_cpfs, start=1):
            row = rows_by_cpf.get(cpf)
            record: dict[str, Any] = {
                "cpf": cpf,
                "has_plano": cpf in corrigiveis_plano,
                "has_pensao": cpf in corrigiveis_pensao,
                "generated": False,
                "xml": "",
            }
            if not row:
                record["reason"] = "CPF nao encontrado nos erros atuais do V2"
                generated.append(record)
                continue
            try:
                xml_old = read_xml_event(conn, row)
                campos = extrair_s1210(xml_old)
                ir_before = count_ir(campos.get("info_ir_complem"))
                record.update({
                    "evento_id": row.get("id"),
                    "item_id": row.get("item_id"),
                    "codigo": row.get("erro_codigo"),
                    "status": row.get("item_status"),
                    "nr_recibo": row.get("nr_recibo") or campos.get("nr_recibo_atual"),
                    "info_pgtos": len(campos.get("info_pgtos") or []),
                    "ir_before": ir_before,
                    "plan_entries": len(plan_map.get(cpf, [])),
                    "pensao_entries": len(pensao_map.get(cpf, [])),
                })
                info_ir = campos.get("info_ir_complem")
                if cpf in corrigiveis_pensao:
                    info_ir, warning = merge_pensao(info_ir, pensao_map[cpf])
                    if warning:
                        record["reason"] = warning
                        generated.append(record)
                        continue
                plan_saude = plan_map.get(cpf) if cpf in corrigiveis_plano else campos.get("plan_saude")
                nr_recibo = row.get("nr_recibo") or campos.get("nr_recibo_atual")
                if not nr_recibo:
                    record["reason"] = "Sem nrRecibo atual para retificar"
                    generated.append(record)
                    continue
                xml_new = S1210XMLGenerator.gerar(
                    empregador=campos["empregador"],
                    beneficiario=campos["beneficiario"],
                    info_pgtos=campos["info_pgtos"],
                    per_apur=campos["per_apur"],
                    ind_retif="2",
                    nr_recibo=nr_recibo,
                    info_ir_complem=info_ir,
                    plan_saude=plan_saude,
                    seq=seq,
                    tp_amb="1",
                )
                if eventos_iguais(xml_old, xml_new):
                    record["reason"] = "XML novo ficou identico ao atual"
                    generated.append(record)
                    continue
                out_xml = XML_DIR / f"S1210_{PER_APUR}_{cpf}_retif_unsigned.xml"
                out_xml.write_bytes(xml_new)
                record.update({
                    "generated": True,
                    "xml": str(out_xml),
                    "reason": "OK",
                    "ir_after": count_ir(info_ir),
                })
                generated.append(record)
            except Exception as exc:  # noqa: BLE001
                record["reason"] = f"{type(exc).__name__}: {exc}"
                generated.append(record)
    finally:
        conn.close()

    summary = {
        "empresa_id": EMPRESA_ID,
        "per_apur": PER_APUR,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "files": {
            "jaque": str(JAQUE_XLSX),
            "modelo": str(MODELO_XLSX),
            "out_dir": str(OUT_DIR),
        },
        "jaque": {
            "plano_rows": plano["row_count"],
            "plano_cpfs": plano["cpf_count"],
            "plano_entries": plano["entry_count"],
            "plano_exact_duplicate_rows": plano["exact_duplicate_rows"],
            "plano_invalid": plano["invalid"],
            "pensao_rows": pensao["row_count"],
            "pensao_cpfs": pensao["cpf_count"],
            "pensao_entries": pensao["entry_count"],
            "pensao_warnings": pensao["warnings"],
        },
        "current_error_counts": {key: len(value) for key, value in sorted(sets.items())},
        "coverage": {
            "modelo_plano_cpfs": len(original_plano),
            "modelo_plano_missing_in_jaque": [
                {"cpf": cpf, "nome": original_cpfs.get(cpf, {}).get("nome", "")}
                for cpf in missing_plano
            ],
            "plano_error_cpfs": len(plano_error_cpfs),
            "plano_corrigiveis": len(corrigiveis_plano),
            "pensao_error_cpfs": len(pensao_error_cpfs),
            "pensao_corrigiveis": len(corrigiveis_pensao),
            "target_cpfs": len(target_cpfs),
        },
        "generated": generated,
        "plan_saude_por_cpf": {cpf: plan_map[cpf] for cpf in corrigiveis_plano},
        "pensao_por_cpf": {cpf: pensao_map[cpf] for cpf in corrigiveis_pensao},
        "pendentes_depois": {
            "plano_4_sem_resposta_jaque": missing_plano,
            "aviso_202": sorted(sets.get("aviso_202", set())),
            "recibo_459": sorted(sets.get("recibo_459", set())),
            "outros_nao_ok": sorted(
                sets.get("nao_ok", set())
                - set(corrigiveis_plano)
                - set(corrigiveis_pensao)
                - set(missing_plano)
                - sets.get("aviso_202", set())
                - sets.get("recibo_459", set())
            ),
        },
    }

    json_path = OUT_DIR / "preflight_correcao_agosto_jaque.json"
    json_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")

    csv_path = OUT_DIR / "preflight_correcao_agosto_jaque.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        fieldnames = [
            "cpf", "has_plano", "has_pensao", "generated", "reason", "codigo", "status",
            "nr_recibo", "info_pgtos", "plan_entries", "pensao_entries", "xml",
        ]
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(generated)

    return summary


def main() -> int:
    if not JAQUE_XLSX.exists():
        raise SystemExit(f"Arquivo Jaque nao encontrado: {JAQUE_XLSX}")
    if not MODELO_XLSX.exists():
        raise SystemExit(f"Modelo nao encontrado: {MODELO_XLSX}")
    summary = generate_artifacts()
    total_generated = sum(1 for item in summary["generated"] if item.get("generated"))
    total_failed = len(summary["generated"]) - total_generated
    print("=== CORRECAO AGOSTO JAQUE - PREFLIGHT LOCAL ===")
    print(f"Arquivo Jaque: {JAQUE_XLSX}")
    print(f"Saida: {OUT_DIR}")
    print(f"Erros atuais: {summary['current_error_counts']}")
    print(f"Plano corrigiveis: {summary['coverage']['plano_corrigiveis']} de {summary['coverage']['plano_error_cpfs']}")
    print(f"Pensao corrigiveis: {summary['coverage']['pensao_corrigiveis']} de {summary['coverage']['pensao_error_cpfs']}")
    print(f"XMLs locais gerados: {total_generated}; bloqueados/falhas: {total_failed}")
    print("Faltantes Jaque plano:")
    for item in summary["coverage"]["modelo_plano_missing_in_jaque"]:
        print(f"  {item['cpf']} - {item['nome']}")
    blocked = [item for item in summary["generated"] if not item.get("generated")]
    if blocked:
        print("Bloqueios:")
        for item in blocked[:20]:
            print(f"  {item['cpf']} - {item.get('reason')}")
    print("Arquivos:")
    print(f"  {OUT_DIR / 'preflight_correcao_agosto_jaque.json'}")
    print(f"  {OUT_DIR / 'preflight_correcao_agosto_jaque.csv'}")
    print(f"  {XML_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())