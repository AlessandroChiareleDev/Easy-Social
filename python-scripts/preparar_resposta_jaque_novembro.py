from __future__ import annotations

import csv
import json
import re
import unicodedata
from collections import Counter
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
PER_APUR = "2025-11"
XLSX = Path(r"C:\Users\xandao\Downloads\resposta final\2025-11_relatorio_final_jaque.xlsx")
ERRORS_CSV = ROOT / "relatorio_ana" / "GISELE_SX_ERROS_S1210_NOVEMBRO_2025.csv"
AUDIT_DIR = ROOT / "relatorio_ana" / "AUDITORIA_RESPOSTAS_JAQUE_PLANO_PENSAO_2025"
VALID_CSV = AUDIT_DIR / "respostas_validas_final.csv"
MISSING_CSV = AUDIT_DIR / "faltantes_final.csv"


def norm_text(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = "".join(ch for ch in unicodedata.normalize("NFKD", text) if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text)


def only_digits(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\D", "", str(value))


def cpf11(value: Any) -> str:
    digits = only_digits(value)
    if not digits:
        return ""
    return digits[-11:].zfill(11)


def cnpj14(value: Any) -> str:
    digits = only_digits(value)
    if not digits:
        return ""
    return digits[-14:].zfill(14)


def parse_number(value: Any) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)).quantize(Decimal("0.01"))
        except InvalidOperation:
            return None
    text = str(value).strip().replace("R$", "").replace(" ", "")
    text = re.sub(r"[^0-9,.-]", "", text)
    if not text:
        return None
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        return Decimal(text).quantize(Decimal("0.01"))
    except InvalidOperation:
        return None


def header_map(ws, row_index: int = 3) -> dict[str, int]:
    return {
        norm_text(ws.cell(row_index, col).value): col
        for col in range(1, ws.max_column + 1)
        if norm_text(ws.cell(row_index, col).value)
    }


def cell_by_header(ws, row: int, headers: dict[str, int], header: str) -> Any:
    col = headers.get(norm_text(header))
    return ws.cell(row, col).value if col else None


def load_targets() -> dict[str, set[str]]:
    targets = {"PLANO": set(), "PENSAO": set(), "DEPENDENTE": set(), "RECIBO": set()}
    with ERRORS_CSV.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle, delimiter=";"):
            cpf = cpf11(row.get("cpf"))
            categoria = norm_text(row.get("categoria"))
            if "plano" in categoria:
                targets["PLANO"].add(cpf)
            elif "pensao" in categoria:
                targets["PENSAO"].add(cpf)
            elif "dependente" in categoria:
                targets["DEPENDENTE"].add(cpf)
            elif "recibo" in categoria:
                targets["RECIBO"].add(cpf)
    return targets


def parse_plan(ws) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    headers = header_map(ws)
    rows: list[dict[str, Any]] = []
    invalid: list[dict[str, Any]] = []
    for row in range(4, ws.max_row + 1):
        cpf = cpf11(cell_by_header(ws, row, headers, "CPF Normalizado") or cell_by_header(ws, row, headers, "CPF"))
        if not cpf:
            continue
        cnpj = cell_by_header(ws, row, headers, "CNPJ Operadora")
        ans = cell_by_header(ws, row, headers, "Registro ANS")
        valor = parse_number(cell_by_header(ws, row, headers, "Valor Titular Descontado em Folha"))
        missing = []
        if not cnpj14(cnpj):
            missing.append("cnpj_operadora")
        if not str(ans or "").strip():
            missing.append("registro_ans")
        if valor is None:
            missing.append("valor_titular")
        if missing:
            invalid.append({"cpf": cpf, "tipo": "PLANO", "motivo": ";".join(missing)})
            continue
        rows.append(
            {
                "per_apur": PER_APUR,
                "cpf": cpf,
                "tipo": "PLANO",
                "file": XLSX.name,
                "sheet": ws.title,
                "data_json": json.dumps(
                    {
                        "cnpj_operadora": cnpj,
                        "registro_ans": ans,
                        "valor_titular": str(valor),
                    },
                    ensure_ascii=False,
                ),
            }
        )
    return rows, invalid


def normalize_tp_rend(value: Any) -> str:
    raw = norm_text(value)
    digits = only_digits(value)
    if digits in {"11", "12", "13", "14", "18"}:
        return digits
    if "mensal" in raw or raw in {"mes", "m"}:
        return "mensal"
    if "13" in raw or "decimo" in raw:
        return "13"
    return str(value or "").strip()


def parse_pension(ws) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    headers = header_map(ws)
    rows: list[dict[str, Any]] = []
    invalid: list[dict[str, Any]] = []
    for row in range(4, ws.max_row + 1):
        cpf = cpf11(cell_by_header(ws, row, headers, "CPF Normalizado") or cell_by_header(ws, row, headers, "CPF"))
        if not cpf:
            continue
        cpf_benef = cpf11(cell_by_header(ws, row, headers, "CPF Beneficiário 1"))
        tipo_rendimento = normalize_tp_rend(cell_by_header(ws, row, headers, "Tipo Rendimento 1"))
        percentual = cell_by_header(ws, row, headers, "Percentual 1")
        valor = parse_number(cell_by_header(ws, row, headers, "Valor Deduzido 1"))
        if not cpf_benef or not tipo_rendimento or valor is None or valor == 0:
            invalid.append({"cpf": cpf, "tipo": "PENSAO", "motivo": "beneficiario_tipo_ou_valor"})
            continue
        rows.append(
            {
                "per_apur": PER_APUR,
                "cpf": cpf,
                "tipo": "PENSAO",
                "file": XLSX.name,
                "sheet": ws.title,
                "data_json": json.dumps(
                    {
                        "beneficiarios": [
                            {
                                "cpf_beneficiario": cpf_benef,
                                "tipo_rendimento": tipo_rendimento,
                                "percentual": str(percentual or ""),
                                "valor_deduzido": str(valor),
                                "valido": True,
                            }
                        ]
                    },
                    ensure_ascii=False,
                ),
            }
        )
    return rows, invalid


def parse_dependent(ws) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    headers = header_map(ws)
    valid: list[dict[str, Any]] = []
    invalid: list[dict[str, Any]] = []
    seen: set[tuple[str, str]] = set()
    for row in range(4, ws.max_row + 1):
        cpf = cpf11(cell_by_header(ws, row, headers, "CPF Trabalhador"))
        if not cpf:
            continue
        cpf_dep = cpf11(cell_by_header(ws, row, headers, "CPF Dependente"))
        key = (cpf, cpf_dep)
        if key in seen:
            continue
        seen.add(key)
        if cpf_dep:
            valid.append({"cpf": cpf, "cpf_dependente": cpf_dep})
        else:
            invalid.append({"cpf": cpf, "tipo": "DEPENDENTE", "motivo": "cpf_dependente_nao_informado"})
    return valid, invalid


def read_existing_csv(path: Path, fieldnames: list[str]) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    if not XLSX.exists():
        raise SystemExit(f"XLSX nao encontrado: {XLSX}")
    targets = load_targets()
    workbook = load_workbook(XLSX, data_only=True, read_only=True)
    try:
        plan_rows, plan_invalid = parse_plan(workbook["Plano de saude"])
        pension_rows, pension_invalid = parse_pension(workbook["Pensao alimenticia"])
        dependent_valid, dependent_invalid = parse_dependent(workbook["Dependente invalido"])
    finally:
        workbook.close()

    valid_rows = plan_rows + pension_rows
    valid_keys = {(row["tipo"], row["cpf"]) for row in valid_rows}
    missing_rows: list[dict[str, Any]] = []
    for tipo in ("PLANO", "PENSAO"):
        valid_cpfs = {cpf for kind, cpf in valid_keys if kind == tipo}
        for cpf in sorted(targets[tipo] - valid_cpfs):
            missing_rows.append({"per_apur": PER_APUR, "cpf": cpf, "tipo": tipo, "linhas_resposta": 0, "motivos_invalidos": "sem_resposta_valida", "erro_mensagem": ""})
    for item in dependent_invalid:
        missing_rows.append({"per_apur": PER_APUR, "cpf": item["cpf"], "tipo": item["tipo"], "linhas_resposta": 1, "motivos_invalidos": item["motivo"], "erro_mensagem": ""})
    for cpf in sorted(targets["RECIBO"]):
        missing_rows.append({"per_apur": PER_APUR, "cpf": cpf, "tipo": "RECIBO", "linhas_resposta": 0, "motivos_invalidos": "recibo_ativo_nao_coberto_pela_planilha_jaque", "erro_mensagem": ""})

    valid_fieldnames = ["per_apur", "cpf", "tipo", "file", "sheet", "data_json"]
    missing_fieldnames = ["per_apur", "cpf", "tipo", "linhas_resposta", "motivos_invalidos", "erro_mensagem"]
    previous_valid = [row for row in read_existing_csv(VALID_CSV, valid_fieldnames) if row.get("per_apur") != PER_APUR]
    previous_missing = [row for row in read_existing_csv(MISSING_CSV, missing_fieldnames) if row.get("per_apur") != PER_APUR]
    write_csv(VALID_CSV, previous_valid + valid_rows, valid_fieldnames)
    write_csv(MISSING_CSV, previous_missing + missing_rows, missing_fieldnames)

    summary = {
        "per_apur": PER_APUR,
        "xlsx": str(XLSX),
        "targets": {key: len(value) for key, value in targets.items()},
        "target_total": sum(len(value) for value in targets.values()),
        "valid_rows_written": len(valid_rows),
        "plan_valid": len(plan_rows),
        "pension_valid": len(pension_rows),
        "dependent_valid": len(dependent_valid),
        "missing_or_not_actionable": len(missing_rows),
        "missing_by_type": dict(Counter(row["tipo"] for row in missing_rows)),
        "invalid_rows": plan_invalid + pension_invalid + dependent_invalid,
        "plan_matches_target": {"missing_in_xlsx": sorted(targets["PLANO"] - {row["cpf"] for row in plan_rows}), "extra_in_xlsx": sorted({row["cpf"] for row in plan_rows} - targets["PLANO"])},
        "pension_matches_target": {"missing_in_xlsx": sorted(targets["PENSAO"] - {row["cpf"] for row in pension_rows}), "extra_in_xlsx": sorted({row["cpf"] for row in pension_rows} - targets["PENSAO"])},
    }
    out_json = AUDIT_DIR / f"auditoria_{PER_APUR}_resposta_jaque.json"
    out_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())