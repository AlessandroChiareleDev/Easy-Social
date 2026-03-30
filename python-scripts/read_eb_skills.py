"""
Import eb_skills_base_legal.xlsx into PostgreSQL.
- Skips column A (visual icons only, always empty in data)
- Uses columns B-K (10 columns) -> col_a through col_j
- Handles merged cells by carrying forward the master cell value
- Skips rows where column B (ID Rub.) is None (continuation/overflow rows)
"""
import openpyxl
import psycopg2

DB_CONFIG = {
    'host': 'localhost',
    'port': 5432,
    'dbname': 'easy_social_db',
    'user': 'easy_social_user',
    'password': 'sua_senha_segura'
}

TABLE_NAME = 'eb_skills_base_legal'

# Column mapping: Excel col -> DB col (file already clean, A-J directly)
# A=ID Rub -> col_a, B=Rubrica -> col_b, C=Cód.Natur -> col_c,
# D=Incid.INSS -> col_d, E=Incid.IRRF -> col_e, F=Incid.FGTS -> col_f,
# G=Análise -> col_g, H=Incid/BL INSS -> col_h, I=Incid/BL IRRF -> col_i,
# J=Incid/BL FGTS -> col_j
EXCEL_COLS = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]  # A through J
DB_COLS = ['col_a', 'col_b', 'col_c', 'col_d', 'col_e', 'col_f', 'col_g', 'col_h', 'col_i', 'col_j']


def get_cell_value(ws, row, col):
    """Get cell value, handling merged cells by finding the master cell."""
    cell = ws.cell(row=row, column=col)
    # Check if this cell is part of a merged range
    for merge_range in ws.merged_cells.ranges:
        if cell.coordinate in merge_range:
            # Return the value of the top-left cell of the merged range
            return ws.cell(row=merge_range.min_row, column=merge_range.min_col).value
    return cell.value


def main():
    print("Loading workbook...")
    wb = openpyxl.load_workbook('../eb_skills_base_legal.xlsx', data_only=True)
    ws = wb.active
    print(f"Sheet: {ws.title}, Rows: {ws.max_row}, Cols: {ws.max_column}")

    # Extract data rows (skip header row 1)
    # Only include rows where column B (ID Rub) has a value
    rows = []
    for row_num in range(2, ws.max_row + 1):
        id_rub = get_cell_value(ws, row_num, 1)  # Column A (ID Rub)
        if id_rub is None:
            continue  # Skip continuation/overflow rows
        
        row_data = []
        for excel_col in EXCEL_COLS:
            val = get_cell_value(ws, row_num, excel_col)
            row_data.append(str(val).strip() if val is not None else None)
        rows.append(row_data)

    print(f"Extracted {len(rows)} data rows (skipped overflow/empty rows)")
    
    # Show first 5 rows
    for i, r in enumerate(rows[:5]):
        print(f"  {r[0]:>5} | {r[1][:25] if r[1] else '':<25} | {r[2][:30] if r[2] else '':<30}")

    # Connect to DB and create table
    conn = psycopg2.connect(**DB_CONFIG)
    cur = conn.cursor()

    # Drop existing table
    cur.execute(f"DROP TABLE IF EXISTS {TABLE_NAME}")
    
    # Create table
    cols_sql = ", ".join([f"{col} TEXT" for col in DB_COLS])
    create_sql = f"""
        CREATE TABLE {TABLE_NAME} (
            id SERIAL PRIMARY KEY,
            upload_id INTEGER,
            row_number INTEGER,
            {cols_sql},
            raw_data JSONB,
            created_at TIMESTAMP DEFAULT NOW()
        )
    """
    cur.execute(create_sql)
    print(f"Created table {TABLE_NAME}")

    # Insert data
    placeholders = ", ".join(["%s"] * (len(DB_COLS) + 2))  # +2 for upload_id and row_number
    insert_cols = ", ".join(["upload_id", "row_number"] + DB_COLS)
    insert_sql = f"INSERT INTO {TABLE_NAME} ({insert_cols}) VALUES ({placeholders})"
    
    for i, row_data in enumerate(rows):
        cur.execute(insert_sql, [None, i + 1] + row_data)

    conn.commit()
    
    # Verify
    cur.execute(f"SELECT count(*) FROM {TABLE_NAME}")
    count = cur.fetchone()[0]
    print(f"Inserted {count} rows into {TABLE_NAME}")
    
    # Show sample
    cur.execute(f"SELECT col_a, col_b, col_c FROM {TABLE_NAME} LIMIT 5")
    for r in cur.fetchall():
        print(f"  {r[0]:>5} | {r[1][:25] if r[1] else '':<25} | {r[2][:40] if r[2] else '':<40}")

    cur.close()
    conn.close()
    print("Done!")


if __name__ == '__main__':
    main()
