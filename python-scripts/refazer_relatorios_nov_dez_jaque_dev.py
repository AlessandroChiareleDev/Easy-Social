from __future__ import annotations

import json
import re
import sys
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
BACKEND_V2 = Path(r"C:\Users\xandao\Documents\GitHub\Easy-eSocial-v2\backend")
if str(ROOT / "python-scripts") not in sys.path:
    sys.path.insert(0, str(ROOT / "python-scripts"))
if str(BACKEND_V2) not in sys.path:
    sys.path.insert(0, str(BACKEND_V2))

from app import db, tenant  # noqa: E402
from operacao_20_05_solucoes import (  # noqa: E402
    EMPRESA_ID,
    REPORT_MONTHS,
    classify_error,
    final_error_rows,
)


OUT_DIR = ROOT / "relatorio_ana" / "OPERACAO_20_05_SOLUCOES" / "RELATORIOS_NOV_DEZ"
PRIOR_REPORT_DIR = Path(r"C:\Users\xandao\Downloads\resposta final")
JAQUE_FILES = {
    "2025-11": OUT_DIR / "2025-11_relatorio_final_jaque.xlsx",
    "2025-12": OUT_DIR / "2025-12_relatorio_final_jaque.xlsx",
}
DEV_FILE = OUT_DIR / "2025-11_2025-12_relatorio_dev.xlsx"
SUMMARY_MD = OUT_DIR / "RELATORIO_NOV_DEZ.md"
SUMMARY_JSON = OUT_DIR / "summary_nov_dez_refeito.json"
LEGACY_SUMMARY_JSON = OUT_DIR / "summary_nov_dez.json"

MONTH_LABELS = {
    "2025-11": "Novembro/2025",
    "2025-12": "Dezembro/2025",
}
SUMMARY_SHEETS = {
    "2025-11": "Novembro",
    "2025-12": "Dezembro",
}

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)
SOFT_FILL = PatternFill("solid", fgColor="EAF3F8")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")
THIN = Side(style="thin", color="D9E2F3")
NOTE_SIDE = Side(style="thin", color="D6B656")
TABLE_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NOTE_BORDER = Border(left=NOTE_SIDE, right=NOTE_SIDE, top=NOTE_SIDE, bottom=NOTE_SIDE)


def only_digits(value: Any) -> str:
    return re.sub(r"\D", "", str(value or ""))


def cpf_digits(value: Any) -> str:
    digits = only_digits(value)
    return digits.zfill(11)[-11:] if digits else ""


def cpf_mask(value: Any) -> str:
    cpf = cpf_digits(value)
    if len(cpf) != 11:
        return str(value or "")
    return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def short_error(value: Any, limit: int = 220) -> str:
    message = clean_text(value)
    if len(message) <= limit:
        return message
    return message[: limit - 1].rstrip() + "..."


def extract_dependents(message: str, worker_cpf: str) -> list[str]:
    patterns = [
        r"CPF do dependente\s+(\d{11})\s+inv",
        r"cpfDep[^0-9]{0,30}(\d{11})",
        r"dependente[^0-9]{0,40}(\d{11})",
    ]
    found: list[str] = []
    for pattern in patterns:
        for value in re.findall(pattern, message or "", flags=re.IGNORECASE):
            cpf = cpf_digits(value)
            if cpf and cpf != worker_cpf and cpf not in found:
                found.append(cpf)
    if found:
        return found
    for value in re.findall(r"\b\d{11}\b", message or ""):
        cpf = cpf_digits(value)
        if cpf and cpf != worker_cpf and cpf not in found:
            found.append(cpf)
    return found


def add_title(ws, title: str, subtitle: str | None = None) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = title
    ws["A1"].font = Font(size=15, bold=True, color="1F4E78")
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = Font(size=10, italic=True, color="666666")


def table_name(base: str) -> str:
    safe = re.sub(r"[^A-Za-z0-9_]", "", base)
    if not safe or safe[0].isdigit():
        safe = f"Tabela{safe}"
    return safe[:240]


def write_table(ws, start_row: int, headers: list[str], rows: list[list[Any]], name: str) -> int:
    for col, header in enumerate(headers, 1):
        cell = ws.cell(start_row, col, header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = TABLE_BORDER
    for row_idx, row_values in enumerate(rows, start_row + 1):
        for col, cell_value in enumerate(row_values, 1):
            cell = ws.cell(row_idx, col, cell_value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = TABLE_BORDER
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F8FBFD")
    end_row = start_row + max(len(rows), 1)
    if rows:
        table = Table(displayName=table_name(name), ref=f"A{start_row}:{get_column_letter(len(headers))}{end_row}")
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    ws.freeze_panes = ws.cell(start_row + 1, 1)
    return end_row


def add_note(ws, note: str, start_col: int, start_row: int = 3, cols: int = 5, rows: int = 8) -> None:
    end_col = start_col + cols - 1
    end_row = start_row + rows - 1
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=end_col)
    title = ws.cell(start_row, start_col, "Por que não passou")
    title.fill = HEADER_FILL
    title.font = HEADER_FONT
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=start_row + 1, start_column=start_col, end_row=end_row, end_column=end_col)
    body = ws.cell(start_row + 1, start_col, note)
    body.fill = NOTE_FILL
    body.border = NOTE_BORDER
    body.alignment = Alignment(wrap_text=True, vertical="top")
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).border = NOTE_BORDER
    for col in range(start_col, end_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = 17
    for row in range(start_row + 1, end_row + 1):
        ws.row_dimensions[row].height = 26


def autosize(ws, widths: dict[int, int]) -> None:
    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths.get(col, 18)


def load_worker_scope(months: list[str]) -> dict[tuple[str, str], dict[str, Any]]:
    internal_empresa_id = tenant.internal_empresa_id(EMPRESA_ID)
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(
                """
                SELECT per_apur, cpf, nome, matricula
                  FROM solucoes.s1210_cpf_scope
                 WHERE empresa_id = %s
                   AND per_apur = ANY(%s)
                """,
                (internal_empresa_id, months),
            )
            result: dict[tuple[str, str], dict[str, Any]] = {}
            for row in cursor.fetchall():
                cpf = cpf_digits(row.get("cpf"))
                if cpf:
                    result[(str(row.get("per_apur")), cpf)] = dict(row)
            return result
    finally:
        conn.close()


def load_prior_report_names() -> dict[str, str]:
    if not PRIOR_REPORT_DIR.exists():
        return {}
    from openpyxl import load_workbook

    names: dict[str, str] = {}
    for path in sorted(PRIOR_REPORT_DIR.glob("*_relatorio_final_jaque.xlsx")):
        workbook = load_workbook(path, data_only=True, read_only=True)
        for sheet_name in workbook.sheetnames:
            if sheet_name not in {"Plano de saude", "Pensao alimenticia"}:
                continue
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=4, values_only=True):
                cpf = cpf_digits(row[1] if len(row) > 1 else row[0] if row else "")
                nome = clean_text(row[2] if len(row) > 2 else "")
                if len(cpf) == 11 and nome and cpf not in names:
                    names[cpf] = nome
    return names


def collect_rows() -> tuple[list[dict[str, Any]], dict[tuple[str, str], dict[str, Any]]]:
    errors_by_month = final_error_rows(REPORT_MONTHS)
    workers = load_worker_scope(REPORT_MONTHS)
    prior_names = load_prior_report_names()
    rows: list[dict[str, Any]] = []
    for per_apur in REPORT_MONTHS:
        for row in errors_by_month.get(per_apur, []):
            cpf = cpf_digits(row.get("cpf"))
            worker = workers.get((per_apur, cpf), {})
            classe = classify_error(row)
            rows.append(
                {
                    **dict(row),
                    "per_apur": per_apur,
                    "cpf": cpf,
                    "classe_operacional": classe,
                    "responsavel": "Jaque" if classe.startswith("Jaque") else "Dev",
                    "nome": clean_text(worker.get("nome")) or prior_names.get(cpf, ""),
                    "matricula": clean_text(worker.get("matricula")),
                }
            )
    return rows, workers


def split_jaque_rows(rows: list[dict[str, Any]], per_apur: str) -> dict[str, list[dict[str, Any]]]:
    selected = [row for row in rows if row["per_apur"] == per_apur and row["responsavel"] == "Jaque"]
    return {
        "plano": [row for row in selected if "plano" in row["classe_operacional"]],
        "pensao": [row for row in selected if "pensao" in row["classe_operacional"]],
        "dependente": [row for row in selected if "dependente" in row["classe_operacional"]],
    }


def build_summary(ws, per_apur: str, grouped: dict[str, list[dict[str, Any]]]) -> None:
    label = MONTH_LABELS[per_apur]
    add_title(ws, f"Relatório final Jaque - {label}", "Somente pendências que dependem de informação da Jaque.")
    ws.merge_cells("A4:F8")
    message = ws["A4"]
    total = sum(len(value) for value in grouped.values())
    message.value = (
        f"{label} ficou com {total} pendências finais do S-1210 que dependem de informação operacional. "
        "Este arquivo não traz erro de recibo nem detalhe técnico de envio: só pede o que precisa ser devolvido "
        "para corrigir plano de saúde, pensão alimentícia e CPF de dependente inválido."
    )
    message.fill = SOFT_FILL
    message.alignment = Alignment(wrap_text=True, vertical="top")
    message.border = TABLE_BORDER
    for row in range(4, 9):
        ws.row_dimensions[row].height = 26
        for col in range(1, 7):
            ws.cell(row, col).border = TABLE_BORDER
            ws.cell(row, col).fill = SOFT_FILL
    summary_rows = [
        ["Plano de saúde", len(grouped["plano"]), "Informar CNPJ operadora, Registro ANS e valor do titular descontado em folha."],
        ["Pensão alimentícia", len(grouped["pensao"]), "Confirmar beneficiário, tipo de rendimento, percentual e valor deduzido real."],
        ["CPF dependente inválido", len(grouped["dependente"]), "Corrigir ou confirmar o CPF do dependente indicado pelo eSocial."],
    ]
    write_table(ws, 11, ["Tipo", "Qtd. trabalhadores", "Próxima ação"], summary_rows, f"Resumo{per_apur}")
    ws["A17"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws["A17"].font = Font(italic=True, color="666666")
    autosize(ws, {1: 26, 2: 16, 3: 78})


def build_plan(ws, per_apur: str, rows: list[dict[str, Any]]) -> None:
    add_title(ws, f"Plano de saúde - {MONTH_LABELS[per_apur]}")
    headers = [
        "CPF",
        "CPF Normalizado",
        "Nome Trabalhador",
        "CNPJ Operadora",
        "Registro ANS",
        "Valor Titular Descontado em Folha",
        "O que falta corrigir",
    ]
    data_rows = [
        [
            cpf_mask(row["cpf"]),
            row["cpf"],
            row.get("nome", ""),
            "",
            "",
            "",
            "Informar CNPJ da operadora, Registro ANS e valor do titular descontado em folha.",
        ]
        for row in rows
    ]
    write_table(ws, 3, headers, data_rows, f"PlanoSaude{per_apur}")
    add_note(
        ws,
        "O eSocial recusou porque o grupo de plano de saúde coletivo não ficou completo/válido. Para estes trabalhadores, a devolução precisa trazer somente os três dados da operadora e o valor descontado do titular.",
        9,
    )
    autosize(ws, {1: 18, 2: 17, 3: 34, 4: 20, 5: 18, 6: 24, 7: 54})


def build_pension(ws, per_apur: str, rows: list[dict[str, Any]]) -> None:
    add_title(ws, f"Pensão alimentícia - {MONTH_LABELS[per_apur]}")
    headers = [
        "CPF",
        "CPF Normalizado",
        "Nome Trabalhador",
        "CPF Beneficiário 1",
        "Tipo Rendimento 1",
        "Percentual 1",
        "Valor Deduzido 1",
        "O que falta corrigir",
    ]
    data_rows = [
        [
            cpf_mask(row["cpf"]),
            row["cpf"],
            row.get("nome", ""),
            "",
            "",
            "",
            "",
            "Informar beneficiário, tipo de rendimento e valor deduzido real. Valor zerado não passa.",
        ]
        for row in rows
    ]
    write_table(ws, 3, headers, data_rows, f"PensaoAlimenticia{per_apur}")
    add_note(
        ws,
        "A pendência é de pensão alimentícia no S-1210. A correção depende de confirmar os dados do beneficiário e o valor efetivamente deduzido; quando o valor vem zero ou incompatível, o XML não passa.",
        10,
    )
    autosize(ws, {1: 18, 2: 17, 3: 34, 4: 20, 5: 18, 6: 14, 7: 16, 8: 54})


def build_dependent(ws, per_apur: str, rows: list[dict[str, Any]]) -> None:
    add_title(ws, f"CPF dependente inválido - {MONTH_LABELS[per_apur]}")
    headers = ["CPF Trabalhador", "CPF Dependente", "Erro eSocial", "Informação necessária"]
    data_rows: list[list[Any]] = []
    for row in rows:
        dependents = extract_dependents(str(row.get("erro_mensagem") or ""), row["cpf"]) or [""]
        for dependent in dependents:
            data_rows.append(
                [
                    cpf_mask(row["cpf"]),
                    cpf_mask(dependent) if dependent else "",
                    short_error(row.get("erro_mensagem"), 180),
                    "Informar o CPF correto do dependente para este trabalhador ou confirmar a retirada do dependente.",
                ]
            )
    write_table(ws, 3, headers, data_rows, f"DependenteInvalido{per_apur}")
    add_note(
        ws,
        "O eSocial recusou o CPF do dependente. A Jaque precisa validar o CPF correto do dependente apontado ou confirmar que ele deve sair do S-1210.",
        6,
    )
    autosize(ws, {1: 18, 2: 18, 3: 58, 4: 52})


def build_jaque_workbook(per_apur: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    grouped = split_jaque_rows(rows, per_apur)
    wb = Workbook()
    wb.remove(wb.active)
    build_summary(wb.create_sheet(SUMMARY_SHEETS[per_apur]), per_apur, grouped)
    build_plan(wb.create_sheet("Plano de saude"), per_apur, grouped["plano"])
    build_pension(wb.create_sheet("Pensao alimenticia"), per_apur, grouped["pensao"])
    build_dependent(wb.create_sheet("Dependente invalido"), per_apur, grouped["dependente"])
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    output = JAQUE_FILES[per_apur]
    wb.save(output)
    return {
        "arquivo": str(output),
        "abas": wb.sheetnames,
        "contagens": {key: len(value) for key, value in grouped.items()},
    }


def dev_problem(row: dict[str, Any]) -> tuple[str, str]:
    message = clean_text(row.get("erro_mensagem"))
    folded = message.lower().translate(str.maketrans("áàãâéêíóôõúüç", "aaaaeeiooouuc"))
    if "459" in folded or "recibo" in folded:
        return (
            "Recibo ativo do S-1210 não localizado",
            "Obter o recibo ativo correto no eSocial antes de retificar. Não é pendência de preenchimento da Jaque.",
        )
    if "106" in folded or "ja existe" in folded or "duplic" in folded:
        return (
            "Evento original já existe no eSocial",
            "Localizar o evento ativo e usar o recibo correto para retificação; envio original duplicado não resolve.",
        )
    return (
        "Erro técnico fora do escopo da Jaque",
        "Analisar XML/regra de envio e corrigir no pipeline antes de devolver para operação.",
    )


def build_dev_workbook(rows: list[dict[str, Any]]) -> dict[str, Any]:
    dev_rows = [row for row in rows if row["responsavel"] == "Dev"]
    wb = Workbook()
    wb.remove(wb.active)

    resumo = wb.create_sheet("Resumo")
    add_title(resumo, "Relatório dev - Novembro/Dezembro 2025", "Erros finais que não devem ir para a Jaque.")
    resumo_rows = []
    by_month = Counter(row["per_apur"] for row in dev_rows)
    for per_apur in REPORT_MONTHS:
        resumo_rows.append([MONTH_LABELS[per_apur], by_month.get(per_apur, 0), "Resolver tecnicamente; não pedir dado operacional para Jaque."])
    write_table(resumo, 4, ["Competência", "Qtd. erros", "Direção"], resumo_rows, "ResumoDevNovDez")
    resumo["A9"] = "Critério: tudo que não é plano de saúde, pensão alimentícia ou CPF dependente inválido ficou neste relatório."
    resumo["A9"].font = Font(italic=True, color="666666")
    autosize(resumo, {1: 20, 2: 14, 3: 76})

    detalhes = wb.create_sheet("Erros dev")
    add_title(detalhes, "Erros para desenvolvimento")
    headers = [
        "Competência",
        "CPF",
        "CPF Normalizado",
        "Nome Trabalhador",
        "Matrícula",
        "Problema",
        "O que precisa fazer",
        "Código eSocial",
        "Erro eSocial resumido",
        "Recibo local usado",
    ]
    detail_rows = []
    for row in dev_rows:
        problem, action = dev_problem(row)
        detail_rows.append(
            [
                MONTH_LABELS.get(row["per_apur"], row["per_apur"]),
                cpf_mask(row["cpf"]),
                row["cpf"],
                row.get("nome", ""),
                row.get("matricula", ""),
                problem,
                action,
                row.get("erro_codigo") or "",
                short_error(row.get("erro_mensagem"), 260),
                row.get("nr_recibo_anterior") or "",
            ]
        )
    write_table(detalhes, 3, headers, detail_rows, "ErrosDevNovDez")
    autosize(detalhes, {1: 18, 2: 18, 3: 17, 4: 34, 5: 16, 6: 34, 7: 64, 8: 14, 9: 74, 10: 28})

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(DEV_FILE)
    return {"arquivo": str(DEV_FILE), "abas": wb.sheetnames, "total": len(dev_rows)}


def write_summary(rows: list[dict[str, Any]], jaque_outputs: dict[str, Any], dev_output: dict[str, Any]) -> dict[str, Any]:
    by_month = Counter(row["per_apur"] for row in rows)
    by_owner = Counter(row["responsavel"] for row in rows)
    by_class = Counter(row["classe_operacional"] for row in rows)
    jaque_by_month: dict[str, dict[str, int]] = {}
    for per_apur in REPORT_MONTHS:
        grouped = split_jaque_rows(rows, per_apur)
        jaque_by_month[per_apur] = {key: len(value) for key, value in grouped.items()}

    summary = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "months": REPORT_MONTHS,
        "total": len(rows),
        "by_month": dict(by_month),
        "jaque": by_owner.get("Jaque", 0),
        "dev": by_owner.get("Dev", 0),
        "by_class": dict(by_class),
        "jaque_by_month": jaque_by_month,
        "files": {"jaque": jaque_outputs, "dev": dev_output},
    }

    lines = [
        "# Relatórios finais Novembro/Dezembro - SOLUÇÕES",
        "",
        f"Gerado em: {summary['generated_at']}",
        "",
        "## Totais finais auditados",
        "",
        f"- Total: {summary['total']}",
        f"- Novembro: {by_month.get('2025-11', 0)}",
        f"- Dezembro: {by_month.get('2025-12', 0)}",
        f"- Jaque: {summary['jaque']}",
        f"- Dev: {summary['dev']}",
        "",
        "## Relatórios para Jaque",
        "",
    ]
    for per_apur in REPORT_MONTHS:
        counts = jaque_by_month[per_apur]
        lines.extend(
            [
                f"### {MONTH_LABELS[per_apur]}",
                "",
                f"- Plano de saúde: {counts['plano']}",
                f"- Pensão alimentícia: {counts['pensao']}",
                f"- CPF dependente inválido: {counts['dependente']}",
                f"- Arquivo: {jaque_outputs[per_apur]['arquivo']}",
                "",
            ]
        )
    lines.extend(
        [
            "## Relatório para dev",
            "",
            f"- Arquivo: {dev_output['arquivo']}",
            "- Conteúdo: somente erros fora do escopo da Jaque, principalmente recibo ativo do S-1210 não localizado.",
            "",
        ]
    )
    SUMMARY_MD.write_text("\n".join(lines), encoding="utf-8")
    SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    LEGACY_SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return summary


def build_reports() -> dict[str, Any]:
    rows, _workers = collect_rows()
    jaque_outputs = {per_apur: build_jaque_workbook(per_apur, rows) for per_apur in REPORT_MONTHS}
    dev_output = build_dev_workbook(rows)
    return write_summary(rows, jaque_outputs, dev_output)


def main() -> int:
    print(json.dumps(build_reports(), ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())