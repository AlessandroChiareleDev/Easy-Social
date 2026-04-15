"""
Gerador de Relatório ANTES vs DEPOIS — S-5002 (Totalizador IRRF)
═══════════════════════════════════════════════════════════════════
Gera um Excel comparando os dados do S-5002 antes e depois da retificação S-1210.

Uso:
  python relatorio_antes_depois.py --run-id 5
  python relatorio_antes_depois.py --periodo 2025-01

Saída: relatorio_ana/ANTES_DEPOIS_S5002_202501_<timestamp>.xlsx
"""

import sys, os, json, argparse, re
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from db_config import DB_CONFIG
import psycopg2
import psycopg2.extras

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERRO: pip install openpyxl")
    sys.exit(1)


# ── Styling ────────────────────────────────────────────────────

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
CHANGE_FILL = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
ERROR_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
OK_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_header(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws, max_col, max_row):
    for col in range(1, max_col + 1):
        max_len = 10
        for row in range(1, min(max_row + 1, 200)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 40)


# ── Data helpers ──────────────────────────────────────────────

def _extract_s5002_fields(dados: dict) -> dict:
    """Extrai campos relevantes do S-5002 datos_json."""
    info_ir = dados.get("infoIR", [])
    vlr_rend_trib = 0
    vlr_deducao = 0
    vlr_irrf = 0
    vlr_cr_men = dados.get("vlrCRMen", 0)

    for item in info_ir:
        tp = str(item.get("tpInfoIR", ""))
        vlr = float(item.get("valor", 0))
        if tp in ("11", "12", "13", "14"):  # rendimento tributável
            vlr_rend_trib += vlr
        elif tp.startswith("79") and tp != "7950":  # deduções
            vlr_deducao += vlr
        elif tp in ("7950", "7951"):  # IRRF
            vlr_irrf += vlr

    return {
        "vlrRendTrib": round(vlr_rend_trib, 2),
        "vlrDeducao": round(vlr_deducao, 2),
        "vlrIRRF": round(vlr_irrf, 2),
        "vlrCRMen": round(float(vlr_cr_men or 0), 2),
        "qtdInfoIR": len(info_ir),
        "tpCR": dados.get("tpCR", ""),
        "infoIR_raw": info_ir,
    }


def _get_conn():
    return psycopg2.connect(**DB_CONFIG)


def gerar_relatorio(run_id: int = None, per_apur: str = None):
    conn = _get_conn()

    # Resolver run_id se não fornecido
    if not run_id and per_apur:
        with conn.cursor() as cur:
            cur.execute(
                "SELECT id FROM pipeline_runs WHERE per_apur = %s ORDER BY id DESC LIMIT 1",
                (per_apur,)
            )
            row = cur.fetchone()
            if row:
                run_id = row[0]
            else:
                print(f"ERRO: Nenhum pipeline_run encontrado para {per_apur}")
                conn.close()
                sys.exit(1)
    elif not run_id:
        print("ERRO: Forneça --run-id ou --periodo")
        sys.exit(1)

    # Buscar info do run
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("SELECT * FROM pipeline_runs WHERE id = %s", (run_id,))
        run_info = cur.fetchone()
        if not run_info:
            print(f"ERRO: Run {run_id} não encontrado")
            conn.close()
            sys.exit(1)
        if not per_apur:
            per_apur = run_info["per_apur"]

    print(f"Gerando relatório para run_id={run_id}, per_apur={per_apur}")

    # Buscar snapshots ANTES
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT cpf, dados_s5002, nr_recibo_s5002
            FROM pipeline_snapshots
            WHERE run_id = %s AND tipo = 'antes'
            ORDER BY cpf
        """, (run_id,))
        snap_antes = {row["cpf"]: row for row in cur.fetchall()}

    # Buscar snapshots DEPOIS
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT cpf, dados_s5002, nr_recibo_s5002
            FROM pipeline_snapshots
            WHERE run_id = %s AND tipo = 'depois'
            ORDER BY cpf
        """, (run_id,))
        snap_depois = {row["cpf"]: row for row in cur.fetchall()}

    # Buscar resultados do pipeline
    with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
        cur.execute("""
            SELECT cpf, status, nr_recibo_original, nr_recibo_novo, erro_descricao
            FROM pipeline_cpf_results
            WHERE run_id = %s
            ORDER BY cpf
        """, (run_id,))
        cpf_results = {row["cpf"]: row for row in cur.fetchall()}

    conn.close()

    all_cpfs = sorted(set(list(snap_antes.keys()) + list(snap_depois.keys()) + list(cpf_results.keys())))
    print(f"  Total CPFs: {len(all_cpfs)}")
    print(f"  Snapshots ANTES: {len(snap_antes)}")
    print(f"  Snapshots DEPOIS: {len(snap_depois)}")
    print(f"  Resultados pipeline: {len(cpf_results)}")

    # ═══════════════════════════════════════════════════════════
    # GERAR EXCEL
    # ═══════════════════════════════════════════════════════════
    wb = Workbook()

    # ── Tab 1: RESUMO ──
    ws = wb.active
    ws.title = "RESUMO"
    info_rows = [
        ("Relatório", "Comparação S-5002 ANTES vs DEPOIS — Retificação S-1210"),
        ("Período", per_apur),
        ("Run ID", run_id),
        ("Status Pipeline", run_info.get("status", "")),
        ("CPFs total", run_info.get("total_cpfs", "")),
        ("CPFs OK", run_info.get("cpfs_ok", "")),
        ("CPFs erro", run_info.get("cpfs_erro", "")),
        ("S-1298 (reabrir)", "SIM" if run_info.get("s1298_done") else "NÃO"),
        ("S-1299 (fechar)", "SIM" if run_info.get("s1299_done") else "NÃO"),
        ("Snapshots ANTES", len(snap_antes)),
        ("Snapshots DEPOIS", len(snap_depois)),
        ("Gerado em", datetime.now().strftime("%Y-%m-%d %H:%M")),
    ]
    for r, (k, v) in enumerate(info_rows, 1):
        ws.cell(row=r, column=1, value=k).font = Font(bold=True)
        ws.cell(row=r, column=2, value=str(v))
    _auto_width(ws, 2, len(info_rows))

    # Contadores de mudança
    total_alterados = 0
    total_sem_mudanca = 0
    total_sem_snapshot = 0

    # ── Tab 2: COMPARAÇÃO DETALHADA ──
    ws2 = wb.create_sheet("COMPARAÇÃO S-5002")
    headers = [
        "CPF", "Status Pipeline", "Recibo Original", "Recibo Novo",
        "ANTES vlrRendTrib", "DEPOIS vlrRendTrib", "Δ vlrRendTrib",
        "ANTES vlrDeducao", "DEPOIS vlrDeducao", "Δ vlrDeducao",
        "ANTES vlrIRRF", "DEPOIS vlrIRRF", "Δ vlrIRRF",
        "ANTES vlrCRMen", "DEPOIS vlrCRMen", "Δ vlrCRMen",
        "ANTES qtdInfoIR", "DEPOIS qtdInfoIR",
        "Observação",
    ]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=1, column=c, value=h)
    _style_header(ws2, 1, len(headers))

    row = 2
    for cpf in all_cpfs:
        result = cpf_results.get(cpf, {})
        antes_row = snap_antes.get(cpf)
        depois_row = snap_depois.get(cpf)

        antes_dados = {}
        depois_dados = {}
        obs = ""

        if antes_row:
            d = antes_row["dados_s5002"]
            if isinstance(d, str):
                d = json.loads(d)
            antes_dados = _extract_s5002_fields(d or {})

        if depois_row:
            d = depois_row["dados_s5002"]
            if isinstance(d, str):
                d = json.loads(d)
            depois_dados = _extract_s5002_fields(d or {})

        if not antes_row and not depois_row:
            obs = "Sem S-5002 ANTES e DEPOIS"
            total_sem_snapshot += 1
        elif not antes_row:
            obs = "Sem S-5002 ANTES (novo?)"
            total_sem_snapshot += 1
        elif not depois_row:
            obs = "Sem S-5002 DEPOIS"
            total_sem_snapshot += 1

        delta_rend = round((depois_dados.get("vlrRendTrib", 0) or 0) - (antes_dados.get("vlrRendTrib", 0) or 0), 2)
        delta_ded = round((depois_dados.get("vlrDeducao", 0) or 0) - (antes_dados.get("vlrDeducao", 0) or 0), 2)
        delta_irrf = round((depois_dados.get("vlrIRRF", 0) or 0) - (antes_dados.get("vlrIRRF", 0) or 0), 2)
        delta_crmen = round((depois_dados.get("vlrCRMen", 0) or 0) - (antes_dados.get("vlrCRMen", 0) or 0), 2)

        has_change = any(abs(d) > 0.001 for d in [delta_rend, delta_ded, delta_irrf, delta_crmen])
        if has_change:
            total_alterados += 1
        elif antes_row and depois_row:
            total_sem_mudanca += 1

        values = [
            cpf,
            result.get("status", "—"),
            result.get("nr_recibo_original", ""),
            result.get("nr_recibo_novo", ""),
            antes_dados.get("vlrRendTrib", ""),
            depois_dados.get("vlrRendTrib", ""),
            delta_rend if has_change else 0,
            antes_dados.get("vlrDeducao", ""),
            depois_dados.get("vlrDeducao", ""),
            delta_ded if has_change else 0,
            antes_dados.get("vlrIRRF", ""),
            depois_dados.get("vlrIRRF", ""),
            delta_irrf if has_change else 0,
            antes_dados.get("vlrCRMen", ""),
            depois_dados.get("vlrCRMen", ""),
            delta_crmen if has_change else 0,
            antes_dados.get("qtdInfoIR", ""),
            depois_dados.get("qtdInfoIR", ""),
            obs,
        ]
        for c, v in enumerate(values, 1):
            cell = ws2.cell(row=row, column=c, value=v)
            cell.border = THIN_BORDER
            if c in (7, 10, 13, 16) and isinstance(v, (int, float)) and abs(v) > 0.001:
                cell.fill = CHANGE_FILL
                cell.font = Font(bold=True)
            if c == 2:
                if v == "ok":
                    cell.fill = OK_FILL
                elif v == "erro":
                    cell.fill = ERROR_FILL

        row += 1

    _auto_width(ws2, len(headers), row)
    ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row - 1}"
    ws2.freeze_panes = "A2"

    # Atualizar resumo com contadores
    ws.cell(row=len(info_rows) + 2, column=1, value="CPFs com alteração S-5002").font = Font(bold=True, color="FF0000")
    ws.cell(row=len(info_rows) + 2, column=2, value=total_alterados)
    ws.cell(row=len(info_rows) + 3, column=1, value="CPFs sem mudança S-5002").font = Font(bold=True)
    ws.cell(row=len(info_rows) + 3, column=2, value=total_sem_mudanca)
    ws.cell(row=len(info_rows) + 4, column=1, value="CPFs sem snapshot").font = Font(bold=True)
    ws.cell(row=len(info_rows) + 4, column=2, value=total_sem_snapshot)

    # ── Tab 3: ERROS ──
    erros = [(cpf, r) for cpf, r in cpf_results.items() if r.get("status") == "erro"]
    if erros:
        ws3 = wb.create_sheet("ERROS")
        err_headers = ["CPF", "Recibo Original", "Erro"]
        for c, h in enumerate(err_headers, 1):
            ws3.cell(row=1, column=c, value=h)
        _style_header(ws3, 1, len(err_headers))
        for i, (cpf, r) in enumerate(sorted(erros), 2):
            ws3.cell(row=i, column=1, value=cpf).border = THIN_BORDER
            ws3.cell(row=i, column=2, value=r.get("nr_recibo_original", "")).border = THIN_BORDER
            ws3.cell(row=i, column=3, value=r.get("erro_descricao", "")).border = THIN_BORDER
        _auto_width(ws3, 3, len(erros) + 1)

    # ── Salvar ──
    os.makedirs("relatorio_ana", exist_ok=True)
    per_key = per_apur.replace("-", "")
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"relatorio_ana/ANTES_DEPOIS_S5002_{per_key}_{ts}.xlsx"
    wb.save(filename)
    print(f"\n✓ Relatório salvo: {filename}")
    print(f"  {total_alterados} CPFs com alteração, {total_sem_mudanca} sem mudança, {total_sem_snapshot} sem snapshot")
    return filename


# ── CLI ───────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Relatório ANTES vs DEPOIS S-5002")
    parser.add_argument("--run-id", type=int, help="ID do pipeline_run")
    parser.add_argument("--periodo", help="Período (AAAA-MM)")
    args = parser.parse_args()

    if not args.run_id and not args.periodo:
        print("ERRO: Forneça --run-id ou --periodo")
        sys.exit(1)

    gerar_relatorio(run_id=args.run_id, per_apur=args.periodo)


if __name__ == "__main__":
    main()
