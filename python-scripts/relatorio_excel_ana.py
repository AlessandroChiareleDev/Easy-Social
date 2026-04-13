"""
RELATÓRIO GERAL EXCEL — Para Ana (RH APPA) enviar ao GI
═════════════════════════════════════════════════════════
Gera um Excel com múltiplas abas:
  1. RESUMO           — Visão geral do que foi feito
  2. RUBRICAS S-1010  — Todas as rubricas alteradas (antes → depois) com recibos
  3. RETIFICAÇÕES     — 7.771 CPFs com S-1210 retificados (recibo antigo → novo)
  4. EVENTOS POR CPF  — Resumo por CPF de todos os eventos
"""
import psycopg2
import psycopg2.extras
import sys
import os
import json
from datetime import datetime
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from db_config import DB_CONFIG

OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "relatorio_ana")
os.makedirs(OUTPUT_DIR, exist_ok=True)
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M")

# ── Styles ────────────────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FILL_GREEN = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
HEADER_FILL_ORANGE = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="404040")
OK_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
ERRO_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
CHANGED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

DESCRICAO_INCIDENCIA_INSS = {
    "00": "00 - Não é base de cálculo",
    "01": "01 - Não é base de cálculo até o limite",
    "11": "11 - Mensal",
    "12": "12 - 13º Salário",
    "13": "13 - Exclusivas empregador (mensal)",
    "14": "14 - Exclusivas empregador (13ºS)",
    "15": "15 - Exclusivas segurado (mensal)",
    "16": "16 - Exclusivas segurado (13ºS)",
    "21": "21 - Salário-maternidade mensal",
    "22": "22 - Salário-maternidade 13ºS",
    "31": "31 - Contrib.desc.empregado (coop/sindical)",
    "51": "51 - Mensal - Atividades concomitantes",
    "91": "91 - Incidência suspensa - decisão judicial",
    "92": "92 - Incidência suspensa - depósito (montante integral)",
    "93": "93 - Incidência suspensa - compensação/liminar",
    "94": "94 - Incidência suspensa - antecipação tutela",
}
DESCRICAO_INCIDENCIA_IRRF = {
    "00": "00 - Não é base de cálculo",
    "01": "01 - Não é base de cálculo - condição",
    "07": "07 - Remuneração Dirigente - Entidade Isenta/Imune",
    "09": "09 - Verba transitada (s/incid.)",
    "11": "11 - Tributável - Mensal",
    "12": "12 - Tributável - 13ºS",
    "13": "13 - Tributável - Férias",
    "14": "14 - Tributável - PLR",
    "15": "15 - Não Tributável - Parc.Isenta 65 anos",
    "31": "31 - RRA - Rendimentos",
    "32": "32 - RRA - 13ºS",
    "33": "33 - RRA - Férias",
    "34": "34 - RRA - PLR",
    "41": "41 - Tributável - Rendimentos",
    "42": "42 - Tributável - 13ºS",
    "43": "43 - Tributável - Férias",
    "46": "46 - Dedução IRRF",
    "51": "51 - Tributável - Outros",
    "61": "61 - Dedução - Previdência Oficial",
    "62": "62 - Dedução - Previdência Privada",
    "63": "63 - Dedução - FAPI",
    "64": "64 - Dedução - Pensão Alimentícia",
    "65": "65 - Dedução - Dependente",
    "66": "66 - Dedução - Previdência Empregador",
    "67": "67 - Dedução - Salário-maternidade",
    "70": "70 - Isenção - Diárias",
    "71": "71 - Isenção - Ajuda de Custo",
    "72": "72 - Isenção - Indenização/Rescisão",
    "73": "73 - Isenção - Abono Pecuniário",
    "74": "74 - Isenção - Ganho Eventual",
    "75": "75 - Isenção - Seguros",
    "76": "76 - Isenção - Moléstia Grave",
    "77": "77 - Isenção - 65 anos",
    "78": "78 - Isenção - Outras",
    "79": "79 - Verba transitada",
    "91": "91 - Incidência suspensa - decisão judicial",
    "92": "92 - Incidência suspensa - depósito",
    "93": "93 - Incidência suspensa - compensação",
    "94": "94 - Incidência suspensa - antecipação tutela",
    "95": "95 - Incidência suspensa - liminar",
}
DESCRICAO_INCIDENCIA_FGTS = {
    "00": "00 - Não é base de cálculo",
    "11": "11 - Base mensal",
    "12": "12 - Base 13ºS",
    "13": "13 - Base rescisória mensal",
    "14": "14 - Base rescisória 13ºS",
    "21": "21 - Base sal.maternidade mensal",
    "22": "22 - Base sal.maternidade 13ºS",
    "91": "91 - Incidência suspensa - decisão judicial",
}


def connect():
    return psycopg2.connect(
        **DB_CONFIG,
        keepalives=1, keepalives_idle=30,
        keepalives_interval=10, keepalives_count=3
    )


def formatar_cpf(cpf):
    if cpf and len(cpf) == 11:
        return f"{cpf[:3]}.{cpf[3:6]}.{cpf[6:9]}-{cpf[9:]}"
    return cpf or ""


def apply_header(ws, row_num, headers, fill=None):
    """Aplica estilo de cabeçalho numa linha."""
    _fill = fill or HEADER_FILL
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_num, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = _fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def auto_width(ws, min_width=10, max_width=50):
    """Ajusta largura das colunas automaticamente."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = width


def desc_incid(code, tipo):
    """Retorna descrição da incidência."""
    if not code:
        return ""
    code = str(code).strip().zfill(2)
    if tipo == "inss":
        return DESCRICAO_INCIDENCIA_INSS.get(code, code)
    elif tipo == "irrf":
        return DESCRICAO_INCIDENCIA_IRRF.get(code, code)
    elif tipo == "fgts":
        return DESCRICAO_INCIDENCIA_FGTS.get(code, code)
    return code


# ══════════════════════════════════════════════════════════════════
# ABA 1: RESUMO
# ══════════════════════════════════════════════════════════════════

def criar_aba_resumo(wb, stats):
    ws = wb.active
    ws.title = "RESUMO"
    ws.sheet_properties.tabColor = "2F5496"

    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="RELATÓRIO GERAL — Correções eSocial APPA")
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    c = ws.cell(row=2, column=1, value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | Empresa: APPA (CNPJ 05.969.071/0001-10)")
    c.font = SUBTITLE_FONT
    c.alignment = Alignment(horizontal="center")

    r = 4
    resumo = [
        ("O QUE FOI FEITO", ""),
        ("", ""),
        ("1. Correção de Rubricas (S-1010)", ""),
        ("   Rubricas com divergência de incidência encontradas", stats.get("total_divergencias", 0)),
        ("   Rubricas corrigidas via webservice (S-1010 alteração)", stats.get("s1010_ok", 0)),
        ("   Rubricas com erro no envio", stats.get("s1010_erro", 0)),
        ("   Rubricas incluídas (S-1010 inclusão)", stats.get("s1010_inclusao_ok", 0)),
        ("", ""),
        ("2. Retificação de Pagamentos (S-1210) — Período 2025-09", ""),
        ("   CPFs retificados com sucesso", stats.get("cpfs_ok", 0)),
        ("   CPFs com erro", stats.get("cpfs_erro", 0)),
        ("   Total de eventos enviados (em lotes de 50)", stats.get("total_eventos_enviados", 0)),
        ("", ""),
        ("3. Eventos de Suporte (S-1298 / S-1299)", ""),
        ("   S-1298 Reabertura de Período", "Executado ✓"),
        ("   S-1299 Fechamento de Período", "Executado ✓"),
        ("", ""),
        ("IMPACTO NO SISTEMA", ""),
        ("   As incidências de INSS, IRRF e FGTS das rubricas foram corrigidas", ""),
        ("   diretamente no eSocial via webservice. As mudanças passam a valer", ""),
        ("   para todos os períodos a partir da validade inicial (iniValid) de", ""),
        ("   cada rubrica. O GI precisa atualizar as mesmas incidências no", ""),
        ("   sistema interno para que os próximos envios reflitam as correções.", ""),
        ("", ""),
        ("ABAS DESTE RELATÓRIO", ""),
        ("   Aba 'RUBRICAS S-1010'", "Todas as rubricas alteradas (antes → depois) com recibos"),
        ("   Aba 'RETIFICAÇÕES S-1210'", "7.771 CPFs com recibo antigo → recibo novo"),
        ("   Aba 'EVENTOS POR CPF'", "Resumo por CPF de todos os eventos no período"),
    ]

    for desc, valor in resumo:
        ws.cell(row=r, column=1, value=desc).font = Font(
            name="Calibri", bold=desc.isupper() or desc.startswith("   Aba"),
            size=11
        )
        ws.cell(row=r, column=4, value=valor if valor else "")
        r += 1

    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["D"].width = 30


# ══════════════════════════════════════════════════════════════════
# ABA 2: RUBRICAS S-1010
# ══════════════════════════════════════════════════════════════════

def criar_aba_rubricas(wb, conn):
    ws = wb.create_sheet("RUBRICAS S-1010")
    ws.sheet_properties.tabColor = "C55A11"
    cur = conn.cursor()

    # Buscar divergências do cruzamento_eb
    cur.execute("""
        SELECT cod_rubrica, descricao, cod_natureza,
               incid_inss, incid_irrf, incid_fgts,
               incid_base_legal_inss, incid_base_legal_irrf, incid_base_legal_fgts,
               corrigido, envio_status, ini_valid_esocial
        FROM cruzamento_eb
        ORDER BY cod_rubrica
    """)
    cruzamento_rows = cur.fetchall()
    cruzamento_cols = [d[0] for d in cur.description]

    # Buscar envios S-1010 com recibo
    cur.execute("""
        SELECT id, modo, status, nr_recibo, rubrica_detalhes, ocorrencias,
               descricao_resposta, created_at
        FROM esocial_envios
        WHERE ambiente = '1' AND modo IN ('alteracao', 'inclusao')
          AND status IN ('processado', 'aceito')
        ORDER BY created_at
    """)
    envios_ok = cur.fetchall()
    envios_cols = [d[0] for d in cur.description]

    # Mapear cod_rubrica → recibo (do envio processado)
    rubrica_recibo = {}
    for row in envios_ok:
        e = dict(zip(envios_cols, row))
        det = e['rubrica_detalhes']
        if det and isinstance(det, list):
            for d in det:
                cod = d.get('codRubr') or d.get('cod_rubrica')
                if cod:
                    rubrica_recibo[str(cod)] = {
                        "recibo": e['nr_recibo'],
                        "modo": e['modo'],
                        "data": e['created_at'],
                        "envio_id": e['id'],
                    }

    # Também buscar rubrica_corrections para valores antes/depois
    cur.execute("""
        SELECT cod_rubrica, descricao,
               inss_antes, irrf_antes, fgts_antes,
               inss_correto, irrf_correto, fgts_correto,
               status, corrigido_em
        FROM rubrica_corrections
        ORDER BY cod_rubrica
    """)
    corrections_rows = cur.fetchall()
    corrections_cols = [d[0] for d in cur.description]
    corrections_map = {}
    for row in corrections_rows:
        r = dict(zip(corrections_cols, row))
        corrections_map[str(r['cod_rubrica'])] = r

    # Título
    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1, value="RUBRICAS ALTERADAS — Incidências INSS / IRRF / FGTS")
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    c = ws.cell(row=2, column=1,
                value="Valores ANTES (no eSocial) → DEPOIS (corrigidos). Cada rubrica foi enviada como S-1010 alteração/inclusão.")
    c.font = SUBTITLE_FONT

    headers = [
        "Cód. Rubrica",
        "Descrição",
        "Natureza",
        "INSS Antes",
        "INSS Correto",
        "IRRF Antes",
        "IRRF Correto",
        "FGTS Antes",
        "FGTS Correto",
        "Status Envio",
        "Nr Recibo eSocial",
        "Data Envio",
    ]
    apply_header(ws, 4, headers, HEADER_FILL_ORANGE)

    row_num = 5
    stats = {"total": 0, "ok": 0, "erro": 0}

    for row in cruzamento_rows:
        r = dict(zip(cruzamento_cols, row))
        cod = str(r['cod_rubrica'])

        # Extrair código correto do texto "11 - Artigo 28..."
        def extract_code(full_text):
            if not full_text:
                return ""
            return full_text.split(" - ")[0].strip() if " - " in full_text else full_text.strip()

        inss_antes = str(r['incid_inss'] or "").strip()
        irrf_antes = str(r['incid_irrf'] or "").strip()
        fgts_antes = str(r['incid_fgts'] or "").strip()
        inss_correto = extract_code(r['incid_base_legal_inss'])
        irrf_correto = extract_code(r['incid_base_legal_irrf'])
        fgts_correto = extract_code(r['incid_base_legal_fgts'])

        # Se existe no corrections, usar seus dados (podem ser mais precisos)
        if cod in corrections_map:
            corr = corrections_map[cod]
            if corr['inss_antes']:
                inss_antes = str(corr['inss_antes']).strip()
            if corr['irrf_antes']:
                irrf_antes = str(corr['irrf_antes']).strip()
            if corr['fgts_antes']:
                fgts_antes = str(corr['fgts_antes']).strip()
            if corr['inss_correto']:
                inss_correto = str(corr['inss_correto']).strip()
            if corr['irrf_correto']:
                irrf_correto = str(corr['irrf_correto']).strip()
            if corr['fgts_correto']:
                fgts_correto = str(corr['fgts_correto']).strip()

        # Checar se houve mudança real
        changed = (inss_antes != inss_correto or irrf_antes != irrf_correto or fgts_antes != fgts_correto)
        if not changed and not r['corrigido']:
            continue  # Sem divergência real, pula

        stats["total"] += 1

        recibo_info = rubrica_recibo.get(cod, {})
        status_envio = r['envio_status'] or "pendente"
        if recibo_info.get("recibo"):
            status_envio = "✓ Processado"
            stats["ok"] += 1
        elif r['corrigido']:
            status_envio = "✓ Corrigido"
            stats["ok"] += 1
        else:
            stats["erro"] += 1

        values = [
            cod,
            r['descricao'] or "",
            r['cod_natureza'] or "",
            f"{inss_antes} → {desc_incid(inss_antes, 'inss')}",
            f"{inss_correto} → {desc_incid(inss_correto, 'inss')}",
            f"{irrf_antes} → {desc_incid(irrf_antes, 'irrf')}",
            f"{irrf_correto} → {desc_incid(irrf_correto, 'irrf')}",
            f"{fgts_antes} → {desc_incid(fgts_antes, 'fgts')}",
            f"{fgts_correto} → {desc_incid(fgts_correto, 'fgts')}",
            status_envio,
            recibo_info.get("recibo", ""),
            recibo_info.get("data", "").strftime('%d/%m/%Y %H:%M') if recibo_info.get("data") else "",
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        # Highlight changed cells
        for col_idx in [4, 5]:  # INSS
            if inss_antes != inss_correto:
                ws.cell(row=row_num, column=col_idx).fill = CHANGED_FILL
        for col_idx in [6, 7]:  # IRRF
            if irrf_antes != irrf_correto:
                ws.cell(row=row_num, column=col_idx).fill = CHANGED_FILL
        for col_idx in [8, 9]:  # FGTS
            if fgts_antes != fgts_correto:
                ws.cell(row=row_num, column=col_idx).fill = CHANGED_FILL

        # Status color
        if "✓" in status_envio:
            ws.cell(row=row_num, column=10).fill = OK_FILL
        else:
            ws.cell(row=row_num, column=10).fill = ERRO_FILL

        row_num += 1

    auto_width(ws, min_width=12, max_width=45)
    ws.column_dimensions["B"].width = 40
    ws.auto_filter.ref = f"A4:L{row_num - 1}"
    ws.freeze_panes = "A5"

    return stats


# ══════════════════════════════════════════════════════════════════
# ABA 3: RETIFICAÇÕES S-1210
# ══════════════════════════════════════════════════════════════════

def criar_aba_retificacoes(wb, conn):
    ws = wb.create_sheet("RETIFICAÇÕES S-1210")
    ws.sheet_properties.tabColor = "548235"
    cur = conn.cursor()

    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1, value="RETIFICAÇÕES S-1210 — Pagamentos de Rendimentos (Período 2025-09)")
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    c = ws.cell(row=2, column=1,
                value="Cada CPF teve seu S-1210 retificado. Recibo Original = evento antigo, Recibo Novo = retificação aceita pelo eSocial.")
    c.font = SUBTITLE_FONT

    headers = [
        "CPF",
        "CPF Formatado",
        "Período",
        "Status",
        "Recibo Original",
        "Recibo Novo (Retificação)",
        "Lote",
        "Processado Em",
        "Erro"
    ]
    apply_header(ws, 4, headers, HEADER_FILL_GREEN)

    cur.execute("""
        SELECT r.per_apur, c.cpf, c.status, c.nr_recibo_original, c.nr_recibo_novo,
               c.lote_num, c.processed_at, c.erro_descricao
        FROM pipeline_cpf_results c
        JOIN pipeline_runs r ON r.id = c.run_id
        ORDER BY c.cpf
    """)
    rows = cur.fetchall()
    cols = [d[0] for d in cur.description]

    stats = {"total": len(rows), "ok": 0, "erro": 0}

    for i, row in enumerate(rows):
        r = dict(zip(cols, row))
        row_num = 5 + i

        if r['status'] == 'ok':
            stats["ok"] += 1
        else:
            stats["erro"] += 1

        values = [
            r['cpf'],
            formatar_cpf(r['cpf']),
            r['per_apur'],
            r['status'].upper(),
            r['nr_recibo_original'] or "",
            r['nr_recibo_novo'] or "",
            r['lote_num'] or "",
            r['processed_at'].strftime('%d/%m/%Y %H:%M') if r['processed_at'] else "",
            r['erro_descricao'] or ""
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

        if r['status'] == 'ok':
            ws.cell(row=row_num, column=4).fill = OK_FILL
        else:
            ws.cell(row=row_num, column=4).fill = ERRO_FILL

    auto_width(ws, min_width=12, max_width=45)
    last_row = 4 + len(rows)
    ws.auto_filter.ref = f"A4:I{last_row}"
    ws.freeze_panes = "A5"

    return stats


# ══════════════════════════════════════════════════════════════════
# ABA 4: EVENTOS POR CPF
# ══════════════════════════════════════════════════════════════════

def criar_aba_eventos_cpf(wb, conn):
    ws = wb.create_sheet("EVENTOS POR CPF")
    ws.sheet_properties.tabColor = "7030A0"
    cur = conn.cursor()

    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value="RESUMO DE EVENTOS POR CPF — Todos os tipos")
    c.font = TITLE_FONT
    c.alignment = Alignment(horizontal="center")

    # Buscar eventos agrupados
    cur.execute("""
        SELECT cpf, tipo_evento, COUNT(*) as qtd
        FROM explorador_eventos
        WHERE cpf IS NOT NULL
        GROUP BY cpf, tipo_evento
        ORDER BY cpf, tipo_evento
    """)
    rows = cur.fetchall()

    cpf_data = defaultdict(lambda: defaultdict(int))
    for cpf, tipo, qtd in rows:
        cpf_data[cpf][tipo] = qtd

    all_tipos = sorted(set(tipo for cpf, tipo, _ in rows))

    headers = ["CPF", "CPF Formatado", "Total Eventos"] + all_tipos
    apply_header(ws, 3, headers)

    row_num = 4
    for cpf in sorted(cpf_data.keys()):
        tipos = cpf_data[cpf]
        total = sum(tipos.values())
        values = [cpf, formatar_cpf(cpf), total] + [tipos.get(t, 0) for t in all_tipos]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal="center" if col_idx > 2 else "left")

        row_num += 1

    auto_width(ws, min_width=8, max_width=20)
    ws.column_dimensions["B"].width = 18
    last_row = 3 + len(cpf_data)
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{last_row}"
    ws.freeze_panes = "A4"

    return {"total_cpfs": len(cpf_data), "total_eventos": len(rows)}


# ══════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════

def main():
    print()
    print("█" * 70)
    print("█  RELATÓRIO EXCEL COMPLETO — ANA (RH APPA) → GI")
    print("█  " + datetime.now().strftime('%d/%m/%Y %H:%M'))
    print("█" * 70)

    wb = Workbook()
    conn = connect()

    print("\n[1/4] Criando aba RUBRICAS S-1010...")
    rub_stats = criar_aba_rubricas(wb, conn)
    print(f"  Total divergências: {rub_stats['total']}, OK: {rub_stats['ok']}, Pendentes: {rub_stats['erro']}")

    print("\n[2/4] Criando aba RETIFICAÇÕES S-1210...")
    ret_stats = criar_aba_retificacoes(wb, conn)
    print(f"  Total: {ret_stats['total']}, OK: {ret_stats['ok']}")

    print("\n[3/4] Criando aba EVENTOS POR CPF...")
    evt_stats = criar_aba_eventos_cpf(wb, conn)
    print(f"  CPFs: {evt_stats['total_cpfs']}")

    # Buscar stats adicionais para o resumo
    cur = conn.cursor()
    cur.execute("""
        SELECT COUNT(*) FROM esocial_envios
        WHERE ambiente='1' AND modo='inclusao' AND status IN ('processado','aceito')
    """)
    inclusao_ok = cur.fetchone()[0]

    print("\n[4/4] Criando aba RESUMO...")
    all_stats = {
        "total_divergencias": rub_stats["total"],
        "s1010_ok": rub_stats["ok"],
        "s1010_erro": rub_stats["erro"],
        "s1010_inclusao_ok": inclusao_ok,
        "cpfs_ok": ret_stats["ok"],
        "cpfs_erro": ret_stats["erro"],
        "total_eventos_enviados": ret_stats["total"] + 2,  # +S1298 +S1299
    }
    criar_aba_resumo(wb, all_stats)

    conn.close()

    # Salvar
    output_file = os.path.join(OUTPUT_DIR, f"RELATORIO_ESOCIAL_APPA_{TIMESTAMP}.xlsx")
    wb.save(output_file)

    print(f"\n{'=' * 70}")
    print(f"  ✅ EXCEL GERADO:")
    print(f"  {output_file}")
    print(f"{'=' * 70}")
    print(f"  Abas: RESUMO | RUBRICAS S-1010 | RETIFICAÇÕES S-1210 | EVENTOS POR CPF")
    print(f"  Pronto para enviar à Ana → GI")


if __name__ == "__main__":
    main()
