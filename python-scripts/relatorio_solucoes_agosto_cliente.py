"""
Relatorio cliente - SOLUCOES Agosto/2025 (V2/F5)

Uma linha por CPF, usando o ultimo status do frontend.
Layout enxuto, sem colunas tecnicas nas abas finais.
Nao consulta eSocial.
"""
from __future__ import annotations

import os
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DSN = os.environ.get(
    "SISTEMA_DB_URL",
    "postgresql://postgres:EsoV2_CoxRHWQ1z6iucG7ZyvdqFIbN@db.kjbgiwnlvqnrfdozjvhq.supabase.co:5432/postgres?sslmode=require",
)
SCHEMA = "solucoes"
PER_APUR = "2025-08"
EMPRESA = "SOLUCOES SERVICOS TERCEIRIZADOS LTDA"
CNPJ = "09.445.502/0001-09"

BASE = Path(__file__).resolve().parent
OUT_DIR = BASE.parent / "relatorio_ana"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT = OUT_DIR / f"RELATORIO_SOLUCOES_AGOSTO_2025_CLIENTE_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

BLUE = "1F4E79"
GREEN = "548235"
YELLOW = "BF9000"
RED = "C00000"
ORANGE = "C55A11"
GRAY = "595959"
WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color=BLUE, size=14)
SUB_FONT = Font(name="Calibri", bold=True, color="404040", size=11)
THIN = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)
FILLS = {
    "ENVIADO_OK": PatternFill("solid", fgColor="E2EFDA"),
    "ENVIADO_COM_ADVERTENCIA": PatternFill("solid", fgColor="FFF2CC"),
    "ERRO_NAO_ENVIADO": PatternFill("solid", fgColor="FCE4EC"),
    "PENDENTE": PatternFill("solid", fgColor="FCE4D6"),
    "SEM_MUDANCA": PatternFill("solid", fgColor="E7E6E6"),
}
SITUACAO = {
    "ENVIADO_OK": "Enviado OK",
    "ENVIADO_COM_ADVERTENCIA": "Enviado com advertência",
    "ERRO_NAO_ENVIADO": "Erro - não enviado",
    "PENDENTE": "Pendente",
    "SEM_MUDANCA": "Sem mudança",
}


def connect():
    conn = psycopg2.connect(DSN)
    with conn.cursor() as cur:
        cur.execute(f'SET search_path TO "{SCHEMA}", public')
    conn.commit()
    return conn


def cpf_fmt(cpf: str | None) -> str:
    digits = "".join(ch for ch in str(cpf or "") if ch.isdigit()).zfill(11)
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return str(cpf or "")


def clean_msg(value: Any) -> str:
    if not value:
        return ""
    msg = str(value).replace(" | ", "\n")
    return re.sub(r"[ \t]+", " ", msg).strip()


def dt_fmt(value: Any) -> str:
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    return str(value)


def classificar(status: str | None, codigo: str | None) -> str:
    status = (status or "").lower()
    codigo = str(codigo or "").strip()
    if status == "sucesso":
        return "ENVIADO_OK"
    if status == "erro_esocial" and codigo == "202":
        return "ENVIADO_COM_ADVERTENCIA"
    if status == "erro_esocial":
        return "ERRO_NAO_ENVIADO"
    if status in {"pendente", "pendente_consulta"}:
        return "PENDENTE"
    if status == "sem_mudanca" or codigo == "SEM_MUDANCA":
        return "SEM_MUDANCA"
    return "PENDENTE"


def fetch_rows() -> list[dict]:
    conn = connect()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT DISTINCT ON (cpf)
                       cpf, nr_recibo AS recibo_xml, dt_processamento, id AS evento_id
                  FROM explorador_eventos
                 WHERE tipo_evento='S-1210'
                   AND per_apur=%s
                   AND cpf IS NOT NULL
                   AND retificado_por_id IS NULL
                 ORDER BY cpf, dt_processamento DESC NULLS LAST, id DESC
                """,
                (PER_APUR,),
            )
            universo = {r["cpf"]: dict(r) for r in cur.fetchall()}

            cur.execute(
                """
                SELECT DISTINCT ON (it.cpf)
                       it.cpf, it.status, it.erro_codigo, it.erro_mensagem,
                       it.nr_recibo_anterior, it.nr_recibo_novo,
                       it.criado_em
                  FROM timeline_envio_item it
                  JOIN timeline_envio te ON te.id = it.timeline_envio_id
                  JOIN timeline_mes tm ON tm.id = te.timeline_mes_id
                 WHERE tm.per_apur=%s
                   AND it.tipo_evento='S-1210'
                   AND it.cpf IS NOT NULL
                 ORDER BY it.cpf, it.criado_em DESC NULLS LAST, it.id DESC
                """,
                (PER_APUR,),
            )
            ultimos = {r["cpf"]: dict(r) for r in cur.fetchall()}
    finally:
        conn.close()

    rows = []
    for cpf in sorted(universo):
        evento = universo[cpf]
        envio = ultimos.get(cpf)
        categoria = classificar(envio.get("status") if envio else None, envio.get("erro_codigo") if envio else None)
        rows.append(
            {
                "cpf": cpf,
                "cpf_formatado": cpf_fmt(cpf),
                "situacao": SITUACAO[categoria],
                "categoria": categoria,
                "codigo": envio.get("erro_codigo") or "" if envio else "",
                "mensagem": clean_msg(envio.get("erro_mensagem") if envio else "Sem envio registrado."),
                "recibo_xml": evento.get("recibo_xml") or "",
                "recibo_usado": envio.get("nr_recibo_anterior") or "" if envio else "",
                "recibo_novo": envio.get("nr_recibo_novo") or "" if envio else "",
                "atualizado_em": dt_fmt(envio.get("criado_em") if envio else evento.get("dt_processamento")),
            }
        )
    return rows


def setup_ws(ws, headers: list[str], fill_color: str, row: int = 1):
    for col, label in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=label)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN
    ws.freeze_panes = f"A{row + 1}"


def write_rows(ws, rows: list[dict], columns: list[tuple[str, str]], fill_color: str, widths: dict[str, int]):
    setup_ws(ws, [label for _, label in columns], fill_color)
    for row_idx, row in enumerate(rows, 2):
        for col_idx, (field, _) in enumerate(columns, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(field, ""))
            cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=field == "mensagem")
            cell.fill = FILLS.get(row["categoria"], PatternFill())
    last_col = get_column_letter(len(columns))
    ws.auto_filter.ref = f"A1:{last_col}{max(1, ws.max_row)}"
    if ws.max_row > 1:
        table = Table(displayName=re.sub(r"[^A-Za-z0-9_]", "_", ws.title)[:25], ref=ws.auto_filter.ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    if any(field == "mensagem" for field, _ in columns):
        for row_idx in range(2, min(ws.max_row + 1, 3000)):
            ws.row_dimensions[row_idx].height = 45


def write_resumo(wb: Workbook, rows: list[dict]):
    ws = wb.active
    ws.title = "RESUMO"
    ws.merge_cells("A1:C1")
    ws["A1"] = "RELATÓRIO SOLUÇÕES - AGOSTO/2025"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:C2")
    ws["A2"] = f"{EMPRESA} | CNPJ {CNPJ}"
    ws["A2"].font = SUB_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    counts = Counter(r["categoria"] for r in rows)
    resumo = [
        ("Total de CPFs", len(rows)),
        ("Enviados OK", counts["ENVIADO_OK"]),
        ("Enviados com advertência", counts["ENVIADO_COM_ADVERTENCIA"]),
        ("Erros - não enviados", counts["ERRO_NAO_ENVIADO"]),
        ("Pendentes", counts["PENDENTE"]),
        ("Sem mudança", counts["SEM_MUDANCA"]),
    ]
    setup_ws(ws, ["Indicador", "Quantidade"], BLUE, row=4)
    for idx, (label, value) in enumerate(resumo, 5):
        ws.cell(idx, 1, label).font = Font(bold=True)
        ws.cell(idx, 1).border = THIN
        ws.cell(idx, 2, value).border = THIN
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    return counts


def gerar():
    rows = fetch_rows()
    grupos = {
        "ENVIADOS_OK": [r for r in rows if r["categoria"] == "ENVIADO_OK"],
        "ADVERTENCIAS": [r for r in rows if r["categoria"] == "ENVIADO_COM_ADVERTENCIA"],
        "ERROS": [r for r in rows if r["categoria"] == "ERRO_NAO_ENVIADO"],
        "PENDENTES": [r for r in rows if r["categoria"] == "PENDENTE"],
        "SEM_MUDANCA": [r for r in rows if r["categoria"] == "SEM_MUDANCA"],
    }

    wb = Workbook()
    counts = write_resumo(wb, rows)

    write_rows(
        wb.create_sheet("TODOS_CPFS"),
        rows,
        [
            ("cpf", "CPF"),
            ("cpf_formatado", "CPF Formatado"),
            ("situacao", "Situação"),
            ("mensagem", "Mensagem"),
            ("recibo_novo", "Recibo novo"),
            ("recibo_usado", "Recibo usado"),
            ("atualizado_em", "Atualizado em"),
        ],
        BLUE,
        {"A": 14, "B": 18, "C": 26, "D": 80, "E": 30, "F": 30, "G": 22},
    )
    write_rows(
        wb.create_sheet("ENVIADOS_OK"),
        grupos["ENVIADOS_OK"],
        [
            ("cpf", "CPF"),
            ("cpf_formatado", "CPF Formatado"),
            ("recibo_novo", "Recibo novo"),
            ("recibo_usado", "Recibo usado"),
            ("atualizado_em", "Atualizado em"),
        ],
        GREEN,
        {"A": 14, "B": 18, "C": 32, "D": 32, "E": 22},
    )
    write_rows(
        wb.create_sheet("ADVERTENCIAS"),
        grupos["ADVERTENCIAS"],
        [
            ("cpf", "CPF"),
            ("cpf_formatado", "CPF Formatado"),
            ("mensagem", "Advertência"),
            ("codigo", "Código"),
            ("recibo_novo", "Recibo novo"),
            ("recibo_usado", "Recibo usado"),
            ("atualizado_em", "Atualizado em"),
        ],
        YELLOW,
        {"A": 14, "B": 18, "C": 90, "D": 10, "E": 32, "F": 32, "G": 22},
    )
    write_rows(
        wb.create_sheet("ERROS"),
        grupos["ERROS"],
        [
            ("cpf", "CPF"),
            ("cpf_formatado", "CPF Formatado"),
            ("mensagem", "Erro"),
            ("codigo", "Código"),
            ("recibo_usado", "Recibo usado"),
            ("recibo_xml", "Recibo no ZIP"),
            ("atualizado_em", "Atualizado em"),
        ],
        RED,
        {"A": 14, "B": 18, "C": 95, "D": 10, "E": 32, "F": 32, "G": 22},
    )
    write_rows(
        wb.create_sheet("PENDENTES"),
        grupos["PENDENTES"],
        [
            ("cpf", "CPF"),
            ("cpf_formatado", "CPF Formatado"),
            ("mensagem", "Observação"),
            ("atualizado_em", "Atualizado em"),
        ],
        ORANGE,
        {"A": 14, "B": 18, "C": 80, "D": 22},
    )
    write_rows(
        wb.create_sheet("SEM_MUDANCA"),
        grupos["SEM_MUDANCA"],
        [
            ("cpf", "CPF"),
            ("cpf_formatado", "CPF Formatado"),
            ("mensagem", "Observação"),
            ("recibo_usado", "Recibo usado"),
            ("atualizado_em", "Atualizado em"),
        ],
        GRAY,
        {"A": 14, "B": 18, "C": 80, "D": 32, "E": 22},
    )

    wb.save(OUT)
    return OUT, counts


if __name__ == "__main__":
    out, counts = gerar()
    print(f"RELATORIO_GERADO={out}")
    for key in ["ENVIADO_OK", "ENVIADO_COM_ADVERTENCIA", "ERRO_NAO_ENVIADO", "PENDENTE", "SEM_MUDANCA"]:
        print(f"{key}={counts[key]}")
