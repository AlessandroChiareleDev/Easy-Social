"""
Relatorio XLSX - SOLUCOES Agosto/2025 (V2/F5)

Fonte: banco V2/F5, schema `solucoes`.
Nao consulta eSocial. Apenas consolida dados ja gravados no frontend/VPS:
- explorador_eventos: universo de S-1210 do ZIP
- timeline_envio / timeline_envio_item: envios, recibos, erros e pendencias

Saida: relatorio_ana/RELATORIO_SOLUCOES_AGOSTO_2025_<timestamp>.xlsx
"""
from __future__ import annotations

import json
import os
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

SISTEMA_DSN = os.environ.get(
    "SISTEMA_DB_URL",
    "postgresql://postgres:EsoV2_CoxRHWQ1z6iucG7ZyvdqFIbN@db.kjbgiwnlvqnrfdozjvhq.supabase.co:5432/postgres?sslmode=require",
)
SCHEMA = "solucoes"
PER_APUR = "2025-08"
EMPRESA = "SOLUCOES SERVICOS TERCEIRIZADOS LTDA"
CNPJ = "09.445.502/0001-09"

BASE = Path(__file__).resolve().parent
OUTPUT_DIR = BASE.parent / "relatorio_ana"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT = OUTPUT_DIR / f"RELATORIO_SOLUCOES_AGOSTO_2025_{TIMESTAMP}.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FILL_GREEN = PatternFill("solid", fgColor="548235")
HEADER_FILL_RED = PatternFill("solid", fgColor="C00000")
HEADER_FILL_ORANGE = PatternFill("solid", fgColor="C55A11")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="404040")
OK_FILL = PatternFill("solid", fgColor="E2EFDA")
WARN_FILL = PatternFill("solid", fgColor="FFF2CC")
ERR_FILL = PatternFill("solid", fgColor="FCE4EC")
PEND_FILL = PatternFill("solid", fgColor="D9EAF7")
THIN = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)


def connect():
    conn = psycopg2.connect(SISTEMA_DSN)
    with conn.cursor() as cur:
        cur.execute(f'SET search_path TO "{SCHEMA}", public')
    conn.commit()
    return conn


def cpf_fmt(cpf: str | None) -> str:
    s = "".join(ch for ch in str(cpf or "") if ch.isdigit()).zfill(11)
    if len(s) == 11:
        return f"{s[:3]}.{s[3:6]}.{s[6:9]}-{s[9:]}"
    return str(cpf or "")


def dt_fmt(v: Any) -> str:
    if not v:
        return ""
    if hasattr(v, "strftime"):
        return v.strftime("%d/%m/%Y %H:%M:%S")
    return str(v)


def limpar_msg(msg: str | None) -> str:
    if not msg:
        return ""
    return re.sub(r"\s+", " ", str(msg).replace(" | ", "\n")).strip()


def classificar(status: str | None, codigo: str | None, msg: str | None) -> str:
    st = (status or "").lower()
    cd = str(codigo or "").strip()
    text = (msg or "").lower()
    if st == "sucesso":
        return "OK"
    if cd == "202":
        return "ADVERTENCIA_202"
    if st == "sem_mudanca" or cd == "SEM_MUDANCA":
        return "SEM_MUDANCA"
    if st in {"pendente", "pendente_consulta"}:
        return "PENDENTE"
    if cd == "401" and "plano de saúde" in text:
        return "ERRO_401_PLANO_SAUDE"
    if cd == "401" and "recibo" in text:
        return "ERRO_401_RECIBO_INATIVO"
    if cd == "401":
        return "ERRO_401"
    if cd:
        return f"ERRO_{cd}"
    if "erro" in st:
        return "ERRO"
    return (status or "SEM_STATUS").upper()


def style_header(ws, row: int, max_col: int, fill=HEADER_FILL):
    for col in range(1, max_col + 1):
        c = ws.cell(row=row, column=col)
        c.font = HEADER_FONT
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN


def auto_width(ws, min_width=10, max_width=55):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = min_width
        for cell in col[:300]:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)) + 2, max_width))
        ws.column_dimensions[letter].width = width


def add_table(ws, name: str, min_row: int, max_row: int, max_col: int):
    if max_row <= min_row:
        return
    ref = f"A{min_row}:{get_column_letter(max_col)}{max_row}"
    safe = re.sub(r"[^A-Za-z0-9_]", "_", name)[:25]
    tab = Table(displayName=safe, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tab)


def write_rows(ws, headers: list[str], rows: list[dict], fields: list[str], row_start=1, fill=HEADER_FILL):
    for i, h in enumerate(headers, 1):
        ws.cell(row=row_start, column=i, value=h)
    style_header(ws, row_start, len(headers), fill)
    r = row_start + 1
    for item in rows:
        categoria = item.get("categoria") or item.get("status") or ""
        for c, field in enumerate(fields, 1):
            cell = ws.cell(row=r, column=c, value=item.get(field, ""))
            cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=field in {"erro_mensagem", "explicacao_erro", "resumo"})
            if categoria == "OK":
                cell.fill = OK_FILL
            elif "ADVERTENCIA" in categoria or "SEM_MUDANCA" in categoria:
                cell.fill = WARN_FILL
            elif "PENDENTE" in categoria:
                cell.fill = PEND_FILL
            elif "ERRO" in categoria:
                cell.fill = ERR_FILL
        r += 1
    ws.freeze_panes = f"A{row_start + 1}"
    ws.auto_filter.ref = f"A{row_start}:{get_column_letter(len(headers))}{max(row_start, r-1)}"
    auto_width(ws)
    return r - 1


def fetch_data():
    conn = connect()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT COUNT(DISTINCT cpf) AS cpfs, COUNT(*) AS eventos,
                       COUNT(*) FILTER (WHERE nr_recibo IS NOT NULL) AS com_recibo
                  FROM explorador_eventos
                 WHERE per_apur=%s AND tipo_evento='S-1210'
                """,
                (PER_APUR,),
            )
            universo = dict(cur.fetchone())

            cur.execute(
                """
                SELECT te.id AS envio_id, te.sequencia, te.tipo, te.status,
                       te.total_tentados, te.total_sucesso, te.total_erro,
                       te.iniciado_em, te.finalizado_em, te.resumo
                  FROM timeline_envio te
                  JOIN timeline_mes tm ON tm.id=te.timeline_mes_id
                 WHERE tm.per_apur=%s
                 ORDER BY te.id
                """,
                (PER_APUR,),
            )
            lotes = [dict(r) for r in cur.fetchall()]

            cur.execute(
                """
                WITH base AS (
                    SELECT it.*, te.id AS envio_id, te.sequencia, te.tipo AS tipo_envio,
                           tm.per_apur
                      FROM timeline_envio_item it
                      JOIN timeline_envio te ON te.id=it.timeline_envio_id
                      JOIN timeline_mes tm ON tm.id=te.timeline_mes_id
                     WHERE tm.per_apur=%s
                )
                SELECT * FROM base ORDER BY id
                """,
                (PER_APUR,),
            )
            items_raw = [dict(r) for r in cur.fetchall()]

            cur.execute(
                """
                SELECT DISTINCT cpf
                  FROM explorador_eventos
                 WHERE per_apur=%s AND tipo_evento='S-1210'
                   AND cpf IS NOT NULL
                 ORDER BY cpf
                """,
                (PER_APUR,),
            )
            universo_cpfs = [r["cpf"] for r in cur.fetchall()]
    finally:
        conn.close()

    items = []
    for r in items_raw:
        categoria = classificar(r.get("status"), r.get("erro_codigo"), r.get("erro_mensagem"))
        erro_msg = limpar_msg(r.get("erro_mensagem"))
        items.append(
            {
                "envio_item_id": r.get("id"),
                "envio_id": r.get("envio_id"),
                "sequencia": r.get("sequencia"),
                "cpf": r.get("cpf"),
                "cpf_formatado": cpf_fmt(r.get("cpf")),
                "per_apur": r.get("per_apur"),
                "tipo_evento": r.get("tipo_evento"),
                "status": r.get("status"),
                "categoria": categoria,
                "recibo_anterior": r.get("nr_recibo_anterior") or "",
                "recibo_novo": r.get("nr_recibo_novo") or "",
                "erro_codigo": r.get("erro_codigo") or "",
                "erro_mensagem": erro_msg,
                "explicacao_erro": erro_msg,
                "duracao_ms": r.get("duracao_ms"),
                "criado_em": dt_fmt(r.get("criado_em")),
                "versao_anterior_id": r.get("versao_anterior_id") or "",
                "versao_nova_id": r.get("versao_nova_id") or "",
            }
        )

    latest_by_cpf = {}
    for item in items:
        cpf = item.get("cpf")
        if cpf:
            latest_by_cpf[cpf] = item

    pendentes = []
    for cpf in universo_cpfs:
        latest = latest_by_cpf.get(cpf)
        if latest is None:
            pendentes.append(
                {
                    "cpf": cpf,
                    "cpf_formatado": cpf_fmt(cpf),
                    "categoria": "PENDENTE",
                    "status": "sem_envio",
                    "explicacao_erro": "CPF possui S-1210 no ZIP de agosto, mas não consta em timeline_envio_item para o mês.",
                }
            )
        elif latest["categoria"] in {"PENDENTE"}:
            row = dict(latest)
            row["explicacao_erro"] = row.get("erro_mensagem") or "Envio pendente/sem retorno conclusivo."
            pendentes.append(row)

    return universo, lotes, items, pendentes


def gerar():
    universo, lotes, items, pendentes = fetch_data()
    wb = Workbook()

    # RESUMO
    ws = wb.active
    ws.title = "RESUMO"
    ws.sheet_properties.tabColor = "1F4E79"
    ws.merge_cells("A1:F1")
    ws["A1"] = "RELATÓRIO S-1210 — SOLUÇÕES — AGOSTO/2025"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:F2")
    ws["A2"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} | Empresa: {EMPRESA} | CNPJ {CNPJ} | Fonte: V2/F5 schema {SCHEMA}"
    ws["A2"].font = SUBTITLE_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    total_items = len(items)
    latest = {}
    for it in items:
        if it.get("cpf"):
            latest[it["cpf"]] = it
    latest_items = list(latest.values())
    latest_counts = Counter((v["categoria"] for v in latest_items))
    all_counts = Counter((v["categoria"] for v in items))
    erro_items = [i for i in latest_items if i.get("status") == "erro_esocial"]
    erro_historico = [i for i in items if i.get("status") == "erro_esocial"]
    ok_items = [i for i in items if i["categoria"] == "OK"]
    recibos_ok = [i for i in ok_items if i.get("recibo_novo")]

    resumo_rows = [
        ("Universo S-1210 no ZIP (CPFs únicos)", universo.get("cpfs", 0)),
        ("Eventos S-1210 no ZIP", universo.get("eventos", 0)),
        ("Eventos S-1210 com recibo no ZIP", universo.get("com_recibo", 0)),
        ("Tentativas registradas na timeline", total_items),
        ("CPFs com último status OK", latest_counts.get("OK", 0)),
        ("Recibos novos OK", len(recibos_ok)),
        ("CPFs com último status erro/adverência", len(erro_items)),
        ("Linhas históricas de erro/adverência (auditoria)", len(erro_historico)),
        ("CPFs pendentes/sem envio", len(pendentes)),
        ("Lotes/envios registrados", len(lotes)),
        ("Observação", "A aba ERROS usa o último status por CPF, igual ao frontend. Todas as tentativas ficam em TODOS_ENVIOS."),
    ]
    r = 4
    for k, v in resumo_rows:
        ws.cell(r, 1, k).font = Font(bold=True)
        ws.cell(r, 2, v)
        r += 1

    r += 1
    ws.cell(r, 1, "Resumo por categoria (último status por CPF)").font = Font(bold=True, color="1F4E79")
    r += 1
    ws.cell(r, 1, "Categoria")
    ws.cell(r, 2, "CPFs")
    style_header(ws, r, 2)
    for cat, n in latest_counts.most_common():
        r += 1
        ws.cell(r, 1, cat)
        ws.cell(r, 2, n)
    r += 2
    ws.cell(r, 1, "Resumo por categoria (todas as tentativas)").font = Font(bold=True, color="1F4E79")
    r += 1
    ws.cell(r, 1, "Categoria")
    ws.cell(r, 2, "Tentativas")
    style_header(ws, r, 2)
    for cat, n in all_counts.most_common():
        r += 1
        ws.cell(r, 1, cat)
        ws.cell(r, 2, n)
    auto_width(ws)

    fields = [
        "envio_item_id", "envio_id", "sequencia", "cpf", "cpf_formatado", "per_apur", "tipo_evento",
        "status", "categoria", "recibo_anterior", "recibo_novo", "erro_codigo",
        "erro_mensagem", "duracao_ms", "criado_em",
    ]
    headers = [
        "ID Item", "ID Envio", "Seq", "CPF", "CPF Formatado", "Período", "Evento",
        "Status", "Categoria", "Recibo Anterior", "Recibo Novo", "Cód. Erro",
        "Mensagem/Explicação do eSocial", "Duração ms", "Criado em",
    ]

    ws_all = wb.create_sheet("TODOS_ENVIOS")
    max_row = write_rows(ws_all, headers, items, fields)
    add_table(ws_all, "TodosEnvios", 1, max_row, len(headers))

    ws_ok = wb.create_sheet("RECIBOS_OK")
    max_row = write_rows(ws_ok, headers, recibos_ok, fields, fill=HEADER_FILL_GREEN)
    add_table(ws_ok, "RecibosOk", 1, max_row, len(headers))

    ws_err = wb.create_sheet("ERROS")
    err_fields = [
        "cpf", "cpf_formatado", "categoria", "erro_codigo", "explicacao_erro",
        "recibo_anterior", "envio_id", "sequencia", "status", "criado_em", "duracao_ms",
    ]
    err_headers = [
        "CPF", "CPF Formatado", "Categoria", "Código", "Explicação do erro na célula",
        "Recibo usado", "ID Envio", "Seq", "Status bruto", "Criado em", "Duração ms",
    ]
    max_row = write_rows(ws_err, err_headers, erro_items, err_fields, fill=HEADER_FILL_RED)
    add_table(ws_err, "Erros", 1, max_row, len(err_headers))

    ws_pend = wb.create_sheet("PENDENTES")
    pend_fields = ["cpf", "cpf_formatado", "status", "categoria", "explicacao_erro", "envio_id", "sequencia", "criado_em"]
    pend_headers = ["CPF", "CPF Formatado", "Status", "Categoria", "Explicação", "ID Envio", "Seq", "Criado em"]
    max_row = write_rows(ws_pend, pend_headers, pendentes, pend_fields, fill=HEADER_FILL_ORANGE)
    add_table(ws_pend, "Pendentes", 1, max_row, len(pend_headers))

    ws_lotes = wb.create_sheet("LOTES")
    lote_rows = []
    for l in lotes:
        resumo = l.get("resumo")
        if isinstance(resumo, (dict, list)):
            resumo = json.dumps(resumo, ensure_ascii=False)
        lote_rows.append(
            {
                "envio_id": l.get("envio_id"),
                "sequencia": l.get("sequencia"),
                "tipo": l.get("tipo"),
                "status": l.get("status"),
                "total_tentados": l.get("total_tentados"),
                "total_sucesso": l.get("total_sucesso"),
                "total_erro": l.get("total_erro"),
                "iniciado_em": dt_fmt(l.get("iniciado_em")),
                "finalizado_em": dt_fmt(l.get("finalizado_em")),
                "resumo": resumo or "",
            }
        )
    lote_fields = ["envio_id", "sequencia", "tipo", "status", "total_tentados", "total_sucesso", "total_erro", "iniciado_em", "finalizado_em", "resumo"]
    lote_headers = ["ID Envio", "Seq", "Tipo", "Status", "Tentados", "Sucesso", "Erro", "Início", "Fim", "Resumo JSON"]
    max_row = write_rows(ws_lotes, lote_headers, lote_rows, lote_fields)
    add_table(ws_lotes, "Lotes", 1, max_row, len(lote_headers))

    # Ajustar altura das linhas de erro
    for wsx in wb.worksheets:
        if wsx.title == "ERROS":
            for rr in range(2, min(wsx.max_row + 1, 5000)):
                wsx.row_dimensions[rr].height = 48

    wb.save(OUTPUT)
    return OUTPUT, {
        "universo_cpfs": universo.get("cpfs", 0),
        "tentativas": total_items,
        "recibos_ok": len(recibos_ok),
        "erros_advertencias": len(erro_items),
        "erros_historico_linhas": len(erro_historico),
        "pendentes": len(pendentes),
        "lotes": len(lotes),
    }


if __name__ == "__main__":
    out, stats = gerar()
    print("RELATORIO_GERADO=", out)
    print(json.dumps(stats, ensure_ascii=False, indent=2))
