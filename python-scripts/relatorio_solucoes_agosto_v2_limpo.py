"""
Relatorio XLSX limpo - SOLUCOES Agosto/2025 (V2/F5)

Fonte: schema `solucoes` do banco V2/F5.
Regra: uma linha por CPF, usando o ultimo status do frontend.
Nao consulta eSocial.
"""
from __future__ import annotations

import os
import re
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any

import psycopg2
import psycopg2.extras
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

DSN = os.environ.get(
    "SISTEMA_DB_URL",
    "postgresql://postgres:EsoV2_CoxRHWQ1z6iucG7ZyvdqFIbN@db.kjbgiwnlvqnrfdozjvhq.supabase.co:5432/postgres?sslmode=require",
)
SCHEMA = "solucoes"
PER_APUR = "2025-08"
EMPRESA = "SOLUCOES SERVICOS TERCEIRIZADOS LTDA"
CNPJ = "09.445.502/0001-09"

BASE = Path(__file__).resolve().parent
OUT_DIR = BASE.parent / "relatorio_ana"
OUT_DIR.mkdir(parents=True, exist_ok=True)
OUT = OUT_DIR / f"RELATORIO_SOLUCOES_AGOSTO_2025_LIMPO_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

BLUE = "1F4E79"
GREEN = "548235"
YELLOW = "BF9000"
RED = "C00000"
ORANGE = "C55A11"
GRAY = "595959"
LIGHT_GREEN = PatternFill("solid", fgColor="E2EFDA")
LIGHT_YELLOW = PatternFill("solid", fgColor="FFF2CC")
LIGHT_RED = PatternFill("solid", fgColor="FCE4EC")
LIGHT_ORANGE = PatternFill("solid", fgColor="FCE4D6")
LIGHT_GRAY = PatternFill("solid", fgColor="E7E6E6")
WHITE_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Calibri", bold=True, color=BLUE, size=14)
SUB_FONT = Font(name="Calibri", bold=True, color="404040", size=11)
THIN = Border(
    left=Side(style="thin", color="D9E2F3"),
    right=Side(style="thin", color="D9E2F3"),
    top=Side(style="thin", color="D9E2F3"),
    bottom=Side(style="thin", color="D9E2F3"),
)

COLUMNS = [
    ("cpf", "CPF"),
    ("cpf_formatado", "CPF Formatado"),
    ("resultado", "Resultado"),
    ("status_bruto", "Status no V2"),
    ("codigo", "Código"),
    ("recibo_xml", "Recibo no ZIP"),
    ("recibo_usado", "Recibo usado"),
    ("recibo_novo", "Recibo novo"),
    ("mensagem", "Mensagem"),
    ("atualizado_em", "Atualizado em"),
]


def connect():
    conn = psycopg2.connect(DSN)
    with conn.cursor() as cur:
        cur.execute(f'SET search_path TO "{SCHEMA}", public')
    conn.commit()
    return conn


def cpf_fmt(cpf: str | None) -> str:
    digits = "".join(ch for ch in str(cpf or "") if ch.isdigit()).zfill(11)
    if len(digits) == 11:
        return f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return str(cpf or "")


def clean_msg(value: Any) -> str:
    if not value:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def dt_fmt(value: Any) -> str:
    if not value:
        return ""
    if hasattr(value, "strftime"):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    return str(value)


def classificar(status: str | None, codigo: str | None) -> str:
    status = (status or "").lower()
    codigo = str(codigo or "").strip()
    if status == "sucesso":
        return "ENVIADO_OK"
    if status == "erro_esocial" and codigo == "202":
        return "ENVIADO_COM_ADVERTENCIA"
    if status == "erro_esocial":
        return "ERRO_NAO_ENVIADO"
    if status in {"pendente", "pendente_consulta"}:
        return "PENDENTE"
    if status == "sem_mudanca" or codigo == "SEM_MUDANCA":
        return "SEM_MUDANCA"
    return "PENDENTE"


def fetch_rows() -> list[dict]:
    conn = connect()
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cur:
            cur.execute(
                """
                SELECT DISTINCT ON (cpf)
                       cpf, nr_recibo AS recibo_xml, referenciado_recibo,
                       dt_processamento, id AS evento_id
                  FROM explorador_eventos
                 WHERE tipo_evento='S-1210'
                   AND per_apur=%s
                   AND cpf IS NOT NULL
                   AND retificado_por_id IS NULL
                 ORDER BY cpf, dt_processamento DESC NULLS LAST, id DESC
                """,
                (PER_APUR,),
            )
            universo = {r["cpf"]: dict(r) for r in cur.fetchall()}

            cur.execute(
                """
                SELECT DISTINCT ON (it.cpf)
                       it.cpf, it.status, it.erro_codigo, it.erro_mensagem,
                       it.nr_recibo_anterior, it.nr_recibo_novo,
                       it.criado_em, it.id AS item_id, it.timeline_envio_id
                  FROM timeline_envio_item it
                  JOIN timeline_envio te ON te.id = it.timeline_envio_id
                  JOIN timeline_mes tm ON tm.id = te.timeline_mes_id
                 WHERE tm.per_apur=%s
                   AND it.tipo_evento='S-1210'
                   AND it.cpf IS NOT NULL
                 ORDER BY it.cpf, it.criado_em DESC NULLS LAST, it.id DESC
                """,
                (PER_APUR,),
            )
            ultimos = {r["cpf"]: dict(r) for r in cur.fetchall()}
    finally:
        conn.close()

    rows: list[dict] = []
    for cpf in sorted(universo):
        ev = universo[cpf]
        it = ultimos.get(cpf)
        resultado = classificar(it.get("status") if it else None, it.get("erro_codigo") if it else None)
        rows.append(
            {
                "cpf": cpf,
                "cpf_formatado": cpf_fmt(cpf),
                "resultado": resultado,
                "status_bruto": it.get("status") if it else "sem_envio",
                "codigo": it.get("erro_codigo") or "" if it else "",
                "recibo_xml": ev.get("recibo_xml") or "",
                "recibo_usado": it.get("nr_recibo_anterior") or "" if it else "",
                "recibo_novo": it.get("nr_recibo_novo") or "" if it else "",
                "mensagem": clean_msg(it.get("erro_mensagem") if it else "Sem envio registrado na timeline do V2."),
                "atualizado_em": dt_fmt(it.get("criado_em") if it else ev.get("dt_processamento")),
            }
        )
    return rows


def header(ws, row: int, fill_color: str):
    for col in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN


def paint(cell, resultado: str):
    if resultado == "ENVIADO_OK":
        cell.fill = LIGHT_GREEN
    elif resultado == "ENVIADO_COM_ADVERTENCIA":
        cell.fill = LIGHT_YELLOW
    elif resultado == "ERRO_NAO_ENVIADO":
        cell.fill = LIGHT_RED
    elif resultado == "PENDENTE":
        cell.fill = LIGHT_ORANGE
    elif resultado == "SEM_MUDANCA":
        cell.fill = LIGHT_GRAY


def write_sheet(wb: Workbook, title: str, rows: list[dict], fill_color: str):
    ws = wb.create_sheet(title)
    for idx, (_, label) in enumerate(COLUMNS, 1):
        ws.cell(row=1, column=idx, value=label)
    header(ws, 1, fill_color)

    for row_idx, item in enumerate(rows, 2):
        for col_idx, (field, _) in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=item.get(field, ""))
            cell.border = THIN
            cell.alignment = Alignment(vertical="top", wrap_text=field == "mensagem")
            paint(cell, item["resultado"])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{max(1, ws.max_row)}"
    if ws.max_row > 1:
        table = Table(displayName=re.sub(r"[^A-Za-z0-9_]", "_", title)[:25], ref=ws.auto_filter.ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
        ws.add_table(table)
    widths = {
        "A": 14,
        "B": 18,
        "C": 24,
        "D": 18,
        "E": 12,
        "F": 30,
        "G": 30,
        "H": 30,
        "I": 80,
        "J": 22,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    if title in {"ERROS_NAO_ENVIADOS", "ADVERTENCIA_202"}:
        for idx in range(2, min(ws.max_row + 1, 2000)):
            ws.row_dimensions[idx].height = 42
    return ws


def write_resumo(wb: Workbook, rows: list[dict]):
    ws = wb.active
    ws.title = "RESUMO"
    ws.merge_cells("A1:D1")
    ws["A1"] = "RELATÓRIO SOLUÇÕES - AGOSTO/2025"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:D2")
    ws["A2"] = f"{EMPRESA} | CNPJ {CNPJ} | Fonte: V2/F5, schema {SCHEMA}"
    ws["A2"].font = SUB_FONT
    ws["A2"].alignment = Alignment(horizontal="center")

    counts = Counter(r["resultado"] for r in rows)
    resumo = [
        ("Total de CPFs no mês", len(rows)),
        ("Enviados OK", counts["ENVIADO_OK"]),
        ("Enviados com advertência 202", counts["ENVIADO_COM_ADVERTENCIA"]),
        ("Erros não enviados", counts["ERRO_NAO_ENVIADO"]),
        ("Pendentes", counts["PENDENTE"]),
        ("Sem mudança", counts["SEM_MUDANCA"]),
    ]
    ws.cell(4, 1, "Indicador")
    ws.cell(4, 2, "Quantidade")
    for col in range(1, 3):
        ws.cell(4, col).fill = PatternFill("solid", fgColor=BLUE)
        ws.cell(4, col).font = WHITE_FONT
        ws.cell(4, col).alignment = Alignment(horizontal="center")
        ws.cell(4, col).border = THIN
    for idx, (label, value) in enumerate(resumo, 5):
        ws.cell(idx, 1, label).border = THIN
        ws.cell(idx, 2, value).border = THIN
        ws.cell(idx, 1).font = Font(bold=True)
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 16
    return counts


def gerar():
    rows = fetch_rows()
    grupos = {
        "ENVIADOS_OK": [r for r in rows if r["resultado"] == "ENVIADO_OK"],
        "ADVERTENCIA_202": [r for r in rows if r["resultado"] == "ENVIADO_COM_ADVERTENCIA"],
        "ERROS_NAO_ENVIADOS": [r for r in rows if r["resultado"] == "ERRO_NAO_ENVIADO"],
        "PENDENTES": [r for r in rows if r["resultado"] == "PENDENTE"],
        "SEM_MUDANCA": [r for r in rows if r["resultado"] == "SEM_MUDANCA"],
    }

    wb = Workbook()
    counts = write_resumo(wb, rows)
    write_sheet(wb, "TODOS_CPFS", rows, BLUE)
    write_sheet(wb, "ENVIADOS_OK", grupos["ENVIADOS_OK"], GREEN)
    write_sheet(wb, "ADVERTENCIA_202", grupos["ADVERTENCIA_202"], YELLOW)
    write_sheet(wb, "ERROS_NAO_ENVIADOS", grupos["ERROS_NAO_ENVIADOS"], RED)
    write_sheet(wb, "PENDENTES", grupos["PENDENTES"], ORANGE)
    write_sheet(wb, "SEM_MUDANCA", grupos["SEM_MUDANCA"], GRAY)
    wb.save(OUT)
    return OUT, counts


if __name__ == "__main__":
    out, counts = gerar()
    print(f"RELATORIO_GERADO={out}")
    for key in ["ENVIADO_OK", "ENVIADO_COM_ADVERTENCIA", "ERRO_NAO_ENVIADO", "PENDENTE", "SEM_MUDANCA"]:
        print(f"{key}={counts[key]}")
