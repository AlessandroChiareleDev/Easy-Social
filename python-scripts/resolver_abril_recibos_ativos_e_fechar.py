from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any

import psycopg2.extras
from lxml import etree
from openpyxl import load_workbook


ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
BACKEND_V2 = Path(r"C:\Users\xandao\Documents\GitHub\Easy-eSocial-v2\backend")
if str(ROOT / "python-scripts") not in sys.path:
    sys.path.insert(0, str(ROOT / "python-scripts"))
if str(BACKEND_V2) not in sys.path:
    sys.path.insert(0, str(BACKEND_V2))

import corrigir_mes_respostas_jaque_plano_pensao as correcao_base  # noqa: E402
import operacao_20_05_solucoes as op  # noqa: E402
from app import db  # noqa: E402
from app.xml_extractor import extrair_s1210  # noqa: E402
from app.xml_s1210 import S1210XMLGenerator  # noqa: E402


PER_APUR = "2025-04"
CONFIRM_TOKEN = "RESOLVER_ABRIL_RECIBOS_ATIVOS"
INPUT_XLSX = op.OUT_BASE / "ABRIL_RECIBOS" / "2025-04_recibos_ativos_s1210_preencher.xlsx"
OUT_BASE = op.OUT_BASE / "S1210_CORRECOES_RECIBOS_ATIVOS"
OUT_DIR = OUT_BASE / PER_APUR
XML_DIR = OUT_DIR / "xml_unsigned"
RESULT_JSON = OUT_DIR / "resultado_abril_recibos_ativos.json"


def clean(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def cpf11(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits.zfill(11)[-11:] if digits else ""


def setup_correcao_base() -> None:
    correcao_base.OUT_BASE = OUT_BASE
    correcao_base.POLL_TENTATIVAS = 12
    correcao_base.POLL_INTERVALO_S = 8


def manifest_path() -> Path:
    setup_correcao_base()
    return correcao_base.manifest_path(PER_APUR)


def load_receipts() -> dict[str, str]:
    if not INPUT_XLSX.exists():
        raise RuntimeError(f"planilha de recibos nao encontrada: {INPUT_XLSX}")
    workbook = load_workbook(INPUT_XLSX, data_only=True)
    try:
        if "Recibos Abril" not in workbook.sheetnames:
            raise RuntimeError(f"aba 'Recibos Abril' nao encontrada: {workbook.sheetnames}")
        sheet = workbook["Recibos Abril"]
        headers = [clean(sheet.cell(4, col).value) for col in range(1, sheet.max_column + 1)]
        try:
            cpf_idx = headers.index("CPF Normalizado")
            recibo_idx = headers.index("Recibo ativo S-1210 (preencher)")
        except ValueError as exc:
            raise RuntimeError(f"cabecalhos esperados ausentes: {headers}") from exc

        receipts: dict[str, str] = {}
        missing: list[str] = []
        invalid: list[dict[str, str]] = []
        for row in sheet.iter_rows(min_row=5, values_only=True):
            if not row or not row[0]:
                continue
            cpf = cpf11(row[cpf_idx] if cpf_idx < len(row) else "")
            recibo = clean(row[recibo_idx] if recibo_idx < len(row) else "")
            if not cpf:
                continue
            if not recibo:
                missing.append(cpf)
                continue
            if not re.fullmatch(r"1\.1\.\d{19}", recibo):
                invalid.append({"cpf": cpf, "recibo": recibo})
                continue
            receipts[cpf] = recibo
        if missing:
            raise RuntimeError(f"recibos vazios para CPFs: {missing}")
        if invalid:
            raise RuntimeError(f"recibos com formato inesperado: {invalid}")
        if len(receipts) != 10:
            raise RuntimeError(f"esperado 10 recibos preenchidos; encontrado {len(receipts)}")
        return receipts
    finally:
        workbook.close()


def load_plan_map() -> dict[str, list[dict[str, str]]]:
    parsed = op.parse_workbooks()
    plan_map: dict[str, list[dict[str, str]]] = defaultdict(list)
    for item in parsed[PER_APUR]["plano"]:
        plan_map[item["cpf"]].append(
            {
                "cnpjOper": item["cnpj_operadora"],
                "regANS": item["registro_ans"],
                "vlrSaudeTit": item["valor_titular"],
            }
        )
    return dict(plan_map)


def generate_manifest() -> dict[str, Any]:
    setup_correcao_base()
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    XML_DIR.mkdir(parents=True, exist_ok=True)
    for old_xml in XML_DIR.glob("*.xml"):
        old_xml.unlink()

    receipts = load_receipts()
    plan_map = load_plan_map()
    missing_plan = sorted(set(receipts) - set(plan_map))
    if missing_plan:
        raise RuntimeError(f"CPFs com recibo mas sem dados de plano no XLSX resposta final: {missing_plan}")

    current_rows = correcao_base.load_current_rows(PER_APUR, sorted(receipts))
    targets: list[dict[str, Any]] = []
    conn = db.connect(empresa_id=op.EMPRESA_ID)
    try:
        for seq, cpf in enumerate(sorted(receipts), start=1):
            record: dict[str, Any] = {
                "cpf": cpf,
                "generated": False,
                "has_plano": True,
                "has_pensao": False,
                "nr_recibo": receipts[cpf],
                "plan_entries": len(plan_map.get(cpf, [])),
            }
            row = current_rows.get(cpf)
            if not row:
                record["reason"] = "CPF sem XML fonte local"
                targets.append(record)
                continue
            record.update(
                {
                    "evento_id": row.get("id"),
                    "item_id": row.get("item_id"),
                    "envio_id": row.get("envio_id"),
                    "codigo": row.get("erro_codigo"),
                    "status": row.get("item_status"),
                    "erro_mensagem": row.get("erro_mensagem"),
                    "nr_recibo_local_anterior": row.get("nr_recibo") or row.get("nr_recibo_anterior"),
                }
            )
            try:
                xml_old = correcao_base.read_xml_event(conn, row)
                campos = extrair_s1210(xml_old)
                if campos.get("per_apur") != PER_APUR:
                    raise RuntimeError(f"perApur do XML fonte divergente: {campos.get('per_apur')}")
                if cpf11(campos.get("beneficiario", {}).get("cpfBenef")) != cpf:
                    raise RuntimeError("cpfBenef do XML fonte diverge do alvo")
                xml_new = S1210XMLGenerator.gerar(
                    empregador=campos["empregador"],
                    beneficiario=campos["beneficiario"],
                    info_pgtos=campos["info_pgtos"],
                    per_apur=PER_APUR,
                    ind_retif="2",
                    nr_recibo=receipts[cpf],
                    info_ir_complem=campos.get("info_ir_complem"),
                    plan_saude=plan_map[cpf],
                    seq=seq,
                    tp_amb=correcao_base.TP_AMB,
                )
                out_xml = XML_DIR / f"S1210_{PER_APUR}_{cpf}_recibo_ativo_unsigned.xml"
                out_xml.write_bytes(xml_new)
                record.update(
                    {
                        "generated": True,
                        "xml": str(out_xml),
                        "reason": "OK",
                        "info_pgtos": len(campos.get("info_pgtos") or []),
                    }
                )
            except Exception as exc:  # noqa: BLE001
                record["reason"] = f"{type(exc).__name__}: {exc}"
            targets.append(record)
    finally:
        conn.close()

    blocked = [item for item in targets if not item.get("generated")]
    manifest = {
        "empresa_id": op.EMPRESA_ID,
        "per_apur": PER_APUR,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "regra": "retificacao S-1210 abril com recibos ativos preenchidos pelo usuario",
        "input_xlsx": str(INPUT_XLSX),
        "out_dir": str(OUT_DIR),
        "target_cpfs": len(receipts),
        "xmls_generated": sum(1 for item in targets if item.get("generated")),
        "blocked_count": len(blocked),
        "xml_type_counts": {"plano": sum(1 for item in targets if item.get("generated")), "pensao": 0},
        "targets": targets,
    }
    manifest_path().write_text(json.dumps(manifest, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return manifest


def validate_manifest(manifest: dict[str, Any]) -> dict[str, Any]:
    items = []
    receipt_map = load_receipts()
    for item in manifest.get("targets") or []:
        if not item.get("generated"):
            continue
        root = etree.fromstring(Path(item["xml"]).read_bytes())
        items.append(
            {
                "cpf": item["cpf"],
                "indRetif": root.xpath('string(//*[local-name()="indRetif"])'),
                "nrRecibo": root.xpath('string(//*[local-name()="nrRecibo"])'),
                "recibo_esperado": receipt_map.get(item["cpf"]),
                "perApur": root.xpath('string(//*[local-name()="perApur"])'),
                "planSaude": len(root.xpath('//*[local-name()="planSaude"]')),
                "signature": bool(root.xpath('//*[local-name()="Signature"]')),
            }
        )
    wrong = [
        item
        for item in items
        if item["indRetif"] != "2"
        or item["nrRecibo"] != item["recibo_esperado"]
        or item["perApur"] != PER_APUR
        or item["planSaude"] < 1
        or item["signature"]
    ]
    return {"total_validated": len(items), "wrong": wrong, "sample": items[:10]}


def execute() -> dict[str, Any]:
    manifest = generate_manifest()
    validation = validate_manifest(manifest)
    if validation["wrong"]:
        raise RuntimeError(f"validacao falhou: {validation['wrong'][:5]}")
    if manifest["blocked_count"]:
        raise RuntimeError(f"preflight bloqueado: {manifest['blocked_count']}")
    result_envio = correcao_base.execute_s1210(
        manifest,
        correcao_base.DEFAULT_CERT,
        correcao_base.read_password(),
        {"persisted": {"codigo": "LOCAL_OPEN", "nr_recibo": None, "origem": "abril_recibos_ativos"}},
    )
    after = op.audit_counts([PER_APUR])[PER_APUR]
    closing = None
    if after["erros"] == 0:
        closing = op.close_month(PER_APUR, True)
    output = {
        "manifest": str(manifest_path()),
        "validation": validation,
        "execute": result_envio,
        "after": after,
        "closing": closing,
    }
    RESULT_JSON.write_text(json.dumps(output, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return output


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--confirmar", default="")
    args = parser.parse_args()
    if not args.execute:
        manifest = generate_manifest()
        print(json.dumps({"manifest": manifest, "validation": validate_manifest(manifest)}, ensure_ascii=False, indent=2, default=str))
        return 0
    if args.confirmar != CONFIRM_TOKEN:
        raise SystemExit(f"Para executar, use --confirmar {CONFIRM_TOKEN}")
    print(json.dumps(execute(), ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())