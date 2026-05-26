from __future__ import annotations

import argparse
import json
import re
import sys
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from lxml import etree
from openpyxl import load_workbook


ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
BACKEND_V2 = Path(r"C:\Users\xandao\Documents\GitHub\Easy-eSocial-v2\backend")
if str(ROOT / "python-scripts") not in sys.path:
    sys.path.insert(0, str(ROOT / "python-scripts"))
if str(BACKEND_V2) not in sys.path:
    sys.path.insert(0, str(BACKEND_V2))

import corrigir_mes_respostas_jaque_plano_pensao as correcao_base  # noqa: E402
import enviar_correcao_agosto_jaque as envio_base  # noqa: E402
from retry_mes_pensao_infodep_1861 import add_info_dep  # noqa: E402
from app import db, esocial_client  # noqa: E402
from app.xml_diff import eventos_iguais  # noqa: E402
from app.xml_extractor import extrair_s1210  # noqa: E402
from app.xml_s1210 import S1210XMLGenerator  # noqa: E402
from app.xml_signer import S1010XMLSigner  # noqa: E402


EMPRESA_ID = 2
PER_APUR = "2025-03"
CNPJ = "09445502000109"
TP_AMB = "1"
CONFIRM_TOKEN = "RETRY_MARCO_PENSAO_TPREND_11"
RESPONSE_XLSX = ROOT / "relatorio_ana" / "RESPOSTAS_JAQUE_PLANO_PENSAO_2025_FEV_OUT" / "SOLUCOES_MARCO_2025_JAQUE_PLANO_PENSAO_PREENCHER.xlsx"
TARGET_CPFS = ["82541787472", "94881898434"]
OUT_DIR = ROOT / "relatorio_ana" / "CORRECAO_JAQUE_PLANO_PENSAO_2025" / PER_APUR / "retry_pensao_tipo_rend_11"
XML_DIR = OUT_DIR / "xml_unsigned"
MANIFEST = OUT_DIR / "manifest_retry_pensao_tipo_rend_11.json"


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def cpf11(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    return digits.zfill(11)[-11:] if digits else ""


def money(value: Any) -> Decimal:
    if value is None or clean_text(value) == "":
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, int):
        return Decimal(value).quantize(Decimal("0.01"))
    if isinstance(value, float):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    raw = clean_text(value).replace("R$", "").replace(" ", "")
    raw = re.sub(r"[^0-9,.-]", "", raw)
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif "," in raw:
        raw = raw.replace(",", ".")
    try:
        return Decimal(raw).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


def money_str(value: Any) -> str:
    return f"{money(value):.2f}"


def load_pension_rows() -> dict[str, dict[str, Any]]:
    workbook = load_workbook(RESPONSE_XLSX, data_only=True, read_only=True)
    try:
        output: dict[str, dict[str, Any]] = {}
        for ws in workbook.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header_index = None
            for index, row in enumerate(rows[:20]):
                normalized = [clean_text(value).lower() for value in row]
                if "cpf beneficiario 1" in normalized and "valor deduzido 1" in normalized:
                    header_index = index
                    break
            if header_index is None:
                continue
            headers = [clean_text(value) or f"col_{index + 1}" for index, value in enumerate(rows[header_index])]
            for row in rows[header_index + 1:]:
                data = {headers[index]: row[index] if index < len(row) else None for index in range(len(headers))}
                cpf = cpf11(data.get("CPF Normalizado") or data.get("CPF"))
                if cpf not in TARGET_CPFS:
                    continue
                cpf_dependente = cpf11(data.get("CPF Beneficiario 1"))
                valor = money_str(data.get("Valor Deduzido 1"))
                if not cpf_dependente or money(valor) <= 0:
                    continue
                output[cpf] = {
                    "cpf": cpf,
                    "nome": clean_text(data.get("Nome Trabalhador")),
                    "cpf_dependente": cpf_dependente,
                    "tpRend": "11",
                    "vlrDedPenAlim": valor,
                    "linha_resposta": data,
                }
        return output
    finally:
        workbook.close()


def validate_current_errors(cpfs: list[str]) -> dict[str, dict[str, Any]]:
    latest = correcao_base.latest_status_summary(PER_APUR, cpfs)
    output: dict[str, dict[str, Any]] = {}
    for row in latest.get("rows") or []:
        message = str(row.get("erro_mensagem") or "")
        if row.get("status") == "erro_esocial" and str(row.get("erro_codigo") or "") == "401" and "pensão alimentícia" in message:
            output[row["cpf"]] = row
    return output


def generate_manifest() -> dict[str, Any]:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    XML_DIR.mkdir(parents=True, exist_ok=True)
    for old_xml in XML_DIR.glob("*.xml"):
        old_xml.unlink()

    pension_rows = load_pension_rows()
    current_errors = validate_current_errors(TARGET_CPFS)
    current_rows = correcao_base.load_current_rows(PER_APUR, sorted(pension_rows))
    targets: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        for index, cpf in enumerate(TARGET_CPFS, start=1):
            pension = pension_rows.get(cpf)
            if not pension:
                skipped.append({"cpf": cpf, "reason": "sem linha de pensao com beneficiario e valor"})
                continue
            if cpf not in current_errors:
                skipped.append({"cpf": cpf, "reason": "ultimo status nao e erro atual de pensao"})
                continue
            row = current_rows.get(cpf)
            if not row:
                skipped.append({"cpf": cpf, "reason": "sem S-1210 ativo local com XML"})
                continue
            try:
                xml_old = correcao_base.read_xml_event(conn, row)
                campos = extrair_s1210(xml_old)
                info_ir, warning = correcao_base.merge_pensao(
                    campos.get("info_ir_complem"),
                    [{"tpRend": pension["tpRend"], "cpfDep": pension["cpf_dependente"], "vlrDedPenAlim": pension["vlrDedPenAlim"]}],
                )
                if warning:
                    skipped.append({"cpf": cpf, "reason": warning})
                    continue
                nr_recibo = row.get("nr_recibo") or campos.get("nr_recibo_atual")
                if not nr_recibo:
                    skipped.append({"cpf": cpf, "reason": "sem nrRecibo ativo"})
                    continue
                xml_new = S1210XMLGenerator.gerar(
                    empregador=campos["empregador"],
                    beneficiario=campos["beneficiario"],
                    info_pgtos=campos["info_pgtos"],
                    per_apur=PER_APUR,
                    ind_retif="2",
                    nr_recibo=nr_recibo,
                    info_ir_complem=info_ir,
                    plan_saude=campos.get("plan_saude"),
                    seq=index,
                    tp_amb=TP_AMB,
                )
                xml_new = add_info_dep(xml_new, PER_APUR, pension["cpf_dependente"])
                if eventos_iguais(xml_old, xml_new):
                    skipped.append({"cpf": cpf, "reason": "XML novo ficou identico ao atual"})
                    continue
                out_xml = XML_DIR / f"S1210_{PER_APUR}_{cpf}_pensao_tprend11_unsigned.xml"
                out_xml.write_bytes(xml_new)
                targets.append(
                    {
                        "cpf": cpf,
                        "cpf_dependente": pension["cpf_dependente"],
                        "tpRend": pension["tpRend"],
                        "vlrDedPenAlim": pension["vlrDedPenAlim"],
                        "xml": str(out_xml),
                        "evento_id": row.get("id"),
                        "nr_recibo": nr_recibo,
                        "has_pensao": True,
                    }
                )
            except Exception as exc:
                conn.rollback()
                skipped.append({"cpf": cpf, "reason": f"{type(exc).__name__}: {exc}"})
    finally:
        conn.close()

    manifest = {
        "empresa_id": EMPRESA_ID,
        "per_apur": PER_APUR,
        "regra": "pensão março com Tipo Rendimento em branco: usar tpRend=11 mensal e inserir infoDep",
        "resposta": str(RESPONSE_XLSX),
        "total_retry": len(targets),
        "total_skipped": len(skipped),
        "targets": targets,
        "skipped": skipped,
    }
    MANIFEST.write_text(json.dumps(manifest, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return manifest


def validate_manifest(manifest: dict[str, Any]) -> dict[str, Any]:
    rows = []
    for item in manifest.get("targets") or []:
        xml_bytes = Path(item["xml"]).read_bytes()
        root = etree.fromstring(xml_bytes)
        rows.append(
            {
                "cpf": item["cpf"],
                "id": esocial_client._extrair_id(xml_bytes),
                "perApur": root.xpath('string(//*[local-name()="perApur"])'),
                "penAlimCpf": root.xpath('string(//*[local-name()="penAlim"]/*[local-name()="cpfDep"])'),
                "tpRend": root.xpath('string(//*[local-name()="penAlim"]/*[local-name()="tpRend"])'),
                "vlrDedPenAlim": root.xpath('string(//*[local-name()="penAlim"]/*[local-name()="vlrDedPenAlim"])'),
                "infoDep": [node.xpath('string(./*[local-name()="cpfDep"])') for node in root.xpath('//*[local-name()="infoDep"]')],
                "signature": bool(root.xpath('//*[local-name()="Signature"]')),
            }
        )
    wrong = [
        row for row in rows
        if row["perApur"] != PER_APUR or row["tpRend"] != "11" or row["penAlimCpf"] not in row["infoDep"] or row["signature"]
    ]
    return {"total": len(rows), "wrong": wrong, "sample": rows[:10]}


def sign_targets(targets: list[dict[str, Any]], senha: str) -> list[dict[str, Any]]:
    pfx_data = correcao_base.DEFAULT_CERT.read_bytes()
    signed: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for item in targets:
        unsigned_xml = Path(item["xml"]).read_bytes()
        xml_assinado = S1010XMLSigner.assinar(unsigned_xml, pfx_data, senha)
        signed_id = esocial_client._extrair_id(xml_assinado)
        if not signed_id:
            raise RuntimeError(f"Id assinado ausente para {item['cpf']}")
        if signed_id in seen_ids:
            raise RuntimeError(f"Id duplicado {signed_id}")
        seen_ids.add(signed_id)
        signed.append({**item, "xml_assinado": xml_assinado, "id_evento_assinado": signed_id})
    return signed


def execute() -> dict[str, Any]:
    manifest = generate_manifest()
    validation = validate_manifest(manifest)
    if validation["wrong"]:
        raise RuntimeError(f"validacao falhou: {validation['wrong'][:5]}")
    if not manifest["targets"]:
        raise RuntimeError("nenhum alvo para retry")
    senha = correcao_base.read_password()
    signed = sign_targets(manifest["targets"], senha)

    conn_db = db.connect(empresa_id=EMPRESA_ID)
    conn_w = db.connect(empresa_id=EMPRESA_ID)
    try:
        envio_base.PER_APUR = PER_APUR
        envio_base.PREFLIGHT = MANIFEST
        envio_base.CFG_GRUPO = correcao_base.GRUPO
        envio_base.POLL_TENTATIVAS = correcao_base.POLL_TENTATIVAS
        envio_base.POLL_INTERVALO_S = correcao_base.POLL_INTERVALO_S
        envio_base._verificar_estado_atual(conn_db, signed)
        envio_id, mes_id = correcao_base._create_timeline_envio(conn_db, PER_APUR, len(signed), MANIFEST)
        print(f"=> retry pensao tpRend=11 {PER_APUR}: envio_id={envio_id} timeline_mes={mes_id} targets={len(signed)}")
        item_ids = envio_base._criar_items(conn_db, envio_id, signed)
        envio_base._persistir_xmls_assinados(conn_db, conn_w, signed, item_ids)

        resultado = envio_base._processar_lote(
            signed,
            item_ids,
            cert_path=correcao_base.DEFAULT_CERT,
            senha=senha,
            cnpj=CNPJ,
            conn_db=conn_db,
            conn_w=conn_w,
        )
        sucesso = int(resultado["sucesso"])
        erro = int(resultado["erro"])
        histograma = dict(resultado.get("histograma") or {})
        protocolo = resultado.get("protocolo")
        envio_base._atualizar_envio(
            conn_db,
            envio_id,
            status="concluido",
            sucesso=sucesso,
            erro=erro,
            resumo_extra={
                "rotulo_final": "retry_marco_pensao_tipo_rend_11",
                "manifest": str(MANIFEST),
                "protocolo": protocolo,
                "histograma_erros": histograma,
            },
        )
        latest = correcao_base.latest_status_summary(PER_APUR, [item["cpf"] for item in manifest["targets"]])
        result = {
            "per_apur": PER_APUR,
            "envio_id": envio_id,
            "sucesso": sucesso,
            "erro": erro,
            "protocolos": [str(protocolo)] if protocolo else [],
            "histograma": histograma,
            "manifest": str(MANIFEST),
            "latest": latest,
        }
        (OUT_DIR / "resultado_retry_pensao_tipo_rend_11.json").write_text(json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        return result
    finally:
        conn_db.close()
        conn_w.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--confirmar", default="")
    args = parser.parse_args()
    if not args.execute:
        manifest = generate_manifest()
        print(json.dumps({"manifest": manifest, "validation": validate_manifest(manifest)}, ensure_ascii=False, indent=2, default=str))
        return 0
    if args.confirmar != CONFIRM_TOKEN:
        raise SystemExit(f"Para executar, use --confirmar {CONFIRM_TOKEN}")
    print(json.dumps(execute(), ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())