from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

import psycopg2.extras
from lxml import etree
from openpyxl import load_workbook


ROOT = Path(r"C:\Users\xandao\Documents\GitHub\Easy-Social")
BACKEND_V2 = Path(r"C:\Users\xandao\Documents\GitHub\Easy-eSocial-v2\backend")
if str(ROOT / "python-scripts") not in sys.path:
    sys.path.insert(0, str(ROOT / "python-scripts"))
if str(BACKEND_V2) not in sys.path:
    sys.path.insert(0, str(BACKEND_V2))

import corrigir_mes_respostas_jaque_plano_pensao as correcao_base  # noqa: E402
import enviar_correcao_agosto_jaque as envio_base  # noqa: E402
from app import db, esocial_client, tenant  # noqa: E402
from app.xml_s1210 import NS as S1210_NS, _gerar_id  # noqa: E402
from app.xml_signer import S1010XMLSigner  # noqa: E402


EMPRESA_ID = 2
CNPJ = "09445502000109"
MONTH_NAMES = {
    "01": "JANEIRO",
    "02": "FEVEREIRO",
    "03": "MARCO",
    "04": "ABRIL",
    "05": "MAIO",
    "06": "JUNHO",
    "07": "JULHO",
    "08": "AGOSTO",
    "09": "SETEMBRO",
    "10": "OUTUBRO",
    "11": "NOVEMBRO",
    "12": "DEZEMBRO",
}


def qname(tag: str) -> str:
    return f"{{{S1210_NS}}}{tag}"


def text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def digits(value: Any) -> str:
    return re.sub(r"\D", "", text(value))


def cpf11(value: Any) -> str:
    only_digits = digits(value)
    return only_digits.zfill(11)[-11:] if only_digits else ""


def cnpj14(value: Any) -> str:
    only_digits = digits(value)
    return only_digits.zfill(14)[-14:] if only_digits else ""


def money(value: Any) -> Decimal:
    if value is None or text(value) == "":
        return Decimal("0.00")
    if isinstance(value, Decimal):
        return value.quantize(Decimal("0.01"))
    if isinstance(value, int):
        return Decimal(value).quantize(Decimal("0.01"))
    if isinstance(value, float):
        return Decimal(str(value)).quantize(Decimal("0.01"))
    raw = text(value).replace("R$", "").replace(" ", "")
    raw = re.sub(r"[^0-9,.-]", "", raw)
    if "," in raw and "." in raw:
        raw = raw.replace(".", "").replace(",", ".")
    elif "," in raw:
        raw = raw.replace(",", ".")
    try:
        return Decimal(raw).quantize(Decimal("0.01"))
    except InvalidOperation:
        return Decimal("0.00")


def money_str(value: Decimal) -> str:
    return f"{value.quantize(Decimal('0.01')):.2f}"


def confirm_token(per_apur: str) -> str:
    return f"RETRY_{per_apur.replace('-', '_')}_PLANSAUDE_1867"


def month_dir(per_apur: str) -> Path:
    return ROOT / "relatorio_ana" / "CORRECAO_JAQUE_PLANO_PENSAO_2025" / per_apur


def retry_dir(per_apur: str) -> Path:
    return month_dir(per_apur) / "retry_plansaude_1867"


def xml_dir(per_apur: str) -> Path:
    return retry_dir(per_apur) / "xml_unsigned"


def manifest_path(per_apur: str) -> Path:
    return retry_dir(per_apur) / "manifest_retry_plansaude_1867.json"


def response_workbook(per_apur: str) -> Path:
    month = MONTH_NAMES[per_apur[-2:]]
    path = ROOT / "relatorio_ana" / "RESPOSTAS_JAQUE_PLANO_PENSAO_2025_FEV_OUT" / f"SOLUCOES_{month}_2025_JAQUE_PLANO_PENSAO_PREENCHER.xlsx"
    if not path.exists():
        raise RuntimeError(f"planilha resposta nao encontrada: {path}")
    return path


def current_1867_rows(per_apur: str) -> list[dict[str, Any]]:
    internal_empresa_id = tenant.internal_empresa_id(EMPRESA_ID)
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        with conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor) as cursor:
            cursor.execute(
                """
                WITH latest AS (
                    SELECT DISTINCT ON (it.cpf)
                           it.id AS item_id, te.id AS envio_id, it.cpf, it.status,
                           it.erro_codigo, it.erro_mensagem, it.nr_recibo_anterior,
                           it.nr_recibo_novo, it.versao_anterior_id, it.xml_enviado_oid,
                           it.criado_em
                      FROM timeline_envio_item it
                      JOIN timeline_envio te ON te.id = it.timeline_envio_id
                      JOIN timeline_mes tm ON tm.id = te.timeline_mes_id
                     WHERE tm.empresa_id = %s
                       AND tm.per_apur = %s
                       AND it.tipo_evento = 'S-1210'
                     ORDER BY it.cpf, it.criado_em DESC NULLS LAST, it.id DESC
                )
                SELECT * FROM latest
                 WHERE status = 'erro_esocial'
                   AND erro_codigo = '401'
                   AND erro_mensagem LIKE '%%1867%%'
                 ORDER BY cpf
                """,
                (internal_empresa_id, per_apur),
            )
            return [dict(row) for row in cursor.fetchall()]
    finally:
        conn.close()


def read_large_object(conn, oid: int) -> bytes:
    with conn.cursor() as cursor:
        cursor.execute("SELECT lo_get(%s)", (oid,))
        row = cursor.fetchone()
    if not row or row[0] is None:
        raise RuntimeError(f"large object nao encontrado: {oid}")
    return bytes(row[0])


def load_raw_plan_entries(per_apur: str, cpfs: set[str]) -> dict[str, list[dict[str, str]]]:
    totals: dict[str, dict[tuple[str, str], Decimal]] = defaultdict(lambda: defaultdict(lambda: Decimal("0.00")))
    workbook = load_workbook(response_workbook(per_apur), data_only=True, read_only=True)
    try:
        for ws in workbook.worksheets:
            if ws.title.strip().lower() != "plano saude - ok":
                continue
            for row in ws.iter_rows(values_only=True):
                cpf = cpf11(row[1] if len(row) > 1 else None)
                if cpf not in cpfs:
                    continue
                cnpj = cnpj14(row[4] if len(row) > 4 else None)
                reg_ans = digits(row[5] if len(row) > 5 else None)
                valor = money(row[6] if len(row) > 6 else None)
                if not cnpj or not reg_ans or valor <= 0:
                    continue
                totals[cpf][(cnpj, reg_ans)] += valor
    finally:
        workbook.close()

    output: dict[str, list[dict[str, str]]] = {}
    for cpf, by_operadora in totals.items():
        output[cpf] = [
            {"cnpjOper": cnpj, "regANS": reg_ans, "vlrSaudeTit": money_str(valor)}
            for (cnpj, reg_ans), valor in sorted(by_operadora.items())
            if valor > 0
        ]
    return output


def direct_child(parent: etree._Element, tag: str) -> etree._Element | None:
    found = parent.xpath(f'./*[local-name()="{tag}"]')
    return found[0] if found else None


def set_child(parent: etree._Element, tag: str, value: str, after_tag: str | None = None) -> None:
    found = direct_child(parent, tag)
    if found is not None:
        found.text = value
        return
    node = etree.Element(qname(tag))
    node.text = value
    insert_at = len(parent)
    if after_tag:
        for index, child in enumerate(parent):
            if etree.QName(child).localname == after_tag:
                insert_at = index + 1
                break
    parent.insert(insert_at, node)


def sub(parent: etree._Element, tag: str, value: str) -> None:
    node = etree.SubElement(parent, qname(tag))
    node.text = value


def inner_s1210(xml_bytes: bytes) -> etree._Element:
    parser = etree.XMLParser(remove_blank_text=True)
    root = etree.fromstring(xml_bytes, parser=parser)
    candidates = root.xpath('//*[local-name()="eSocial" and ./*[local-name()="evtPgtos"]]')
    if candidates:
        root = candidates[0]
    inner = etree.fromstring(etree.tostring(root), parser=parser)
    for signature in inner.xpath('.//*[local-name()="Signature"]'):
        parent = signature.getparent()
        if parent is not None:
            parent.remove(signature)
    return inner


def ensure_info_ir(inner: etree._Element) -> etree._Element:
    ide_benef_nodes = inner.xpath('//*[local-name()="evtPgtos"]/*[local-name()="ideBenef"]')
    if not ide_benef_nodes:
        raise RuntimeError("ideBenef nao encontrado")
    ide_benef = ide_benef_nodes[0]
    info_ir = direct_child(ide_benef, "infoIRComplem")
    if info_ir is None:
        info_ir = etree.Element(qname("infoIRComplem"))
        ide_benef.append(info_ir)
    return info_ir


def replace_plan_saude(info_ir: etree._Element, plan_entries: list[dict[str, str]]) -> None:
    for existing in list(info_ir.xpath('./*[local-name()="planSaude"]')):
        info_ir.remove(existing)
    for plan in plan_entries:
        plan_node = etree.Element(qname("planSaude"))
        sub(plan_node, "cnpjOper", plan["cnpjOper"])
        sub(plan_node, "regANS", plan["regANS"])
        sub(plan_node, "vlrSaudeTit", plan["vlrSaudeTit"])
        info_ir.append(plan_node)


def prepare_xml(xml_bytes: bytes, per_apur: str, receipt: str, plan_entries: list[dict[str, str]]) -> tuple[bytes, dict[str, Any]]:
    inner = inner_s1210(xml_bytes)
    event_nodes = inner.xpath('./*[local-name()="evtPgtos"]')
    if not event_nodes:
        raise RuntimeError("evtPgtos ausente")
    event = event_nodes[0]
    old_id = event.get("Id") or ""
    tp_insc = int(inner.xpath('string(//*[local-name()="ideEmpregador"]/*[local-name()="tpInsc"])') or "1")
    nr_insc = inner.xpath('string(//*[local-name()="ideEmpregador"]/*[local-name()="nrInsc"])')
    event.set("Id", _gerar_id(tp_insc, nr_insc))
    ide_evento = event.xpath('./*[local-name()="ideEvento"]')[0]
    set_child(ide_evento, "indRetif", "2")
    set_child(ide_evento, "nrRecibo", receipt, after_tag="indRetif")
    info_ir = ensure_info_ir(inner)
    replace_plan_saude(info_ir, plan_entries)

    if inner.xpath('string(//*[local-name()="perApur"])') != per_apur:
        raise RuntimeError("perApur divergente")
    if inner.xpath('.//*[local-name()="Signature"]'):
        raise RuntimeError("Signature antiga permaneceu")
    if any(money(node.xpath('string(.)')) <= 0 for node in inner.xpath('//*[local-name()="planSaude"]/*[local-name()="vlrSaudeTit"]')):
        raise RuntimeError("vlrSaudeTit zerado permaneceu")
    output = etree.tostring(inner, xml_declaration=True, encoding="UTF-8", pretty_print=False)
    return output, {"source_id": old_id, "new_id": event.get("Id"), "plan_entries": plan_entries}


def generate_manifest(per_apur: str) -> dict[str, Any]:
    retry_dir(per_apur).mkdir(parents=True, exist_ok=True)
    xml_dir(per_apur).mkdir(parents=True, exist_ok=True)
    for old_xml in xml_dir(per_apur).glob("*.xml"):
        old_xml.unlink()

    rows = current_1867_rows(per_apur)
    raw_plans = load_raw_plan_entries(per_apur, {cpf11(row.get("cpf")) for row in rows})
    targets: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    conn = db.connect(empresa_id=EMPRESA_ID)
    try:
        for row in rows:
            cpf = cpf11(row.get("cpf"))
            plan_entries = raw_plans.get(cpf) or []
            if not plan_entries:
                skipped.append({**row, "reason": "sem planSaude positivo na aba plano saude - ok"})
                continue
            if not row.get("xml_enviado_oid"):
                skipped.append({**row, "reason": "sem xml_enviado_oid no ultimo erro"})
                continue
            if not row.get("nr_recibo_anterior"):
                skipped.append({**row, "reason": "sem nr_recibo_anterior"})
                continue
            try:
                source_xml = read_large_object(conn, int(row["xml_enviado_oid"]))
                xml_out, meta = prepare_xml(source_xml, per_apur, str(row["nr_recibo_anterior"]), plan_entries)
            except Exception as exc:
                conn.rollback()
                skipped.append({**row, "reason": f"{type(exc).__name__}: {exc}"})
                continue
            out_xml = xml_dir(per_apur) / f"S1210_{per_apur}_{cpf}_plansaude_1867_unsigned.xml"
            out_xml.write_bytes(xml_out)
            targets.append(
                {
                    "cpf": cpf,
                    "xml": str(out_xml),
                    "evento_id": row.get("versao_anterior_id"),
                    "nr_recibo": str(row["nr_recibo_anterior"]),
                    "id_evento": meta["new_id"],
                    "latest_item_id": row.get("item_id"),
                    "latest_envio_id": row.get("envio_id"),
                    "plan_entries": plan_entries,
                    "validation": meta,
                }
            )
    finally:
        conn.close()

    manifest = {
        "empresa_id": EMPRESA_ID,
        "per_apur": per_apur,
        "regra": "retry 401/1867 de planSaude: substituir planSaude zerado por valores positivos da aba plano saude - ok",
        "workbook": str(response_workbook(per_apur)),
        "total_current_1867": len(rows),
        "total_retry": len(targets),
        "total_skipped": len(skipped),
        "targets": targets,
        "skipped": skipped,
    }
    manifest_path(per_apur).write_text(json.dumps(manifest, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    return manifest


def validate_manifest(manifest: dict[str, Any]) -> dict[str, Any]:
    rows = []
    for item in manifest.get("targets") or []:
        xml_bytes = Path(item["xml"]).read_bytes()
        root = etree.fromstring(xml_bytes)
        values = [node.xpath('string(.)') for node in root.xpath('//*[local-name()="planSaude"]/*[local-name()="vlrSaudeTit"]')]
        rows.append(
            {
                "cpf": item["cpf"],
                "id": esocial_client._extrair_id(xml_bytes),
                "perApur": root.xpath('string(//*[local-name()="perApur"])'),
                "nrRecibo": root.xpath('string(//*[local-name()="nrRecibo"])'),
                "planSaude": len(root.xpath('//*[local-name()="planSaude"]')),
                "vlrSaudeTit": values,
                "signature": bool(root.xpath('//*[local-name()="Signature"]')),
            }
        )
    wrong = [
        row for row in rows
        if row["perApur"] != manifest["per_apur"] or row["signature"] or not row["planSaude"] or any(money(value) <= 0 for value in row["vlrSaudeTit"])
    ]
    return {"total": len(rows), "wrong": wrong, "sample": rows[:10]}


def sign_targets(targets: list[dict[str, Any]], senha: str) -> list[dict[str, Any]]:
    pfx_data = correcao_base.DEFAULT_CERT.read_bytes()
    signed: list[dict[str, Any]] = []
    seen_ids: set[str] = set()
    for item in targets:
        unsigned_xml = Path(item["xml"]).read_bytes()
        xml_assinado = S1010XMLSigner.assinar(unsigned_xml, pfx_data, senha)
        signed_id = esocial_client._extrair_id(xml_assinado)
        if not signed_id:
            raise RuntimeError(f"Id assinado ausente para {item['cpf']}")
        if signed_id in seen_ids:
            raise RuntimeError(f"Id duplicado {signed_id}")
        seen_ids.add(signed_id)
        signed.append({**item, "xml_assinado": xml_assinado, "id_evento_assinado": signed_id})
    return signed


def execute(per_apur: str) -> dict[str, Any]:
    manifest = generate_manifest(per_apur)
    validation = validate_manifest(manifest)
    if validation["wrong"]:
        raise RuntimeError(f"validacao falhou: {validation['wrong'][:5]}")
    if not manifest["targets"]:
        raise RuntimeError("nenhum alvo para retry")
    senha = correcao_base.read_password()
    signed = sign_targets(manifest["targets"], senha)

    conn_db = db.connect(empresa_id=EMPRESA_ID)
    conn_w = db.connect(empresa_id=EMPRESA_ID)
    try:
        envio_base.PER_APUR = per_apur
        envio_base.PREFLIGHT = manifest_path(per_apur)
        envio_base.CFG_GRUPO = correcao_base.GRUPO
        envio_base.POLL_TENTATIVAS = correcao_base.POLL_TENTATIVAS
        envio_base.POLL_INTERVALO_S = correcao_base.POLL_INTERVALO_S
        envio_base._verificar_estado_atual(conn_db, signed)
        envio_id, mes_id = correcao_base._create_timeline_envio(conn_db, per_apur, len(signed), manifest_path(per_apur))
        print(f"=> retry planSaude 1867 {per_apur}: envio_id={envio_id} timeline_mes={mes_id} targets={len(signed)}")
        item_ids = envio_base._criar_items(conn_db, envio_id, signed)
        envio_base._persistir_xmls_assinados(conn_db, conn_w, signed, item_ids)

        sucesso_total = 0
        erro_total = 0
        protocolos: list[str] = []
        histograma: dict[str, int] = {}
        for index in range(0, len(signed), correcao_base.CFG_LOTE_MAX):
            lote = signed[index:index + correcao_base.CFG_LOTE_MAX]
            resultado = envio_base._processar_lote(
                lote,
                item_ids,
                cert_path=correcao_base.DEFAULT_CERT,
                senha=senha,
                cnpj=CNPJ,
                conn_db=conn_db,
                conn_w=conn_w,
            )
            sucesso_total += int(resultado["sucesso"])
            erro_total += int(resultado["erro"])
            if resultado.get("protocolo"):
                protocolos.append(str(resultado["protocolo"]))
            for codigo, total in (resultado.get("histograma") or {}).items():
                histograma[codigo] = histograma.get(codigo, 0) + int(total)

        envio_base._atualizar_envio(
            conn_db,
            envio_id,
            status="concluido",
            sucesso=sucesso_total,
            erro=erro_total,
            resumo_extra={
                "rotulo_final": "retry_mes_plansaude_1867_raw_ok",
                "per_apur": per_apur,
                "manifest": str(manifest_path(per_apur)),
                "protocolos": protocolos,
                "histograma_erros": histograma,
            },
        )
        target_cpfs = [item["cpf"] for item in manifest["targets"]]
        latest = correcao_base.latest_status_summary(per_apur, target_cpfs)
        result = {
            "per_apur": per_apur,
            "envio_id": envio_id,
            "sucesso": sucesso_total,
            "erro": erro_total,
            "protocolos": protocolos,
            "histograma": histograma,
            "manifest": str(manifest_path(per_apur)),
            "latest": latest,
        }
        (retry_dir(per_apur) / "resultado_retry_plansaude_1867.json").write_text(
            json.dumps(result, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        return result
    finally:
        conn_db.close()
        conn_w.close()


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--per-apur", required=True)
    parser.add_argument("--execute", action="store_true")
    parser.add_argument("--confirmar", default="")
    args = parser.parse_args()

    if not re.fullmatch(r"20\d{2}-\d{2}", args.per_apur):
        raise SystemExit(f"per_apur invalido: {args.per_apur}")
    if not args.execute:
        manifest = generate_manifest(args.per_apur)
        print(json.dumps({"manifest": manifest, "validation": validate_manifest(manifest)}, ensure_ascii=False, indent=2, default=str))
        return 0
    expected_token = confirm_token(args.per_apur)
    if args.confirmar != expected_token:
        raise SystemExit(f"Para executar, use --confirmar {expected_token}")
    print(json.dumps(execute(args.per_apur), ensure_ascii=False, indent=2, default=str))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())