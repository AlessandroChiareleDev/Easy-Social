"""
Easy e-Social - Excel Parser
Developed By Xandao

Módulo para análise e extração de dados do arquivo DIRF.xlsx
"""
import pandas as pd
import openpyxl
from typing import Dict, List
import os


class DIRFExcelParser:
    def __init__(self, file_path: str):
        self.file_path = file_path
        self.workbook = None
        self.tables_info = {}

    def analyze_dirf(self) -> Dict:
        """Analisa o arquivo DIRF.xlsx e retorna informações das 6 tabelas"""
        try:
            # Ler arquivo com openpyxl para metadados
            self.workbook = openpyxl.load_workbook(self.file_path, data_only=True)

            expected_tables = [
                'ANALISE NATUREZA',
                'Dinamica',
                'Tabela Eventos GI',
                'Tabela EB'
            ]

            analysis = {
                'file_name': os.path.basename(self.file_path),
                'file_size': self._get_file_size(),
                'tables': [],
                'total_sheets': len(self.workbook.sheetnames),
                'analysis_date': pd.Timestamp.now().isoformat()
            }

            for table_name in expected_tables:
                if table_name in self.workbook.sheetnames:
                    table_info = self._extract_table_info(table_name)
                    analysis['tables'].append(table_info)

            return analysis

        except Exception as e:
            raise Exception(f"Erro ao analisar DIRF.xlsx: {str(e)}")

    def _extract_table_info(self, table_name: str) -> Dict:
        """Extrai informações de uma tabela específica"""
        df = pd.read_excel(self.file_path, sheet_name=table_name)

        column_letters = self._generate_column_letters(len(df.columns))

        return {
            'name': table_name,
            'sheet_name': table_name,
            'row_count': len(df),
            'column_count': len(df.columns),
            'columns': df.columns.tolist(),
            'column_letters': column_letters,
            'data_types': df.dtypes.astype(str).to_dict()
        }

    def extract_table_data(self, table_name: str, limit: int = None) -> pd.DataFrame:
        """Extrai dados de uma tabela específica"""
        df = pd.read_excel(self.file_path, sheet_name=table_name)
        return df.head(limit) if limit else df

    def _generate_column_letters(self, count: int) -> List[str]:
        """Gera referências de letras (A, B, C, ..., Z, AA, AB, etc.)"""
        letters = []
        for i in range(count):
            letter = ''
            num = i
            while num >= 0:
                letter = chr(65 + (num % 26)) + letter
                num = num // 26 - 1
            letters.append(letter)
        return letters

    def _get_file_size(self) -> int:
        """Retorna tamanho do arquivo em bytes"""
        return os.path.getsize(self.file_path)

    def validate_data(self, table_name: str) -> Dict:
        """Valida integridade dos dados de uma tabela"""
        df = pd.read_excel(self.file_path, sheet_name=table_name)

        validation = {
            'table_name': table_name,
            'total_rows': len(df),
            'total_columns': len(df.columns),
            'null_values': df.isnull().sum().to_dict(),
            'duplicate_rows': len(df[df.duplicated()]),
            'is_valid': True
        }

        # Verificar se há valores nulos críticos
        if df.isnull().sum().sum() > 0:
            validation['is_valid'] = False

        return validation
