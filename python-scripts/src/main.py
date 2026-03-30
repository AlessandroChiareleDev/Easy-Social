"""
Easy Social - Python Scripts
Developed By Xandao

Módulo de processamento de dados DIRF e web scraping para eSocial.
"""
from fastapi import FastAPI, Query
from fastapi.middleware.cors import CORSMiddleware
import uvicorn
import openpyxl
import json
import os

app = FastAPI(
    title="Easy Social - Python API",
    description="API para processamento de dados DIRF e web scraping eSocial",
    version="1.0.0",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:5174", "http://localhost:3333"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Diretório de uploads do backend
UPLOADS_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'backend', 'uploads'))


@app.get("/health")
async def health_check():
    return {"status": "ok", "system": "Easy Social - Python API", "version": "1.0.0"}


@app.get("/sheet/info")
async def get_sheet_info(file_path: str = Query(...), sheet_name: str = Query(...)):
    """Retorna informações de uma aba específica via openpyxl"""
    resolved = os.path.abspath(file_path)
    if not resolved.startswith(UPLOADS_DIR):
        return {"error": "Caminho não permitido"}, 400

    wb = openpyxl.load_workbook(resolved, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            return {"error": f"Aba '{sheet_name}' não encontrada"}

        ws = wb[sheet_name]
        row_count = ws.max_row or 0
        col_count = ws.max_column or 0

        # Ler cabeçalho (primeira linha)
        headers = []
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if first_row:
            headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(first_row)]

        # Gerar letras das colunas
        column_letters = []
        for i in range(col_count):
            letter = ''
            num = i
            while num >= 0:
                letter = chr(65 + (num % 26)) + letter
                num = num // 26 - 1
            column_letters.append(letter)

        return {
            "name": sheet_name,
            "sheetName": sheet_name,
            "rowCount": row_count,
            "columnCount": col_count,
            "columns": headers,
            "columnLetters": column_letters
        }
    finally:
        wb.close()


@app.get("/sheet/data")
async def get_sheet_data(
    file_path: str = Query(...),
    sheet_name: str = Query(...),
    limit: int = Query(default=0),
    offset: int = Query(default=0)
):
    """Extrai dados de uma aba como lista de objetos JSON (com paginação)"""
    resolved = os.path.abspath(file_path)
    if not resolved.startswith(UPLOADS_DIR):
        return {"error": "Caminho não permitido"}, 400

    wb = openpyxl.load_workbook(resolved, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name]

        rows_iter = ws.iter_rows(values_only=True)
        # Primeira linha = cabeçalho
        header_row = next(rows_iter, None)
        if not header_row:
            return {"data": [], "total": 0}

        headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_row)]

        # Pular linhas do offset
        skipped = 0
        while skipped < offset:
            if next(rows_iter, None) is None:
                return {"data": [], "total": ws.max_row - 1 if ws.max_row else 0}
            skipped += 1

        data = []
        count = 0
        for row in rows_iter:
            obj = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    # Converter datetime para string ISO
                    if hasattr(val, 'isoformat'):
                        obj[headers[i]] = val.isoformat()
                    else:
                        obj[headers[i]] = val if val is not None else None
            data.append(obj)
            count += 1
            if limit > 0 and count >= limit:
                break

        return {"data": data, "total": ws.max_row - 1 if ws.max_row else 0}
    finally:
        wb.close()


if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
